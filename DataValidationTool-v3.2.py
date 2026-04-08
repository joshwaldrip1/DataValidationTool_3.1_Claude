# Data Validation Tool v3.2  (Windows)
# - Multi-file drag & drop (CSV + FXL together)
# - Single Excel instance/workbook; single-run validation
# - Full VBA module included (no missing methods)
# - Restored "Validate Entire Sheet (Excel)" button
# - Safer COM startup (auto-cleans gen_py cache on failure)
# - FXL/META kept VeryHidden; Data is the only visible sheet

import os
import sys
import csv
import re
import json
import uuid
import atexit
import shutil
import struct
import hashlib
import tempfile
import platform
import datetime
import subprocess
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import xml.etree.ElementTree as ET
from typing import Any, cast, Sequence, Protocol, runtime_checkable
from urllib.request import Request, urlopen
from urllib.error import URLError

# ── App version & update config ──────────────────────────────────
APP_VERSION = "3.2"
_GITHUB_OWNER = "joshwaldrip1"
_GITHUB_REPO  = "DataValidationTool"
# Fine-grained PAT with read-only Contents access to the private repo.
# Generate at: GitHub → Settings → Developer settings → Fine-grained tokens
_GITHUB_TOKEN = "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"

# --- Minimal Protocols to type Excel COM objects (win32com) ---
@runtime_checkable
class RangeLike(Protocol):
    @property
    def Value(self) -> Any: ...


@runtime_checkable
class WorksheetLike(Protocol):
    @property
    def UsedRange(self) -> RangeLike: ...
    def Range(self, address: str) -> RangeLike: ...
    def Activate(self) -> Any: ...
    def Copy(self) -> Any: ...
    @property
    def Visible(self) -> Any: ...
    @Visible.setter
    def Visible(self, v: Any) -> None: ...
    @property
    def CodeName(self) -> str: ...


@runtime_checkable
class WorkbooksLike(Protocol):
    def Open(self, *, Filename: str) -> "WorkbookLike": ...
    @property
    def Count(self) -> int: ...


@runtime_checkable
class WorkbookLike(Protocol):
    @property
    def Name(self) -> str: ...
    def SaveCopyAs(self, path: str) -> Any: ...
    def SaveAs(self, Filename: str, FileFormat: int, *args: Any, **kwargs: Any) -> Any: ...
    def Close(self, SaveChanges: Any = ...) -> Any: ...
    def Worksheets(self, name: str) -> WorksheetLike: ...
    @property
    def VBProject(self) -> Any: ...


@runtime_checkable
class ExcelApplicationLike(Protocol):
    def Run(self, target: str) -> Any: ...
    @property
    def Workbooks(self) -> WorkbooksLike: ...
    @property
    def ActiveWorkbook(self) -> WorkbookLike | None: ...
    @property
    def DisplayAlerts(self) -> Any: ...
    @DisplayAlerts.setter
    def DisplayAlerts(self, v: Any) -> None: ...
    @property
    def EnableEvents(self) -> Any: ...
    @EnableEvents.setter
    def EnableEvents(self, v: Any) -> None: ...
    @property
    def Visible(self) -> Any: ...
    @Visible.setter
    def Visible(self, v: Any) -> None: ...
    @property
    def WindowState(self) -> Any: ...
    @WindowState.setter
    def WindowState(self, v: Any) -> None: ...
    def Quit(self) -> Any: ...

# ---- Tooltip helper ----
class ToolTip:
    """Simple hover tooltip for tkinter widgets."""
    def __init__(self, widget: tk.Widget, text: str) -> None:
        self._widget = widget
        self._text = text
        self._tip: tk.Toplevel | None = None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)

    def _show(self, _event: Any = None) -> None:
        if self._tip:
            return
        x = self._widget.winfo_rootx() + 20
        y = self._widget.winfo_rooty() + self._widget.winfo_height() + 4
        self._tip = tk.Toplevel(self._widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        lbl = tk.Label(
            self._tip, text=self._text, justify="left",
            background="#ffffe0", relief="solid", borderwidth=1,
            wraplength=280, padx=4, pady=2,
        )
        lbl.pack()

    def _hide(self, _event: Any = None) -> None:
        if self._tip:
            self._tip.destroy()
            self._tip = None


# ---- Optional deps ----
try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import win32com.client as win32  # Windows only
except Exception:
    win32 = None

# Drag & Drop (optional)
# Provide a typed default base class; replace with TkinterDnD.Tk if available.
TkBase: type[tk.Tk] = tk.Tk
tkdnd_available = False
_dnd_files: Any = None
try:
    # pip install tkinterdnd2
    from tkinterdnd2 import TkinterDnD, DND_FILES as _TK_DND_FILES  # type: ignore[reportMissingTypeStubs]
    TkBase = cast(type[tk.Tk], TkinterDnD.Tk)
    _dnd_files = _TK_DND_FILES
    tkdnd_available = True
except Exception:
    # Keep graceful fallback without redefining an UPPER_CASE constant name
    TkBase = tk.Tk
    _dnd_files = None
    tkdnd_available = False


class DataValidationTool(TkBase):
    def __init__(self):
        super().__init__()
        self.title(f"Data Validation Tool v{APP_VERSION}")
        _place_window(self, 820, 520)

        # state
        self.fxl_data: dict[str, list[dict[str, Any]]] = {}
        # Persist the first FXL seen to reuse for CSV-only drops
        self._initial_fxl_data: dict[str, list[dict[str, Any]]] | None = None
        self.df: pd.DataFrame | None = None
        self.csv_path: str | None = None
        self.fxl_path: str | None = None
        self._initial_fxl_path: str | None = None
        self._initial_csv_path: str | None = None

        self.has_station: bool | None = None
        self.mapping: dict[str, int | None] | None = None
        self.attr_indices: list[int] = []

        # Optional MTR spreadsheet state
        self.mtr_df: pd.DataFrame | None = None
        self.mtr_path: str | None = None

        # temp workspace (Excel files + sentinel)
        self._tmpdir = tempfile.mkdtemp(prefix="dvt_")
        atexit.register(shutil.rmtree, self._tmpdir, True)  # ensure cleanup on crash
        self._temp_xlsx = None
        self._temp_xlsm = None
        self._sentinel = os.path.join(self._tmpdir, "saveflag.txt")

        # COM handles / flags
        self._excel: ExcelApplicationLike | None = None
        self._wb_com: WorkbookLike | None = None
        self._excel_opened = False  # guard double-open
        # All open workbooks this session: list of (excel, wb_com, csv_path)
        self._all_workbooks: list[tuple[Any, Any, str]] = []
        # Capture last COM error text for diagnostics in UI
        self._last_com_error: str | None = None

        # Behavior flags (can be overridden via config.json)
        # Default to a single Excel instance to avoid DispatchEx issues
        self.single_excel_instance: bool = True
        self._pending_fxl_path: str | None = None      # auto-load from last_fxl_path in config
        self.numeric_bounds: dict[str, list[float]] = {}  # from config numeric_bounds section
        self._error_filter_on: bool = False             # toggle state for Show Errors Only button
        self.fxl_library_path: str = r"S:\TOPOGRAPHIC DATA\TOPOGRAPHIC STANDARDS\DATA DICTIONARY - GEOID FILES"
        self._fxl_library_cache: list[str] | None = None  # lazily populated on first use
        self.jxl_path: str | None = None               # companion .jxl path when detected
        self._jxl_data: dict[str, Any] | None = None   # parsed JXL data (cached per session)
        self._jxl_parse_cache: dict[str, dict[str, Any]] = {}  # {abs_path: parsed result}
        self._media_index_cache: dict[str, dict[str, str]] = {}  # {sync_root: {basename_lower: full_path}}
        self.crdb_path: str | None = None              # loaded Carlson .crdb path
        try:
            self._load_app_config()
        except Exception:
            pass

        # Patch standard messagebox functions so they always use this window as parent
        # and briefly flash to the front on Windows (fixes dialogs opening behind other apps).
        _app = self
        _orig_info     = messagebox.showinfo
        _orig_warn     = messagebox.showwarning
        _orig_error    = messagebox.showerror
        _orig_yesno    = messagebox.askyesno
        _orig_okcancel = messagebox.askokcancel

        def _mb_lift() -> None:
            try:
                _app.lift()  # type: ignore[reportUnknownMemberType]
                _app.attributes("-topmost", True)   # type: ignore[reportUnknownMemberType]
                _app.update_idletasks()
                _app.attributes("-topmost", False)  # type: ignore[reportUnknownMemberType]
            except Exception:
                pass

        def _info(title: Any = None, message: Any = None, **kw: Any) -> Any:
            kw.setdefault("parent", _app); _mb_lift(); return _orig_info(title, message, **kw)
        def _warn(title: Any = None, message: Any = None, **kw: Any) -> Any:
            kw.setdefault("parent", _app); _mb_lift(); return _orig_warn(title, message, **kw)
        def _error(title: Any = None, message: Any = None, **kw: Any) -> Any:
            kw.setdefault("parent", _app); _mb_lift(); return _orig_error(title, message, **kw)
        def _yesno(title: Any = None, message: Any = None, **kw: Any) -> Any:
            kw.setdefault("parent", _app); _mb_lift(); return _orig_yesno(title, message, **kw)
        def _okcancel(title: Any = None, message: Any = None, **kw: Any) -> Any:
            kw.setdefault("parent", _app); _mb_lift(); return _orig_okcancel(title, message, **kw)

        messagebox.showinfo    = _info      # type: ignore[method-assign]
        messagebox.showwarning = _warn      # type: ignore[method-assign]
        messagebox.showerror   = _error     # type: ignore[method-assign]
        messagebox.askyesno    = _yesno     # type: ignore[method-assign]
        messagebox.askokcancel = _okcancel  # type: ignore[method-assign]

        # Deferred startup: show any pending watch-list notifications
        self.after(800, self._check_pending_notifications)  # type: ignore[attr-defined]
        # Deferred startup: check watched CRDBs for changes (catches missed scheduled runs)
        self.after(2000, self._startup_watchlist_check)  # type: ignore[attr-defined]

        # UI
        self._build_ui()
        self._wire_dnd()

        # graceful shutdown
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---------- UI ----------
    def _build_ui(self):
        # Menu bar
        menubar = tk.Menu(self)
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Check for Updates…", command=self._check_for_updates)
        help_menu.add_separator()
        help_menu.add_command(label=f"About Data Validation Tool v{APP_VERSION}", command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)
        self.config(menu=menubar)

        info = tk.Label(
            self,
            text=("Drag & drop files to get started. Supported: CSV, FXL, JXL, CRDB.\n"
                  "Each file type opens an action dialog with available options."),
            justify="left",
        )
        info.pack(pady=8)

        btns = tk.Frame(self)
        btns.pack(pady=6)

        self.btn_save_all = tk.Button(
            btns, text="Save All Outputs", width=20,
            state="disabled", command=self.export_all_outputs,
            bg="#1F4E79", fg="white"
        )
        self.btn_save_all.grid(row=0, column=0, padx=8)
        ToolTip(self.btn_save_all,
                "Saves 3 files to the original CSV/JXL folder:\n"
                "  • _Error_Report.xlsm — report as-is\n"
                "  • _corrected.csv     — Data sheet, no headers, NEZ at 10 decimal places\n"
                "  • _GNSS.csv          — Geodetic Info sheet (header rows stripped)")

        self.btn_email = tk.Button(
            btns, text="Email Report…", width=20,
            state="disabled", command=self.email_report
        )
        self.btn_email.grid(row=0, column=1, padx=8)
        ToolTip(self.btn_email, "Validate the sheet first, then use this to email the error report via Outlook.")

        self.btn_missing_heats = tk.Button(
            btns, text="Email Missing Heats", width=22,
            state="disabled", command=self.email_missing_heats
        )
        self.btn_missing_heats.grid(row=0, column=2, padx=8)
        ToolTip(self.btn_missing_heats, "Load an MTR spreadsheet (drag & drop an Excel file with 'MTR' in the name) to enable.")

        self.btn_watchlist = tk.Button(
            btns, text="Watch List…", width=14,
            command=self._show_watchlist_dialog  # type: ignore[attr-defined]
        )
        self.btn_watchlist.grid(row=0, column=3, padx=8)
        ToolTip(self.btn_watchlist,
                "View and manage the list of CRDB files being monitored for changes. "
                "Entries are checked daily at 2 AM and removed after 3 months of no new uploads.")

        self.btn_schemas = tk.Button(
            btns, text="Client Schemas…", width=14,
            command=self._show_schema_manager  # type: ignore[attr-defined]
        )
        self.btn_schemas.grid(row=0, column=4, padx=8)
        ToolTip(self.btn_schemas,
                "Manage client GDB schemas. Import a .gdb file to capture a client's\n"
                "field structure, then select the schema when exporting CRDB data.\n"
                "You can also drag & drop a .gdb folder directly onto the app.")

        self.status = tk.Label(self, text="Ready. Drop a file to begin.", anchor="w", relief="groove")
        self.status.pack(fill="x", pady=6)

        # Indeterminate progress bar — hidden until an operation is running
        self.progress = ttk.Progressbar(self, mode="indeterminate", length=300)
        self.progress.pack(pady=2)
        self.progress.pack_forget()

        self.dropbox = tk.Text(self, height=8, wrap="word")
        self.dropbox.insert("1.0",
                            "Drop files here to begin:\n"
                            "  • CSV — Validate or Generate GNSS Report\n"
                            "  • JXL — Validate, Rename Photos, or Generate GNSS Report\n"
                            "  • CRDB — Export Data (GPKG, CSV, Shapefile, LandXML, KMZ, GNSS Report)\n"
                            "  • FXL — Load field definitions (required for CSV/JXL validation)")
        self.dropbox.config(state="disabled")
        self.dropbox.pack(fill="both", expand=True, padx=8, pady=8)

        # Auto-load last FXL after the UI event loop starts (non-blocking)
        _pfxl = self._pending_fxl_path
        if _pfxl:
            self.after(200, lambda p=_pfxl: self._load_fxl_path(p, silent=True))

    def _wire_dnd(self):
        if not tkdnd_available:
            self.status.config(text="Drag & drop not available (install 'tkinterdnd2').")
            return
        self.drop_target_register(_dnd_files)  # type: ignore[arg-type]
        self.dnd_bind("<<Drop>>", self._on_drop)  # type: ignore[misc]
        self.dropbox.drop_target_register(_dnd_files)  # type: ignore[arg-type]
        self.dropbox.dnd_bind("<<Drop>>", self._on_drop)  # type: ignore[misc]

    # ---------- Update checker ----------
    def _show_about(self) -> None:
        messagebox.showinfo(
            "About",
            f"Data Validation Tool v{APP_VERSION}\n"
            f"Topographic Land Surveyors\n\n"
            f"Python {platform.python_version()}\n"
            f"{platform.system()} {platform.release()}",
        )

    def _check_for_updates(self) -> None:
        """Check GitHub Releases for a newer version (runs in a background thread)."""
        if not _GITHUB_TOKEN:
            messagebox.showwarning(
                "Update Check",
                "No GitHub token configured.\n\n"
                "A personal access token is required to check for updates "
                "from the private repository. Contact your administrator.",
            )
            return
        self.status.config(text="Checking for updates…")
        self.update_idletasks()
        threading.Thread(target=self._do_update_check, daemon=True).start()

    def _do_update_check(self) -> None:
        """Background thread: query GitHub Releases API and compare versions."""
        try:
            url = f"https://api.github.com/repos/{_GITHUB_OWNER}/{_GITHUB_REPO}/releases/latest"
            req = Request(url)
            req.add_header("Authorization", f"Bearer {_GITHUB_TOKEN}")
            req.add_header("Accept", "application/vnd.github+json")
            with urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))

            tag: str = str(data.get("tag_name", ""))
            # Strip leading 'v' from tags like "v3.3"
            remote_ver = tag.lstrip("vV").strip()
            if not remote_ver:
                self.after(0, lambda: self._update_result("error", "Could not read version from latest release."))
                return

            if self._version_newer(remote_ver, APP_VERSION):
                # Find the setup EXE asset
                download_url: str = ""
                asset_name: str = ""
                for asset in data.get("assets", []):
                    name: str = str(asset.get("name", ""))
                    if name.lower().endswith(".exe") and "setup" in name.lower():
                        # Private repos need the API URL, not browser_download_url
                        download_url = str(asset.get("url", ""))
                        asset_name = name
                        break
                self.after(0, lambda: self._update_result(
                    "available", f"Version {remote_ver} is available (you have {APP_VERSION}).",
                    download_url, asset_name,
                ))
            else:
                self.after(0, lambda: self._update_result("current", f"You are running the latest version ({APP_VERSION})."))

        except URLError as e:
            self.after(0, lambda: self._update_result("error", f"Could not reach GitHub:\n{e.reason}"))
        except Exception as e:
            self.after(0, lambda: self._update_result("error", f"Update check failed:\n{e}"))

    @staticmethod
    def _version_newer(remote: str, local: str) -> bool:
        """Return True if remote version is strictly newer than local."""
        def _parts(v: str) -> list[int]:
            return [int(x) for x in re.findall(r"\d+", v)]
        return _parts(remote) > _parts(local)

    def _update_result(self, status: str, message: str,
                       download_url: str = "", asset_name: str = "") -> None:
        """Handle update check result on the main thread."""
        self.status.config(text="Ready")
        if status == "available" and download_url:
            if messagebox.askyesno("Update Available",
                                   f"{message}\n\nDownload and install now?"):
                self._download_and_install(download_url, asset_name)
        elif status == "available":
            messagebox.showinfo("Update Available",
                                f"{message}\n\nNo installer asset found in the release.\n"
                                "Please download manually from GitHub.")
        elif status == "current":
            messagebox.showinfo("Up to Date", message)
        else:
            messagebox.showwarning("Update Check", message)

    def _download_and_install(self, url: str, asset_name: str) -> None:
        """Download the setup EXE from GitHub and launch it."""
        self.status.config(text=f"Downloading {asset_name}…")
        self.progress.pack(pady=2)
        self.progress.start(15)
        self.update_idletasks()
        threading.Thread(
            target=self._do_download, args=(url, asset_name), daemon=True,
        ).start()

    def _do_download(self, url: str, asset_name: str) -> None:
        """Background thread: download asset and launch installer."""
        try:
            req = Request(url)
            req.add_header("Authorization", f"Bearer {_GITHUB_TOKEN}")
            req.add_header("Accept", "application/octet-stream")
            dest = os.path.join(tempfile.gettempdir(), asset_name)
            with urlopen(req, timeout=120) as resp:
                with open(dest, "wb") as f:
                    while True:
                        chunk = resp.read(65536)
                        if not chunk:
                            break
                        f.write(chunk)
            self.after(0, lambda: self._download_complete(dest))
        except Exception as e:
            self.after(0, lambda: self._download_failed(str(e)))

    def _download_complete(self, installer_path: str) -> None:
        self.progress.stop()
        self.progress.pack_forget()
        self.status.config(text="Download complete.")
        if messagebox.askyesno("Download Complete",
                               f"Installer saved to:\n{installer_path}\n\n"
                               "Launch the installer now?\n"
                               "(The app will close.)"):
            try:
                os.startfile(installer_path)  # type: ignore[attr-defined]
                self._on_close()
            except Exception as e:
                messagebox.showerror("Launch Error", f"Could not launch installer:\n{e}")

    def _download_failed(self, error: str) -> None:
        self.progress.stop()
        self.progress.pack_forget()
        self.status.config(text="Ready")
        messagebox.showerror("Download Failed", f"Could not download the update:\n{error}")

    def _load_app_config(self) -> None:
        # Resolve config.json relative to the app/exe, not the working directory
        _base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
        cfg_path = os.path.join(_base, "config.json")
        if not os.path.isfile(cfg_path):
            return
        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg_any: Any = json.load(f)
            if isinstance(cfg_any, dict):
                cfg = cast(dict[str, Any], cfg_any)
                raw: Any = cfg.get("single_excel_instance", None)
                if isinstance(raw, bool):
                    self.single_excel_instance = raw
                elif isinstance(raw, str):
                    self.single_excel_instance = raw.strip().lower() in {"1", "true", "yes", "y", "on"}
                elif raw is not None:
                    self.single_excel_instance = bool(raw)
                # Org-wide FXL library root for fallback auto-detection
                raw_lib: Any = cfg.get("fxl_library_path", "")
                if isinstance(raw_lib, str) and raw_lib.strip():
                    self.fxl_library_path = raw_lib.strip()
                # Remember last FXL path across sessions
                raw_fxl: Any = cfg.get("last_fxl_path", None)
                if isinstance(raw_fxl, str) and raw_fxl.strip() and os.path.isfile(raw_fxl.strip()):
                    self._pending_fxl_path = raw_fxl.strip()
                # Numeric bounds for attribute range checking
                raw_bounds: Any = cfg.get("numeric_bounds", {})
                if isinstance(raw_bounds, dict):
                    typed_bounds = cast(dict[str, Any], raw_bounds)
                    self.numeric_bounds = {
                        k.strip().upper(): cast(list[float], v)
                        for k, v in typed_bounds.items()
                        if isinstance(v, list) and len(v) == 2  # type: ignore[arg-type]
                    }
        except Exception:
            pass

    def _save_app_config(self, updates: dict[str, Any]) -> None:
        """Persist key/value pairs to config.json next to the app/exe."""
        _base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
        cfg_path = os.path.join(_base, "config.json")
        cfg: dict[str, Any] = {}
        try:
            if os.path.isfile(cfg_path):
                with open(cfg_path, "r", encoding="utf-8") as f:
                    cfg = cast(dict[str, Any], json.load(f))
        except Exception:
            pass
        cfg.update(updates)
        try:
            with open(cfg_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
        except Exception:
            pass

    # ---------- Drag & drop ----------
    def _on_drop(self, event: Any) -> None:
        """Accept multiple files; load CSV/FXL without prompting if both were dropped."""
        raw = event.data
        if not raw:
            return

        # Accept single/multiple files; support {...} paths with spaces
        paths: list[str] = []
        for m in re.finditer(r"\{[^}]+\}|[^\s]+", raw):
            p = m.group(0)
            if p.startswith("{") and p.endswith("}"):
                p = p[1:-1]
            paths.append(p)

        csvs: list[str] = [p for p in paths if os.path.splitext(p)[1].lower() == ".csv"]
        fxls: list[str] = [p for p in paths if os.path.splitext(p)[1].lower() in {".fxl", ".xml"}]
        crdbs: list[str] = [p for p in paths if os.path.splitext(p)[1].lower() == ".crdb"]
        jxls: list[str] = [p for p in paths if os.path.splitext(p)[1].lower() == ".jxl"]
        # Any Excel whose filename contains 'MTR' is treated as the MTR spreadsheet
        mtrs: list[str] = [
            p for p in paths
            if os.path.splitext(p)[1].lower() in {".xlsx", ".xlsm", ".xls"}
            and ("mtr" in os.path.basename(p).lower())
        ]

        # GDB folders → import client schema
        gdbs: list[str] = [p for p in paths if p.lower().endswith(".gdb") and os.path.isdir(p)]
        if gdbs:
            for gdb_path in gdbs:
                self._show_gdb_import_dialog(gdb_path)  # type: ignore[attr-defined]
            return

        # CRDB files → go straight to export (includes GNSS report)
        if crdbs:
            self._show_crdb_action_dialog(crdbs)  # type: ignore[attr-defined]
            return

        # JXL-only drop → ask what to do (rename photos or GNSS report)
        if jxls and not csvs:
            self._show_jxl_action_dialog(jxls)  # type: ignore[attr-defined]
            return

        # CSV drop → ask what to do before running anything
        if csvs:
            csv_action = self._show_csv_action_dialog(len(csvs))  # type: ignore[attr-defined]
            if not csv_action:
                return
            if csv_action == "gnss":
                jxl_by_stem_quick: dict[str, str] = {
                    os.path.splitext(os.path.basename(p))[0].lower(): p for p in jxls
                }
                jxl_paths_for_report: list[str] = []
                for csv_p in csvs:
                    csv_stem = os.path.splitext(os.path.basename(csv_p))[0].lower()
                    jxl_p = jxl_by_stem_quick.get(csv_stem) or self._find_jxl_alongside_csv(csv_p)  # type: ignore[attr-defined]
                    if jxl_p and jxl_p not in jxl_paths_for_report:
                        jxl_paths_for_report.append(jxl_p)
                if not jxl_paths_for_report:
                    messagebox.showwarning(
                        "No JXL Found",
                        "No JXL file was found alongside the dropped CSV(s).\n"
                        "Cannot generate a GNSS Report without a JXL.",
                    )
                    return
                self._generate_gnss_report(jxl_paths_for_report)  # type: ignore[attr-defined]
                return
            # else: csv_action == "validate" — fall through to current validation flow

        # Determine which FXL to use for this drop
        drop_fxl: str | None = None
        if fxls:
            drop_fxl = fxls[0]

        # If we have CSV files, process each using the chosen FXL (if available)
        if csvs:
            if drop_fxl is None:
                # Auto-detect: JXL companion → local folder → parent → library → previous FXL
                if not self._ensure_fxl_after_csv(csvs[0]):
                    return
                drop_fxl = self.fxl_path
            if drop_fxl is None:
                # Still no FXL resolved; abort gracefully
                return
            # If an MTR file was included, load and keep it for this session
            if mtrs:
                try:
                    self._load_mtr_path(mtrs[0])
                except Exception:
                    pass
            # Build stem→path map for any JXL files dropped alongside the CSVs so that
            # cross-folder pairs (CSV in WORKING, JXL in SYNC) are matched correctly.
            jxl_by_stem: dict[str, str] = {
                os.path.splitext(os.path.basename(p))[0].lower(): p
                for p in jxls
            }
            # Process each CSV with the selected FXL, opening a new Excel instance each time
            batch_results: list[tuple[str, str]] = []
            for csv_p in csvs:
                fname = os.path.basename(csv_p)
                csv_stem = os.path.splitext(fname)[0].lower()
                jxl_override = jxl_by_stem.get(csv_stem)
                try:
                    self._process_pair(csv_p, drop_fxl, jxl_override=jxl_override)  # type: ignore[call-arg]
                    batch_results.append((fname, "OK — opened in Excel"))
                except Exception as e:
                    self.status.config(text=f"Error processing {fname}: {e}")
                    batch_results.append((fname, f"Error: {e}"))
            if len(batch_results) > 1:
                self._show_batch_summary(batch_results)
        else:
            # No CSVs in this drop; if an FXL was dropped, load it as the (new) available FXL
            if drop_fxl:
                try:
                    self._load_fxl_path(drop_fxl)
                except Exception as e:
                    self.status.config(text=f"FXL load failed: {e}")
            # If only an MTR was dropped, load it now and keep for later
            if (not csvs) and mtrs:
                try:
                    self._load_mtr_path(mtrs[0])
                except Exception as e:
                    self.status.config(text=f"MTR load failed: {e}")

    # ---------- Public actions ----------
    def validate_in_excel(self):
        """Run the full-sheet validator inside the already-open Excel workbook."""
        try:
            if not (self._excel and self._wb_com):
                messagebox.showwarning("Excel not ready", "Open a CSV (and FXL) first so Excel can launch.")
                return
            wbname = self._wb_com.Name
            self._progress_start("Running full validation…")
            try:
                for target in [
                    f"'{wbname}'!ValidationModule.ValidateSheetAll",
                    f"'{wbname}'!ValidateSheetAll",
                    "ValidationModule.ValidateSheetAll",
                    "ValidateSheetAll",
                ]:
                    try:
                        self._excel.Run(target)
                        break
                    except Exception:
                        pass
            finally:
                self._progress_stop()
                self.status.config(text="Validation complete.")
        except Exception as e:
            self._progress_stop()
            messagebox.showerror("Error", f"Failed to run validation in Excel:\n{e}")

    # Sheets stripped from the email copy — keep only Data + Error Count for recipients
    _EMAIL_EXCLUDE_SHEETS: frozenset[str] = frozenset({"Geodetic Info"})

    def _save_wb_for_email(self, wb_com: Any, csv_path: str) -> str | None:
        """Save an email-ready copy of wb_com with excluded sheets removed.

        The copy goes to the temp directory so it never collides with the workbook's
        own save path (which is now in the source folder as *_Error_Report.xlsm).
        """
        csv_name = os.path.splitext(os.path.basename(csv_path or "Data"))[0]
        out_path = os.path.join(self._tmpdir, f"{csv_name}_Error_Report_email.xlsm")
        try:
            try:
                if os.path.exists(out_path):
                    os.remove(out_path)
            except Exception:
                pass
            wb_com.SaveCopyAs(out_path)
        except Exception:
            try:
                wb_com.SaveAs(out_path, 52)
            except Exception:
                return None

        # Open the saved copy in a hidden background instance and strip excluded sheets
        if win32 is not None:
            xl_bg: Any = None
            try:
                xl_bg = win32.DispatchEx("Excel.Application")  # type: ignore[union-attr]
                xl_bg.Visible = False  # type: ignore[union-attr]
                xl_bg.DisplayAlerts = False  # type: ignore[union-attr]
                wb_copy: Any = xl_bg.Workbooks.Open(out_path)  # type: ignore[union-attr]
                for sheet_name in self._EMAIL_EXCLUDE_SHEETS:
                    try:
                        wb_copy.Worksheets(sheet_name).Delete()  # type: ignore[union-attr]
                    except Exception:
                        pass
                wb_copy.Save()  # type: ignore[union-attr]
                wb_copy.Close(False)  # type: ignore[union-attr]
            except Exception:
                pass  # leave the copy intact if stripping fails
            finally:
                try:
                    if xl_bg is not None:
                        xl_bg.Quit()  # type: ignore[union-attr]
                except Exception:
                    pass

        return out_path

    def _compose_email(self, attach_paths: list[str], subject: str) -> None:
        """Open a new Outlook email with the given attachments and subject."""
        if win32 is None:
            messagebox.showinfo("Not available", "Email compose requires Outlook/Windows (win32com).")
            return
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        try:
            mail.BodyFormat = 2  # olFormatHTML
        except Exception:
            pass
        mail.Subject = subject
        for p in attach_paths:
            if p and os.path.exists(p):
                mail.Attachments.Add(p)
        mail.Display()

    def email_report(self):
        """Create a blank email with validation report(s) attached."""
        try:
            # Prune stale COM handles — a closed workbook raises when accessed
            live: list[tuple[Any, Any, str]] = []
            for exc, wb, cp in self._all_workbooks:
                try:
                    _ = wb.Name  # probe: raises if workbook was closed
                    live.append((exc, wb, cp))
                except Exception:
                    pass
            self._all_workbooks = live

            if not live:
                messagebox.showwarning("Excel not ready", "Open/validate a file first so there's something to email.")
                return

            if len(live) == 1:
                # Single workbook — original behaviour
                _, wb, cp = live[0]
                attach_path = self._save_wb_for_email(wb, cp)
                if not attach_path:
                    messagebox.showwarning("Save failed", "Could not save the report file.")
                    return
                csv_name = os.path.splitext(os.path.basename(cp or "Data"))[0]
                self._compose_email([attach_path], f"{csv_name}_Error_Report")
                return

            # Multiple workbooks — let the user pick which to email
            self._show_email_picker(live)

        except Exception as e:
            messagebox.showerror("Email error", f"Couldn't prepare the email:\n{e}")

    def _show_email_picker(self, live: list[tuple[Any, Any, str]]) -> None:
        """Dialog: checkboxes for each open report; emails all selected ones in one message."""
        dlg = tk.Toplevel(self)
        dlg.title("Email Reports")
        dlg.resizable(False, False)
        dlg.grab_set()
        _raise_window(dlg)

        tk.Label(dlg, text="Select reports to attach to one email:", font=("", 10, "bold")).pack(anchor="w", padx=12, pady=(10, 4))

        vars_: list[tk.BooleanVar] = []
        for _, _, cp in live:
            var = tk.BooleanVar(value=True)
            vars_.append(var)
            name = os.path.basename(cp) if cp else "Unknown"
            tk.Checkbutton(dlg, text=name, variable=var, anchor="w").pack(fill="x", padx=20, pady=1)

        def _send():
            selected = [(wb, cp) for (_, wb, cp), var in zip(live, vars_) if var.get()]
            if not selected:
                messagebox.showwarning("Nothing selected", "Check at least one report to email.", parent=dlg)
                return
            dlg.destroy()
            attach_paths: list[str] = []
            failed: list[str] = []
            for wb, cp in selected:
                p = self._save_wb_for_email(wb, cp)
                if p:
                    attach_paths.append(p)
                else:
                    failed.append(os.path.basename(cp or "?"))
            if failed:
                messagebox.showwarning("Save failed", f"Could not save:\n" + "\n".join(failed))
            if attach_paths:
                names = ", ".join(os.path.splitext(os.path.basename(p))[0] for p in attach_paths)
                self._compose_email(attach_paths, names)

        btn_frame = tk.Frame(dlg)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Email Selected", width=18, command=_send).grid(row=0, column=0, padx=6)
        tk.Button(btn_frame, text="Cancel", width=10, command=dlg.destroy).grid(row=0, column=1, padx=6)

    def _norm_key(self, s: str) -> str:
        return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())

    def _fxl_attrs_for_fc(self, fc: str) -> list[dict[str, Any]] | None:
        if not self.fxl_data:
            return None
        key = (fc or "").strip()
        if key in self.fxl_data:
            return self.fxl_data[key]
        u = key.upper()
        for k in self.fxl_data.keys():
            if k.strip().upper() == u:
                return self.fxl_data[k]
        return None

    def _py_heat_attr_index_for_fc(self, fc: str) -> int | None:
        """Return 1-based attribute index for Heat in the given Field Code using FXL data."""
        attrs = self._fxl_attrs_for_fc(fc)
        if not attrs:
            return None
        aliases = {
            "HEAT", "HEATNUMBER", "HEATNO", "HEAT#", "HEATNUM", "HEATID", "HEATIDENT", "HEATIDENTIFICATION"
        }
        for i, a in enumerate(attrs, start=1):
            nm = (a.get("name") or "").strip()
            if self._norm_key(nm) in aliases:
                return i
        for i, a in enumerate(attrs, start=1):
            nm = (a.get("name") or "").strip()
            u = self._norm_key(nm)
            if "HEAT" in u and ("BACK" not in u and "BEHIND" not in u):
                return i
        return None

    def _collect_csv_heats(self) -> set[str]:
        used: set[str] = set()
        try:
            if self.df is None or self.df.empty or not self.mapping or not self.attr_indices:
                return used
            fc_idx = self.mapping.get("fc")
            if fc_idx is None:
                return used
            for row in self.df.itertuples(index=False, name=None):  # type: ignore[reportOptionalMemberAccess]
                try:
                    fc = str(row[fc_idx] or "").strip()
                except Exception:
                    fc = ""
                if not fc:
                    continue
                h_idx1 = self._py_heat_attr_index_for_fc(fc)
                if not h_idx1 or h_idx1 < 1 or h_idx1 > len(self.attr_indices):
                    continue
                col = self.attr_indices[h_idx1 - 1]
                try:
                    hv = ("" if row[col] is None else str(row[col])).strip().upper()
                except Exception:
                    hv = ""
                if hv:
                    used.add(hv)
        except Exception:
            pass
        return used

    def email_missing_heats(self) -> None:
        try:
            if self.mtr_df is None or self.mtr_df.empty:
                messagebox.showwarning("MTR required", "Load an MTR spreadsheet first to compare Heat numbers.")
                return
            used_heats = self._collect_csv_heats()
            mtr_heats: set[str] = set(
                ("" if pd.isna(x) else str(x)).strip().upper() for x in self.mtr_df.get("HEAT", [])  # type: ignore[reportUnknownArgumentType]
            )
            mtr_heats.discard("")
            missing = sorted([h for h in used_heats if h not in mtr_heats], key=lambda s: (len(s), s))
            if not missing:
                messagebox.showinfo("Missing Heats", "No missing Heat numbers found. All used heats are present in the MTR.")
                return

            if win32 is None:
                messagebox.showinfo("Not available", "Email compose requires Outlook/Windows (win32com).")
                return

            csv_name = os.path.splitext(os.path.basename(self.csv_path or "Data"))[0]
            mtr_name = os.path.basename(self.mtr_path or "MTR")
            # Compose HTML body directly (replaces prior plain-text body)

            outlook = win32.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            # Use HTML format and set an HTML body
            try:
                mail.BodyFormat = 2  # olFormatHTML
            except Exception:
                pass
            mail.Subject = "Missing Heats"
            # Build a simple HTML body from our values
            html_lines: list[str] = []
            html_lines.append(f"<p><b>CSV:</b> {csv_name}<br><b>MTR:</b> {mtr_name}</p>")
            html_lines.append(f"<p><b>Missing Heat Numbers</b> (count {len(missing)}):</p>")
            html_lines.append("<ul>")
            for h in missing:
                html_lines.append(f"<li>{h}</li>")
            html_lines.append("</ul>")
            mail.HTMLBody = "".join(html_lines)
            mail.Display()
        except Exception as e:
            messagebox.showerror("Email error", f"Couldn't prepare the Missing Heats email:\n{e}")

    def export_corrected_csv(self, show_popup: bool = True) -> tuple[bool, str | None, str | None]:
        """Export the current 'Data' sheet to a CSV named '<csv>_corrected.csv' in the CSV's folder.
        Returns (ok, path, error). If show_popup is False, suppress popups."""
        try:
            if not (self._excel and self._wb_com):
                if show_popup:
                    messagebox.showwarning("Excel not ready", "Open/validate a file first so there is data to export.")
                return False, None, "excel_not_ready"
            csv_dir = os.path.dirname(self.csv_path) if self.csv_path else os.getcwd()
            csv_name = os.path.splitext(os.path.basename(self.csv_path or "data"))[0]
            out_path = os.path.join(csv_dir, f"{csv_name}_corrected.csv")
            try:
                data_ws: WorksheetLike = self._wb_com.Worksheets("Data")
            except Exception as e_ws:
                if show_popup:
                    messagebox.showerror("Export error", f"Could not access 'Data' sheet in Excel:\n{e_ws}")
                return False, None, str(e_ws)
            try:
                if os.path.exists(out_path):
                    os.remove(out_path)
            except Exception:
                pass
            new_wb: WorkbookLike | None = None
            try:
                # Prevent sheet events/macros from firing in the temporary workbook copy
                try:
                    self._excel.EnableEvents = False
                except Exception:
                    pass
                data_ws.Copy()
                new_wb = self._excel.ActiveWorkbook
                try:
                    self._excel.DisplayAlerts = False
                except Exception:
                    pass
                # Primary path: use Excel SaveAs to CSV
                if new_wb is None:
                    raise RuntimeError("Workbook copy was not created")
                new_wb.SaveAs(out_path, 6)  # 6 = xlCSV
                try:
                    self._excel.DisplayAlerts = True
                except Exception:
                    pass
                new_wb.Close(False)
                try:
                    self._excel.EnableEvents = True
                except Exception:
                    pass
                if show_popup:
                    messagebox.showinfo("Exported", f"Saved corrected CSV to:\n{out_path}")
                return True, out_path, None
            except Exception as e:
                # Fallback: close temp workbook and export via Python CSV writer from live Data sheet values
                try:
                    try:
                        self._excel.DisplayAlerts = True
                    except Exception:
                        pass
                    try:
                        if new_wb is not None:
                            new_wb.Close(False)
                    except Exception:
                        pass
                    try:
                        self._excel.EnableEvents = True
                    except Exception:
                        pass
                except Exception:
                    pass
                try:
                    ws_data: WorksheetLike = self._wb_com.Worksheets("Data")
                    used: RangeLike = ws_data.UsedRange
                    values: Any = used.Value
                    # Normalize values to a 2D iterable
                    rows_list: list[list[Any]]
                    if values is None:
                        rows_list = []
                    elif isinstance(values, (list, tuple)):
                        values_seq = cast(Sequence[Any], values)
                        if len(values_seq) > 0 and isinstance(values_seq[0], (list, tuple)):
                            rows_list = [list(inner) for inner in values_seq]  # type: ignore[reportUnknownArgumentType]
                        else:
                            rows_list = [list(values_seq)]
                    else:
                        rows_list = [[values]]
                    os.makedirs(os.path.dirname(out_path), exist_ok=True)
                    # Write as legacy Windows CSV (ANSI/Windows-1252) to match
                    # Excel's "CSV (Comma delimited)" format.
                    with open(out_path, "w", newline="", encoding="cp1252", errors="replace") as f:
                        writer = csv.writer(f)
                        for r in rows_list:
                            writer.writerow(["" if v is None else v for v in r])
                    if show_popup:
                        messagebox.showinfo(
                            "Exported",
                            f"Saved corrected CSV to (fallback):\n{out_path}\n\nNote: Used ANSI (Windows-1252) CSV to match Excel's 'CSV (Comma delimited)'."
                        )
                    return True, out_path, None
                except Exception as e2:
                    if show_popup:
                        messagebox.showerror(
                            "Export error",
                            f"Failed to export CSV:\nExcel SaveAs error: {e}\nFallback writer error: {e2}"
                        )
                    return False, None, f"Excel SaveAs error: {e}; Fallback writer error: {e2}"
        except Exception as e_outer:
            if show_popup:
                messagebox.showerror("Export error", str(e_outer))
            return False, None, str(e_outer)

    def export_error_report(self, show_popup: bool = True) -> tuple[bool, str | None, str | None]:
        """Save the current workbook as an Error Report xlsm in the initial CSV folder.
        Returns (ok, path, error). If show_popup is False, suppress popups."""
        try:
            if not (self._excel and self._wb_com):
                if show_popup:
                    messagebox.showwarning("Excel not ready", "Open/validate a file first so there is a workbook to export.")
                return False, None, "excel_not_ready"
            # Prefer the workbook's META sheet (A1=dir, A2=corrected name) for the active workbook
            csv_dir = os.path.dirname(self.csv_path) if self.csv_path else os.getcwd()
            csv_name = os.path.splitext(os.path.basename(self.csv_path or "Data"))[0]
            try:
                meta_ws: WorksheetLike = self._wb_com.Worksheets("META")
                meta_dir = str(meta_ws.Range("A1").Value).strip()
                meta_name = str(meta_ws.Range("A2").Value).strip()
                if meta_dir:
                    csv_dir = meta_dir
                if meta_name:
                    base = meta_name
                    if base.lower().endswith("_corrected.csv"):
                        base = base[: -len("_corrected.csv")]
                    else:
                        base = os.path.splitext(base)[0]
                    if base:
                        csv_name = base
            except Exception:
                pass
            out_path = os.path.join(csv_dir, f"{csv_name}_Error_Report.xlsm")

            # Remove any existing file to avoid SaveCopyAs errors
            try:
                if os.path.exists(out_path):
                    os.remove(out_path)
            except Exception:
                pass

            # Try SaveCopyAs first; if the workbook is already at out_path (re-homed by
            # _export_and_open_excel), fall back to Save() which just flushes in place.
            try:
                self._wb_com.SaveCopyAs(out_path)
            except Exception:
                try:
                    # Workbook may already be at out_path — try an in-place save
                    try:
                        wb_full = str(self._wb_com.FullName).strip()  # type: ignore[union-attr]
                    except Exception:
                        wb_full = ""
                    if wb_full.lower() == out_path.lower():
                        self._wb_com.Save()  # type: ignore[union-attr]
                    else:
                        self._wb_com.SaveAs(out_path, 52)  # 52 = xlOpenXMLWorkbookMacroEnabled
                except Exception as e:
                    if show_popup:
                        messagebox.showerror("Export error", f"Failed to export Error Report:\n{e}")
                    return False, None, str(e)

            if show_popup:
                messagebox.showinfo("Exported", f"Saved Error Report to:\n{out_path}")
            return True, out_path, None
        except Exception as e:
            if show_popup:
                messagebox.showerror("Export error", f"Failed to export Error Report:\n{e}")
            return False, None, str(e)

    def export_outputs(self, show_summary: bool = True) -> None:
        """Export both the corrected CSV and the Error Report xlsm.
        If show_summary is True, show one popup and update the status bar with saved paths."""
        if not (self._excel and self._wb_com):
            messagebox.showwarning("Excel not ready", "Open/validate a file first so there is data to export.")
            return
        ok_csv, path_csv, err_csv = self.export_corrected_csv(show_popup=False)
        ok_rep, path_rep, err_rep = self.export_error_report(show_popup=False)

        lines: list[str] = []
        if ok_csv and path_csv:
            lines.append(f"CSV saved: {path_csv}")
        else:
            lines.append(f"CSV failed: {err_csv or 'unknown error'}")
        if ok_rep and path_rep:
            lines.append(f"Report saved: {path_rep}")
        else:
            lines.append(f"Report failed: {err_rep or 'unknown error'}")

        # Update status bar
        try:
            self.status.config(text=";  ".join(lines))
        except Exception:
            pass

        if show_summary:
            if ok_csv and ok_rep:
                messagebox.showinfo("Exported", "\n".join(lines))
            else:
                messagebox.showwarning("Export partial", "\n".join(lines))

    def _close_workbook_if_open(self, path: str) -> bool:
        """Close *path* in any reachable Excel instance (without saving).

        Returns True if the workbook was found and closed.  Call this before
        overwriting a file that Excel may have locked.
        Matches by full path first, then by filename alone (handles OneDrive
        path variations where FullName may differ from the local abspath).
        """
        norm = os.path.normcase(os.path.abspath(path))
        target_name = os.path.basename(path).lower()

        def _try_close(wb: Any) -> bool:
            try:
                full = str(wb.FullName)
                if (os.path.normcase(os.path.abspath(full)) == norm
                        or full.lower().endswith(os.sep + target_name)
                        or full.lower().endswith("/" + target_name)
                        or full.lower() == target_name):
                    wb.Close(False)
                    return True
            except Exception:
                pass
            return False

        def _scan_xl(xl_app: Any) -> bool:
            try:
                for _i in range(int(xl_app.Workbooks.Count), 0, -1):  # type: ignore[reportUnknownMemberType]
                    if _try_close(xl_app.Workbooks(_i)):  # type: ignore[reportUnknownMemberType]
                        return True
            except Exception:
                pass
            return False

        # Check workbooks tracked from this session
        for _, _wb, _ in list(self._all_workbooks):
            if _try_close(_wb):
                return True

        # Check current COM handle
        if self._wb_com is not None and _try_close(self._wb_com):
            self._wb_com = None
            return True

        # Check the Excel instance we launched (scan all its workbooks)
        if self._excel is not None and _scan_xl(self._excel):
            return True

        # Last resort: attach to whatever Excel instance Windows reports as active
        try:
            import win32com.client as _wc  # type: ignore[import]
            _xl_active: Any = _wc.GetActiveObject("Excel.Application")  # type: ignore[reportUnknownMemberType]
            if _scan_xl(_xl_active):
                return True
        except Exception:
            pass

        return False

    def export_all_outputs(self) -> None:
        """Save all three outputs to the original CSV/JXL source folder:
          • _Error_Report.xlsm  — current workbook state
          • _corrected.csv      — Data sheet, no header row, NEZ at 10 decimal places
          • _GNSS.csv           — Geodetic Info sheet with metadata rows 1-9 stripped
        """
        if not (self._excel and self._wb_com):
            messagebox.showwarning("Excel not ready", "Open/validate a file first.")
            return

        # Source folder = folder the original CSV or JXL came from
        src_path = self.csv_path or self.jxl_path
        if not src_path:
            messagebox.showwarning("No source file", "Cannot determine source folder.")
            return
        src_dir = os.path.dirname(os.path.abspath(src_path))
        csv_name = os.path.splitext(os.path.basename(src_path))[0]

        saved: list[str] = []
        failed: list[str] = []

        # ── Pre-read sheet data before any SaveCopyAs call ───────────────────
        # SaveCopyAs on a macro-enabled workbook can trigger VBA events that
        # invalidate the COM proxy.  Capture all needed values into plain Python
        # objects now so the CSV steps never have to touch the COM object again.
        _ws_data_vals: Any = None
        _ws_geo_vals: Any = None
        try:
            _ws_data_vals = self._wb_com.Worksheets("Data").UsedRange.Value
        except Exception:
            pass
        try:
            _ws_geo_vals = self._wb_com.Worksheets("Geodetic Info").UsedRange.Value
        except Exception:
            pass

        # ── 1. Error Report xlsm ──────────────────────────────────────────────
        xlsm_path = os.path.join(src_dir, f"{csv_name}_Error_Report.xlsm")
        try:
            # SaveCopyAs to the temp dir first — that path is never open in Excel.
            # Then copy the temp file to the destination.  If the destination is
            # locked (open in another Excel session) fall back to a timestamped name.
            _tmp_copy = os.path.join(self._tmpdir, f"{csv_name}_report_copy_{uuid.uuid4().hex[:6]}.xlsm")
            self._wb_com.SaveCopyAs(_tmp_copy)
            _dest = xlsm_path
            try:
                shutil.copy2(_tmp_copy, _dest)
            except OSError:
                # Destination is locked — close the open workbook then retry
                self._close_workbook_if_open(_dest)
                try:
                    shutil.copy2(_tmp_copy, _dest)
                except OSError as _e2:
                    raise OSError(f"Could not overwrite {os.path.basename(_dest)}: {_e2}") from _e2
            saved.append(f"  {os.path.basename(_dest)}")
        except Exception as e:
            failed.append(f"  Error Report: {e}")

        # ── 2. Corrected CSV — no header row, NEZ columns at 10 decimal places ─
        corrected_path = os.path.join(src_dir, f"{csv_name}_corrected.csv")
        try:
            all_vals: Any = _ws_data_vals
            if all_vals is None:
                raise ValueError("Data sheet is empty")
            # Normalise to list-of-lists
            if isinstance(all_vals, (list, tuple)) and all_vals and isinstance(all_vals[0], (list, tuple)):
                rows_raw: list[list[Any]] = [list(r) for r in all_vals]  # type: ignore[reportUnknownVariableType,reportUnknownArgumentType]
            elif isinstance(all_vals, (list, tuple)):
                rows_raw = [list(all_vals)]  # type: ignore[reportUnknownArgumentType]
            else:
                rows_raw = [[all_vals]]

            # Identify NEZ column positions from header row
            nez_cols: set[int] = set()
            if rows_raw:
                header_lower = [str(v).strip().lower() if v is not None else "" for v in rows_raw[0]]
                _nez_names = {"northing", "n", "easting", "e", "elevation", "z", "elev", "height"}
                for col_i, h in enumerate(header_lower):
                    if h in _nez_names:
                        nez_cols.add(col_i)
            # Also trust self.mapping column indices (0-based, same as Data sheet columns)
            if self.mapping:
                for key in ("n", "e", "z"):
                    idx = self.mapping.get(key)
                    if idx is not None:
                        nez_cols.add(idx)

            with open(corrected_path, "w", newline="", encoding="cp1252", errors="replace") as fh:
                writer = csv.writer(fh)
                for row in rows_raw[1:]:   # skip row 0 = header
                    out_row: list[Any] = []
                    for col_i, val in enumerate(row):
                        if col_i in nez_cols and val is not None and val != "":
                            try:
                                out_row.append(f"{float(val):.10f}")
                            except (ValueError, TypeError):
                                out_row.append("" if val is None else val)
                        else:
                            out_row.append("" if val is None else val)
                    writer.writerow(out_row)
            saved.append(f"  {os.path.basename(corrected_path)}")
        except Exception as e:
            failed.append(f"  Corrected CSV: {e}")

        # ── 3. GNSS CSV — Geodetic Info sheet, skip metadata rows 1-9 ──────────
        gnss_path = os.path.join(src_dir, f"{csv_name}_GNSS.csv")
        try:
            geo_vals: Any = _ws_geo_vals
            if geo_vals is None:
                raise ValueError("Geodetic Info sheet is empty")
            if isinstance(geo_vals, (list, tuple)) and geo_vals and isinstance(geo_vals[0], (list, tuple)):
                geo_rows: list[list[Any]] = [list(r) for r in geo_vals]  # type: ignore[reportUnknownVariableType,reportUnknownArgumentType]
            elif isinstance(geo_vals, (list, tuple)):
                geo_rows = [list(geo_vals)]  # type: ignore[reportUnknownArgumentType]
            else:
                geo_rows = [[geo_vals]]

            with open(gnss_path, "w", newline="", encoding="cp1252", errors="replace") as fh:
                writer = csv.writer(fh)
                for row in geo_rows[9:]:   # skip rows 1-9 (0-based indices 0-8)
                    writer.writerow(["" if v is None else v for v in row])
            saved.append(f"  {os.path.basename(gnss_path)}")
        except Exception as e:
            failed.append(f"  GNSS CSV: {e}")

        # ── Summary ──────────────────────────────────────────────────────────────
        msg = f"Saved to:\n{src_dir}\n\n" + "\n".join(saved)
        if failed:
            msg += "\n\nFailed:\n" + "\n".join(failed)
            messagebox.showwarning("Export — partial", msg)
        else:
            messagebox.showinfo("Export Complete", msg)
        self.status.config(text=f"Saved outputs to: {src_dir}")

    # ---------- helpers ----------
    @staticmethod
    def _norm_token(s: str) -> str:
        return re.sub(r"[\s\-_\.#]+", "", str(s or "")).upper()

    def _is_float(self, s: Any) -> bool:
        try:
            float(str(s).strip())
            return True
        except Exception:
            return False

    def _ensure_fxl_after_csv(self, csv_path: str) -> bool:
        folder = os.path.dirname(csv_path)
        # 0) JXL companion file — most reliable FXL source; records exact FXL used in field.
        #    Always check this FIRST, even if an FXL is already loaded from the previous session,
        #    because the JXL-specified FXL must override the session default (last_fxl_path).
        jxl_cand: str | None = self._find_jxl_alongside_csv(csv_path)  # type: ignore[attr-defined]
        if jxl_cand:
            self.jxl_path = jxl_cand
            jxl_info: dict[str, Any] = self._parse_jxl(jxl_cand)  # type: ignore[attr-defined]
            self._jxl_data = jxl_info
            fxl_from_jxl: str = cast(str, jxl_info.get("fxl_filename") or "")  # type: ignore[union-attr]
            if fxl_from_jxl:
                # Already have the right FXL loaded — no reload needed, no confirm needed
                if self.fxl_path and os.path.basename(self.fxl_path).lower() == fxl_from_jxl.lower():
                    return True
                _jxl_intro = f"JXL companion detected.\nThis job was collected with:\n  {fxl_from_jxl}\n\nFound it at the path below."
                # a) Check same folder
                cand_local = os.path.join(folder, fxl_from_jxl)
                if os.path.isfile(cand_local):
                    confirmed = self._confirm_fxl(cand_local, intro=_jxl_intro)  # type: ignore[attr-defined]
                    if confirmed is None:
                        return False
                    try:
                        self._load_fxl_path(confirmed)
                        return True
                    except Exception:
                        pass
                # b) Search library for exact filename match
                lib_cands = self._get_fxl_library_candidates()
                fxl_lower: str = fxl_from_jxl.lower()
                for lp in lib_cands:
                    if os.path.basename(lp).lower() == fxl_lower:
                        confirmed = self._confirm_fxl(lp, intro=_jxl_intro)  # type: ignore[attr-defined]
                        if confirmed is None:
                            return False
                        try:
                            self._load_fxl_path(confirmed)
                            return True
                        except Exception:
                            pass
        # JXL either not found, had no FXL reference, or the referenced FXL couldn't be located.
        # FXL files live in the org-wide library, not alongside the CSV — go straight there.
        # 2) Search the org-wide FXL library (lazy-cached recursive walk)
        if self.fxl_library_path and os.path.isdir(self.fxl_library_path):
            lib_cands = self._get_fxl_library_candidates()
            if lib_cands:
                # Prefer any file whose stem partially matches the CSV basename — confirm before loading
                csv_stem = os.path.splitext(os.path.basename(csv_path))[0].lower()
                name_matches = [p for p in lib_cands if csv_stem in os.path.splitext(os.path.basename(p))[0].lower()]
                if len(name_matches) == 1:
                    confirmed = self._confirm_fxl(name_matches[0])
                    if confirmed is None:
                        return False
                    try:
                        self._load_fxl_path(confirmed)
                        return True
                    except Exception:
                        pass
                # Let the user pick from the library list
                chosen = self._pick_fxl_from_library(lib_cands, csv_path)
                if chosen:
                    try:
                        self._load_fxl_path(chosen)
                        return True
                    except Exception:
                        pass

        # 4) Fall back to previously-loaded session FXL — nothing better was found; confirm with user
        if self.fxl_data and self.fxl_path:
            confirmed = self._confirm_fxl(  # type: ignore[attr-defined]
                self.fxl_path,
                intro="No FXL was found automatically for this CSV.\nUsing the previously loaded FXL:",
            )
            if confirmed is None:
                return False
            if os.path.abspath(confirmed) != os.path.abspath(self.fxl_path):
                try:
                    self._load_fxl_path(confirmed)
                except Exception:
                    pass
            return True

        # 5) Ask the user via file browser
        p = filedialog.askopenfilename(
            title="Select FXL",
            initialdir=self.fxl_library_path if self.fxl_library_path and os.path.isdir(self.fxl_library_path) else folder,
            filetypes=[("FXL files", "*.fxl"), ("XML files", "*.xml")]
        )
        if p:
            self._load_fxl_path(p)
            return True
        return False

    def _get_fxl_library_candidates(self) -> list[str]:
        """Return all .fxl files under fxl_library_path, cached for the session."""
        if self._fxl_library_cache is not None:
            return self._fxl_library_cache
        found: list[str] = []
        try:
            self.status.config(text="Scanning FXL library…")
            self.update_idletasks()
            for dirpath, _dirs, files in os.walk(self.fxl_library_path):
                for fname in files:
                    if fname.lower().endswith((".fxl", ".xml")):
                        found.append(os.path.join(dirpath, fname))
        except Exception:
            pass
        self._fxl_library_cache = sorted(found, key=lambda p: os.path.basename(p).lower())
        return self._fxl_library_cache

    def _find_fxl_for_jxl(self, jxl_path: str, fxl_filename: str) -> str | None:
        """Search for fxl_filename relative to a JXL file, then in the FXL library.

        Search order:
          1. Same directory as the JXL
          2. Up to 5 parent directories
          3. FXL library candidates (self.fxl_library_path)
        Returns the full path if found, None otherwise.
        """
        if not fxl_filename:
            return None
        fname_lower = fxl_filename.lower()
        # Walk up the directory tree from the JXL location
        check_dir = os.path.dirname(os.path.abspath(jxl_path))
        for _ in range(6):
            cand = os.path.join(check_dir, fxl_filename)
            if os.path.isfile(cand):
                return cand
            parent = os.path.dirname(check_dir)
            if parent == check_dir:
                break
            check_dir = parent
        # Search the FXL library
        for lib_path in self._get_fxl_library_candidates():
            if os.path.basename(lib_path).lower() == fname_lower:
                return lib_path
        return None

    def _confirm_fxl(self, fxl_path: str, intro: str | None = None) -> str | None:
        """Ask the user to confirm an auto-selected FXL.

        Returns the confirmed path (same or user-chosen replacement), or None
        if the user cancelled without choosing anything.
        """
        name = os.path.basename(fxl_path)
        msg_intro = intro if intro else "No FXL was specified in the job file."
        use_it = messagebox.askyesno(
            "Confirm FXL",
            f"{msg_intro}\n\n"
            f"Auto-selected:  {name}\n\n"
            f"Use this FXL?",
            icon="question",
        )
        if use_it:
            return fxl_path
        # User wants a different one — show library picker if available, else file browser
        lib_cands = self._get_fxl_library_candidates() if (self.fxl_library_path and os.path.isdir(self.fxl_library_path)) else []
        if lib_cands:
            return self._pick_fxl_from_library(lib_cands, "")
        p = filedialog.askopenfilename(
            title="Select FXL",
            initialdir=os.path.dirname(fxl_path),
            filetypes=[("FXL files", "*.fxl"), ("XML files", "*.xml")],
        )
        return p if p else None

    def _pick_fxl_from_library(self, candidates: list[str], csv_path: str) -> str | None:
        """Show a dialog listing all library FXL files so the user can choose one."""
        win = tk.Toplevel(self)
        win.title("Select FXL from Library")
        _place_window(win, 580, 440)
        win.grab_set()
        win.resizable(True, True)
        _raise_window(win)

        csv_name = os.path.basename(csv_path)
        tk.Label(win, text=f"No FXL found automatically for:  {csv_name}\n"
                           f"Select the correct FXL from the library:",
                 justify="left", wraplength=540).pack(padx=10, pady=(10, 4), anchor="w")

        # Filter entry
        filter_frame = tk.Frame(win)
        filter_frame.pack(fill="x", padx=10, pady=2)
        tk.Label(filter_frame, text="Filter:").pack(side="left")
        filter_var = tk.StringVar()
        filter_entry = tk.Entry(filter_frame, textvariable=filter_var, width=40)
        filter_entry.pack(side="left", padx=4)

        # Listbox
        list_frame = tk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=10, pady=4)
        sb = tk.Scrollbar(list_frame)
        sb.pack(side="right", fill="y")
        lb = tk.Listbox(list_frame, yscrollcommand=sb.set, font=("Consolas", 9), selectmode="browse")
        lb.pack(side="left", fill="both", expand=True)
        sb.config(command=lb.yview)  # type: ignore[arg-type]

        # Populate list showing relative path from library root
        lib_root = self.fxl_library_path
        display: list[str] = []
        for p in candidates:
            try:
                rel = os.path.relpath(p, lib_root)
            except Exception:
                rel = os.path.basename(p)
            display.append(rel)

        def refresh_list(filter_text: str = "") -> None:
            lb.delete(0, "end")
            ft = filter_text.strip().lower()
            for d in display:
                if not ft or ft in d.lower():
                    lb.insert("end", d)

        filter_var.trace_add("write", lambda *_: refresh_list(filter_var.get()))  # type: ignore[arg-type]
        refresh_list()

        result: list[str | None] = [None]

        def on_ok() -> None:
            sel = lb.curselection()  # type: ignore[var-annotated]
            if not sel:
                messagebox.showwarning("No selection", "Please select an FXL file.", parent=win)
                return
            chosen_rel = lb.get(sel[0])  # type: ignore[arg-type]
            # Map back to full path
            for p, d in zip(candidates, display):
                if d == chosen_rel:
                    result[0] = p
                    break
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Use Selected", width=16, command=on_ok).grid(row=0, column=0, padx=6)
        tk.Button(btn_frame, text="Browse…", width=12,
                  command=lambda: [win.destroy()] or [None]).grid(row=0, column=1, padx=6)
        tk.Button(btn_frame, text="Cancel", width=10, command=win.destroy).grid(row=0, column=2, padx=6)

        # Pre-filter by CSV stem to help user find the right FXL quickly
        filter_var.set(os.path.splitext(csv_name)[0])
        filter_entry.select_range(0, "end")
        filter_entry.focus_set()
        win.wait_window()
        return result[0]

    # numeric recognition (string → number)
    _num_re = re.compile(r'^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$')

    @classmethod
    def _looks_like_number(cls, s: str) -> bool:
        s = (s or "").strip()
        return bool(cls._num_re.match(s))

    @classmethod
    def _to_number(cls, s: str) -> float | str:
        try:
            return float(s)
        except Exception:
            return s

    def _station_score(self, series: pd.Series, sample_n: int = 600) -> tuple[int, int]:
        pat = re.compile(r"^\s*\d+\+\d{2,3}(?:\.\d+)?\s*$")
        matches = 0
        total = 0
        for v in series.head(sample_n):
            s = str(v).strip()
            if not s:
                continue
            total += 1
            if pat.match(s):
                matches += 1
        return matches, total

    def _detect_schema(self, df: pd.DataFrame) -> tuple[bool, dict[str, int | None], list[int]]:
        ncols = df.shape[1]
        if ncols < 5:
            raise ValueError("CSV appears to have too few columns (need at least 5).")
        # best N/E/Z window
        best_i = None
        best_score = -1.0
        sample = df.head(1000)
        for i in range(0, ncols - 2):
            ncol, ecol, zcol = sample.iloc[:, i], sample.iloc[:, i + 1], sample.iloc[:, i + 2]
            good = 0
            total = 0
            for nv, ev, zv in zip(ncol, ecol, zcol):
                sn, se, sz = str(nv).strip(), str(ev).strip(), str(zv).strip()
                if sn == "" and se == "":
                    continue
                # Skip zero-coordinate pairs — real survey coordinates are never (0, 0).
                # Attribute columns often contain 0 padding that would otherwise
                # outscore the actual coordinate columns.
                try:
                    if float(sn) == 0.0 and float(se) == 0.0:
                        continue
                except (ValueError, TypeError):
                    pass
                total += 1
                if self._is_float(sn) and self._is_float(se) and (sz == "" or self._is_float(sz)):
                    good += 1
            if total:
                score = good / total
                if score > best_score:
                    best_score, best_i = score, i
        if best_i is None or best_score < 0.6:
            raise ValueError("Couldn't confidently identify Northing/Easting(/Elevation) numeric columns.")
        # station anywhere
        station_idx, station_matches = None, -1
        for j in range(ncols):
            m, t = self._station_score(df.iloc[:, j])
            if t > 0 and (m / t) >= 0.6 and m > station_matches:
                station_idx, station_matches = j, m
        has_station = station_idx is not None
        pn_idx = best_i - 1 if best_i - 1 >= 0 else None
        north_i = best_i
        east_i = best_i + 1
        elev_i = best_i + 2
        fc_i = best_i + 3 if (best_i + 3) < ncols else None
        if pn_idx is None or elev_i >= ncols or fc_i is None:
            raise ValueError("CSV missing required columns near coordinate detection (PN/N/E/Elev/FC).")
        attr_indices: list[int] = list(range(fc_i + 1, ncols))
        mapping: dict[str, int | None] = {
            "station": station_idx,
            "pn": pn_idx,
            "north": north_i,
            "east": east_i,
            "elev": elev_i,
            "fc": fc_i,
        }
        return has_station, mapping, attr_indices

    def _strip_fc_attr_cells_shift_left(self) -> int:
        if self.df is None:
            return 0
        # Match Trimble CODE:ATTRIBUTE_NAME banner cells structurally:
        # Left side (field code): uppercase identifiers only — letters/digits/underscores/
        # hyphens, no spaces, no leading digit.  Dates (2026-02-23) are safe because
        # they start with a digit, not [A-Z].
        # Right side (attribute name): also starts with uppercase but allows spaces and
        # common punctuation (#, ., (), /) since Trimble attribute labels can be
        # multi-word (e.g. "PLF:PIPE DIAMETER", "PLF:HEAT #", "PLF:O.D.").
        banner_re = re.compile(r"^[A-Z][A-Z0-9_-]*:[A-Z][A-Z0-9_ #()./-]*$")
        total_removed = 0
        nrows, ncols = self.df.shape
        # Only strip banner cells from attribute columns (after FC column).
        # The FC column itself may contain CODE:ATTRIBUTE values (e.g. WELD:X-RAY)
        # which are the actual field code — stripping them shifts all attribute data
        # one column left, causing validation to report wrong column positions.
        _fc_col_raw: int | None = self.mapping.get("fc") if self.mapping else None
        fc_col: int = _fc_col_raw if _fc_col_raw is not None else -1
        # Collect all rows as plain Python lists, modify those that need stripping,
        # then rebuild self.df in one shot.  This is immune to pandas Copy-on-Write
        # behaviour (pandas 2.x) where df.iloc[r, :] = value can silently write to
        # an internal copy rather than the original DataFrame.
        new_rows: list[list[str]] = []
        any_changed = False
        for r in range(nrows):
            row_vals: list[str] = ["" if x is None else str(x) for x in self.df.iloc[r, :].tolist()]
            to_remove: list[int] = [
                c for c, s in enumerate(row_vals)
                if c > fc_col and banner_re.match(s.strip())
            ]
            if to_remove:
                for idx in sorted(to_remove, reverse=True):
                    row_vals.pop(idx)
                    row_vals.append("")
                total_removed += len(to_remove)
                if len(row_vals) < ncols:
                    row_vals.extend([""] * (ncols - len(row_vals)))
                elif len(row_vals) > ncols:
                    row_vals = row_vals[:ncols]
                any_changed = True
            new_rows.append(row_vals)
        if any_changed:
            # Replace self.df with the rebuilt data so changes are guaranteed visible
            self.df = pd.DataFrame(new_rows)
        return total_removed

    def _strip_banner_header_rows(self) -> int:
        if self.df is None or self.df.empty:
            return 0
        removed = 0
        pat = re.compile(r'^\s*(JOB|VERSION|UNITS)\s*:', re.IGNORECASE)
        while len(self.df) > 0:
            vals = self.df.iloc[0, :].tolist()
            row = ["" if v is None else str(v) for v in vals]
            hit = False
            for cell in row[:8]:
                if pat.search(cell):
                    hit = True
                    break
            if hit:
                self.df = self.df.iloc[1:].reset_index(drop=True)
                removed += 1
            else:
                break
        return removed

    # ---------- file loaders ----------
    def load_csv(self):
        p = filedialog.askopenfilename(title="Select CSV", filetypes=[("CSV files", "*.csv")])
        if p:
            self._load_csv_path(p, autopick_fxl=True)
            if self.df is not None and self.fxl_data:
                self.validate_and_open()

    def _read_csv_permissive(self, path: str) -> pd.DataFrame:
        """Load a CSV even if rows have inconsistent column counts."""
        rows: list[list[str]] = []
        max_cols = 0
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                rows.append(list(row))
                if len(row) > max_cols:
                    max_cols = len(row)
        if not rows:
            return pd.DataFrame()
        padded = [row + [""] * (max_cols - len(row)) for row in rows]
        df = pd.DataFrame(padded, dtype=str)
        return df

    def _load_csv_path(self, path: str, autopick_fxl: bool = False):
        try:
            parser_note = ""
            try:
                df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", header=None)  # type: ignore[reportUnknownMemberType]
            except pd.errors.ParserError:
                df = self._read_csv_permissive(path)
                parser_note = " (normalized irregular columns)"
            df = df.fillna("")  # type: ignore[reportUnknownMemberType]
            # Strip "JobName Files\photo.jpg" folder prefixes from any photo-type cells.
            # Trimble exports photo attributes as "FolderName\filename.ext"; we only want
            # the bare filename so the Data sheet and comparisons stay consistent.
            _photo_exts = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".heic", ".mp4", ".mov"}
            def _strip_photo_prefix(v: Any) -> Any:
                if not isinstance(v, str) or ("\\" not in v and "/" not in v):
                    return v
                p = v.replace("\\", "/")
                if os.path.splitext(p)[1].lower() in _photo_exts:
                    return os.path.basename(p)
                return v
            df = df.apply(lambda col: col.map(_strip_photo_prefix))  # type: ignore[reportUnknownMemberType]
            self.df = df
            self.csv_path = path
            if self._initial_csv_path is None:
                self._initial_csv_path = path
            banners = self._strip_banner_header_rows()
            if banners:
                self.status.config(text=f"Removed {banners} banner row(s) (Job:/Version:/Units:).")
            # Detect schema FIRST so self.mapping.fc is known before stripping.
            # Stripping uses fc_col to avoid touching the Field Code column itself —
            # if stripping ran before schema detection, self.mapping was stale/None
            # and fc_col defaulted to -1, causing the FC cell (e.g. "WELD:X-RAY") to
            # be stripped out and shifting all attribute values one column left.
            try:
                self.has_station, self.mapping, self.attr_indices = self._detect_schema(self.df)
            except Exception as schema_err:
                self.status.config(text=f"Auto-detection failed: {schema_err}. Manual mapping required.")
                mapped = self._ask_column_mapping(self.df)
                if mapped is not None:
                    self.has_station, self.mapping, self.attr_indices = mapped
                else:
                    self.has_station, self.mapping, self.attr_indices = None, None, []
            # Strip banner cells NOW — after self.mapping is set so fc_col is correct
            if self.fxl_data and self.mapping:
                removed = self._strip_fc_attr_cells_shift_left()
                if removed:
                    messagebox.showinfo("Cleaned", f"Deleted {removed} 'FieldCode:Attribute' cells (shifted left).")
            self.status.config(text=f"Loaded CSV: {os.path.basename(path)}{parser_note}")
            if autopick_fxl and not self.fxl_data:
                self._ensure_fxl_after_csv(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV:\n{e}")

    def load_fxl(self):
        p = filedialog.askopenfilename(title="Select FXL", filetypes=[("FXL files", "*.fxl"), ("XML files", "*.xml")])
        if p:
            self._load_fxl_path(p)
            if self.df is not None and self.fxl_data:
                self.validate_and_open()

    def _load_fxl_path(self, path: str, silent: bool = False):
        try:
            self.fxl_data = self.parse_fxl(path)
            self.fxl_path = path
            # Capture the first FXL seen as the initial default for CSV-only drops
            if self._initial_fxl_data is None:
                self._initial_fxl_data = self.fxl_data.copy()
                self._initial_fxl_path = path
            self.status.config(text=f"Loaded FXL: {os.path.basename(path)}")
            if self.df is not None:
                self.status.config(text=self.status.cget("text") + "  (CSV already loaded → ready)")
            # Persist so next session auto-loads this FXL
            try:
                self._save_app_config({"last_fxl_path": path})
            except Exception:
                pass
            if not silent:
                sample = ", ".join(list(sorted(self.fxl_data.keys()))[:8])
                messagebox.showinfo(
                    "Loaded",
                    f"Loaded FXL: {os.path.basename(path)}\nField codes parsed: {len(self.fxl_data)}\nSample: {sample or '(none)'}",
                )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load FXL:\n{e}")

    def _load_mtr_path(self, path: str) -> None:
        """Load a third Excel file (MTR) and normalize its headers for cross-checking.
        Expected columns (case/spacing agnostic):
        HEAT/HEAT NUMBER, MANUFACTURER, NOM DIAMETER, OUT DIAMETER, WALL THICKNESS, GRADE, PIPE SPEC, SEAM TYPE."""
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext not in {".xlsx", ".xlsm", ".xls"}:
                raise ValueError("MTR file must be an Excel workbook")
            df_raw = pd.read_excel(path, dtype=str, header=0)  # type: ignore[reportUnknownMemberType]
            df_raw = df_raw.fillna("")  # type: ignore[reportUnknownMemberType]

            def n(s: str) -> str:
                return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())

            col_lookup = {n(c): c for c in df_raw.columns}
            mapping: dict[str, list[str]] = {
                "HEAT": [
                    "HEAT", "HEATNUMBER", "HEATNO", "HEAT#",
                    "HEATID", "HEATIDENT", "HEATIDENTIFICATION", "HEATNUM"
                ],
                "MANUFACTURER": [
                    "MANUFACTURER", "MFR", "MFG", "MANUFACTURERNAME", "MAKE", "MAKER",
                    "MFRNAME", "MANUFACTURERSNAME"
                ],
                # Include size-based aliases for nominal diameter
                "NOM DIAMETER": [
                    "NOMDIAMETER", "NOMINALDIAMETER", "NOMDIAM", "NOMDIA",
                    "NOMINALSIZE", "NOMSIZE", "NOMINALPIPESIZE", "NOMPIPE", "NPS",
                    "NOMINALBORE", "NB", "DN"
                ],
                # Include outside-dia variants
                "OUT DIAMETER": [
                    "OUTDIAMETER", "OD", "OUTERDIAMETER", "OUTSIDEDIAMETER", "OUTDIA", "OUTSIDEDIA",
                    "OUTDIAM", "OUTERDIA", "ODIN", "ODMM"
                ],
                # Common wall thickness variants
                "WALL THICKNESS": ["WALLTHICKNESS", "WALL", "WT", "WALLTHK", "THICKNESS", "THK"],
                "GRADE": ["GRADE"],
                # Pipe spec variants
                "PIPE SPEC": [
                    "PIPESPEC", "SPEC", "SPECIFICATION", "PIPESPECIFICATION", "STANDARD", "PIPESTD",
                    "PIPESTANDARD", "STD", "SPECSTD", "SPECSTD."
                ],
                # Seam/weld variants
                "SEAM TYPE": ["SEAMTYPE", "SEAM", "WELDTYPE", "WELD", "WELDSEAM", "WELDEDSEAM"],
            }

            resolved: dict[str, str] = {}
            for out, keys in mapping.items():
                src = None
                for k in keys:
                    if k in col_lookup:
                        src = col_lookup[k]
                        break
                if src is None:
                    # Fallback by position as last resort (based on documented order)
                    order = [
                        "HEAT", "MANUFACTURER", "NOM DIAMETER", "OUT DIAMETER",
                        "WALL THICKNESS", "GRADE", "PIPE SPEC", "SEAM TYPE",
                    ]
                    try:
                        pos = order.index(out)
                        if pos < len(df_raw.columns):
                            src = df_raw.columns[pos]
                    except Exception:
                        pass
                if src is None:
                    raise ValueError(f"Missing MTR column: {out}")
                resolved[out] = src

            df_norm = pd.DataFrame({k: df_raw[v] for k, v in resolved.items()})
            for c in df_norm.columns:
                df_norm[c] = df_norm[c].str.strip()  # fillna("") already applied above
            # Normalize HEAT to uppercase so matching is case-insensitive
            if "HEAT" in df_norm.columns:
                df_norm["HEAT"] = df_norm["HEAT"].str.upper()

            self.mtr_df = df_norm
            self.mtr_path = path
            self.status.config(text=f"Loaded MTR: {os.path.basename(path)}  ({len(df_norm)} rows)")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load MTR spreadsheet:\n{e}")

    def parse_fxl(self, path: str) -> dict[str, list[dict[str, Any]]]:
        def ln(tag: str) -> str:
            return tag.split("}", 1)[-1]
        def attr_type_from_tag(local: str) -> str:
            u = local.lower()
            if u.startswith("list"):
                return "list"
            if u.startswith("double") or u.startswith("integer") or u.startswith("numeric"):
                return "number"
            if u.startswith("string"):
                return "text"
            if u.startswith("photo") or u.startswith("image"):
                return "photo"
            return "text"
        tree = ET.parse(path)
        root = tree.getroot()
        data: dict[str, list[dict[str, Any]]] = {}
        found_any = False
        feature_nodes: list[ET.Element] = []
        for el in root.iter():
            t = ln(el.tag).lower()
            if t in (
                "pointfeaturedefinition",
                "linefeaturedefinition",
                "polygonfeaturedefinition",
                "feature",
                "featuredefinition",
                "featureclass",
                "code",
                "surveycode",
                "pointcode",
            ):
                feature_nodes.append(el)
        for f in feature_nodes:
            code = (f.get("Code") or f.get("code") or f.get("Name") or f.get("name") or f.get("id") or "").strip()
            if not code:
                for ch in f:
                    if ln(ch.tag).lower() == "code":
                        t = (ch.text or "").strip()
                        if t:
                            code = t
                            break
            if not code:
                continue
            found_any = True
            attrs: list[dict[str, Any]] = []
            for a in f.iter():
                at = ln(a.tag).lower()
                if at.endswith("attribute"):
                    name = (a.get("Name") or a.get("name") or a.get("Label") or a.get("label") or "").strip()
                    typ = (a.get("type") or attr_type_from_tag(at)).strip().lower()
                    entry = (a.get("EntryMethod") or a.get("entrymethod") or "").strip().lower()
                    required = entry.startswith("req")
                    items: list[str] = []
                    for it in a.iter():
                        itn = ln(it.tag).lower()
                        if itn in ("item", "value", "listitem", "option"):
                            iv = (it.text or "").strip()
                            if iv:
                                items.append(iv)
                    attrs.append({"name": name, "type": typ, "items": items, "required": required})
            if code in data:
                if len(attrs) > len(data[code]):
                    data[code] = attrs
            else:
                data[code] = attrs
        if not found_any and not data:
            codes: set[str] = set()
            for el in root.iter():
                if ln(el.tag).lower() == "code":
                    t = (el.text or "").strip()
                    if t:
                        codes.add(t)
            for c in sorted(codes):
                data[c] = []
        return data

    # ---------- Excel export & VBA ----------
    def _nuke_genpy_cache(self):
        try:
            import win32com
            p = getattr(win32com, "__gen_path__", None)
            if p and os.path.isdir(p):
                shutil.rmtree(p, ignore_errors=True)
        except Exception:
            pass

    def _get_excel(self) -> ExcelApplicationLike | None:
        if platform.system() != "Windows" or win32 is None:
            return None
        excel: ExcelApplicationLike | None = None
        try:
            try:
                from win32com.client import gencache
                excel = cast(ExcelApplicationLike, gencache.EnsureDispatch("Excel.Application"))
            except Exception:
                # Record reason, then retry after nuking cache
                try:
                    import traceback as _tb
                    self._last_com_error = _tb.format_exc()
                except Exception:
                    self._last_com_error = "EnsureDispatch failed"
                self._nuke_genpy_cache()
                try:
                    from win32com.client import gencache
                    excel = cast(ExcelApplicationLike, gencache.EnsureDispatch("Excel.Application"))
                except Exception:
                    try:
                        import traceback as _tb
                        self._last_com_error = _tb.format_exc()
                    except Exception:
                        self._last_com_error = "EnsureDispatch after cache clear failed"
                    excel = cast(ExcelApplicationLike, win32.Dispatch("Excel.Application"))  
        except Exception:
            try:
                excel = cast(ExcelApplicationLike, win32.Dispatch("Excel.Application"))  
            except Exception:
                try:
                    import traceback as _tb
                    self._last_com_error = _tb.format_exc()
                except Exception:
                    self._last_com_error = "Dispatch failed"
                excel = None
        return excel

    def _resolve_csv_photo_names(self) -> None:
        """Replace stale photo filenames in self.df with actual on-disk names.

        When photos are renamed (e.g., IMG_2.jpg → TW50777_..._20260221.jpg) the CSV
        still carries the original Trimble filenames.  Two-pass scan:
          Pass 1 — exact filename match: file wasn't renamed, nothing to do.
          Pass 2 — prefix scan: look for a file starting with "{pt_name}_" on disk.
        Must be called after self.mapping and self.attr_indices are set.
        """
        if self.df is None or self.mapping is None or not self.attr_indices or not self.csv_path:
            return
        pn_col: int | None = self.mapping.get("pn")
        if pn_col is None:
            return
        _media_exts = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".heic", ".mp4", ".mov"}
        csv_dir = os.path.dirname(os.path.abspath(self.csv_path))
        _sync_root = self._find_sync_folder(csv_dir)
        _scan_roots: list[str] = ([_sync_root] if _sync_root else []) + [csv_dir]
        _media_idx: dict[str, str] = {}
        for _sr in _scan_roots:
            try:
                for _dp, _, _fns in os.walk(_sr):
                    for _fn in _fns:
                        if os.path.splitext(_fn)[1].lower() in _media_exts:
                            _k = _fn.lower()
                            if _k not in _media_idx:
                                _media_idx[_k] = os.path.join(_dp, _fn)
            except OSError:
                pass
        if not _media_idx:
            return
        for row_idx in range(len(self.df)):
            pt_name: str = str(self.df.iat[row_idx, pn_col]).strip()
            if not pt_name:
                continue
            _prefix: str = pt_name.lower() + "_"
            for col_idx in self.attr_indices:
                if col_idx >= self.df.shape[1]:
                    continue
                cell_val: str = str(self.df.iat[row_idx, col_idx]).strip()
                if not cell_val or os.path.splitext(cell_val.lower())[1] not in _media_exts:
                    continue
                # Pass 1: exact match — photo still has its original name on disk
                if cell_val.lower() in _media_idx:
                    continue
                # Pass 2: renamed file — scan for any file starting with "{pt_name}_"
                for _key in _media_idx:
                    if _key.startswith(_prefix):
                        self.df.iat[row_idx, col_idx] = os.path.basename(_media_idx[_key])
                        break

    def _apply_jxl_photo_corrections(
        self, jxl_data: dict[str, Any]
    ) -> tuple[int, dict[str, str]]:
        """Overwrite stale CSV photo names with the correct names stored in the JXL.

        The JXL is the authoritative record — it holds the final photo filename
        assigned in the field.  The CSV often still carries the generic Trimble
        export name (e.g. 'IMG_2.jpg').  This updates self.df in-place before
        the Data sheet is written so the Error Report shows the right names.

        Returns (number_of_cells_updated, {pt_name: old_csv_basename}) so the
        caller can locate the old files on disk and offer to rename them.
        """
        if self.df is None or self.mapping is None or not self.attr_indices:
            return 0, {}
        pn_col: int | None = self.mapping.get("pn")
        if pn_col is None:
            return 0, {}
        _media_exts = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".heic", ".mp4", ".mov"}

        # Build {pt_name_lower: correct_bare_filename} from JXL point records
        jxl_photos: dict[str, str] = {}
        for pt_name, pt_data in jxl_data.get("points", {}).items():
            photo = str(pt_data.get("photo_name", "")).strip()
            if photo:
                jxl_photos[pt_name.strip().lower()] = photo
        if not jxl_photos:
            return 0, {}

        updated = 0
        old_names: dict[str, str] = {}  # pt_name → old csv basename (before correction)
        for row_idx in range(len(self.df)):  # type: ignore[arg-type]
            pt_name_val: str = str(self.df.iat[row_idx, pn_col]).strip()
            if not pt_name_val:
                continue
            correct = jxl_photos.get(pt_name_val.lower())
            if not correct:
                continue
            for col_idx in self.attr_indices:
                if col_idx >= self.df.shape[1]:
                    continue
                cell_val: str = str(self.df.iat[row_idx, col_idx]).strip()
                if not cell_val:
                    continue
                if os.path.splitext(cell_val.lower())[1] not in _media_exts:
                    continue
                if cell_val.lower() != correct.lower():
                    old_names[pt_name_val] = cell_val  # capture before overwriting
                    self.df.iat[row_idx, col_idx] = correct
                    updated += 1
        return updated, old_names

    def validate_and_open(self, *, open_new_excel_instance: bool = False):
        if self.df is None:
            messagebox.showwarning("No CSV", "Please load a CSV first.")
            return
        if not self.fxl_data:
            messagebox.showwarning("No FXL", "Please load an FXL first.")
            return
        self._strip_banner_header_rows()
        # Detect schema before stripping so fc_col is correct
        try:
            self.has_station, self.mapping, self.attr_indices = self._detect_schema(self.df)
        except Exception as e:
            messagebox.showerror("Format Error", str(e))
            return
        removed = self._strip_fc_attr_cells_shift_left()
        if removed:
            messagebox.showinfo("Cleaned", f"Deleted {removed} 'FieldCode:Attribute' cells (shifted left) before validation.")
        # Re-detect schema after stripping in case column layout shifted
        try:
            self.has_station, self.mapping, self.attr_indices = self._detect_schema(self.df)
        except Exception as e:
            messagebox.showerror("Format Error", str(e))
            return
        # Resolve stale photo filenames (e.g. IMG_2.jpg → renamed on-disk file)
        self._resolve_csv_photo_names()
        self._export_and_open_excel(open_new_excel_instance=open_new_excel_instance)

    def _export_and_open_excel(self, *, open_new_excel_instance: bool = False) -> None:
        # Guard for type checkers; validate_and_open enforces these at runtime already
        if self.df is None or self.mapping is None:
            return
        # Do not close any previously opened workbook when we're about to
        # launch a new Excel instance for the next validation. This preserves
        # earlier sessions as requested. Only close when reusing the same
        # instance/flow.
        if not open_new_excel_instance:
            try:
                if self._wb_com:
                    self._wb_com.Close(False)
                    self._wb_com = None
            except Exception:
                pass

        if openpyxl is None:
            messagebox.showerror("Missing openpyxl", "openpyxl is required to export Excel.")
            return

        csv_dir = os.path.dirname(self.csv_path) if self.csv_path else os.getcwd()
        csv_name = os.path.splitext(os.path.basename(self.csv_path or "data"))[0]

        unique = uuid.uuid4().hex[:8]
        temp_xlsx = os.path.join(self._tmpdir, f"{csv_name}_temp_{unique}.xlsx")
        # Save the final xlsm to the user's Desktop so it's easy to find.
        # Intermediate xlsx stays in temp.
        _desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        _save_dir = _desktop if os.path.isdir(_desktop) else csv_dir
        xlsm_file = os.path.join(_save_dir, f"{csv_name}_Error_Report.xlsm")
        # Close any open copy in Excel then remove it so COM SaveAs doesn't fail
        try:
            if os.path.exists(xlsm_file):
                self._close_workbook_if_open(xlsm_file)
                os.remove(xlsm_file)
        except Exception:
            pass
        self._temp_xlsx = temp_xlsx
        self._temp_xlsm = xlsm_file

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"

        order: list[tuple[str, int | None]] = []
        if self.has_station and self.mapping["station"] is not None:
            order.append(("Station", self.mapping["station"]))
        order += [
            ("Point Number", self.mapping["pn"]),
            ("Northing", self.mapping["north"]),
            ("Easting", self.mapping["east"]),
            ("Elevation", self.mapping["elev"]),
            ("Field Code", self.mapping["fc"]),
        ]
        for k, idx_src in enumerate(self.attr_indices, start=1):
            order.append((f"Attr{k}", idx_src))
        # Fast header write
        header_row = [name for (name, _) in order]
        ws.append(header_row)

        # Skip the first row if it looks like the CSV header that was loaded with header=None.
        # When pandas reads with header=None the "Point Number, Northing, ..." row becomes
        # self.df row 0 as plain data.  Writing it after the Python-generated header row
        # produces a duplicate header at Excel row 2, shifting all VBA validation by one row.
        df_start = 0
        if len(self.df) > 0:  # type: ignore[reportOptionalMemberAccess]
            r0 = [str(v).strip().lower() for v in self.df.iloc[0, :8].tolist()]  # type: ignore[reportOptionalMemberAccess]
            if "northing" in r0 or "point number" in r0:
                df_start = 1

        # Apply JXL photo name corrections before writing rows — the JXL holds the
        # authoritative filename (e.g. JP5145_JPA_020426_TTG_20260204.jpg) while
        # the CSV may still carry the original generic name (IMG.jpg / IMG_1.jpg).
        # Resolve the JXL early if not already done (handles CSV-only drop where the
        # JXL sits alongside the CSV but was not explicitly dropped by the user).
        if self._jxl_data is None and self.csv_path:
            _early_jxl: str | None = self._find_jxl_alongside_csv(self.csv_path)  # type: ignore[attr-defined]
            if _early_jxl:
                self.jxl_path = _early_jxl
                self._jxl_data = self._parse_jxl(_early_jxl)  # type: ignore[attr-defined]
        _old_photo_names: dict[str, str] = {}
        if self._jxl_data:
            _, _old_photo_names = self._apply_jxl_photo_corrections(self._jxl_data)

        # Fast row writes via itertuples + append (significantly faster than per-cell writes)
        for tup in self.df.iloc[df_start:].itertuples(index=False, name=None):  # type: ignore[reportOptionalMemberAccess]
            out_row: list[Any] = []
            for _, idx_src in order:
                raw = "" if idx_src is None else ("" if tup[idx_src] is None else str(tup[idx_src]))
                out_row.append(self._to_number(raw) if raw != "" and self._looks_like_number(raw) else raw)
            ws.append(out_row)

        ws_fxl = wb.create_sheet("FXL")
        assert ws_fxl is not None
        ws_fxl.append(["Code", "AttrIdx", "AttrName", "AttrType", "AllowedItemsCSV", "Required"])
        for code, attrs in sorted(self.fxl_data.items(), key=lambda kv: kv[0].strip().upper()):
            code_key = code.strip()
            if not attrs:
                ws_fxl.append([code_key, 0, "", "", "", ""])
            else:
                for idx, a in enumerate(attrs, start=1):
                    aname = (a.get("name") or "").strip()
                    atyp = (a.get("type") or "").strip().lower()
                    req = "Required" if a.get("required") else "Optional"
                    allowed: list[str] = []
                    for it in a.get("items", []):
                        s = str(it).strip()
                        if s:
                            allowed.append(s)
                    ws_fxl.append([code_key, idx, aname, atyp, ",".join(sorted(set(allowed))), req])
        ws_fxl.sheet_state = "veryHidden"

        # Optional: embed MTR sheet if provided
        if self.mtr_df is not None and not self.mtr_df.empty:
            ws_mtr = wb.create_sheet("MTR")
            assert ws_mtr is not None
            mtr_cols = [
                "HEAT", "MANUFACTURER", "NOM DIAMETER", "OUT DIAMETER",
                "WALL THICKNESS", "GRADE", "PIPE SPEC", "SEAM TYPE",
            ]
            ws_mtr.append(mtr_cols)
            for _, r in self.mtr_df[mtr_cols].iterrows():  # type: ignore[reportUnknownMemberType]
                ws_mtr.append(["" if pd.isna(r[c]) else str(r[c]).strip() for c in mtr_cols])
            ws_mtr.sheet_state = "veryHidden"

        ws_meta = wb.create_sheet("META")
        assert ws_meta is not None
        ws_meta["A1"] = _save_dir  # report save dir (Desktop)
        ws_meta["A2"] = f"{csv_name}_corrected.csv"
        ws_meta["A3"] = self._sentinel
        ws_meta.sheet_state = "veryHidden"

        # Optional: numeric bounds for attribute range checking (from config)
        if self.numeric_bounds:
            ws_bounds = wb.create_sheet("BOUNDS")
            assert ws_bounds is not None
            ws_bounds.append(["AttrName", "Min", "Max"])
            for attr_nm, mnmx in self.numeric_bounds.items():
                ws_bounds.append([attr_nm, mnmx[0], mnmx[1]])
            ws_bounds.sheet_state = "veryHidden"

        # JXL companion sheets (Geodetic Info + hidden JXL_AUDIT)
        _jxl: dict[str, Any] | None = self._jxl_data
        if _jxl is None and self.csv_path:
            _jxl_cand: str | None = self._find_jxl_alongside_csv(self.csv_path)  # type: ignore[attr-defined]
            if _jxl_cand:
                self.jxl_path = _jxl_cand
                _jxl = self._parse_jxl(_jxl_cand)  # type: ignore[attr-defined]
                self._jxl_data = _jxl
        # Photo discovery — done here (before wb.save) so MEDIA_AUDIT sheet can be embedded
        self._pending_photo_found: dict[str, str] = {}
        self._pending_photo_missing: dict[str, str] = {}
        if _jxl and self.jxl_path:
            try:
                self._pending_photo_found, self._pending_photo_missing = (
                    self._find_jxl_photos(self.jxl_path, _jxl)  # type: ignore[attr-defined]
                )
            except Exception:
                pass

        # When _apply_jxl_photo_corrections renamed a CSV cell (e.g. IMG.jpg →
        # JP5145_….jpg), the disk file still carries the old generic name.
        # _find_jxl_photos couldn't find the JXL-authoritative name on disk, so
        # those points landed in _pending_photo_missing.  Scan for the old names
        # now and promote any matches to _pending_photo_found so _offer_photo_rename
        # can present them for disk renaming.
        if _old_photo_names and self.jxl_path:
            _old_idx = self._build_jxl_media_index(  # type: ignore[attr-defined]
                [self.jxl_path],
                {os.path.abspath(self.jxl_path): self._jxl_data} if self._jxl_data else None,
            )
            for _pt2, _old_bn in _old_photo_names.items():
                # Skip if already found under the JXL-authoritative name
                if _pt2 in self._pending_photo_found:
                    continue
                if _old_bn.lower() in _old_idx:
                    self._pending_photo_found[_pt2] = _old_idx[_old_bn.lower()]
                    self._pending_photo_missing.pop(_pt2, None)

        if _jxl:
            self._write_geodetic_sheet(wb, _jxl, self.df)  # type: ignore[attr-defined]
            self._write_jxl_audit_sheet(wb, _jxl, self.df)  # type: ignore[attr-defined]
        if self._pending_photo_found or self._pending_photo_missing:
            self._write_media_audit_sheet(  # type: ignore[attr-defined]
                wb, self._pending_photo_found, self._pending_photo_missing
            )

        wb.save(temp_xlsx)

        # === FULL VBA MODULE ===
        vbcode: str = r"""
Option Explicit
' ---------------- Safe helpers ----------------
Public Function VCount(ByVal v As Variant) As Long
    On Error Resume Next
    If IsArray(v) Then
        VCount = UBound(v) - LBound(v) + 1
        If Err.Number <> 0 Then VCount = 0: Err.Clear
    Else
        VCount = 0
    End If
End Function
Public Function HasItems(arr As Variant) As Boolean: HasItems = (VCount(arr) > 0): End Function
Public Function NormName(ByVal s As String) As String
    s = UCase$(s)
    s = Replace(s, "-", "")
    s = Replace(s, "_", "")
    s = Replace(s, " ", "")
    s = Replace(s, ".", "")
    s = Replace(s, "#", "")
    ' Strip common punctuation/symbols often present in units/labels
    s = Replace(s, "(", "")
    s = Replace(s, ")", "")
    s = Replace(s, "/", "")
    s = Replace(s, ":", "")
    s = Replace(s, ",", "")
    s = Replace(s, "'", "")
    s = Replace(s, ChrW(216), "O") ' Ø to O
    s = Replace(s, ChrW(248), "o") ' ø to o
    NormName = s
End Function
Private Function HasLetters(ByVal s As String) As Boolean
    Dim i As Long
    For i = 1 To Len(s)
        If Mid$(s, i, 1) Like "[A-Za-z]" Then
            HasLetters = True
            Exit Function
        End If
    Next i
End Function
Private Function IsAllCaps(ByVal s As String) As Boolean
    If Not HasLetters(s) Then
        IsAllCaps = True
    Else
        IsAllCaps = (UCase$(s) = s)
    End If
End Function
Private Function IsMediaFilename(ByVal s As String) As Boolean
    ' Returns True if s looks like a photo/media file (has a known image/video extension).
    ' Used to exempt renamed media files from the all-caps check regardless of FXL type.
    Dim low As String: low = LCase$(Trim$(s))
    IsMediaFilename = (Right$(low, 4) = ".jpg"  Or Right$(low, 5) = ".jpeg" Or _
                       Right$(low, 4) = ".png"  Or Right$(low, 4) = ".tif"  Or _
                       Right$(low, 5) = ".tiff" Or Right$(low, 4) = ".bmp"  Or _
                       Right$(low, 5) = ".heic" Or Right$(low, 4) = ".mp4"  Or _
                       Right$(low, 4) = ".mov")
End Function
Private Function InvalidToken(ByVal s As String) As String
    Dim up As String: up = UCase$(Trim$(s))
    If up = "" Then InvalidToken = "Invalid Blank": Exit Function
    If up = "N/A" Or up = "NA" Or up = "UNK" Or up = "UNKNOWN" Or up = "UNDETERMINED" _
       Or up = "_" Or up = "-" Or up = "." Or up = "\" Or up = "/" _
       Or up = "NAN" Then InvalidToken = "Invalid " & up: Exit Function
    InvalidToken = ""
End Function
Private Function StationPatternOK(ByVal s As String) As Boolean
    Dim re As Object: Set re = CreateObject("VBScript.RegExp")
    re.Pattern = "^\s*\d+\+\d{2,3}(\.\d+)?\s*$": re.IgnoreCase = True
    StationPatternOK = re.Test(CStr(s))
End Function
Private Function IsHorizAngleName(ByVal nm As String) As Boolean
    Dim u As String: u = NormName(nm)
    IsHorizAngleName = (InStr(1, u, "HORIZANGLE", vbTextCompare) > 0) _
                       Or (InStr(1, u, "HORIZONTALANGLE", vbTextCompare) > 0) _
                       Or (u = "HANGLE" Or u = "HORZANG" Or u = "HORZ" Or u = "H")
End Function
Private Function IsVertAngleName(ByVal nm As String) As Boolean
    Dim u As String: u = NormName(nm)
    IsVertAngleName = (InStr(1, u, "VERTANGLE", vbTextCompare) > 0) _
                      Or (InStr(1, u, "VERTICALANGLE", vbTextCompare) > 0) _
                      Or (u = "VANGLE" Or u = "VERT" Or u = "V")
End Function
Private Function IsRadiusName(ByVal nm As String) As Boolean
    Dim u As String: u = NormName(nm)
    IsRadiusName = (u = "BENDRADIUS" Or u = "RADIUS" Or u = "RAD")
End Function
' ========= Coloring & comments =========
Private Sub FlagCell(ByVal r As Long, ByVal c As Long, ByVal msg As String, _
                     Optional ByVal rgbR As Long = 255, Optional ByVal rgbG As Long = 200, Optional ByVal rgbB As Long = 200)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim header As String: header = Trim$(CStr(ws.Cells(1, c).Value))
    Dim display As String: display = header

    If Len(header) > 0 And UCase$(Left$(header, 4)) = "ATTR" Then
        Dim cFC As Long: cFC = ColOf("Field Code")
        If cFC > 0 Then
            Dim fc As String: fc = Trim$(CStr(ws.Cells(r, cFC).Value))
            If fc <> "" Then
                Dim idx As Long: idx = c - cFC
                Dim nm As String: nm = AttrNameForFCIndex(fc, idx)
                If Len(Trim$(nm)) > 0 Then display = nm
            End If
        End If
    End If

    Dim full As String: full = display & " — " & msg
    ws.Cells(r, c).Interior.Color = RGB(rgbR, rgbG, rgbB)
    On Error Resume Next
    If ws.Cells(r, c).Comment Is Nothing Then
        ws.Cells(r, c).AddComment full
    Else
        ws.Cells(r, c).Comment.Text Text:=full
    End If
    On Error GoTo 0
End Sub

' === Coloring helpers (replace/add) ===
Private Sub FlagRedStrong(ByVal r As Long, ByVal c As Long, ByVal msg As String)
    ' heavy red
    FlagCell r, c, msg, 255, 120, 120
End Sub

Private Sub FlagRed(ByVal r As Long, ByVal c As Long, ByVal msg As String)
    ' wrapper for legacy calls
    FlagRedStrong r, c, msg
End Sub

Private Sub FlagOrange(ByVal r As Long, ByVal c As Long, ByVal msg As String)
    FlagCell r, c, msg, 255, 170, 100
End Sub

Private Sub FlagYellow(ByVal r As Long, ByVal c As Long, ByVal msg As String)
    FlagCell r, c, msg, 255, 230, 150
End Sub

Private Sub FlagPurple(ByVal r As Long, ByVal c As Long, ByVal msg As String)
    FlagCell r, c, msg, 200, 150, 255
End Sub

Private Sub FlagLightGreen(ByVal r As Long, ByVal c As Long, ByVal msg As String)
    FlagCell r, c, msg, 200, 255, 200
End Sub

Private Sub FlagTeal(ByVal r As Long, ByVal c As Long, ByVal msg As String):
    FlagCell r, c, msg, 60, 190, 190:
End Sub

Private Sub FlagPink(ByVal r As Long, ByVal c As Long, ByVal msg As String):
    FlagCell r, c, msg, 240, 90, 170:
End Sub

Private Sub ClearRowFlags(ByVal r As Long)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastCol As Long: lastCol = LastHeaderCol()
    If lastCol < 1 Then Exit Sub
    ' Clear the entire row in two range operations instead of one-cell-at-a-time loop.
    ' For a 15-column sheet this cuts ~30 object calls down to 2 per row.
    Dim rng As Range: Set rng = ws.Range(ws.Cells(r, 1), ws.Cells(r, lastCol))
    rng.Interior.ColorIndex = xlNone
    On Error Resume Next
    rng.ClearComments
    On Error GoTo 0
End Sub

Private Sub ClearMarkIfContains(ws As Worksheet, ByVal r As Long, ByVal c As Long, ByVal key As String, Optional ByVal rgbR As Long = -1, Optional ByVal rgbG As Long = -1, Optional ByVal rgbB As Long = -1)
    On Error Resume Next
    If Not ws.Cells(r, c).Comment Is Nothing Then
        If InStr(1, ws.Cells(r, c).Comment.Text, key, vbTextCompare) > 0 Then
            ws.Cells(r, c).Comment.Delete
            If rgbR >= 0 Then
                If ws.Cells(r, c).Interior.Color = RGB(rgbR, rgbG, rgbB) Then ws.Cells(r, c).Interior.ColorIndex = xlNone
            Else
                ws.Cells(r, c).Interior.ColorIndex = xlNone
            End If
        End If
    End If
    On Error GoTo 0
End Sub

' ---------------- Utilities ----------------
Public Function ColOf(headerName As String) As Long
    Static cache As Object
    If cache Is Nothing Then Set cache = CreateObject("Scripting.Dictionary")
    Dim want As String: want = Trim$(UCase$(headerName))
    If cache.Exists(want) Then ColOf = cache(want): Exit Function
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastCol As Long: lastCol = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
    Dim c As Long
    For c = 1 To lastCol
        If Trim$(UCase$(ws.Cells(1, c).Value)) = want Then
            cache(want) = c
            ColOf = c
            Exit Function
        End If
    Next c
    cache(want) = 0
    ColOf = 0
End Function
Private Function LastHeaderCol() As Long: LastHeaderCol = Worksheets("Data").Cells(1, Worksheets("Data").Columns.Count).End(xlToLeft).Column: End Function
Private Function CellText(ws As Worksheet, ByVal r As Long, ByVal c As Long) As String
    On Error GoTo SafeExit
    If c <= 0 Then CellText = "": Exit Function
    CellText = Trim$(CStr(ws.Cells(r, c).Value))
    Exit Function
SafeExit: CellText = ""
End Function

' Case-insensitive membership check for Variant arrays (e.g., list items)
Public Function InArrayCI(ByVal items As Variant, ByVal value As String) As Boolean
    On Error GoTo done
    If Not IsArray(items) Then Exit Function
    Dim tgt As String: tgt = UCase$(Trim$(CStr(value)))
    Dim i As Long
    For i = LBound(items) To UBound(items)
        If UCase$(Trim$(CStr(items(i)))) = tgt Then
            InArrayCI = True
            Exit Function
        End If
    Next i
done:
End Function

' ---------------- FXL lookups ----------------
Private Function GetFXLIndex() As Object
    Static dict As Object
    If dict Is Nothing Then
        Set dict = CreateObject("Scripting.Dictionary")
        Dim fws As Worksheet: Set fws = Worksheets("FXL")
        Dim lastRow As Long: lastRow = fws.Cells(fws.Rows.Count, 1).End(xlUp).Row
        Dim r As Long
        Dim code As String, key As String, idx As Long, aname As String, atype As String
        Dim allowed As String, req As String
        Dim entry As Object, perCode As Object
        For r = 2 To lastRow
            code = CStr(fws.Cells(r, 1).Value)
            key = UCase$(Trim$(code))
            If Len(key) > 0 Then
                idx = CLng(Val(fws.Cells(r, 2).Value))
                aname = Trim$(CStr(fws.Cells(r, 3).Value))
                atype = UCase$(Trim$(CStr(fws.Cells(r, 4).Value)))
                allowed = CStr(fws.Cells(r, 5).Value)
                req = UCase$(Trim$(CStr(fws.Cells(r, 6).Value)))
                If Not dict.Exists(key) Then
                    Set perCode = CreateObject("Scripting.Dictionary")
                    dict.Add key, perCode
                Else
                    Set perCode = dict(key)
                End If
                Set entry = CreateObject("Scripting.Dictionary")
                entry("name") = aname
                entry("type") = atype
                entry("allowed_csv") = allowed
                entry("required") = (req = "REQUIRED")
                Set perCode(CStr(idx)) = entry
            End If
        Next r
    End If
    Set GetFXLIndex = dict
End Function

' ---------------- FXL lookups (legacy helper retained for fallback) ----------------
Private Sub FindCodeBounds(ByVal code As String, ByRef r1 As Long, ByRef r2 As Long)
    Dim fws As Worksheet: Set fws = Worksheets("FXL")
    r1 = 0: r2 = -1
    Dim lastRow As Long: lastRow = fws.Cells(fws.Rows.Count, 1).End(xlUp).Row
    Dim tgt As String: tgt = UCase$(Trim$(code))
    Dim r As Long
    For r = 2 To lastRow
        If UCase$(Trim$(CStr(fws.Cells(r, 1).Value))) = tgt Then
            If r1 = 0 Then r1 = r
            r2 = r
        ElseIf r1 > 0 And r2 >= r1 Then
            Exit For
        End If
    Next r
End Sub
Public Function FCExists(ByVal code As String) As Boolean
    Dim idx As Object: Set idx = GetFXLIndex()
    FCExists = idx.Exists(UCase$(Trim$(code)))
End Function
Public Function AttrTypeForFCIndex(ByVal code As String, ByVal idx As Long) As String
    Dim key As String: key = UCase$(Trim$(code))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then AttrTypeForFCIndex = "": Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim e As Object
    On Error Resume Next: Set e = perCode(CStr(idx)): On Error GoTo 0
    If e Is Nothing Then AttrTypeForFCIndex = "" Else AttrTypeForFCIndex = CStr(e("type"))
End Function
Public Function AttrNameForFCIndex(ByVal code As String, ByVal idx As Long) As String
    Dim key As String: key = UCase$(Trim$(code))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then AttrNameForFCIndex = "": Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim e As Object
    On Error Resume Next: Set e = perCode(CStr(idx)): On Error GoTo 0
    If e Is Nothing Then AttrNameForFCIndex = "" Else AttrNameForFCIndex = CStr(e("name"))
End Function
Public Function AllowedItemsForFCIndex(ByVal code As String, ByVal idx As Long) As Variant
    Dim key As String: key = UCase$(Trim$(code))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then AllowedItemsForFCIndex = Array(): Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim e As Object
    On Error Resume Next: Set e = perCode(CStr(idx)): On Error GoTo 0
    If e Is Nothing Then
        AllowedItemsForFCIndex = Array()
    Else
        Dim s As String: s = CStr(e("allowed_csv"))
        If Len(Trim$(s)) = 0 Then AllowedItemsForFCIndex = Array() Else AllowedItemsForFCIndex = Split(s, ",")
    End If
End Function
Public Function AttrIsOptionalForFCIndex(ByVal code As String, ByVal idx As Long) As Boolean
    Dim key As String: key = UCase$(Trim$(code))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then AttrIsOptionalForFCIndex = True: Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim e As Object
    On Error Resume Next: Set e = perCode(CStr(idx)): On Error GoTo 0
    If e Is Nothing Then
        AttrIsOptionalForFCIndex = True
    Else
        AttrIsOptionalForFCIndex = (Not CBool(e("required")))
    End If
End Function
Public Function AttrIndexByAliases(ByVal code As String, ByVal aliases As Variant) As Long
    AttrIndexByAliases = 0
    Dim key As String: key = UCase$(Trim$(code))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim k As Variant, nm As String, i As Long, e As Object
    For Each k In perCode.Keys
        On Error Resume Next
        Set e = perCode(CStr(k))
        If Err.Number <> 0 Then
            Err.Clear
            nm = CStr(perCode(CStr(k)))
        Else
            If Not e Is Nothing Then nm = CStr(e("name")) Else nm = ""
        End If
        On Error GoTo 0
        For i = LBound(aliases) To UBound(aliases)
            If NormName(nm) = NormName(CStr(aliases(i))) Then AttrIndexByAliases = CLng(Val(CStr(k))): Exit Function
        Next i
    Next k
End Function

' ---------------- Context Menu ----------------
Public Sub ContextMenu_Add()
    On Error Resume Next
    Dim bar As Object: Set bar = Application.CommandBars("Cell")
    If bar Is Nothing Then Exit Sub
    ' Clean any existing items first
    ContextMenu_Remove
    ' Add a simple caption-only button (Type:=1 = msoControlButton, Style:=2 = msoButtonCaption)
    Dim btn As Object
    Set btn = bar.Controls.Add(Type:=1, Temporary:=True)
    btn.Caption = "Clear Validation Flag(s)"
    btn.Tag = "DVT_ClearFlags"
    btn.Style = 2
    btn.OnAction = "ValidationModule.ClearValidationFlagsSelection"
    ' Add ignore-by-type button
    Dim btn2 As Object
    Set btn2 = bar.Controls.Add(Type:=1, Temporary:=True)
    btn2.Caption = "Ignore All Errors of This Type"
    btn2.Tag = "DVT_IgnoreType"
    btn2.Style = 2
    btn2.OnAction = "ValidationModule.IgnoreErrorTypeSelection"
    ' Add MTR helpers if MTR sheet exists
    If HasMTR() Then
        Dim btn3 As Object
        Set btn3 = bar.Controls.Add(Type:=1, Temporary:=True)
        btn3.Caption = "Use MTR value for this cell"
        btn3.Tag = "DVT_MTR_ApplyCell"
        btn3.Style = 2
        btn3.OnAction = "ValidationModule.ApplyMTRValueCell"

        Dim btn4 As Object
        Set btn4 = bar.Controls.Add(Type:=1, Temporary:=True)
        btn4.Caption = "Use MTR value for all associated Heats"
        btn4.Tag = "DVT_MTR_ApplyAll"
        btn4.Style = 2
        btn4.OnAction = "ValidationModule.ApplyMTRValueAttrAll"
    End If
End Sub

Public Sub ContextMenu_Remove()
    On Error Resume Next
    Dim bar As Object: Set bar = Application.CommandBars("Cell")
    If Not bar Is Nothing Then
        ' Remove by tag if available
        Dim ctl As Object
        Set ctl = bar.FindControl(Tag:="DVT_ClearFlags", Recursive:=True)
        If Not ctl Is Nothing Then ctl.Delete
        ' Also try by caption as fallback
        bar.Controls("Clear Validation Flag(s)").Delete
        Set ctl = bar.FindControl(Tag:="DVT_IgnoreType", Recursive:=True)
        If Not ctl Is Nothing Then ctl.Delete
        bar.Controls("Ignore All Errors of This Type").Delete
        ' Remove optional MTR helpers
        Set ctl = bar.FindControl(Tag:="DVT_MTR_ApplyCell", Recursive:=True)
        If Not ctl Is Nothing Then ctl.Delete
        On Error Resume Next: bar.Controls("Use MTR value for this cell").Delete: On Error GoTo 0
        Set ctl = bar.FindControl(Tag:="DVT_MTR_ApplyAll", Recursive:=True)
        If Not ctl Is Nothing Then ctl.Delete
        On Error Resume Next: bar.Controls("Use MTR value for all associated Heats").Delete: On Error GoTo 0
    End If
End Sub

Public Sub ClearValidationFlagsSelection()
    On Error Resume Next
    Dim rng As Range: Set rng = Selection
    On Error GoTo 0
    If rng Is Nothing Then Exit Sub
    Dim c As Range
    For Each c In rng.Cells
        ' Remove fill and any comment
        c.Interior.ColorIndex = xlNone
        On Error Resume Next
        If Not c.Comment Is Nothing Then c.Comment.Delete
        On Error GoTo 0
    Next c
    ' Refresh the Error Count sheet after clearing flags
    RebuildErrorCount
End Sub

' Ignore all errors matching the selected cell's error type
Private Function ExtractErrorMsg(ByVal t As String) As String
    Dim msg As String: msg = Trim$(t)
    On Error Resume Next
    If InStr(1, t, " – ", vbTextCompare) > 0 Then
        Dim p1() As String: p1 = Split(t, " – ")
        ExtractErrorMsg = Trim$(p1(UBound(p1)))
        Exit Function
    End If
    If InStr(1, t, " - ", vbTextCompare) > 0 Then
        Dim p2() As String: p2 = Split(t, " - ")
        ExtractErrorMsg = Trim$(p2(UBound(p2)))
        Exit Function
    End If
    ExtractErrorMsg = msg
End Function
Private Function ErrorTypeKey(ByVal msg As String) As String
    Dim s As String: s = Trim$(msg)
    Dim p As Long
    p = InStr(1, s, "(")
    If p > 1 Then s = Trim$(Left$(s, p - 1))
    If InStr(1, s, "Duplicate XRAY", vbTextCompare) = 1 Then s = "Duplicate XRAY"
    ErrorTypeKey = s
End Function
Public Sub IgnoreErrorTypeSelection()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim rng As Range
    On Error Resume Next
    Set rng = Selection
    On Error GoTo 0
    If rng Is Nothing Then Exit Sub

    Dim keys As Object: Set keys = CreateObject("Scripting.Dictionary")
    Dim cell As Range, src As String, key As String, keyNorm As String
    For Each cell In rng.Cells
        If cell.Row >= 2 Then
            If Not cell.Comment Is Nothing Then
                src = cell.Comment.Text
                key = ErrorTypeKey(ExtractErrorMsg(src))
                keyNorm = UCase$(Trim$(key))
                If Len(keyNorm) > 0 Then
                    If Not keys.Exists(keyNorm) Then keys.Add keyNorm, keyNorm
                End If
            End If
        End If
    Next cell
    If keys.Count = 0 Then Exit Sub

    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim lastCol As Long: lastCol = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
    Dim rr As Long, cc As Long, tt As String, ttKey As String, ttNorm As String
    Dim ttUpper As String, ttMsgUpper As String, ttExtract As String
    Dim matched As Boolean, k As Variant
    Dim arrKeys As Variant: arrKeys = keys.Keys

    For rr = 2 To lastRow
        For cc = 1 To lastCol
            On Error Resume Next
            If Not ws.Cells(rr, cc).Comment Is Nothing Then
                tt = ws.Cells(rr, cc).Comment.Text
                ttExtract = ExtractErrorMsg(tt)
                ttKey = ErrorTypeKey(ttExtract)
                ttNorm = UCase$(Trim$(ttKey))
                ttUpper = UCase$(tt)
                ttMsgUpper = UCase$(ttExtract)
                matched = False
                For Each k In arrKeys
                    If StrComp(ttNorm, CStr(k), vbTextCompare) = 0 _
                       Or InStr(1, ttUpper, CStr(k), vbTextCompare) > 0 _
                       Or InStr(1, ttMsgUpper, CStr(k), vbTextCompare) > 0 Then
                        matched = True
                        Exit For
                    End If
                Next k
                If matched Then
                    ws.Cells(rr, cc).Comment.Delete
                    ws.Cells(rr, cc).Interior.ColorIndex = xlNone
                End If
            End If
            On Error GoTo 0
        Next cc
    Next rr
    RebuildErrorCount
End Sub

' ---------------- Normalize view & formatting ----------------
Public Sub NormalizeView()
    On Error Resume Next
    Application.ScreenUpdating = True
    Application.DisplayFullScreen = False
    Application.EnableEvents = True
    Worksheets("Data").Visible = xlSheetVisible
    Worksheets("Data").Activate
    Worksheets("Data").Cells.EntireRow.Hidden = False
    Worksheets("Data").Cells.EntireColumn.Hidden = False
    With ActiveWindow
        .DisplayGridlines = True
        .DisplayHeadings = True
        .FreezePanes = False
        .SplitRow = 0
        .SplitColumn = 0
        .View = xlNormalView
        .Zoom = 100
        .ScrollRow = 1
        .ScrollColumn = 1
    End With
    Worksheets("Data").Range("A1").Select
    ' Ensure context menu is available on this sheet
    ContextMenu_Add
    ' Auto-fit columns/rows and left-justify on all sheets
    AutoFitAllSheets
End Sub
Public Sub AutoFitAllSheets()
    ' Autofit only the used range (not the entire 1M-row sheet) — dramatically faster.
    ' Row autofit is skipped; column widths are all that matter for readability.
    On Error Resume Next
    Dim ws As Worksheet
    For Each ws In Worksheets
        ws.Cells.HorizontalAlignment = xlLeft
        ws.UsedRange.EntireColumn.AutoFit
    Next ws
    On Error GoTo 0
End Sub
Public Sub FormatNEZ8dp()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cN As Long: cN = ColOf("Northing")
    Dim cE As Long: cE = ColOf("Easting")
    Dim cZ As Long: cZ = ColOf("Elevation")
    On Error Resume Next
    If cN > 0 Then ws.Columns(cN).NumberFormat = "0.00000000"
    If cE > 0 Then ws.Columns(cE).NumberFormat = "0.00000000"
    If cZ > 0 Then ws.Columns(cZ).NumberFormat = "0.00000000"
    On Error GoTo 0
End Sub

' ---------------- Data Validation (lists) ----------------
Public Sub ApplyDVForRow(ByVal r As Long, Optional ByVal fcVal As String = "")
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub
    Dim lastC As Long: lastC = LastHeaderCol()
    Dim fc As String
    If Len(fcVal) > 0 Then fc = fcVal Else fc = Trim$(ws.Cells(r, cFC).Value)
    Dim c As Long, idx As Long, atype As String
    Dim allowed As Variant, csv As String
    Dim i As Long, s As String, cnt As Long
    Dim mustUseRange As Boolean
    For c = cFC + 1 To lastC
        idx = c - cFC
        On Error Resume Next
        ws.Cells(r, c).Validation.Delete
        On Error GoTo 0
        atype = UCase$(AttrTypeForFCIndex(fc, idx))
        If atype Like "*LIST*" Then
            allowed = AllowedItemsForFCIndex(fc, idx)
            If HasItems(allowed) Then
                csv = "": mustUseRange = False
                On Error Resume Next
                For i = LBound(allowed) To UBound(allowed)
                    s = Trim$(CStr(allowed(i)))
                    If InStr(1, s, ",") > 0 Or InStr(1, s, vbLf) > 0 Or InStr(1, s, vbCr) > 0 Then mustUseRange = True: Exit For
                    If csv <> "" Then
                        If Len(csv) + 1 + Len(s) > 255 Then mustUseRange = True: Exit For
                        csv = csv & "," & s
                    Else
                        If Len(s) > 255 Then mustUseRange = True: Exit For
                        csv = s
                    End If
                Next i
                On Error GoTo 0
                If Not mustUseRange And Len(csv) > 0 Then
                    ws.Cells(r, c).Validation.Add Type:=xlValidateList, AlertStyle:=xlValidAlertStop, Formula1:=csv
                Else
                    cnt = VCount(allowed)
                    If cnt > 0 Then
                        WriteAllowedToLive r, idx, allowed
                        ws.Cells(r, c).Validation.Add Type:=xlValidateList, AlertStyle:=xlValidAlertStop, Formula1:="=" & LiveBlockAddress(r, idx, cnt)
                    End If
                End If
                ws.Cells(r, c).Validation.InCellDropdown = True
                ws.Cells(r, c).Validation.IgnoreBlank = True
            End If
        End If
    Next c
End Sub
Public Sub ApplyFCValidation()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    On Error Resume Next
    ws.Range(ws.Cells(2, cFC), ws.Cells(lastRow, cFC)).Validation.Delete
    On Error GoTo 0
    Dim fws As Worksheet: Set fws = Worksheets("FXL")
    Dim lastFXL As Long: lastFXL = fws.Cells(fws.Rows.Count, 1).End(xlUp).Row
    Dim addr As String: addr = fws.Range(fws.Cells(2, 1), fws.Cells(lastFXL, 1)).Address(True, True, xlA1, True)
    With ws.Range(ws.Cells(2, cFC), ws.Cells(lastRow, cFC)).Validation
        .Add Type:=xlValidateList, AlertStyle:=xlValidAlertStop, Formula1:="=" & addr
        .InCellDropdown = True: .IgnoreBlank = True
    End With
End Sub
Private Function EnsureLiveSheet() As Worksheet
    Dim ws As Worksheet
    On Error Resume Next: Set ws = Worksheets("FXL_LIVE"): On Error GoTo 0
    If ws Is Nothing Then
        Set ws = Worksheets.Add(After:=Worksheets(Worksheets.Count))
        ws.Name = "FXL_LIVE": ws.Visible = xlSheetVeryHidden
    End If
    Set EnsureLiveSheet = ws
End Function
Private Sub WriteAllowedToLive(ByVal r As Long, ByVal idx As Long, ByVal items As Variant)
    Dim wsL As Worksheet: Set wsL = EnsureLiveSheet()
    Const BLOCK As Long = 100
    Dim startCol As Long: startCol = 2 + (idx - 1) * BLOCK
    Dim i As Long
    wsL.Range(wsL.Cells(r, startCol), wsL.Cells(r, startCol + BLOCK - 1)).ClearContents
    On Error Resume Next
    For i = LBound(items) To UBound(items): wsL.Cells(r, startCol + (i - LBound(items))).Value = Trim$(CStr(items(i))): Next i
    On Error GoTo 0
End Sub
Private Function LiveBlockAddress(ByVal r As Long, ByVal idx As Long, ByVal count As Long) As String
    Dim wsL As Worksheet: Set wsL = EnsureLiveSheet()
    Const BLOCK As Long = 100
    Dim startCol As Long: startCol = 2 + (idx - 1) * BLOCK
    Dim endCol As Long: endCol = startCol + IIf(count <= 0, 0, count - 1)
    Dim rng As Range: Set rng = wsL.Range(wsL.Cells(r, startCol), wsL.Cells(r, endCol))
    LiveBlockAddress = rng.Address(True, True, xlA1, True)
End Function

' ---------------- Per-row validation & bend rules ----------------
Public Sub ValidateRow(ByVal r As Long)
    If r < 2 Then Exit Sub
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    ClearRowFlags r
    Dim cPN As Long, cN As Long, cE As Long, cZ As Long, cFC As Long, cStation As Long
    cPN = ColOf("Point Number"): cN = ColOf("Northing"): cE = ColOf("Easting"): cZ = ColOf("Elevation"): cFC = ColOf("Field Code"): cStation = ColOf("Station")
    ' Bulk-read the entire row in one COM call — much faster than per-cell reads.
    ' rowArr(1, c) gives the value at column c for row r.
    Dim lastCol As Long: lastCol = LastHeaderCol()
    Dim rowArr As Variant
    rowArr = ws.Range(ws.Cells(r, 1), ws.Cells(r, lastCol)).Value
    Dim valN As String:  valN = Trim$(CStr(rowArr(1, cN)))
    Dim valE As String:  valE = Trim$(CStr(rowArr(1, cE)))
    Dim valZ As String:  valZ = Trim$(CStr(rowArr(1, cZ)))
    Dim valFC As String: valFC = Trim$(CStr(rowArr(1, cFC)))
    If cStation > 0 Then
        Dim valSt As String: valSt = Trim$(CStr(rowArr(1, cStation)))
        If valSt <> "" And Not StationPatternOK(valSt) Then FlagTeal r, cStation, "Station must look like 0+00 or 0+00.00"
    End If
    If Not IsNumeric(valN) Then FlagRed r, cN, "Northing must be numeric"
    If Not IsNumeric(valE) Then FlagRed r, cE, "Easting must be numeric"
    If Not (valZ = "" Or IsNumeric(valZ)) Then FlagRed r, cZ, "Elevation must be numeric or blank"
    If valFC <> "" And Not FCExists(valFC) Then FlagOrange r, cFC, "Field Code Not in FXL"
    ' Field Code must be all caps
    If valFC <> "" And Not IsAllCaps(valFC) Then FlagRedStrong r, cFC, "Not All Caps"
    ' Apply list validations first (pass valFC so it avoids re-reading the cell)
    ApplyDVForRow r, valFC

    Dim c As Long, valAttr As String, idx As Long
    Dim atype As String, allowed As Variant, inv As String
    Dim nm As String, nmNorm As String, isOptional As Boolean
    Dim isCommentish As Boolean, skipCaps As Boolean, zeroAllowedHere As Boolean
    Dim didListErr As Boolean
    For c = cFC + 1 To lastCol
        didListErr = False
        idx = c - cFC
        nm = AttrNameForFCIndex(valFC, idx)
        nmNorm = NormName(nm)
        isOptional = AttrIsOptionalForFCIndex(valFC, idx)
        isCommentish = (InStr(1, nmNorm, "COMMENT", vbTextCompare) > 0) Or (InStr(1, nmNorm, "REMARK", vbTextCompare) > 0)
        valAttr = Trim$(CStr(rowArr(1, c)))
        If nm = "" And valAttr = "" Then GoTo NextC
        atype = UCase$(AttrTypeForFCIndex(valFC, idx))
        allowed = AllowedItemsForFCIndex(valFC, idx)

        If atype Like "*LIST*" And HasItems(allowed) Then
            If valAttr = "" Then
                didListErr = True
                FlagLightGreen r, c, "Required (pick from list)"
                GoTo NextC
            ElseIf Not InArrayCI(allowed, valAttr) Then
                didListErr = True
                FlagLightGreen r, c, "Value Not in FXL List"
            Else
                skipCaps = True    ' valid list value → do not check caps
            End If
        Else
            skipCaps = False
        End If

        ' Free-text attributes (non-list, non-comment) must be all caps.
        ' Photo/media filenames are exempt — extensions like .jpg are intentionally lowercase,
        ' and the FXL may not always mark the attribute type as Photo.
        If valAttr <> "" And Not skipCaps And Not (atype Like "*PHOTO*") And Not IsMediaFilename(valAttr) Then
            If Not IsAllCaps(valAttr) Then FlagRedStrong r, c, "Not All Caps"
        End If

        ' Unknown/NA tokens; respects optionality for blank fields
        zeroAllowedHere = False
        If (atype Like "*LIST*") And HasItems(allowed) Then If InArrayCI(allowed, "0") Then zeroAllowedHere = True
        If IsHorizAngleName(nm) Or IsVertAngleName(nm) Then zeroAllowedHere = True
        If valAttr = "0" And Not zeroAllowedHere Then FlagRed r, c, "Invalid 0": GoTo NextC
        inv = InvalidToken(valAttr)
        If inv = "Invalid Blank" Then
            If (nm = "") Or isOptional Or isCommentish Then inv = ""
        End If
        If inv <> "" Then
            If (atype Like "*LIST*") And HasItems(allowed) Then
                FlagLightGreen r, c, inv
            Else
                FlagRed r, c, inv
            End If
        End If
NextC:
    Next c
    ValidateBendForRow r
    ' MTR cross-checks if available
    On Error Resume Next
    If HasMTR() Then ValidateMTRForRow r
    On Error GoTo 0
    ' Numeric bounds check (from BOUNDS sheet / config)
    CheckBoundsForRow r
End Sub

' ========= Error Count sheet =========
Private Sub LegendRow(ws As Worksheet, ByVal row As Long, ByVal name As String, _
                      ByVal desc As String, ByVal colorRGB As Long)
    ws.Cells(row, 1).Value = name
    ws.Cells(row, 2).Value = desc
    ws.Cells(row, 3).Interior.Color = colorRGB
End Sub

Public Sub EnsureErrorCountSheet()
    Dim ws As Worksheet
    On Error Resume Next
    Set ws = Worksheets("Error Count")
    On Error GoTo 0
    If ws Is Nothing Then
        Set ws = Worksheets.Add(After:=Worksheets(Worksheets.Count))
        ws.Name = "Error Count"
    End If

    With ws
        .Cells.Clear
        .Range("A1").Value = "Legend"
        .Range("A1").Font.Bold = True

        .Range("G1").Value = "Summary"
        .Range("G1").Font.Bold = True
        .Range("G3").Value = "Cells Checked"
        .Range("G4").Value = "Cells With Errors"
        .Range("G5").Value = "Cells Correct"
        .Range("G6").Value = "Percent Correct"
        .Range("H3:H5").NumberFormat = "0"
        .Range("H6").NumberFormat = "0.00%"

        LegendRow ws, 3, "Red",           "Hard errors (not all caps, invalid token, numeric required, bend rules)", RGB(255, 120, 120)
        LegendRow ws, 4, "Orange",        "Field Code not in FXL / directional conflicts",                         RGB(255, 170, 100)
        LegendRow ws, 5, "Yellow",        "Duplicates (Point Number / Station / XRAY)",                             RGB(255, 230, 150)
        LegendRow ws, 6, "Purple",        "List mismatch (not in FXL list)",                                        RGB(200, 150, 255)
        LegendRow ws, 7, "Light Green",   "Unusual value / NA-type in LIST attributes",                             RGB(200, 255, 200)
        LegendRow ws, 8, "Teal",          "Station format issues",                                                   RGB(150, 220, 220)
        LegendRow ws, 9, "Pink",          "Joint length outliers / invalid",                                        RGB(255, 170, 220)

        .Range("A11").Value = "Error Type"
        .Range("B11").Value = "Count"
        .Range("A11:B11").Font.Bold = True
        ' Header for Heats list
        .Range("D11").Value = "Heats Used"
        .Range("E11").Value = "Status"
        .Range("D11:E11").Font.Bold = True
        ' Tip row: instructions for heat-fix feature
        With .Range("D10")
            .Value = "* Edit a Missing heat number here to update all matching cells in the Data sheet."
            .Font.Italic = True
            .Font.Color = RGB(100, 100, 100)
            .Font.Size = 9
        End With
    End With
End Sub

Public Sub RebuildErrorCount()
    Dim wsD As Worksheet: Set wsD = Worksheets("Data")
    Call EnsureErrorCountSheet
    Dim wsE As Worksheet: Set wsE = Worksheets("Error Count")

    Dim lastRow As Long: lastRow = wsD.Cells(wsD.Rows.Count, 1).End(xlUp).Row
    Dim lastCol As Long: lastCol = LastHeaderCol()

    Dim dict As Object: Set dict = CreateObject("Scripting.Dictionary")

    Dim cNorth As Long: cNorth = ColOf("Northing")
    Dim cEast As Long: cEast = ColOf("Easting")
    Dim cElev As Long: cElev = ColOf("Elevation")
    Dim totalCells As Long: totalCells = 0
    Dim errorCells As Long: errorCells = 0

    Dim r As Long, c As Long, t As String, msg As String, parts() As String
    Dim cellVal As String
    Dim cellHasComment As Boolean
    For r = 2 To lastRow
        For c = 1 To lastCol
            cellHasComment = False
            If Not wsD.Cells(r, c).Comment Is Nothing Then
                cellHasComment = True
                t = wsD.Cells(r, c).Comment.Text
                parts = Split(t, " — ")
                If UBound(parts) >= 1 Then
                    msg = Trim$(parts(1))
                Else
                    msg = Trim$(t)
                End If
                If dict.Exists(msg) Then
                    dict(msg) = dict(msg) + 1
                Else
                    dict.Add msg, 1
                End If
            End If
            If Not (c = cNorth Or c = cEast Or c = cElev) Then
                cellVal = Trim$(CStr(wsD.Cells(r, c).Value))
                If cellVal <> "" Or cellHasComment Then
                    totalCells = totalCells + 1
                    If cellHasComment Then
                        errorCells = errorCells + 1
                    End If
                End If
            End If
        Next c
    Next r

    Dim correctCells As Long: correctCells = totalCells - errorCells
    wsE.Cells(3, 8).Value = totalCells
    wsE.Cells(4, 8).Value = errorCells
    wsE.Cells(5, 8).Value = correctCells
    If totalCells > 0 Then
        wsE.Cells(6, 8).Value = correctCells / totalCells
    Else
        wsE.Cells(6, 8).Value = ""
    End If

    ' Write and sort the table
    Dim startRow As Long: startRow = 12
    wsE.Rows(CStr(startRow) & ":" & CStr(wsE.Rows.Count)).ClearContents

    Dim i As Long: i = 0
    Dim k As Variant
    For Each k In dict.Keys
        wsE.Cells(startRow + i, 1).Value = CStr(k)
        wsE.Cells(startRow + i, 2).Value = dict(k)
        i = i + 1
    Next k

    If i > 0 Then
        wsE.Range(wsE.Cells(startRow, 1), wsE.Cells(startRow + i - 1, 2)) _
            .Sort Key1:=wsE.Cells(startRow, 2), Order1:=xlDescending, Header:=xlNo
    End If

    ' ---- Heats used vs MTR ----
    Dim usedHeats As Object: Set usedHeats = CreateObject("Scripting.Dictionary")
    Dim cFC As Long: cFC = ColOf("Field Code")
    Dim fc As String, heatIdx As Long, hv As String
    If cFC > 0 Then
        For r = 2 To lastRow
            fc = UCase$(Trim$(wsD.Cells(r, cFC).Value))
            If fc <> "" Then
                heatIdx = HeatAttrIndexForFC(fc)
                If heatIdx > 0 Then
                    hv = Trim$(CStr(wsD.Cells(r, cFC + heatIdx).Value))
                    If hv <> "" Then If Not usedHeats.Exists(hv) Then usedHeats.Add hv, True
                End If
            End If
        Next r
    End If

    ' Build MTR heat set if available (avoid name clash with HasMTR function)
    Dim hasMTRFlag As Boolean: hasMTRFlag = HasMTR()
    Dim mtrSet As Object: Set mtrSet = CreateObject("Scripting.Dictionary")
    If hasMTRFlag Then
        On Error Resume Next
        Dim mws As Worksheet: Set mws = Worksheets("MTR")
        On Error GoTo 0
        If Not mws Is Nothing Then
            Dim mLast As Long: mLast = mws.Cells(mws.Rows.Count, 1).End(xlUp).Row
            For r = 2 To mLast
                hv = Trim$(CStr(mws.Cells(r, 1).Value))
                If hv <> "" Then mtrSet(hv) = True
            Next r
        End If
    End If

    ' Write heats list starting at D12
    Dim rowH As Long: rowH = 12
    Dim hk As Variant
    For Each hk In usedHeats.Keys
        wsE.Cells(rowH, 4).Value = CStr(hk)
        If hasMTRFlag Then
            If mtrSet.Exists(CStr(hk)) Then
                wsE.Cells(rowH, 5).Value = "OK"
                wsE.Cells(rowH, 4).Interior.ColorIndex = xlNone
                wsE.Cells(rowH, 5).Interior.ColorIndex = xlNone
            Else
                wsE.Cells(rowH, 5).Value = "Missing"
                wsE.Cells(rowH, 4).Interior.Color = RGB(255, 170, 100)
                wsE.Cells(rowH, 5).Interior.Color = RGB(255, 170, 100)
            End If
        Else
            wsE.Cells(rowH, 5).Value = "MTR not loaded"
            wsE.Cells(rowH, 4).Interior.ColorIndex = xlNone
            wsE.Cells(rowH, 5).Interior.ColorIndex = xlNone
        End If
        rowH = rowH + 1
    Next hk

    ' Sort heats by value if any
    If rowH > 12 Then
        wsE.Range(wsE.Cells(12, 4), wsE.Cells(rowH - 1, 5)) _
            .Sort Key1:=wsE.Cells(12, 4), Order1:=xlAscending, Header:=xlNo
    End If
End Sub

Private Sub ValidateBendForRow(ByVal r As Long)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub

    Dim fc As String: fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
    If fc = "" Then Exit Sub

    ' Codes that imply a bend/fitting
    Dim bendFC As Variant
    bendFC = Array( _
        "PIPE_BEND", "PT_OF_INTERSECTION", "BURIED_FITTINGS", "BEND", "FITTING", _
        "FIELD BEND", "ELBOW", "FACTORY BEND", "BENDS", "ELBOWS", "FITTINGS", "PI", "ELL" _
    )

    ' ---- FIXED: multi-line For/If; no inline colon chain ----
    Dim isBend As Boolean, i As Long
    For i = LBound(bendFC) To UBound(bendFC)
        If InStr(1, fc, UCase$(CStr(bendFC(i))), vbTextCompare) > 0 Then
            isBend = True
            Exit For
        End If
    Next i
    If Not isBend Then Exit Sub

    Dim cFCcol As Long: cFCcol = ColOf("Field Code")

    ' Locate attribute indices by aliases
    Dim dirIdx As Long, hzIdx As Long, vtIdx As Long, rdIdx As Long
    dirIdx = AttrIndexByAliases(fc, Array("PIPE_BEND_TYPE_D", "BEND_DIRECTION", "BEND DIRECTION", "DIRECTION", "TURN_DIRECTION", _
                                          "BEND TYPE", "BENDTYPE", "TYPE", "BEND DIR", "BENDDIR"))
    hzIdx  = AttrIndexByAliases(fc, Array("HORIZ_ANGLE", "HORIZONTAL ANGLE", "HORIZ ANGLE", "HORIZONTAL_ANGLE", "H ANGLE", "HORZ ANG", "HORZ", "H"))
    vtIdx  = AttrIndexByAliases(fc, Array("VERT_ANGLE", "VERTICAL ANGLE", "VERT ANGLE", "VERTICAL_ANGLE", "V ANGLE", "VERT", "V"))
    rdIdx  = AttrIndexByAliases(fc, Array("BEND_RADIUS", "RADIUS", "BEND RADIUS", "RAD"))

    Dim cDir As Long, cHz As Long, cVt As Long, cRd As Long
    If dirIdx > 0 Then cDir = cFCcol + dirIdx
    If hzIdx  > 0 Then cHz  = cFCcol + hzIdx
    If vtIdx  > 0 Then cVt  = cFCcol + vtIdx
    If rdIdx  > 0 Then cRd  = cFCcol + rdIdx

    Dim dir As String: dir = UCase$(CellText(ws, r, cDir))
    Dim vH  As String: vH  = CellText(ws, r, cHz)
    Dim vV  As String: vV  = CellText(ws, r, cVt)
    Dim rad As String: rad = UCase$(CellText(ws, r, cRd))

    If cRd > 0 And rad <> "" Then
        If Not (rad = "LR" Or rad = "SR" Or rad = "3D" Or rad = "3R") Then
            FlagRedStrong r, cRd, "Bend Radius must be LR, SR, 3D, or 3R"
        End If
    End If

    Dim isVertical As Boolean, isHorizontal As Boolean, isCombo As Boolean
    If InStr(1, dir, "SAG", vbTextCompare) > 0 _
       Or InStr(1, dir, "OVB", vbTextCompare) > 0 _
       Or InStr(1, dir, "OVERBEND", vbTextCompare) > 0 _
       Or InStr(1, dir, "SAG_BEND", vbTextCompare) > 0 _
       Or InStr(1, dir, "OVER", vbTextCompare) > 0 _
       Or InStr(1, dir, "VERT", vbTextCompare) > 0 Then
        isVertical = True
    End If

    If InStr(1, dir, "LEFT", vbTextCompare) > 0 _
       Or InStr(1, dir, "RIGHT", vbTextCompare) > 0 _
       Or InStr(1, dir, "LFT", vbTextCompare) > 0 _
       Or InStr(1, dir, "RGT", vbTextCompare) > 0 _
       Or InStr(1, dir, "HORZ", vbTextCompare) > 0 _
       Or InStr(1, dir, "HORIZ", vbTextCompare) > 0 Then
        isHorizontal = True
    End If

    Dim vHOK As Boolean, vVOK As Boolean
    Dim vHVal As Double, vVVal As Double
    If vH <> "" And IsNumeric(vH) Then vHVal = CDbl(vH): vHOK = True
    If vV <> "" And IsNumeric(vV) Then vVVal = CDbl(vV): vVOK = True

    If Not (isVertical Or isHorizontal) Then
        If (vVOK And vVVal > 0) And ((Not vHOK) Or vHVal = 0) Then
            isVertical = True
        ElseIf (vHOK And vHVal > 0) And ((Not vVOK) Or vVVal = 0) Then
            isHorizontal = True
        ElseIf (vVOK And vVVal > 0) And (vHOK And vHVal > 0) Then
            isCombo = True
        End If
    End If

    If (isVertical And isHorizontal) Then isCombo = True

    If isCombo Then
        If cVt > 0 And Not (vVOK And vVVal > 0) Then FlagRedStrong r, cVt, "Combo Bend: Vertical angle must be > 0"
        If cHz > 0 And Not (vHOK And vHVal > 0) Then FlagRedStrong r, cHz, "Combo Bend: Horizontal angle must be > 0"
        Exit Sub
    End If

    If isVertical Then
        If cVt > 0 And Not (vVOK And vVVal > 0) Then FlagRedStrong r, cVt, "Vertical Bend: Vertical angle must be > 0"
        If cHz > 0 Then
            If vH <> "" Then
                If vHOK Then
                    If vHVal <> 0 Then FlagRedStrong r, cHz, "Vertical Bend: Horizontal angle must be 0 or blank"
                Else
                    FlagRedStrong r, cHz, "Vertical Bend: Horizontal angle must be 0 or blank"
                End If
            End If
        End If
        Exit Sub
    End If

    If isHorizontal Then
        If cHz > 0 And Not (vHOK And vHVal > 0) Then FlagRedStrong r, cHz, "Horizontal Bend: Horizontal angle must be > 0"
        If cVt > 0 Then
            If vV <> "" Then
                If vVOK Then
                    If vVVal <> 0 Then FlagRedStrong r, cVt, "Horizontal Bend: Vertical angle must be 0 or blank"
                Else
                    FlagRedStrong r, cVt, "Horizontal Bend: Vertical angle must be 0 or blank"
                End If
            End If
        End If
        Exit Sub
    End If
End Sub

'' ---------------- MTR Cross-Check ----------------
Private Function SheetExists(ByVal nm As String) As Boolean
    On Error Resume Next
    SheetExists = Not (Worksheets(nm) Is Nothing)
    If Err.Number <> 0 Then SheetExists = False: Err.Clear
    On Error GoTo 0
End Function
Public Function HasMTR() As Boolean
    HasMTR = SheetExists("MTR")
End Function
Private Function MTRColIndex(ByVal fieldNm As String) As Long
    Dim u As String: u = NormName(fieldNm)
    If u = "HEAT" Or u = "HEATNUMBER" Or u = "HEATNO" Or u = "HEAT#" Then MTRColIndex = 1: Exit Function
    If u = "MANUFACTURER" Or u = "MFR" Or u = "MFG" Then MTRColIndex = 2: Exit Function
    If u = "NOMDIAMETER" Or u = "NOMINALDIAMETER" Or u = "NOMDIAM" Or u = "NOMDIA" Then MTRColIndex = 3: Exit Function
    If u = "OUTDIAMETER" Or u = "OD" Or u = "OUTERDIAMETER" Then MTRColIndex = 4: Exit Function
    If u = "WALLTHICKNESS" Or u = "WALL" Or u = "WT" Then MTRColIndex = 5: Exit Function
    If u = "GRADE" Then MTRColIndex = 6: Exit Function
    If u = "PIPESPEC" Or u = "SPEC" Or u = "SPECIFICATION" Then MTRColIndex = 7: Exit Function
    If u = "SEAMTYPE" Or u = "SEAM" Or u = "WELDTYPE" Then MTRColIndex = 8: Exit Function
End Function
Private Function MTRRowForHeat(ByVal heat As String) As Long
    Dim ws As Worksheet: Set ws = Worksheets("MTR")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim r As Long
    Dim h As String: h = Trim$(CStr(heat))
    For r = 2 To lastRow
        If StrComp(Trim$(CStr(ws.Cells(r, 1).Value)), h, vbTextCompare) = 0 Then MTRRowForHeat = r: Exit Function
    Next r
    MTRRowForHeat = 0
End Function
Private Function MTRGetWithDiameterFallback(ByVal heat As String, ByVal mtrName As String) As String
    Dim v As String: v = MTRGet(heat, mtrName)
    If v = "" Then
        If mtrName = "NOM DIAMETER" Then v = MTRGet(heat, "OUT DIAMETER")
        ElseIf mtrName = "OUT DIAMETER" Then v = MTRGet(heat, "NOM DIAMETER")
    End If
    MTRGetWithDiameterFallback = v
End Function
Public Function MTRGet(ByVal heat As String, ByVal fieldNm As String) As String
    Dim rr As Long: rr = MTRRowForHeat(heat)
    If rr = 0 Then Exit Function
    Dim cc As Long: cc = MTRColIndex(fieldNm)
    If cc <= 0 Then Exit Function
    MTRGet = Trim$(CStr(Worksheets("MTR").Cells(rr, cc).Value))
End Function
Private Function HeatAttrIndexForFC(ByVal fc As String) As Long
    ' First try common exact aliases
    Dim idx As Long
    idx = AttrIndexByAliases(fc, Array("HEAT", "HEAT NUMBER", "HEAT_NO", "HEATNO", "HEAT#", "HEATNUMBER", "HEAT AHEAD"))
    If idx > 0 Then HeatAttrIndexForFC = idx: Exit Function
    ' Fallback: any attribute containing HEAT (excluding HEAT BACK/HEAT BEHIND)
    On Error Resume Next
    Dim key As String: key = UCase$(Trim$(fc))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim k As Variant, e As Object, nm As String, nmU As String, nn As String
    For Each k In perCode.Keys
        Set e = perCode(CStr(k))
        If Err.Number <> 0 Then Err.Clear: Set e = Nothing
        If Not e Is Nothing Then
            nm = CStr(e("name"))
            nmU = UCase$(Trim$(nm))
            If InStr(1, nmU, "HEAT", vbTextCompare) > 0 Then
                nn = NormName(nmU)
                If InStr(1, nn, "HEATBACK", vbTextCompare) = 0 And InStr(1, nn, "HEATBEHIND", vbTextCompare) = 0 Then
                    HeatAttrIndexForFC = CLng(Val(CStr(k)))
                    Exit Function
                End If
            End If
        End If
    Next k
    On Error GoTo 0
End Function
Private Function AttrAliasesToMTRName(ByVal nm As String) As String
    Dim u As String: u = NormName(nm)
    If u = "MANUFACTURER" Or u = "MFR" Or u = "MFG" Or u = "MANF" Or u = "MFGR" _
       Or u = "MANUFACTURERNAME" Or u = "MAKE" Or u = "MAKER" Or u = "MFRNAME" Or u = "MANUFACTURERSNAME" Then AttrAliasesToMTRName = "MANUFACTURER": Exit Function
    ' Heuristic: any attribute that mentions diameter/dia and size likely refers to nominal size
    If (InStr(1, u, "DIAMETER", vbTextCompare) > 0 Or InStr(1, u, "DIA", vbTextCompare) > 0) _
       And InStr(1, u, "SIZE", vbTextCompare) > 0 Then
        AttrAliasesToMTRName = "NOM DIAMETER": Exit Function
    End If
    ' Nominal diameter, including size-based aliases
    If u = "NOMDIAMETER" Or u = "NOMINALDIAMETER" Or u = "NOMDIAM" Or u = "NOMDIA" Or u = "NOMINALDIA" Or u = "NOMD" _
       Or u = "NOMINALSIZE" Or u = "NOMSIZE" Or u = "NOMINALPIPESIZE" Or u = "NOMPIPE" Or u = "NPS" _
       Or u = "NOMINALBORE" Or u = "NB" Or u = "DN" _
       Or u = "DIAMETERSIZE" Or u = "DIASIZE" Or u = "PIPESIZE" Then AttrAliasesToMTRName = "NOM DIAMETER": Exit Function
    ' Outside diameter, including outside-dia alias
    If u = "OUTDIAMETER" Or u = "OD" Or u = "OUTERDIAMETER" Or u = "OUTDIA" Or u = "OUTSIDEDIAMETER" Or u = "OUTSIDEDIA" _
       Or u = "OUTDIAM" Or u = "OUTERDIA" Or u = "ODIN" Or u = "ODMM" Then AttrAliasesToMTRName = "OUT DIAMETER": Exit Function
    If u = "WALLTHICKNESS" Or u = "WALL" Or u = "WT" Or u = "WALLTHK" Or u = "THICKNESS" Or u = "WALLTHICK" Or u = "THK" Then AttrAliasesToMTRName = "WALL THICKNESS": Exit Function
    If u = "GRADE" Or u = "MATERIALGRADE" Or u = "MATLGRADE" Or u = "PIPEGRADE" Then AttrAliasesToMTRName = "GRADE": Exit Function
    If u = "PIPESPEC" Or u = "SPEC" Or u = "SPECIFICATION" Or u = "PIPESPECIFICATION" Or u = "STANDARD" Or u = "PIPESTD" _
       Or u = "PIPESTANDARD" Or u = "STD" Or u = "SPECSTD" Then AttrAliasesToMTRName = "PIPE SPEC": Exit Function
    If u = "SEAMTYPE" Or u = "SEAM" Or u = "WELDTYPE" Or u = "WELD" Or u = "WELDSEAM" Or u = "WELDEDSEAM" Then AttrAliasesToMTRName = "SEAM TYPE": Exit Function
    ' Generic contains-based fallback using alias lists (handles names like PIPE_SEG_PIPE_GRADE_D)
    Dim fields As Variant: fields = Array("MANUFACTURER", "NOM DIAMETER", "OUT DIAMETER", "WALL THICKNESS", "GRADE", "PIPE SPEC", "SEAM TYPE")
    Dim f As Variant, aliases As Variant, i As Long, a As String
    For Each f In fields
        aliases = MTRFieldAliases(CStr(f))
        On Error Resume Next
        For i = LBound(aliases) To UBound(aliases)
            a = NormName(CStr(aliases(i)))
            ' Only use contains on reasonably specific tokens to avoid accidental matches
            If Len(a) >= 3 Then
                If InStr(1, u, a, vbTextCompare) > 0 Then AttrAliasesToMTRName = CStr(f): Exit Function
            End If
        Next i
        On Error GoTo 0
    Next f
    AttrAliasesToMTRName = ""
End Function
Private Function MTRFieldAliases(ByVal key As String) As Variant
    Dim k As String: k = UCase$(Trim$(key))
    Select Case k
        Case "MANUFACTURER": MTRFieldAliases = Array("MANUFACTURER", "MFR", "MFG", "MANF", "MFGR", "MANUFACTURER NAME", "MAKE", "MAKER", "MFR NAME", "MANUFACTURERS NAME")
        Case "NOM DIAMETER": MTRFieldAliases = Array("NOM DIAMETER", "NOMDIAMETER", "NOMINAL DIAMETER", "NOMINALDIAMETER", "NOM DIA", "NOMDIA", "NOMD", "NOMINAL SIZE", "NOM SIZE", "NOMINAL PIPE SIZE", "NOM PIPE", "NPS", "NOMINAL BORE", "NB", "DN")
        Case "OUT DIAMETER": MTRFieldAliases = Array("OUT DIAMETER", "OUTDIAMETER", "OD", "OUT DIA", "OUTER DIAMETER", "OUTERDIAMETER", "OUTSIDE DIAMETER", "OUTSIDE DIA", "OUT DIAM", "OUTER DIA", "OD IN", "OD MM")
        Case "WALL THICKNESS": MTRFieldAliases = Array("WALL THICKNESS", "WALLTHICKNESS", "WALL", "WT", "WALL THK", "WALLTHK", "THICKNESS", "THK")
        Case "GRADE": MTRFieldAliases = Array("GRADE", "PIPE GRADE", "MATERIAL GRADE", "MATL GRADE", "MATLGRADE", "PIPEGRADE")
        Case "PIPE SPEC": MTRFieldAliases = Array("PIPE SPEC", "PIPESPEC", "SPEC", "SPECIFICATION", "PIPE SPECIFICATION", "STANDARD", "PIPE STD", "PIPE STANDARD", "STD", "SPEC STD", "SPEC STD.")
        Case "SEAM TYPE": MTRFieldAliases = Array("SEAM TYPE", "SEAMTYPE", "SEAM", "WELD TYPE", "WELDTYPE", "WELD", "WELD SEAM", "WELDED SEAM")
        Case Else: MTRFieldAliases = Array(k)
    End Select
End Function
Private Function MTRAttrIndexByField(ByVal fc As String, ByVal key As String) As Long
    ' First try exact/alias match
    MTRAttrIndexByField = AttrIndexByAliases(fc, MTRFieldAliases(key))
    If MTRAttrIndexByField > 0 Then Exit Function
    ' Fallback: contains-based match on normalized names
    On Error Resume Next
    Dim code As String: code = UCase$(Trim$(fc))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(code) Then Exit Function
    Dim perCode As Object: Set perCode = map(code)
    Dim k As Variant, nm As String, e As Object
    Dim aliases As Variant, i As Long, a As String, uattr As String
    aliases = MTRFieldAliases(key)
    For Each k In perCode.Keys
        Set e = Nothing
        Err.Clear
        Set e = perCode(CStr(k))
        If Err.Number <> 0 Then Err.Clear
        If Not e Is Nothing Then nm = CStr(e("name")) Else nm = ""
        uattr = NormName(nm)
        For i = LBound(aliases) To UBound(aliases)
            a = NormName(CStr(aliases(i)))
            If Len(a) >= 3 Then
                If InStr(1, uattr, a, vbTextCompare) > 0 Then MTRAttrIndexByField = CLng(Val(CStr(k))): Exit Function
            End If
        Next i
    Next k
    On Error GoTo 0
End Function
Private Function CleanNumber(ByVal s As String) As Double
    Dim i As Long, ch As String, t As String
    For i = 1 To Len(s)
        ch = Mid$(s, i, 1)
        If (ch >= "0" And ch <= "9") Or ch = "." Then t = t & ch
    Next i
    If Len(Trim$(t)) = 0 Then
        CleanNumber = 0
    Else
        CleanNumber = Val(t)
    End If
End Function
Private Function IsWholeNumber(ByVal d As Double) As Boolean
    IsWholeNumber = (Abs(d - Round(d, 0)) < 0.0005)
End Function
Private Function FindSizeLikeAttrIndex(ByVal fc As String) As Long
    On Error Resume Next
    Dim key As String: key = UCase$(Trim$(fc))
    Dim map As Object: Set map = GetFXLIndex()
    If Not map.Exists(key) Then Exit Function
    Dim perCode As Object: Set perCode = map(key)
    Dim k As Variant, e As Object, nm As String, nn As String
    For Each k In perCode.Keys
        Set e = perCode(CStr(k))
        If Err.Number <> 0 Then Err.Clear: Set e = Nothing
        If Not e Is Nothing Then
            nm = CStr(e("name"))
            nn = NormName(nm)
            If InStr(1, nn, "DIAMETER", vbTextCompare) > 0 Or _
               InStr(1, nn, "DIA", vbTextCompare) > 0 Or _
               InStr(1, nn, "SIZE", vbTextCompare) > 0 Then
                FindSizeLikeAttrIndex = CLng(Val(CStr(k)))
                Exit Function
            End If
        End If
    Next k
    On Error GoTo 0
End Function
Public Sub ValidateMTRForRow(ByVal r As Long)
    If r < 2 Then Exit Sub
    If Not HasMTR() Then Exit Sub
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub
    Dim fc As String: fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
    If fc = "" Then Exit Sub
    Dim heatIdx As Long: heatIdx = HeatAttrIndexForFC(fc)
    If heatIdx <= 0 Then Exit Sub
    Dim heatVal As String: heatVal = Trim$(CStr(ws.Cells(r, cFC + heatIdx).Value))
    If heatVal = "" Then Exit Sub
    ' Fallback: if neither nominal nor outside diameter attributes are defined for this FC,
    ' locate a generic size/diameter attribute and compare its number to the appropriate MTR value
    Dim idxNom As Long: idxNom = MTRAttrIndexByField(fc, "NOM DIAMETER")
    Dim idxOD As Long: idxOD = MTRAttrIndexByField(fc, "OUT DIAMETER")
    If idxNom <= 0 And idxOD <= 0 Then
        Dim idxSize As Long: idxSize = FindSizeLikeAttrIndex(fc)
        If idxSize > 0 Then
            Dim cAny As Long: cAny = cFC + idxSize
            Dim sv As String: sv = Trim$(CStr(ws.Cells(r, cAny).Value))
            If sv <> "" Then
                Dim dv As Double: dv = CleanNumber(sv)
                Dim mtrNom As String: mtrNom = MTRGet(heatVal, "NOM DIAMETER")
                Dim mtrOD As String: mtrOD = MTRGet(heatVal, "OUT DIAMETER")
                If mtrNom <> "" And IsWholeNumber(dv) Then
                    If Round(dv, 0) <> Round(CleanNumber(mtrNom), 0) Then
                        FlagOrange r, cAny, "MTR mismatch (NOM DIAMETER) expected: " & mtrNom
                    End If
                ElseIf mtrOD <> "" Then
                    If Abs(dv - CleanNumber(mtrOD)) > 0.005 Then
                        FlagOrange r, cAny, "MTR mismatch (OUT DIAMETER) expected: " & mtrOD
                    End If
                End If
            End If
        End If
    End If
    Dim fields As Variant: fields = Array("MANUFACTURER", "WALL THICKNESS", "GRADE", "PIPE SPEC", "SEAM TYPE")
    Dim i As Long
    For i = LBound(fields) To UBound(fields)
        Dim target As String: target = CStr(fields(i))
        Dim idx As Long: idx = MTRAttrIndexByField(fc, target)
        If idx > 0 Then
            Dim nm As String: nm = AttrNameForFCIndex(fc, idx)
            Dim mtrName As String: mtrName = AttrAliasesToMTRName(nm)
            If mtrName <> "" Then
                Dim c As Long: c = cFC + idx
                Dim val As String: val = Trim$(CStr(ws.Cells(r, c).Value))
                Dim mtrVal As String: mtrVal = MTRGet(heatVal, mtrName)
                If mtrVal <> "" And val <> "" Then
                    If StrComp(NormName(val), NormName(mtrVal), vbTextCompare) <> 0 Then
                        FlagOrange r, c, "MTR mismatch (" & mtrName & ") expected: " & mtrVal
                    End If
                End If
            End If
        End If
    Next i
    ' Numeric comparison for diameters (use tolerant numeric compare instead of string)
    Dim idxD As Long, nmD As String, mtrNm As String, cD As Long, v As String, mv As String
    ' Nominal Diameter
    idxD = MTRAttrIndexByField(fc, "NOM DIAMETER")
    If idxD > 0 Then
        nmD = AttrNameForFCIndex(fc, idxD)
        mtrNm = AttrAliasesToMTRName(nmD)
        If mtrNm = "NOM DIAMETER" Then
            cD = cFC + idxD
            v = Trim$(CStr(ws.Cells(r, cD).Value))
            mv = MTRGet(heatVal, "NOM DIAMETER")
            If mv <> "" And v <> "" Then
                If Round(CleanNumber(v), 0) <> Round(CleanNumber(mv), 0) Then
                    FlagOrange r, cD, "MTR mismatch (NOM DIAMETER) expected: " & mv
                End If
            End If
        End If
    End If
    ' Outside Diameter
    idxD = MTRAttrIndexByField(fc, "OUT DIAMETER")
    If idxD > 0 Then
        nmD = AttrNameForFCIndex(fc, idxD)
        mtrNm = AttrAliasesToMTRName(nmD)
        If mtrNm = "OUT DIAMETER" Then
            cD = cFC + idxD
            v = Trim$(CStr(ws.Cells(r, cD).Value))
            mv = MTRGet(heatVal, "OUT DIAMETER")
            If mv <> "" And v <> "" Then
                If Abs(CleanNumber(v) - CleanNumber(mv)) > 0.005 Then
                    FlagOrange r, cD, "MTR mismatch (OUT DIAMETER) expected: " & mv
                End If
            End If
        End If
    End If
End Sub
Public Sub ApplyMTRValueCell()
    If Not HasMTR() Then Exit Sub
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim r As Long, c As Long
    On Error Resume Next: r = Selection.Row: c = Selection.Column: On Error GoTo 0
    If r < 2 Then Exit Sub
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Or c <= cFC Then Exit Sub
    Dim fc As String: fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
    Dim idx As Long: idx = c - cFC
    Dim heatIdx As Long: heatIdx = HeatAttrIndexForFC(fc)
    If heatIdx <= 0 Then Exit Sub
    Dim heatVal As String: heatVal = Trim$(CStr(ws.Cells(r, cFC + heatIdx).Value))
    If heatVal = "" Then Exit Sub
    Dim nm As String: nm = AttrNameForFCIndex(fc, idx)
    Dim mtrName As String: mtrName = AttrAliasesToMTRName(nm)
    Dim mtrVal As String
    ' If we couldn't map the attribute directly, try a size/diameter heuristic
    If mtrName = "" Then
        Dim nu As String: nu = NormName(nm)
        If (InStr(1, nu, "DIAMETER", vbTextCompare) > 0 Or InStr(1, nu, "DIA", vbTextCompare) > 0 Or InStr(1, nu, "SIZE", vbTextCompare) > 0) Then
            mtrVal = MTRGetWithDiameterFallback(heatVal, "NOM DIAMETER")
        End If
    Else
        mtrVal = MTRGetWithDiameterFallback(heatVal, mtrName)
    End If
    If mtrVal = "" Then Exit Sub
    ws.Cells(r, c).Value = mtrVal
    RecalcFormatForEdit r, c
    ' Clear any prior MTR mismatch flag on this cell now that it matches
    ClearMarkIfContains ws, r, c, "MTR mismatch"
End Sub
Public Sub ApplyMTRValueAttrAll()
    If Not HasMTR() Then Exit Sub
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim r0 As Long, c0 As Long
    On Error Resume Next: r0 = Selection.Row: c0 = Selection.Column: On Error GoTo 0
    If r0 < 2 Then Exit Sub
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Or c0 <= cFC Then Exit Sub
    Dim fc0 As String: fc0 = UCase$(Trim$(ws.Cells(r0, cFC).Value))
    Dim idx0 As Long: idx0 = c0 - cFC
    Dim nm0 As String: nm0 = AttrNameForFCIndex(fc0, idx0)
    If nm0 = "" Then Exit Sub
    Dim targetNm As String: targetNm = nm0
    ' Determine the selected row's Heat value so we only update associated heats
    Dim heatIdx0 As Long: heatIdx0 = HeatAttrIndexForFC(fc0)
    If heatIdx0 <= 0 Then Exit Sub
    Dim selHeat As String: selHeat = Trim$(CStr(ws.Cells(r0, cFC + heatIdx0).Value))
    If selHeat = "" Then Exit Sub
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim rr As Long, fc As String, heatIdx As Long, aIdx As Long
    Dim heatVal As String, mtrName As String, mtrVal As String, cc As Long
    mtrName = AttrAliasesToMTRName(targetNm)
    If mtrName = "" Then
        Dim targetNu As String: targetNu = NormName(targetNm)
        If (InStr(1, targetNu, "DIAMETER", vbTextCompare) > 0 Or InStr(1, targetNu, "DIA", vbTextCompare) > 0 Or InStr(1, targetNu, "SIZE", vbTextCompare) > 0) Then
            mtrName = "NOM DIAMETER"
        End If
    End If
    If mtrName = "" Then Exit Sub
    For rr = 2 To lastRow
        fc = UCase$(Trim$(ws.Cells(rr, cFC).Value))
        If fc <> "" Then
            heatIdx = HeatAttrIndexForFC(fc)
            If heatIdx > 0 Then
                heatVal = Trim$(CStr(ws.Cells(rr, cFC + heatIdx).Value))
                If heatVal <> "" And heatVal = selHeat Then
                    aIdx = AttrIndexByAliases(fc, Array(targetNm))
                    If aIdx > 0 Then
                        cc = cFC + aIdx
                        mtrVal = MTRGetWithDiameterFallback(heatVal, mtrName)
                        If mtrVal <> "" Then
                            ws.Cells(rr, cc).Value = mtrVal
                            RecalcFormatForEdit rr, cc
                            ClearMarkIfContains ws, rr, cc, "MTR mismatch"
                        End If
                    End If
                End If
            End If
        End If
    Next rr
End Sub

' ---------------- Duplicates (PN, Station) ----------------
Public Sub RecalcDuplicates()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim cPN As Long: cPN = ColOf("Point Number")
    Dim cStation As Long: cStation = ColOf("Station")
    If cPN = 0 Then Exit Sub

    Dim dictPN As Object: Set dictPN = CreateObject("Scripting.Dictionary")
    Dim rr As Long
    For rr = 2 To lastRow
        ws.Cells(rr, cPN).Interior.ColorIndex = xlNone
        On Error Resume Next: If Not ws.Cells(rr, cPN).Comment Is Nothing Then ws.Cells(rr, cPN).Comment.Delete: On Error GoTo 0
        If cStation > 0 Then
            ws.Cells(rr, cStation).Interior.ColorIndex = xlNone
            On Error Resume Next: If Not ws.Cells(rr, cStation).Comment Is Nothing Then ws.Cells(rr, cStation).Comment.Delete: On Error GoTo 0
        End If
    Next rr

    Dim r As Long, val As String, firstr As Long
    For r = 2 To lastRow
        val = Trim$(ws.Cells(r, cPN).Value)
        If val <> "" Then
            If dictPN.Exists(val) Then
                ws.Cells(r, cPN).Interior.Color = RGB(255, 230, 150)
                firstr = dictPN(val)
                ws.Cells(firstr, cPN).Interior.Color = RGB(255, 230, 150)
                FlagYellow r, cPN, "Duplicate Point Number"
                FlagYellow firstr, cPN, "Duplicate Point Number"
            Else
                dictPN(val) = r
            End If
        End If
    Next r

    If cStation > 0 Then
        Dim dictST2 As Object: Set dictST2 = CreateObject("Scripting.Dictionary")
        Dim firsts As Long
        For r = 2 To lastRow
            val = Trim$(ws.Cells(r, cStation).Value)
            If val <> "" Then
                If dictST2.Exists(val) Then
                    ws.Cells(r, cStation).Interior.Color = RGB(255, 230, 150)
                    firsts = dictST2(val)
                    ws.Cells(firsts, cStation).Interior.Color = RGB(255, 230, 150)
                    FlagYellow r, cStation, "Duplicate Station"
                    FlagYellow firsts, cStation, "Duplicate Station"
                Else
                    dictST2(val) = r
                End If
            End If
        Next r
    End If
End Sub

Public Sub RecalcColumnDuplicates(ByVal c As Long)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim hdr As String: hdr = Trim$(CStr(ws.Cells(1, c).Value))

    Dim rr As Long

    ' ---- HARD CLEAR of previous duplicate marks on this column ----
    With ws
        ' clear any duplicate-colored fills on the column (safe for PN/Station)
        .Range(.Cells(2, c), .Cells(lastRow, c)).Interior.ColorIndex = xlNone

        ' remove duplicate comments
        For rr = 2 To lastRow
            On Error Resume Next
            If Not .Cells(rr, c).Comment Is Nothing Then
                If InStr(1, .Cells(rr, c).Comment.Text, "Duplicate ", vbTextCompare) > 0 Then
                    .Cells(rr, c).Comment.Delete
                End If
            End If
            On Error GoTo 0
        Next rr
    End With

    ' ---- Recompute duplicates ----
    Dim dict As Object: Set dict = CreateObject("Scripting.Dictionary")
    Dim v As String, firstR As Long
    For rr = 2 To lastRow
        v = Trim$(CStr(ws.Cells(rr, c).Value))
        If v <> "" Then
            If dict.Exists(v) Then
                firstR = dict(v)
                FlagYellow rr, c, "Duplicate " & hdr
                FlagYellow firstR, c, "Duplicate " & hdr
            Else
                dict.Add v, rr
            End If
        End If
    Next rr
End Sub

' ---------------- XRAY duplicate detection ----------------
Private Function IsXRAYAttrName(ByVal nm As String) As Boolean
    Dim u As String
    u = NormName(nm)                 ' removes spaces, dashes, underscores, dots, #
    IsXRAYAttrName = (InStr(1, u, "XRAY", vbTextCompare) > 0)
End Function
' Back-compat alias (some places may call IsXRAYName):
Private Function IsXRAYName(ByVal nm As String) As Boolean
    IsXRAYName = IsXRAYAttrName(nm)
End Function
Private Sub ClearXRAYMark(ws As Worksheet, ByVal r As Long, ByVal c As Long)
    On Error Resume Next
    If Not ws.Cells(r, c).Comment Is Nothing Then
        If InStr(1, ws.Cells(r, c).Comment.Text, "XRAY", vbTextCompare) > 0 Then
            ws.Cells(r, c).Comment.Delete
            ws.Cells(r, c).Interior.ColorIndex = xlNone
        End If
    End If
    If ws.Cells(r, c).Interior.Color = RGB(255, 230, 150) Then ws.Cells(r, c).Interior.ColorIndex = xlNone
    On Error GoTo 0
End Sub
Public Sub RecalcXRAYDupForEdit(ByVal r As Long, ByVal c As Long)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub

    Dim fc As String: fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
    If c > cFC Then
        Dim idx As Long: idx = c - cFC
        Dim nm As String: nm = AttrNameForFCIndex(fc, idx)
        If nm <> "" And IsXRAYAttrName(nm) Then
            RecalcXRAYGroup NormName(nm)  ' same-attribute duplicates
            Exit Sub
        End If
    End If

    Dim cN As Long: cN = ColOf("Northing")
    Dim cE As Long: cE = ColOf("Easting")
    If c = cN Or c = cE Or c > cFC Then
        RecalcXRAYByXY
    End If
End Sub
Private Sub RecalcXRAYGroup(ByVal groupNorm As String)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim cFC As Long: cFC = ColOf("Field Code")
    Dim lastC As Long: lastC = LastHeaderCol()
    If cFC = 0 Then Exit Sub
    Dim r As Long, i As Long, cc As Long, nm As String, v As String

    For r = 2 To lastRow
        For i = 1 To lastC - cFC
            nm = AttrNameForFCIndex(UCase$(Trim$(ws.Cells(r, cFC).Value)), i)
            If nm <> "" And NormName(nm) = groupNorm Then
                cc = cFC + i
                ClearXRAYMark ws, r, cc
            End If
        Next i
    Next r

    Dim dict As Object: Set dict = CreateObject("Scripting.Dictionary")
    Dim firstR As Long, firstC As Long, key As String
    For r = 2 To lastRow
        For i = 1 To lastC - cFC
            nm = AttrNameForFCIndex(UCase$(Trim$(ws.Cells(r, cFC).Value)), i)
            If nm <> "" And NormName(nm) = groupNorm Then
                cc = cFC + i
                v = UCase$(Trim$(ws.Cells(r, cc).Value))
                If v <> "" Then
                    key = v
                    If dict.Exists(key) Then
                        firstR = dict(key)(0): firstC = dict(key)(1)
                        FlagYellow r, cc, "Duplicate XRAY (" & nm & ")"
                        FlagYellow firstR, firstC, "Duplicate XRAY (" & nm & ")"
                    Else
                        dict.Add key, Array(r, cc)
                    End If
                End If
            End If
        Next i
    Next r
End Sub
Private Function XYKeyFromRow(ByVal r As Long, ByVal cN As Long, ByVal cE As Long) As String
    Dim n As String: n = Trim$(Worksheets("Data").Cells(r, cN).Value)
    Dim e As String: e = Trim$(Worksheets("Data").Cells(r, cE).Value)
    If Not IsNumeric(n) Or Not IsNumeric(e) Then XYKeyFromRow = "": Exit Function
    Dim rn As Double, re As Double
    rn = Application.WorksheetFunction.Round(CDbl(n), 2)
    re = Application.WorksheetFunction.Round(CDbl(e), 2)
    XYKeyFromRow = CStr(rn) & "|" & CStr(re)
End Function
Public Sub RecalcXRAYByXY()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim cFC As Long: cFC = ColOf("Field Code")
    Dim cN As Long:  cN  = ColOf("Northing")
    Dim cE As Long:  cE  = ColOf("Easting")
    Dim lastC As Long: lastC = LastHeaderCol()
    If cFC = 0 Or cN = 0 Or cE = 0 Then Exit Sub

    Dim r As Long, i As Long, idIdx As Long, fc As String, nm As String, idCol As Long

    ' Clear previous XY marks
    For r = 2 To lastRow
        ClearXRAYMark ws, r, cN
        ClearXRAYMark ws, r, cE
    Next r

    ' Map row -> column of first XRAY-named attribute
    Dim idOfRow As Object: Set idOfRow = CreateObject("Scripting.Dictionary")
    For r = 2 To lastRow
        fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
        If InStr(1, fc, "XRAY", vbTextCompare) > 0 Then
            idIdx = 0
            For i = 1 To lastC - cFC
                nm = AttrNameForFCIndex(fc, i)
                If nm <> "" And IsXRAYAttrName(nm) Then
                    idIdx = i
                    Exit For
                End If
            Next i
            If idIdx > 0 Then idOfRow(r) = cFC + idIdx
        End If
    Next r

    ' XRAY rows with BLANK ID → duplicates by XY(2dp)
    Dim dict As Object: Set dict = CreateObject("Scripting.Dictionary")
    Dim key As String, firstR As Long
    For r = 2 To lastRow
        If idOfRow.Exists(r) Then
            idCol = idOfRow(r)
            If Trim$(ws.Cells(r, idCol).Value) = "" Then
                key = XYKeyFromRow(r, cN, cE)
                If key <> "" Then
                    If dict.Exists(key) Then
                        firstR = dict(key)
                        FlagYellow r, cN, "Duplicate XRAY (by XY)"
                        FlagYellow r, cE, "Duplicate XRAY (by XY)"
                        FlagYellow firstR, cN, "Duplicate XRAY (by XY)"
                        FlagYellow firstR, cE, "Duplicate XRAY (by XY)"
                    Else
                        dict.Add key, r
                    End If
                End If
            End If
        End If
    Next r
End Sub
Public Sub RecalcXRAYDuplicates()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim lastC As Long:  lastC  = LastHeaderCol()
    Dim cFC As Long:    cFC    = ColOf("Field Code")
    If cFC = 0 Then Exit Sub

    Dim r As Long, ii As Long, fc As String, nm As String, v As String, cc As Long
    For r = 2 To lastRow
        fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
        For ii = 1 To lastC - cFC
            nm = AttrNameForFCIndex(fc, ii)
            If nm <> "" And IsXRAYAttrName(nm) Then
                cc = cFC + ii
                ClearXRAYMark ws, r, cc
            End If
        Next ii
    Next r

    Dim dict As Object: Set dict = CreateObject("Scripting.Dictionary")
    Dim dictVals As Object, key As String, valKey As String
    Dim firstR As Long, firstC As Long
    For r = 2 To lastRow
        fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
        For ii = 1 To lastC - cFC
            nm = AttrNameForFCIndex(fc, ii)
            If nm <> "" And IsXRAYAttrName(nm) Then
                cc = cFC + ii
                v = UCase$(Trim$(ws.Cells(r, cc).Value))
                key = NormName(nm)
                If Not dict.Exists(key) Then Set dict(key) = CreateObject("Scripting.Dictionary")
                Set dictVals = dict(key)
                If v <> "" Then
                    valKey = v
                    If dictVals.Exists(valKey) Then
                        firstR = dictVals(valKey)(0): firstC = dictVals(valKey)(1)
                        FlagYellow r, cc, "Duplicate XRAY (" & nm & ")"
                        FlagYellow firstR, firstC, "Duplicate XRAY (" & nm & ")"
                    Else
                        dictVals.Add valKey, Array(r, cc)
                    End If
                End If
            End If
        Next ii
    Next r
End Sub

' ---------------- Attribute format consistency ----------------
Public Function XRAYPatternOf(ByVal s As String) As String
    Dim i As Long, ch As String, out As String
    s = Trim$(CStr(s))
    For i = 1 To Len(s)
        ch = Mid$(s, i, 1)
        If ch Like "[0-9]" Then
            out = out & "9"
        ElseIf ch Like "[A-Za-z]" Then
            out = out & "A"
        Else
            out = out & ch
        End If
    Next i
    XRAYPatternOf = out
End Function
Private Function NormTokenStrict(ByVal s As String) As String
    s = UCase$(Trim$(s))
    s = Replace(s, " ", ""): s = Replace(s, "-", ""): s = Replace(s, "_", "")
    s = Replace(s, ".", ""): s = Replace(s, "#", "")
    NormTokenStrict = s
End Function
Private Function IsJointLengthName(ByVal nm As String) As Boolean
    Dim u As String: u = NormName(nm)
    If InStr(u, "JOINTLENGTH") > 0 Then IsJointLengthName = True: Exit Function
    If InStr(u, "JOINTLEN") > 0 Then IsJointLengthName = True: Exit Function
    If u = "JOINTLT" Or u = "JNTLENGTH" Then IsJointLengthName = True: Exit Function
    If InStr(u, "JOINT") > 0 And (InStr(u, "LENGTH") > 0 Or InStr(u, "LEN") > 0 Or Right$(u, 2) = "LT") Then IsJointLengthName = True: Exit Function
End Function
Private Function CanonicalAttrGroupName(ByVal nm As String) As String
    Dim u As String: u = NormName(nm)
    If u = "" Then CanonicalAttrGroupName = "": Exit Function
    If IsJointLengthName(nm) Then CanonicalAttrGroupName = "": Exit Function
    If InStr(u, "JOINT") > 0 Or InStr(u, "TALLY") > 0 Then CanonicalAttrGroupName = "JOINT_ID": Exit Function
    If InStr(u, "HEAT") > 0 Then CanonicalAttrGroupName = "HEAT_ID": Exit Function
    CanonicalAttrGroupName = u
End Function
Private Sub ClearFormatFlagIfAny(ws As Worksheet, ByVal r As Long, ByVal c As Long)
    On Error Resume Next
    If Not ws.Cells(r, c).Comment Is Nothing Then
        If InStr(1, ws.Cells(r, c).Comment.Text, "Format inconsistent", vbTextCompare) > 0 _
           Or InStr(1, ws.Cells(r, c).Comment.Text, "Unusual value", vbTextCompare) > 0 Then
            ws.Cells(r, c).Comment.Delete
            If ws.Cells(r, c).Interior.Color = RGB(170, 200, 255) Then ws.Cells(r, c).Interior.ColorIndex = xlNone
        End If
    End If
    On Error GoTo 0
End Sub
Private Sub RecalcFormatGroup(ByVal groupKey As String)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim lastC As Long:  lastC  = LastHeaderCol()

    Dim occ As New Collection
    Dim shapeCount As Object: Set shapeCount = CreateObject("Scripting.Dictionary")
    Dim tokenCount As Object: Set tokenCount = CreateObject("Scripting.Dictionary")

    Dim r As Long, i As Long, fc As String, nm As String, v As String
    Dim shape As String, tok As String, cAttr As Long, gk As String

    For r = 2 To lastRow
        fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
        If fc <> "" Then
            For i = 1 To lastC - cFC
                nm = AttrNameForFCIndex(fc, i)
                If nm <> "" Then
                    gk = CanonicalAttrGroupName(nm)
                    If gk = groupKey And gk <> "" Then
                        cAttr = cFC + i
                        ClearFormatFlagIfAny ws, r, cAttr
                        v = Trim$(ws.Cells(r, cAttr).Value)
                        If v <> "" Then
                            shape = XRAYPatternOf(v)
                            tok = NormTokenStrict(v)
                            occ.Add Array(r, cAttr, shape, tok)
                            If shapeCount.Exists(shape) Then shapeCount(shape) = shapeCount(shape) + 1 Else shapeCount.Add shape, 1
                            If tokenCount.Exists(tok) Then tokenCount(tok) = tokenCount(tok) + 1 Else tokenCount.Add tok, 1
                        End If
                    End If
                End If
            Next i
        End If
    Next r
    If occ.Count = 0 Then Exit Sub

    Dim bestShape As String: bestShape = "": Dim bestShapeN As Long: bestShapeN = -1
    Dim k As Variant
    For Each k In shapeCount.Keys
        If CLng(shapeCount(k)) > bestShapeN Then bestShapeN = CLng(shapeCount(k)): bestShape = CStr(k)
    Next k
    Dim useShape As Boolean: useShape = (bestShapeN >= CLng(0.6 * occ.Count))

    Dim majors As Object: Set majors = CreateObject("Scripting.Dictionary")
    Dim cover As Long: cover = 0
    For Each k In tokenCount.Keys
        If CLng(tokenCount(k)) >= 2 Then majors.Add CStr(k), True: cover = cover + CLng(tokenCount(k))
    Next k
    Dim useSingletons As Boolean: useSingletons = (cover >= CLng(0.5 * occ.Count))

    Dim idx As Long, rr As Long, cc As Long, sh As String, tk As String
    If useShape Then
        For idx = 1 To occ.Count
            rr = occ(idx)(0): cc = occ(idx)(1): sh = occ(idx)(2)
            If sh <> bestShape Then FlagRedStrong rr, cc, "Format inconsistent (expected like '" & bestShape & "')"
        Next idx
    End If
    If useSingletons Then
        For idx = 1 To occ.Count
            rr = occ(idx)(0): cc = occ(idx)(1): tk = occ(idx)(3)
If Not majors.Exists(tk) And tokenCount(tk) = 1 Then
    ' was: FlagRedStrong rr, cc, "Unusual value for this attribute"
    FlagLightGreen rr, cc, "Unusual value for this attribute"
End If
        Next idx
    End If
End Sub
Public Sub RecalcFormatForEdit(ByVal r As Long, ByVal c As Long)
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Or c <= cFC Then Exit Sub
    Dim fc As String: fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
    Dim idx As Long: idx = c - cFC
    Dim nm As String: nm = AttrNameForFCIndex(fc, idx)
    If nm = "" Then Exit Sub
    Dim atype As String: atype = UCase$(AttrTypeForFCIndex(fc, idx))
    If atype Like "*NUMBER*" Then Exit Sub
    Dim allowed As Variant: allowed = AllowedItemsForFCIndex(fc, idx)
    If HasItems(allowed) Then Exit Sub
    Dim gkey As String: gkey = CanonicalAttrGroupName(nm)
    If gkey <> "" Then RecalcFormatGroup gkey
End Sub
Public Sub RecalcAttributeFormatConsistency()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim lastC As Long:  lastC  = LastHeaderCol()
    Dim groups As Object: Set groups = CreateObject("Scripting.Dictionary")
    Dim r As Long, i As Long, fc As String, nm As String
    Dim atype As String, allowed As Variant, gk As String
    For r = 2 To lastRow
        fc = UCase$(Trim$(ws.Cells(r, cFC).Value))
        If fc <> "" Then
            For i = 1 To lastC - cFC
                nm = AttrNameForFCIndex(fc, i)
                If nm <> "" Then
                    atype = UCase$(AttrTypeForFCIndex(fc, i))
                    If Not (atype Like "*NUMBER*") Then
                        allowed = AllowedItemsForFCIndex(fc, i)
                        If Not HasItems(allowed) Then
                            gk = CanonicalAttrGroupName(nm)
                            If gk <> "" Then groups(gk) = True
                        End If
                    End If
                End If
            Next i
        End If
    Next r
    Dim g As Variant
    For Each g In groups.Keys
        RecalcFormatGroup CStr(g)
    Next g
End Sub

' ---------------- Joint length outliers ----------------
Public Sub RecalcJointLengthOutliers()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim cFC As Long: cFC = ColOf("Field Code")
    If cFC = 0 Then Exit Sub
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim lastC As Long:  lastC  = LastHeaderCol()
    ' Use a Collection to avoid O(n^2) ReDim Preserve in a loop
    Dim occ As New Collection
    Dim r As Long, c As Long, idx As Long
    Dim fc As String, nm As String, v As String
    For r = 2 To lastRow
        fc = Trim$(ws.Cells(r, cFC).Value)
        If fc <> "" Then
            For c = cFC + 1 To lastC
                idx = c - cFC
                nm = AttrNameForFCIndex(fc, idx)
                If NormName(nm) Like "*JOINTLENGTH*" Or NormName(nm) = "JLENGTH" Or Right$(NormName(nm), 2) = "LT" Then
                    v = Trim$(ws.Cells(r, c).Value)
                    If v <> "" Then
                        If IsNumeric(v) Then
                            occ.Add Array(CDbl(v), r, c)
                        Else
                            FlagPink r, c, "Invalid Joint Length"
                        End If
                    End If
                End If
            Next c
        End If
    Next r
    Dim cnt As Long: cnt = occ.Count
    If cnt < 3 Then Exit Sub
    ' Copy into arrays for arithmetic
    Dim vals() As Double, rowIx() As Long, colIx() As Long
    ReDim vals(cnt - 1): ReDim rowIx(cnt - 1): ReDim colIx(cnt - 1)
    Dim i As Long
    For i = 1 To cnt
        vals(i - 1) = occ(i)(0)
        rowIx(i - 1) = occ(i)(1)
        colIx(i - 1) = occ(i)(2)
    Next i
    Dim mu As Double, sd As Double, varsum As Double
    For i = 0 To cnt - 1: mu = mu + vals(i): Next i
    mu = mu / cnt
    For i = 0 To cnt - 1: varsum = varsum + (vals(i) - mu) * (vals(i) - mu): Next i
    sd = Sqr(varsum / (cnt - 1))
    If sd = 0 Then Exit Sub
    For i = 0 To cnt - 1
        If Abs(vals(i) - mu) > 2 * sd Then
            FlagPink rowIx(i), colIx(i), "Inconsistent Joint Length"
        End If
    Next i
End Sub

' ---------------- Full-sheet validation ----------------
Public Sub ValidateSheetAll()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastRow As Long: lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim r As Long

    ' Performance guard: turn off expensive Excel features during bulk validation
    Dim prevCalc As Long
    Dim prevSU As Boolean, prevEvt As Boolean, prevDisp As Boolean
    On Error Resume Next
    prevCalc = Application.Calculation
    prevSU = Application.ScreenUpdating
    prevEvt = Application.EnableEvents
    prevDisp = Application.DisplayAlerts
    Application.ScreenUpdating = False
    Application.EnableEvents = False
    Application.DisplayAlerts = False
    Application.Calculation = xlCalculationManual
    On Error GoTo 0

    ' FXL lookup cache builds on first use via GetFXLIndex

    ValidationModule.ApplyFCValidation
    ValidationModule.FormatNEZ8dp

    For r = 2 To lastRow
        ValidationModule.ValidateRow r
    Next r

    ValidationModule.RecalcDuplicates
    ValidationModule.RecalcXRAYDuplicates
    ValidationModule.RecalcAttributeFormatConsistency
    ValidationModule.RecalcJointLengthOutliers
    ValidationModule.CheckJXLAudit

    ' Refresh the Error Count sheet
    ValidationModule.RebuildErrorCount
    ValidationModule.ReportJXLMissingPoints
    ValidationModule.ReportMediaMissing

    ' Restore Excel settings
    On Error Resume Next
    Application.Calculation = prevCalc
    Application.ScreenUpdating = prevSU
    Application.EnableEvents = prevEvt
    Application.DisplayAlerts = prevDisp
    On Error GoTo 0
End Sub

' ---------------- Numeric bounds checking (from BOUNDS sheet) ----------------
Private Function GetBoundsCache() As Object
    Static cache As Object
    Static loaded As Boolean
    If Not loaded Then
        loaded = True
        Set cache = CreateObject("Scripting.Dictionary")
        On Error Resume Next
        Dim ws As Worksheet
        Set ws = Nothing
        Set ws = ThisWorkbook.Worksheets("BOUNDS")
        If Not ws Is Nothing Then
            Dim lr As Long: lr = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            Dim r As Long
            For r = 2 To lr
                Dim nm As String: nm = Trim$(UCase$(CStr(ws.Cells(r, 1).Value)))
                Dim mn As Double: mn = CDbl(ws.Cells(r, 2).Value)
                If Err.Number <> 0 Then Err.Clear: GoTo NextBound
                Dim mx As Double: mx = CDbl(ws.Cells(r, 3).Value)
                If Err.Number <> 0 Then Err.Clear: GoTo NextBound
                If nm <> "" Then cache(nm) = Array(mn, mx)
NextBound:
                Err.Clear
            Next r
        End If
        On Error GoTo 0
    End If
    Set GetBoundsCache = cache
End Function

Private Sub CheckBoundsForRow(ByVal r As Long)
    Dim bc As Object: Set bc = GetBoundsCache()
    If bc Is Nothing Then Exit Sub
    If bc.Count = 0 Then Exit Sub
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lastCol As Long: lastCol = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
    Dim c As Long
    For c = 1 To lastCol
        Dim hdr As String: hdr = Trim$(UCase$(CStr(ws.Cells(1, c).Value)))
        If bc.Exists(hdr) Then
            Dim cellVal As String: cellVal = Trim$(CStr(ws.Cells(r, c).Value))
            If cellVal <> "" Then
                On Error Resume Next
                Dim numVal As Double: numVal = CDbl(cellVal)
                If Err.Number = 0 Then
                    Dim bounds As Variant: bounds = bc(hdr)
                    If numVal < bounds(0) Or numVal > bounds(1) Then
                        FlagOrange r, c, "Out of range (" & bounds(0) & Chr(8211) & bounds(1) & ")"
                    End If
                End If
                On Error GoTo 0
            End If
        End If
    Next c
End Sub

' ---------------- Error row filter ----------------
Public Sub FilterErrorRows()
    Dim ws As Worksheet: Set ws = Worksheets("Data")
    Dim lr As Long: lr = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    Dim lc As Long: lc = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
    Application.ScreenUpdating = False
    Dim r As Long
    For r = 2 To lr
        Dim hasErr As Boolean: hasErr = False
        Dim c As Long
        For c = 1 To lc
            If ws.Cells(r, c).Interior.ColorIndex <> xlNone Then
                hasErr = True
                Exit For
            End If
        Next c
        ws.Rows(r).Hidden = Not hasErr
    Next r
    Application.ScreenUpdating = True
End Sub

Public Sub ShowAllRows()
    Worksheets("Data").Rows.Hidden = False
End Sub

' ---------------- VBA project protection ----------------
' ---------------- JXL Audit ----------------
Public Sub CheckJXLAudit()
    ' Read JXL_AUDIT sheet and flag deleted / not-in-JXL rows in the Data sheet
    On Error Resume Next
    Dim wsAudit As Worksheet
    Set wsAudit = Nothing
    Set wsAudit = ThisWorkbook.Worksheets("JXL_AUDIT")
    On Error GoTo 0
    If wsAudit Is Nothing Then Exit Sub

    Dim wsData As Worksheet: Set wsData = Worksheets("Data")
    Dim cPN As Long: cPN = ColOf("Point Number")
    If cPN = 0 Then Exit Sub

    ' Build issue lookup: point name -> issue type (skip NOT_IN_CSV — those are handled separately)
    Dim issues As Object: Set issues = CreateObject("Scripting.Dictionary")
    Dim lr As Long: lr = wsAudit.Cells(wsAudit.Rows.Count, 1).End(xlUp).Row
    Dim r As Long
    For r = 2 To lr
        Dim ptName As String: ptName = Trim$(CStr(wsAudit.Cells(r, 1).Value))
        Dim issueType As String: issueType = Trim$(CStr(wsAudit.Cells(r, 2).Value))
        If ptName <> "" And issueType <> "NOT_IN_CSV" Then
            issues(ptName) = issueType
        End If
    Next r

    If issues.Count = 0 Then Exit Sub

    Dim lastDataRow As Long: lastDataRow = wsData.Cells(wsData.Rows.Count, 1).End(xlUp).Row
    For r = 2 To lastDataRow
        Dim pn As String: pn = Trim$(CStr(wsData.Cells(r, cPN).Value))
        If issues.Exists(pn) Then
            Dim issue As String: issue = CStr(issues(pn))
            If issue = "DELETED_IN_FIELD" Then
                FlagRed r, cPN, "DELETED IN FIELD — Point was marked deleted in the Trimble job file."
            ElseIf issue = "NOT_IN_JXL" Then
                FlagOrange r, cPN, "NOT IN JXL — No matching point record found in the Trimble job file."
            End If
        End If
    Next r
End Sub

Public Sub ReportJXLMissingPoints()
    ' Append three sections to the Error Count sheet from JXL_AUDIT:
    '   1. Points deleted in the field
    '   2. JXL points not exported to CSV
    '   3. CSV points with no matching JXL record
    On Error Resume Next
    Dim wsAudit As Worksheet
    Set wsAudit = Nothing
    Set wsAudit = ThisWorkbook.Worksheets("JXL_AUDIT")
    On Error GoTo 0
    If wsAudit Is Nothing Then Exit Sub

    On Error Resume Next
    Dim wsE As Worksheet: Set wsE = Nothing
    Set wsE = ThisWorkbook.Worksheets("Error Count")
    On Error GoTo 0
    If wsE Is Nothing Then Exit Sub

    Dim lr As Long: lr = wsAudit.Cells(wsAudit.Rows.Count, 1).End(xlUp).Row
    Dim r As Long

    ' ── Collect all three issue types ──────────────────────────────────────────
    Dim delNames()  As String: Dim delReasons()  As String: Dim nDel  As Long: nDel  = 0
    Dim csvNames()  As String: Dim csvReasons()  As String: Dim nCSV  As Long: nCSV  = 0
    Dim jxlNames()  As String: Dim jxlReasons()  As String: Dim nJXL  As Long: nJXL  = 0

    For r = 2 To lr
        Dim ptN   As String: ptN   = Trim$(CStr(wsAudit.Cells(r, 1).Value))
        Dim issue As String: issue = Trim$(CStr(wsAudit.Cells(r, 2).Value))
        Dim rsn   As String: rsn   = Trim$(CStr(wsAudit.Cells(r, 3).Value))
        If ptN = "" Then GoTo NextRow

        Select Case issue
        Case "DELETED_IN_FIELD"
            If rsn = "" Then rsn = "Point was deleted in the field (marked deleted in Trimble job file)"
            ReDim Preserve delNames(nDel): ReDim Preserve delReasons(nDel)
            delNames(nDel) = ptN: delReasons(nDel) = rsn: nDel = nDel + 1
        Case "NOT_IN_CSV"
            If rsn = "" Then rsn = "Present in JXL fieldbook but not found in CSV"
            ReDim Preserve csvNames(nCSV): ReDim Preserve csvReasons(nCSV)
            csvNames(nCSV) = ptN: csvReasons(nCSV) = rsn: nCSV = nCSV + 1
        Case "NOT_IN_JXL"
            If rsn = "" Then rsn = "Point is in CSV but has no matching record in the JXL fieldbook"
            ReDim Preserve jxlNames(nJXL): ReDim Preserve jxlReasons(nJXL)
            jxlNames(nJXL) = ptN: jxlReasons(nJXL) = rsn: nJXL = nJXL + 1
        End Select
        NextRow:
    Next r

    If nDel = 0 And nCSV = 0 And nJXL = 0 Then Exit Sub

    Dim lastRow As Long: lastRow = wsE.Cells(wsE.Rows.Count, 1).End(xlUp).Row
    Dim i As Long

    ' Helper: write a section header + rows
    ' ── Section 1: Deleted in field ────────────────────────────────────────────
    If nDel > 0 Then
        lastRow = lastRow + 2
        wsE.Cells(lastRow, 1).Value = "POINTS DELETED IN FIELD"
        wsE.Cells(lastRow, 1).Font.Bold = True
        wsE.Cells(lastRow, 1).Font.Color = RGB(180, 0, 0)
        lastRow = lastRow + 1
        wsE.Cells(lastRow, 1).Value = "Point Name": wsE.Cells(lastRow, 1).Font.Bold = True
        wsE.Cells(lastRow, 2).Value = "Note":       wsE.Cells(lastRow, 2).Font.Bold = True
        For i = 0 To nDel - 1
            lastRow = lastRow + 1
            wsE.Cells(lastRow, 1).Value = delNames(i)
            wsE.Cells(lastRow, 2).Value = delReasons(i)
        Next i
    End If

    ' ── Section 2: JXL points not in CSV ───────────────────────────────────────
    If nCSV > 0 Then
        lastRow = lastRow + 2
        wsE.Cells(lastRow, 1).Value = "JXL POINTS NOT EXPORTED TO CSV"
        wsE.Cells(lastRow, 1).Font.Bold = True
        wsE.Cells(lastRow, 1).Font.Color = RGB(180, 60, 0)
        lastRow = lastRow + 1
        wsE.Cells(lastRow, 1).Value = "Point Name": wsE.Cells(lastRow, 1).Font.Bold = True
        wsE.Cells(lastRow, 2).Value = "Note":       wsE.Cells(lastRow, 2).Font.Bold = True
        For i = 0 To nCSV - 1
            lastRow = lastRow + 1
            wsE.Cells(lastRow, 1).Value = csvNames(i)
            wsE.Cells(lastRow, 2).Value = csvReasons(i)
        Next i
    End If

    ' ── Section 3: CSV points not in JXL ───────────────────────────────────────
    If nJXL > 0 Then
        lastRow = lastRow + 2
        wsE.Cells(lastRow, 1).Value = "CSV POINTS NOT IN JXL FIELDBOOK"
        wsE.Cells(lastRow, 1).Font.Bold = True
        wsE.Cells(lastRow, 1).Font.Color = RGB(0, 60, 140)
        lastRow = lastRow + 1
        wsE.Cells(lastRow, 1).Value = "Point Name": wsE.Cells(lastRow, 1).Font.Bold = True
        wsE.Cells(lastRow, 2).Value = "Note":       wsE.Cells(lastRow, 2).Font.Bold = True
        For i = 0 To nJXL - 1
            lastRow = lastRow + 1
            wsE.Cells(lastRow, 1).Value = jxlNames(i)
            wsE.Cells(lastRow, 2).Value = jxlReasons(i)
        Next i
    End If
End Sub

Public Sub ReportMediaMissing()
    ' Append a MISSING MEDIA FILES section to Error Count from the hidden MEDIA_AUDIT sheet.
    On Error Resume Next
    Dim wsM As Worksheet
    Set wsM = Nothing
    Set wsM = ThisWorkbook.Worksheets("MEDIA_AUDIT")
    On Error GoTo 0
    If wsM Is Nothing Then Exit Sub

    Dim lr As Long: lr = wsM.Cells(wsM.Rows.Count, 1).End(xlUp).Row
    Dim ptNames() As String
    Dim fileNames() As String
    Dim nMissing As Long: nMissing = 0
    Dim r As Long
    For r = 2 To lr
        If Trim$(CStr(wsM.Cells(r, 3).Value)) = "MISSING" Then
            ReDim Preserve ptNames(nMissing)
            ReDim Preserve fileNames(nMissing)
            ptNames(nMissing) = Trim$(CStr(wsM.Cells(r, 1).Value))
            fileNames(nMissing) = Trim$(CStr(wsM.Cells(r, 2).Value))
            nMissing = nMissing + 1
        End If
    Next r
    If nMissing = 0 Then Exit Sub

    On Error Resume Next
    Dim wsE As Worksheet: Set wsE = Nothing
    Set wsE = ThisWorkbook.Worksheets("Error Count")
    On Error GoTo 0
    If wsE Is Nothing Then Exit Sub

    Dim lastRow As Long: lastRow = wsE.Cells(wsE.Rows.Count, 1).End(xlUp).Row
    lastRow = lastRow + 2
    wsE.Cells(lastRow, 1).Value = "MISSING MEDIA FILES"
    wsE.Cells(lastRow, 1).Font.Bold = True
    wsE.Cells(lastRow, 1).Font.Color = RGB(180, 60, 0)
    lastRow = lastRow + 1
    wsE.Cells(lastRow, 1).Value = "Point Name"
    wsE.Cells(lastRow, 2).Value = "Expected File"
    wsE.Cells(lastRow, 1).Font.Bold = True
    wsE.Cells(lastRow, 2).Font.Bold = True
    Dim i As Long
    For i = 0 To nMissing - 1
        lastRow = lastRow + 1
        wsE.Cells(lastRow, 1).Value = ptNames(i)
        wsE.Cells(lastRow, 2).Value = fileNames(i)
    Next i
End Sub

Public Sub ProtectVBAProject()
    ' VBProject.Protection is a read-only enum, not a settable object.
    ' VBA project password protection cannot be set programmatically via the
    ' public object model — this sub is intentionally left empty.
End Sub

' ---------------- Heat-number fix helper ----------------
Public Sub FixHeatInDataSheet(ByVal oldHeat As String, ByVal newHeat As String)
    ' Called from ThisWorkbook.Workbook_SheetChange when user edits a Missing heat in Error Count.
    oldHeat = Trim$(oldHeat)
    newHeat = Trim$(newHeat)
    If oldHeat = "" Or newHeat = "" Then Exit Sub
    If UCase$(oldHeat) = UCase$(newHeat) Then Exit Sub

    Dim wsD As Worksheet: Set wsD = Worksheets("Data")
    Dim lr As Long: lr = wsD.Cells(wsD.Rows.Count, 1).End(xlUp).Row
    Dim lc As Long: lc = wsD.Cells(1, wsD.Columns.Count).End(xlToLeft).Column

    Dim prevCalc As Long: prevCalc = Application.Calculation
    Dim prevSU As Boolean: prevSU = Application.ScreenUpdating
    Application.Calculation = -4135  ' xlCalculationManual
    Application.ScreenUpdating = False
    Application.EnableEvents = False

    Dim changed As Long: changed = 0
    Dim r As Long, c As Long
    For r = 2 To lr
        For c = 1 To lc
            If UCase$(Trim$(CStr(wsD.Cells(r, c).Value))) = UCase$(oldHeat) Then
                wsD.Cells(r, c).Value = newHeat
                changed = changed + 1
            End If
        Next c
    Next r

    Application.EnableEvents = True
    Application.Calculation = prevCalc
    Application.ScreenUpdating = prevSU

    If changed > 0 Then
        ValidateSheetAll
        MsgBox "Updated " & changed & " cell(s): [" & oldHeat & "] to [" & newHeat & "].", _
               vbInformation, "Heat Updated"
    Else
        MsgBox "No cells found with heat [" & oldHeat & "] in the Data sheet.", _
               vbExclamation, "Heat Not Found"
    End If
End Sub

' ---------------- Save/close helper ----------------
Public Sub PromptSaveAndClose()
    Dim meta As Worksheet: Set meta = Worksheets("META")
    Dim defPath As String: defPath = CStr(meta.Range("A1").Value)
    Dim defName As String: defName = CStr(meta.Range("A2").Value)
    Dim sentinel As String: sentinel = CStr(meta.Range("A3").Value)
    Dim rv As VbMsgBoxResult
    rv = MsgBox("Export corrected 'Data' sheet as CSV before closing?", vbYesNoCancel + vbQuestion, "Save?")
    If rv = vbCancel Then Exit Sub
    Dim savedPath As String: savedPath = ""
    If rv = vbYes Then
        Dim initFull As String: initFull = defPath
        If Len(initFull) > 0 Then If Right$(initFull, 1) <> "\" Then initFull = initFull & "\"
        initFull = initFull & defName
        ' Ensure the dialog opens in the CSV folder (not the temp workbook folder)
        On Error Resume Next
        If Len(defPath) > 0 Then
            ChDrive Left$(defPath, 1)
            ChDir defPath
        End If
        On Error GoTo 0
        Dim outPath As String
        ' Prefer FileDialog SaveAs so Excel starts in the target folder
        On Error Resume Next
        Dim fd As Object: Set fd = Application.FileDialog(2) ' msoFileDialogSaveAs
        On Error GoTo 0
        If Not fd Is Nothing Then
            fd.InitialFileName = initFull
            If fd.Show = -1 Then outPath = CStr(fd.SelectedItems(1))
        End If
        If outPath = "" Then
            Dim fname As Variant
            fname = Application.GetSaveAsFilename(InitialFileName:=initFull, FileFilter:="CSV (Comma delimited) (*.csv), *.csv")
            If fname <> False Then outPath = CStr(fname)
        End If
        If outPath <> "" Then
            On Error Resume Next
            Worksheets("Data").Copy
            Application.DisplayAlerts = False
            If LCase$(Right$(outPath, 4)) <> ".csv" Then outPath = outPath & ".csv"
            ActiveWorkbook.SaveAs outPath, 6  ' xlCSV = 6
            Application.DisplayAlerts = True
            ActiveWorkbook.Close SaveChanges:=False
            On Error GoTo 0
            savedPath = outPath
        End If
    End If
    On Error Resume Next
    Dim f As Integer: f = FreeFile
    Open sentinel For Output As #f
    If savedPath <> "" Then
        Print #f, "saved:" & savedPath
    Else
        If rv = vbNo Then Print #f, "discard" Else Print #f, "cancel"
    End If
    Close #f
    On Error GoTo 0
    Application.DisplayAlerts = False
    ThisWorkbook.Close SaveChanges:=False
End Sub
"""
        # Sanitize any accidental odd separators injected near MTR mismatch message
        try:
            vbcode = re.sub(
                r"FlagOrange r, c, \"MTR mismatch \(\" \& mtrName \& \"\)[^\n]*\& mtrVal",
                'FlagOrange r, c, "MTR mismatch (" & mtrName & ") expected: " & mtrVal',
                vbcode,
            )
        except Exception:
            pass

        sheet_module_code = r"""
Option Explicit
Private Sub Worksheet_Change(ByVal Target As Range)
    On Error GoTo Done
    If Target Is Nothing Then Exit Sub
    If Target.Row < 2 Then Exit Sub

    Application.EnableEvents = False

    Dim r As Long: r = Target.Row
    Dim c As Long: c = Target.Column
    Dim cPN As Long: cPN = ValidationModule.ColOf("Point Number")
    Dim cSt As Long: cSt = ValidationModule.ColOf("Station")
    Dim cFC As Long: cFC = ValidationModule.ColOf("Field Code")
    Dim cN  As Long: cN  = ValidationModule.ColOf("Northing")
    Dim cE  As Long: cE  = ValidationModule.ColOf("Easting")

    ValidationModule.ValidateRow r
    If c = cFC Then ValidationModule.ApplyDVForRow r

    If c = cPN Then ValidationModule.RecalcColumnDuplicates cPN
    If cSt > 0 And c = cSt Then ValidationModule.RecalcColumnDuplicates cSt
    If c >= cFC Or c = cN Or c = cE Then ValidationModule.RecalcXRAYDupForEdit r, c

    If c > cFC Then
        ValidationModule.RecalcFormatForEdit r, c
    ElseIf c = cFC Then
        Dim lastC As Long: lastC = Cells(1, Columns.Count).End(xlToLeft).Column
        Dim cc As Long
        For cc = cFC + 1 To lastC
            If Trim$(Cells(r, cc).Value) <> "" Then ValidationModule.RecalcFormatForEdit r, cc
        Next cc
    End If

    ' NEW: live update Error Count
    ValidationModule.RebuildErrorCount

Done:
    Application.EnableEvents = True
End Sub

Private Sub Worksheet_Activate()
    On Error Resume Next
    ValidationModule.ContextMenu_Add
End Sub
"""

        # ThisWorkbook module: intercepts edits on the Error Count "Missing" heat rows
        # and propagates the corrected heat number to every matching cell in the Data sheet.
        wb_module_code = r"""
Option Explicit
Private prevHeatVal As String
Private prevHeatRow As Long

Private Sub Workbook_SheetSelectionChange(ByVal Sh As Object, ByVal Target As Range)
    If Sh.Name <> "Error Count" Then Exit Sub
    On Error Resume Next
    If Target.Column = 4 And Target.Row >= 12 And Target.Rows.Count = 1 Then
        ' Capture the old value only when the row is a Missing heat
        Dim statusVal As String
        statusVal = Trim$(CStr(Sh.Cells(Target.Row, 5).Value))
        If statusVal = "Missing" Then
            prevHeatVal = Trim$(CStr(Target.Value))
            prevHeatRow = Target.Row
        Else
            prevHeatVal = ""
            prevHeatRow = 0
        End If
    Else
        prevHeatVal = ""
        prevHeatRow = 0
    End If
    On Error GoTo 0
End Sub

Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)
    If Sh.Name <> "Error Count" Then Exit Sub
    If Target.Column <> 4 Then Exit Sub
    If Target.Row < 12 Then Exit Sub
    If Target.Rows.Count <> 1 Then Exit Sub
    On Error GoTo Done
    Dim newHeat As String: newHeat = Trim$(CStr(Target.Value))
    If newHeat = "" Or prevHeatVal = "" Then GoTo Done
    If UCase$(newHeat) = UCase$(prevHeatVal) Then GoTo Done
    If prevHeatRow <> Target.Row Then GoTo Done
    Dim oldH As String: oldH = prevHeatVal
    prevHeatVal = ""
    prevHeatRow = 0
    ValidationModule.FixHeatInDataSheet oldH, newHeat
Done:
    Exit Sub
End Sub
"""

        # === VBA & Excel automation (Windows only) ===
        if platform.system() == "Windows" and win32:
            excel = self._get_excel() if not open_new_excel_instance else self._get_excel_new_instance()
            if excel is None:
                # Fallback: open the generated XLSX without macros so user can still view data
                try:
                    xlsx_abs = os.path.abspath(temp_xlsx)
                    try:
                        os.startfile(xlsx_abs)  # type: ignore[attr-defined]
                        detail: str = (self._last_com_error or "Unknown COM error")
                        detail = detail.strip()
                        if len(detail) > 1200:
                            detail = detail[:1200] + "..."
                        messagebox.showwarning(
                            "Excel (no macros)",
                            "Could not start Excel via COM. Opened the workbook without macros.\n"
                            "Validation macros were not injected; you can still review the Data sheet.\n\n"
                            f"Details:\n{detail}"
                        )
                    except Exception:
                        messagebox.showerror(
                            "Excel error",
                            "Could not start Excel via COM, and auto-open fallback failed.\n"
                            f"You can open this file manually:\n{xlsx_abs}"
                        )
                except Exception:
                    messagebox.showerror("Excel error", "Could not start Excel via COM.")
                return

            try:
                excel.Visible = False
            except Exception:
                pass

            xlsx_abs = os.path.abspath(temp_xlsx)
            xlsm_abs = os.path.abspath(xlsm_file)

            self._progress_start(f"Opening {csv_name} in Excel…")
            try:
                wb_com = excel.Workbooks.Open(Filename=xlsx_abs)

                data_sheet = None
                try:
                    data_sheet = wb_com.Worksheets("Data")
                    data_sheet.Visible = 1  # xlSheetVisible
                    data_sheet.Activate()
                except Exception:
                    data_sheet = None

                # Inject macros
                try:
                    # Replace any prior ValidationModule to avoid duplicate procedures
                    try:
                        old_mod = wb_com.VBProject.VBComponents("ValidationModule")
                        wb_com.VBProject.VBComponents.Remove(old_mod)
                    except Exception:
                        pass
                    mod_obj = wb_com.VBProject.VBComponents.Add(1)  # StdModule
                    try:
                        mod_obj.Name = "ValidationModule"
                    except Exception:
                        pass
                    mod_obj.CodeModule.AddFromString(vbcode)
                    if data_sheet is not None:
                        wb_com.VBProject.VBComponents(data_sheet.CodeName).CodeModule.AddFromString(sheet_module_code)
                    wb_com.VBProject.VBComponents("ThisWorkbook").CodeModule.AddFromString(wb_module_code)
                except Exception as e_vba:
                    messagebox.showwarning(
                        "VBA injection skipped",
                        "Could not inject macros (Trust Center setting). "
                        "You can still review the sheet.\n\nDetails:\n" + str(e_vba)
                    )

                # Save as .xlsm
                try:
                    wb_com.SaveAs(xlsm_abs, 52)  # xlOpenXMLWorkbookMacroEnabled
                except Exception:
                    xlsm_abs = xlsx_abs

                # Show Excel and run validation once
                self._progress_stop()
                self._progress_start(f"Validating {csv_name}…")
                try:
                    excel.Visible = True
                    excel.WindowState = -4137  # xlMaximized
                except Exception:
                    pass

                for target in [
                    f"'{wb_com.Name}'!ValidationModule.NormalizeView",
                    "ValidationModule.NormalizeView",
                    "NormalizeView",
                ]:
                    try:
                        excel.Run(target)
                        break
                    except Exception:
                        pass

                for target in [
                    f"'{wb_com.Name}'!ValidationModule.ValidateSheetAll",
                    f"'{wb_com.Name}'!ValidateSheetAll",
                    "ValidationModule.ValidateSheetAll",
                    "ValidateSheetAll",
                ]:
                    try:
                        excel.Run(target)
                        break
                    except Exception:
                        pass

                # Run autofit/left-justify after validation to catch new sheets
                for target in [
                    f"'{wb_com.Name}'!ValidationModule.AutoFitAllSheets",
                    "ValidationModule.AutoFitAllSheets",
                    "AutoFitAllSheets",
                ]:
                    try:
                        excel.Run(target)
                        break
                    except Exception:
                        pass

                # Protect VBA project from casual editing
                for target in [
                    f"'{wb_com.Name}'!ValidationModule.ProtectVBAProject",
                    "ValidationModule.ProtectVBAProject",
                ]:
                    try:
                        excel.Run(target)
                        break
                    except Exception:
                        pass

                try:
                    wb_com.Worksheets("FXL").Visible = 2   # xlSheetVeryHidden
                except Exception:
                    pass
                try:
                    wb_com.Worksheets("META").Visible = 2  # xlSheetVeryHidden
                except Exception:
                    pass


                try:
                    self.btn_save_all.config(state="normal")
                    self.btn_email.config(state="normal")
                    # Enable Missing Heats email only if an MTR is loaded
                    if self.mtr_df is not None and not self.mtr_df.empty:
                        self.btn_missing_heats.config(state="normal")
                    else:
                        self.btn_missing_heats.config(state="disabled")
                except Exception:
                    pass

                # Auto-export feature removed per request; exports are user-triggered via the button

                # Track the last opened Excel instance for on-demand actions
                self._excel = excel
                self._wb_com = wb_com
                self._all_workbooks.append((excel, wb_com, self.csv_path or ""))
                self._excel_opened = True

                # Log this validation run
                self._write_validation_log()

                # Photo rename offer — uses results discovered before wb.save()
                _jxl_data_dict: dict[str, Any] | None = cast(dict[str, Any] | None, self._jxl_data)  # type: ignore[arg-type]
                _found = getattr(self, "_pending_photo_found", {})
                _missing = getattr(self, "_pending_photo_missing", {})
                if _found and _jxl_data_dict:
                    try:
                        self._offer_photo_rename(  # type: ignore[attr-defined]
                            _jxl_data_dict, _found, self.jxl_path or ""
                        )
                    except Exception:
                        pass
                if _missing:
                    _missing_lines = "\n".join(
                        f"  {pt}: {fn}" for pt, fn in sorted(_missing.items())
                    )
                    messagebox.showwarning(
                        "Missing Media Files",
                        f"{len(_missing)} photo file(s) could not be located in the SYNC folder "
                        f"or any standard sub-folder.\n\n"
                        f"They are listed in the 'Error Count' sheet.\n\n"
                        f"{_missing_lines}",
                    )

                self._progress_stop()
                self.status.config(text=f"Ready — {csv_name} validated.")

                try:
                    if os.path.exists(xlsm_abs) and xlsm_abs.lower() != xlsx_abs.lower() and os.path.exists(xlsx_abs):
                        os.remove(xlsx_abs)
                except Exception:
                    pass

            except Exception as e:
                self._progress_stop()
                messagebox.showerror("Excel open error", f"Failed to open workbook in Excel:\n{e}\n\nPath:\n{xlsx_abs}")
                try:
                    os.startfile(xlsx_abs)
                except Exception:
                    pass
        else:
            messagebox.showinfo("Exported", f"Excel saved at {temp_xlsx} (no VBA on this OS). Open it manually.")

    def _get_excel_new_instance(self) -> ExcelApplicationLike | None:
        if platform.system() != "Windows" or win32 is None:
            return None
        try:
            # DispatchEx starts a brand new Excel instance
            excel: ExcelApplicationLike = cast(ExcelApplicationLike, getattr(win32, "DispatchEx")("Excel.Application"))
            return excel
        except Exception:
            # First failure: clear pywin32 cache and retry DispatchEx
            try:
                self._nuke_genpy_cache()
            except Exception:
                pass
            try:
                excel = cast(ExcelApplicationLike, getattr(win32, "DispatchEx")("Excel.Application"))
                return excel
            except Exception:
                pass
            try:
                # Fallback to normal dispatch (may attach to existing instance)
                return cast(ExcelApplicationLike, getattr(win32, "Dispatch")("Excel.Application"))
            except Exception:
                try:
                    import traceback as _tb
                    self._last_com_error = _tb.format_exc()
                except Exception:
                    self._last_com_error = "DispatchEx/Dispatch failed"
                return None

    def _process_pair(self, csv_path: str, fxl_path: str,
                      jxl_override: str | None = None) -> None:
        """Load specific CSV+FXL and open validation in a new Excel instance, without altering initial FXL."""
        csv_name = os.path.basename(csv_path)

        # Check for a JXL companion — it records the exact FXL used in the field and
        # must override the session default (last_fxl_path / _initial_fxl_path).
        # jxl_override is set when the JXL was dropped alongside the CSV from a different folder.
        jxl_cand: str | None = jxl_override or self._find_jxl_alongside_csv(csv_path)  # type: ignore[attr-defined]
        if jxl_cand:
            self.jxl_path = jxl_cand
            jxl_info: dict[str, Any] = self._parse_jxl(jxl_cand)  # type: ignore[attr-defined]
            self._jxl_data = jxl_info
            fxl_from_jxl: str = str(jxl_info.get("fxl_filename") or "")
            if fxl_from_jxl and os.path.basename(fxl_path).lower() != fxl_from_jxl.lower():
                # Session default is a different FXL — search library for the JXL-specified one
                lib_cands = self._get_fxl_library_candidates()
                for lp in lib_cands:
                    if os.path.basename(lp).lower() == fxl_from_jxl.lower():
                        fxl_path = lp
                        break

        try:
            # Load the FXL silently in batch mode (popup fires once for the first interactive load)
            self._load_fxl_path(fxl_path, silent=True)
        except Exception as e:
            self.status.config(text=f"FXL load failed for {csv_name}: {e}")
            return
        try:
            # Load CSV without triggering FXL auto-pick prompts
            self._load_csv_path(csv_path, autopick_fxl=False)
        except Exception as e:
            self.status.config(text=f"CSV load failed for {csv_name}: {e}")
            return
        # Validate and open using a separate Excel instance
        use_new = (not self.single_excel_instance)
        self.validate_and_open(open_new_excel_instance=use_new)

    # ---------- progress bar ----------
    def _progress_start(self, msg: str = "Working…") -> None:
        self.status.config(text=msg)
        self.progress.pack(pady=2)
        self.progress.start(12)
        self.update_idletasks()

    def _progress_stop(self) -> None:
        self.progress.stop()
        self.progress.pack_forget()
        self.update_idletasks()

    # ---------- batch summary ----------
    def _show_batch_summary(self, results: list[tuple[str, str]]) -> None:
        """Show a Toplevel table of (filename, result) after batch processing."""
        if not results:
            return
        win = tk.Toplevel(self)
        win.title("Batch Processing Summary")
        win.resizable(True, True)
        _raise_window(win)
        tk.Label(win, text=f"Processed {len(results)} file(s):", font=("", 10, "bold")).pack(anchor="w", padx=10, pady=(8, 2))
        frame = tk.Frame(win)
        frame.pack(fill="both", expand=True, padx=10, pady=4)
        sb = tk.Scrollbar(frame)
        sb.pack(side="right", fill="y")
        cols = ("File", "Result")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=min(len(results), 15), yscrollcommand=sb.set)
        tree.heading("File", text="File")
        tree.heading("Result", text="Result")
        tree.column("File", width=220, anchor="w")
        tree.column("Result", width=340, anchor="w")
        sb.config(command=tree.yview)  # type: ignore[arg-type]
        for fname, status in results:
            tag = "ok" if status.lower().startswith("ok") else "err"
            tree.insert("", "end", values=(fname, status), tags=(tag,))
        tree.tag_configure("ok", foreground="#006600")
        tree.tag_configure("err", foreground="#cc0000")
        tree.pack(side="left", fill="both", expand=True)
        tk.Button(win, text="Close", command=win.destroy).pack(pady=8)
        win.grab_set()
        _raise_window(win)

    # ---------- process folder ----------
    def process_folder(self) -> None:
        """Walk a chosen folder, pair each CSV with the best FXL, and validate all."""
        folder = filedialog.askdirectory(title="Select folder to process")
        if not folder:
            return
        csvs = sorted(
            os.path.join(folder, f) for f in os.listdir(folder)
            if f.lower().endswith(".csv") and not f.lower().endswith("_corrected.csv")
        )
        if not csvs:
            messagebox.showinfo("No CSVs", "No CSV files found in the selected folder.")
            return
        fxls = sorted(os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith((".fxl", ".xml")))
        if not fxls and self._initial_fxl_path:
            fxls = [self._initial_fxl_path]
        if not fxls:
            p = filedialog.askopenfilename(
                title="Select FXL for this folder",
                initialdir=folder,
                filetypes=[("FXL files", "*.fxl"), ("XML files", "*.xml")]
            )
            if not p:
                return
            fxls = [p]
        fxl_to_use = fxls[0]
        if len(fxls) > 1:
            messagebox.askyesno(
                "Multiple FXLs",
                f"Found {len(fxls)} FXL files.\nUsing first: '{os.path.basename(fxls[0])}'",
            )
        if not messagebox.askyesno(
            "Process Folder",
            f"Process {len(csvs)} CSV file(s) in:\n{folder}\n\nUsing FXL: {os.path.basename(fxl_to_use)}\n\nContinue?"
        ):
            return
        batch_results: list[tuple[str, str]] = []
        for csv_p in csvs:
            fname = os.path.basename(csv_p)
            self._progress_start(f"Processing {fname}…")
            try:
                self._process_pair(csv_p, fxl_to_use)
                batch_results.append((fname, "OK — opened in Excel"))
            except Exception as e:
                batch_results.append((fname, f"Error: {e}"))
            finally:
                self._progress_stop()
        self._show_batch_summary(batch_results)

    # ---------- error filter toggle ----------
    def toggle_error_filter(self) -> None:
        """Toggle Show Errors Only / Show All Rows via VBA."""
        if not (self._excel and self._wb_com):
            return
        self._error_filter_on = not self._error_filter_on
        macro = "FilterErrorRows" if self._error_filter_on else "ShowAllRows"
        wbname = self._wb_com.Name
        for target in [f"'{wbname}'!ValidationModule.{macro}", f"ValidationModule.{macro}", macro]:
            try:
                self._excel.Run(target)
                break
            except Exception:
                pass
        pass  # filter button removed

    # ---------- help panel ----------
    def show_help(self) -> None:
        """Open a non-modal help window with color legend and usage notes."""
        win = tk.Toplevel(self)
        win.title("Data Validation Tool — Help")
        _place_window(win, 560, 540)
        win.resizable(True, True)
        text = tk.Text(win, wrap="word", padx=10, pady=8, font=("Consolas", 9))
        sb = tk.Scrollbar(win, command=text.yview)  # type: ignore[arg-type]
        text.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        text.pack(fill="both", expand=True)
        help_content = (
            "DATA VALIDATION TOOL v3.2 — Quick Reference\n"
            "═══════════════════════════════════════════\n\n"
            "HOW TO USE\n"
            "──────────\n"
            "1. Drag & drop a CSV file (field survey data) onto the window.\n"
            "2. Drag & drop an FXL file (Trimble feature library) — or the tool\n"
            "   will auto-detect one in the same/parent folder.\n"
            "3. Excel opens automatically and runs validation.\n"
            "4. Review colored cells. Hover for error details (comment bubble).\n"
            "5. Fix errors in Excel or export and send back to field staff.\n\n"
            "OPTIONAL: Drag an MTR Excel file (filename must contain 'MTR') to\n"
            "enable material test record cross-checking.\n\n"
            "BUTTONS\n"
            "───────\n"
            "Validate Entire Sheet   Re-run all checks after editing cells.\n"
            "Email Report…           Attach the .xlsm report to an Outlook email.\n"
            "Export CSV + Report     Save corrected CSV and .xlsm to the CSV folder.\n"
            "Email Missing Heats     Email a list of heats not found in the MTR.\n"
            "Process Folder…         Batch-validate all CSVs in a chosen folder.\n"
            "Show Errors Only        Hide rows with no errors (click again to restore).\n\n"
            "COLOR LEGEND (Excel cells)\n"
            "──────────────────────────\n"
            "RED (strong)     Not ALL CAPS; invalid token (NA, UNK, -, _)\n"
            "RED              Primary validation failure\n"
            "ORANGE           Field Code not in FXL; MTR mismatch; out-of-range\n"
            "YELLOW           Duplicate Point Number, Station, or coordinates\n"
            "PURPLE           Value not in FXL allowed list\n"
            "LIGHT GREEN      Unusual/NA value in a required list attribute\n"
            "TEAL             Station format error (expected 0+00 or 0+00.00)\n"
            "PINK             Joint length statistical outlier (>2 std devs)\n\n"
            "RIGHT-CLICK MENU (in Excel)\n"
            "───────────────────────────\n"
            "Clear Validation Flag(s)       Remove color/comment from selection.\n"
            "Ignore All Errors of This Type Remove all instances of that error type.\n"
            "Use MTR value for this cell    Auto-fill from MTR data.\n"
            "Use MTR value for all Heats    Auto-fill all rows with same heat.\n\n"
            "FIXING MISSING HEATS (Error Count sheet)\n"
            "─────────────────────────────────────────\n"
            "When a heat number in the Error Count sheet shows 'Missing', click\n"
            "the heat cell in column D and type the corrected heat number.\n"
            "All matching cells in the Data sheet update automatically.\n\n"
            "FXL AUTO-DETECTION ORDER\n"
            "────────────────────────\n"
            "When a CSV is dropped without an FXL, the tool searches:\n"
            "  1. JXL companion file (same filename, same folder) — uses the exact FXL\n"
            "     recorded in the JXL from the field.\n"
            "  2. FXL library:  S:\\TOPOGRAPHIC DATA\\TOPOGRAPHIC STANDARDS\\\n"
            "                   DATA DICTIONARY - GEOID FILES  (and subfolders)\n"
            "     → A filter dialog lets you search by name.\n"
            "  3. Previously loaded FXL (silent fallback if library has no match).\n"
            "  4. File browser as a last resort.\n"
            "The library path can be changed in config.json (fxl_library_path).\n\n"
            "FXL FILE FORMAT\n"
            "───────────────\n"
            "Trimble FXL files are XML. Supported feature types:\n"
            "  PointFeatureDefinition, LineFeatureDefinition,\n"
            "  PolygonFeatureDefinition, Feature, SurveyCode, etc.\n"
            "Attribute types: List (dropdown), Text, Number, Photo.\n"
            "Entry method 'Required' = cell must not be blank.\n\n"
            "MTR SPREADSHEET FORMAT\n"
            "──────────────────────\n"
            "Any Excel file with 'MTR' in the filename. Required columns\n"
            "(names are flexible — aliases are recognized):\n"
            "  HEAT / HEAT NUMBER, MANUFACTURER\n"
            "  NOM DIAMETER / OUT DIAMETER, WALL THICKNESS\n"
            "  GRADE, PIPE SPEC, SEAM TYPE\n\n"
            "VALIDATION LOG\n"
            "──────────────\n"
            "Each validation run is logged to validation_log.json in the\n"
            "CSV folder. Records: timestamp, user, CSV, FXL, and MTR files.\n"
        )
        text.insert("1.0", help_content)
        text.config(state="disabled")
        tk.Button(win, text="Close", command=win.destroy).pack(pady=6)

    # ---------- column mapping dialog ----------
    def _ask_column_mapping(self, df: pd.DataFrame) -> tuple[bool, dict[str, int | None], list[int]] | None:
        """Show a dialog to manually map CSV columns to known roles when auto-detection fails.
        Returns (has_station, mapping, attr_indices) or None if cancelled."""
        ncols = df.shape[1]
        headers = [str(df.iloc[0, c]) if not df.empty else f"Col {c}" for c in range(ncols)]
        win = tk.Toplevel(self)
        win.title("Column Mapping")
        win.grab_set()
        win.resizable(False, False)
        _raise_window(win)
        tk.Label(win, text="Auto-detection failed. Map CSV columns to known fields:",
                 wraplength=440, justify="left").pack(padx=12, pady=(10, 4))
        roles = ["Point Number", "Northing", "Easting", "Elevation", "Field Code", "Station (optional)"]
        choices: dict[str, tk.StringVar] = {}
        opts = ["(none)"] + [f"[{i}] {h[:40]}" for i, h in enumerate(headers)]
        frame = tk.Frame(win)
        frame.pack(padx=12, pady=4)
        for r_idx, role in enumerate(roles):
            tk.Label(frame, text=role, width=22, anchor="e").grid(row=r_idx, column=0, padx=4, pady=2)
            var = tk.StringVar(value="(none)")
            choices[role] = var
            om = tk.OptionMenu(frame, var, *opts)
            om.config(width=34)
            om.grid(row=r_idx, column=1, padx=4, pady=2)
        result: list[Any] = [None]

        def idx_of(role: str) -> "int | None":
            v = choices[role].get()
            if v == "(none)":
                return None
            try:
                return int(v.split("]")[0].lstrip("["))
            except Exception:
                return None

        def on_ok() -> None:
            pn = idx_of("Point Number")
            north = idx_of("Northing")
            east = idx_of("Easting")
            elev = idx_of("Elevation")
            fc = idx_of("Field Code")
            station = idx_of("Station (optional)")
            if any(v is None for v in [pn, north, east, elev, fc]):
                messagebox.showwarning("Incomplete", "Please map: Point Number, Northing, Easting, Elevation, Field Code.", parent=win)
                return
            used = {pn, north, east, elev, fc}
            if station is not None:
                used.add(station)
            attrs = [c for c in range(ncols) if c not in used]
            mapping: dict[str, int | None] = {"station": station, "pn": pn, "north": north, "east": east, "elev": elev, "fc": fc}
            result[0] = (station is not None, mapping, attrs)
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=8)
        tk.Button(btn_frame, text="OK", width=10, command=on_ok).grid(row=0, column=0, padx=6)
        tk.Button(btn_frame, text="Cancel", width=10, command=win.destroy).grid(row=0, column=1, padx=6)
        win.wait_window()
        return result[0]  # type: ignore[return-value]

    # ---------- JXL integration ----------

    def _find_jxl_alongside_csv(self, csv_path: str) -> str | None:
        """Return a .jxl path with the same stem as the CSV if it exists."""
        stem = os.path.splitext(csv_path)[0]
        for ext in (".jxl", ".JXL"):
            p = stem + ext
            if os.path.isfile(p):
                return p
        return None

    def _parse_jxl(self, jxl_path: str) -> dict[str, Any]:
        """Parse a Trimble JobXML (.jxl) file and return structured job + point data."""
        cache_key = os.path.abspath(jxl_path)
        if cache_key in self._jxl_parse_cache:
            return self._jxl_parse_cache[cache_key]

        result: dict[str, Any] = {
            "job_name": "", "timestamp": "", "date_str": "",
            "fxl_filename": "", "coordinate_system": "", "zone": "",
            "datum": "", "geoid": "", "distance_units": "", "height_units": "",
            "points": {},
            "attrs_by_pt": {},  # {pt_name: [val1, val2, ...]} — populated alongside points
        }
        try:
            tree = ET.parse(jxl_path)
            root = tree.getroot()
        except Exception:
            return result

        result["job_name"] = root.get("jobName", "")
        raw_ts: str = root.get("TimeStamp", "")
        result["timestamp"] = raw_ts
        try:
            dt = datetime.datetime.fromisoformat(raw_ts)
            result["date_str"] = dt.strftime("%Y%m%d")
        except Exception:
            result["date_str"] = raw_ts[:10].replace("-", "") if len(raw_ts) >= 10 else ""

        fieldbook = root.find("FieldBook")
        if fieldbook is None:
            return result

        for rec in fieldbook.findall("FeatureCodingRecord"):
            sf = (rec.findtext("SourceFilename") or "").strip()
            if sf:
                # SourceFilename is a full path on the Trimble data collector
                # (e.g. "\My Documents\...\DEA_PLF.fxl").  Store only the basename
                # so library/local lookups work correctly on the office PC.
                result["fxl_filename"] = os.path.basename(sf.replace("/", os.sep))

        cs = fieldbook.find("CoordinateSystemRecord")
        if cs is not None:
            result["coordinate_system"] = cs.findtext("SystemName") or ""
            result["zone"] = cs.findtext("ZoneName") or ""
            result["datum"] = cs.findtext("DatumName") or ""

        for va in fieldbook.findall("VerticalAdjustmentRecord"):
            g = va.findtext("GeoidName") or ""
            if g:
                result["geoid"] = g

        for ur in fieldbook.findall("UnitsRecord"):
            du = ur.findtext("DistanceUnits") or ""
            if du:
                result["distance_units"] = du
                result["height_units"] = ur.findtext("HeightUnits") or ""

        def _sfloat(text: str | None) -> float | None:
            try:
                return float(text) if text else None
            except Exception:
                return None

        def _sint(text: str | None) -> int | None:
            try:
                return int(text) if text else None
            except Exception:
                return None

        def _ecef_to_wgs84(x: float, y: float, z: float) -> tuple[float, float, float]:
            """Convert ECEF X/Y/Z (metres) to WGS84 lat, lon, ellipsoidal height (degrees/metres)."""
            import math as _math
            a = 6_378_137.0               # WGS84 semi-major axis (m)
            f = 1.0 / 298.257_223_563     # WGS84 flattening
            b = a * (1.0 - f)
            e2 = 2.0 * f - f * f          # first eccentricity squared
            ep2 = (a * a - b * b) / (b * b)  # second eccentricity squared

            lon_rad = _math.atan2(y, x)
            p = _math.sqrt(x * x + y * y)
            theta = _math.atan2(z * a, p * b)
            lat_rad = _math.atan2(
                z + ep2 * b * _math.sin(theta) ** 3,
                p - e2 * a * _math.cos(theta) ** 3,
            )
            for _ in range(10):
                sin_lat = _math.sin(lat_rad)
                n_rad = a / _math.sqrt(1.0 - e2 * sin_lat * sin_lat)
                lat_new = _math.atan2(z + e2 * n_rad * sin_lat, p)
                if abs(lat_new - lat_rad) < 1e-12:
                    lat_rad = lat_new
                    break
                lat_rad = lat_new
            sin_lat = _math.sin(lat_rad)
            cos_lat = _math.cos(lat_rad)
            n_rad = a / _math.sqrt(1.0 - e2 * sin_lat * sin_lat)
            h = (p / cos_lat - n_rad) if abs(cos_lat) > 1e-10 else (abs(z) / abs(sin_lat) - n_rad * (1.0 - e2))
            return _math.degrees(lat_rad), _math.degrees(lon_rad), h

        def _wgs84_to_ecef(lat_deg: float, lon_deg: float, h: float) -> tuple[float, float, float]:
            """Convert WGS84 lat/lon (degrees) + ellipsoidal height (m) to ECEF X/Y/Z (m)."""
            import math as _math
            a = 6_378_137.0
            f = 1.0 / 298.257_223_563
            e2 = 2.0 * f - f * f
            lat = _math.radians(lat_deg)
            lon = _math.radians(lon_deg)
            sin_lat = _math.sin(lat)
            cos_lat = _math.cos(lat)
            n_r = a / _math.sqrt(1.0 - e2 * sin_lat * sin_lat)
            return (
                (n_r + h) * cos_lat * _math.cos(lon),
                (n_r + h) * cos_lat * _math.sin(lon),
                (n_r * (1.0 - e2) + h) * sin_lat,
            )

        points: dict[str, Any] = {}
        for pr in fieldbook.findall("PointRecord"):
            name = (pr.findtext("Name") or "").strip()
            if not name:
                continue
            deleted = (pr.findtext("Deleted") or "false").strip().lower() == "true"
            pt: dict[str, Any] = {
                "deleted": deleted,
                "code": pr.findtext("Code") or "",
                "method": pr.findtext("Method") or "",
                "survey_method": pr.findtext("SurveyMethod") or "",
                "source": pr.findtext("Source") or "",
                "wgs84_lat": None, "wgs84_lon": None, "wgs84_height": None,
                "grid_north": None, "grid_east": None, "grid_elev": None,
                "h_precision": None, "v_precision": None,
                "pdop": None, "hdop": None, "vdop": None,
                "num_satellites": None, "num_gps_svs": None,
                "num_glonass_svs": None, "num_galileo_svs": None,
                "poor_precision_warning": "", "excess_tilt_warning": "",
                "bad_environment_warning": "", "photo_name": "",
            }
            wgs = pr.find("WGS84")
            if wgs is not None:
                pt["wgs84_lat"] = _sfloat(wgs.findtext("Latitude"))
                pt["wgs84_lon"] = _sfloat(wgs.findtext("Longitude"))
                pt["wgs84_height"] = _sfloat(wgs.findtext("Height"))
            else:
                _rtx1 = pr.find("RTXECEF")
                rtx = _rtx1 if _rtx1 is not None else pr.find("ECEF")
                if rtx is not None:
                    rx = _sfloat(rtx.findtext("X"))
                    ry = _sfloat(rtx.findtext("Y"))
                    rz = _sfloat(rtx.findtext("Z"))
                    if rx is not None and ry is not None and rz is not None:
                        try:
                            lat, lon, h = _ecef_to_wgs84(rx, ry, rz)
                            pt["wgs84_lat"] = lat
                            pt["wgs84_lon"] = lon
                            pt["wgs84_height"] = h
                        except Exception:
                            pass
                else:
                    # RTK rover: position stored as ECEF delta from a named base station.
                    # Resolve in second pass once all points (including the base) are collected.
                    deltas_el = pr.find("ECEFDeltas")
                    if deltas_el is not None:
                        dx = _sfloat(deltas_el.findtext("DeltaX"))
                        dy = _sfloat(deltas_el.findtext("DeltaY"))
                        dz = _sfloat(deltas_el.findtext("DeltaZ"))
                        base_name = (pr.findtext("RTK_Base") or "").strip()
                        if dx is not None and dy is not None and dz is not None and base_name:
                            pt["_ecef_dx"] = dx
                            pt["_ecef_dy"] = dy
                            pt["_ecef_dz"] = dz
                            pt["_rtk_base"] = base_name
            _grid1 = pr.find("ComputedGrid")
            grid = _grid1 if _grid1 is not None else pr.find("Grid")
            if grid is not None:
                pt["grid_north"] = _sfloat(grid.findtext("North"))
                pt["grid_east"] = _sfloat(grid.findtext("East"))
                pt["grid_elev"] = _sfloat(grid.findtext("Elevation"))
            prec = pr.find("Precision")
            if prec is not None:
                pt["h_precision"] = _sfloat(prec.findtext("Horizontal"))
                pt["v_precision"] = _sfloat(prec.findtext("Vertical"))
            qc = pr.find("QualityControl1")
            if qc is not None:
                pt["pdop"] = _sfloat(qc.findtext("PDOP"))
                pt["hdop"] = _sfloat(qc.findtext("HDOP"))
                pt["vdop"] = _sfloat(qc.findtext("VDOP"))
                pt["num_satellites"] = _sint(qc.findtext("NumberOfSatellites"))
                pt["num_gps_svs"] = _sint(qc.findtext("NumGPSSVs"))
                pt["num_glonass_svs"] = _sint(qc.findtext("NumGLONASSSVs"))
                pt["num_galileo_svs"] = _sint(qc.findtext("NumGalileoSVs"))
                warnings_el = qc.find("Warnings")
                if warnings_el is not None:
                    pt["poor_precision_warning"] = warnings_el.findtext("PoorPrecisionsWarning") or ""
                    pt["excess_tilt_warning"] = warnings_el.findtext("ExcessTiltWarning") or ""
                    pt["bad_environment_warning"] = warnings_el.findtext("BadEnvironmentWarning") or ""
            # Single pass: collect all attribute values AND locate the photo attribute.
            # Previously had a bare `break` that exited after the first attribute regardless
            # of type — photo_name was never set unless photo happened to be attr #1.
            attr_vals: list[str] = []
            features_el = pr.find("Features")
            if features_el is not None:
                for feat in features_el.findall("Feature"):
                    for attr in feat.findall("Attribute"):
                        a_type = (attr.findtext("Type") or "").strip().lower()
                        val = (attr.findtext("Value") or "").strip()
                        if a_type == "photo":
                            bare = os.path.basename(val.replace("\\", "/"))
                            attr_vals.append(bare)
                            if val and not pt["photo_name"]:
                                pt["photo_name"] = bare
                                pt["photo_path"] = val
                        else:
                            attr_vals.append(val)
            result["attrs_by_pt"][name] = attr_vals
            points[name] = pt

        # Second pass: resolve RTK rover points that have ECEFDeltas but no WGS84 yet.
        # base station WGS84 → ECEF + delta → ECEF → WGS84
        for pt in points.values():
            if pt.get("wgs84_lat") is not None:
                continue
            dx = pt.pop("_ecef_dx", None)
            dy = pt.pop("_ecef_dy", None)
            dz = pt.pop("_ecef_dz", None)
            base_name: str = pt.pop("_rtk_base", "")
            if dx is None or dy is None or dz is None or not base_name:
                continue
            base_pt = points.get(base_name)
            if base_pt is None:
                continue
            b_lat = base_pt.get("wgs84_lat")
            b_lon = base_pt.get("wgs84_lon")
            b_h = base_pt.get("wgs84_height")
            if b_lat is None or b_lon is None or b_h is None:
                continue
            try:
                bx, by, bz = _wgs84_to_ecef(b_lat, b_lon, b_h)
                rlat, rlon, rh = _ecef_to_wgs84(bx + dx, by + dy, bz + dz)
                pt["wgs84_lat"] = rlat
                pt["wgs84_lon"] = rlon
                pt["wgs84_height"] = rh
            except Exception:
                pass

        result["points"] = points
        self._jxl_parse_cache[cache_key] = result
        return result

    def _write_geodetic_sheet(self, wb: Any, jxl_data: dict[str, Any], df: Any) -> None:
        """Create a visible 'Geodetic Info' sheet with per-point geodetic and GPS quality data."""
        if openpyxl is None:
            return
        ws = wb.create_sheet("Geodetic Info")

        # Job metadata header block (rows 1–8)
        meta_pairs = [
            ("Job Name:", jxl_data.get("job_name", "")),
            ("Job Timestamp:", jxl_data.get("timestamp", "")),
            ("Coordinate System:", jxl_data.get("coordinate_system", "")),
            ("Zone:", jxl_data.get("zone", "")),
            ("Datum:", jxl_data.get("datum", "")),
            ("Geoid:", jxl_data.get("geoid", "")),
            ("Distance Units:", jxl_data.get("distance_units", "")),
            ("FXL Used:", jxl_data.get("fxl_filename", "")),
        ]
        for lbl, val in meta_pairs:
            ws.append([lbl, val])
        ws.append([])  # blank row separator

        # Column headers (row 10)
        headers = [
            "Point Name", "Northing", "Easting", "Elevation", "Code",
            "WGS84 Latitude", "WGS84 Longitude", "WGS84 Height (m)",
            "Grid North (m)", "Grid East (m)", "Grid Elevation (m)",
            "Survey Method", "H Precision", "V Precision",
            "PDOP", "HDOP", "VDOP",
            "Num Satellites", "GPS SVs", "GLONASS SVs", "Galileo SVs",
            "Poor Precision", "Excess Tilt", "Bad Environment",
            "Source", "Photo Name",
        ]
        ws.append(headers)

        # Build CSV coordinate lookup keyed by point name
        csv_coords: dict[str, tuple[str, str, str]] = {}
        if df is not None and self.mapping:
            pn_c = self.mapping.get("pn")
            n_c = self.mapping.get("north")
            e_c = self.mapping.get("east")
            z_c = self.mapping.get("elev")
            if pn_c is not None:
                for tup in df.itertuples(index=False, name=None):  # type: ignore[union-attr]
                    pn_v = str(tup[pn_c]).strip()
                    n_v = "" if n_c is None else str(tup[n_c]).strip()
                    e_v = "" if e_c is None else str(tup[e_c]).strip()
                    z_v = "" if z_c is None else str(tup[z_c]).strip()
                    if pn_v:
                        csv_coords[pn_v] = (n_v, e_v, z_v)

        points: dict[str, Any] = jxl_data.get("points", {})
        for pt_name in sorted(points.keys()):
            pt = points[pt_name]
            if pt.get("deleted"):
                continue  # deleted points excluded from Geodetic Info sheet
            n_s, e_s, z_s = csv_coords.get(pt_name, ("", "", ""))

            def _num(s: str) -> Any:
                try:
                    return float(s) if s else ""
                except Exception:
                    return s

            ws.append([
                pt_name,
                _num(n_s), _num(e_s), _num(z_s),
                pt.get("code", ""),
                pt.get("wgs84_lat"), pt.get("wgs84_lon"), pt.get("wgs84_height"),
                pt.get("grid_north"), pt.get("grid_east"), pt.get("grid_elev"),
                pt.get("survey_method", ""),
                pt.get("h_precision"), pt.get("v_precision"),
                pt.get("pdop"), pt.get("hdop"), pt.get("vdop"),
                pt.get("num_satellites"), pt.get("num_gps_svs"),
                pt.get("num_glonass_svs"), pt.get("num_galileo_svs"),
                pt.get("poor_precision_warning", ""),
                pt.get("excess_tilt_warning", ""),
                pt.get("bad_environment_warning", ""),
                pt.get("source", ""),
                pt.get("photo_name", ""),
            ])

        # Style: bold white header on dark-blue background
        from openpyxl.styles import PatternFill as _PF, Font as _Fnt  # type: ignore[import]
        hdr_fill = _PF(fill_type="solid", fgColor="1F4E79")
        hdr_font = _Fnt(bold=True, color="FFFFFF")
        for cell in ws[10]:
            cell.fill = hdr_fill
            cell.font = hdr_font
        # Bold meta labels
        lbl_font = _Fnt(bold=True)
        for r in range(1, 9):
            ws.cell(row=r, column=1).font = lbl_font
        # Freeze panes below header row
        ws.freeze_panes = ws["A11"]

    def _write_jxl_audit_sheet(self, wb: Any, jxl_data: dict[str, Any], df: Any) -> None:
        """Write a hidden JXL_AUDIT sheet listing deleted / cross-check issues for VBA to flag."""
        if openpyxl is None:
            return
        ws = wb.create_sheet("JXL_AUDIT")

        ws.append(["PointName", "Issue", "Reason"])

        points: dict[str, Any] = jxl_data.get("points", {})

        def _not_in_csv_reason(pt: dict[str, Any]) -> str:
            """Derive a human-readable explanation for why a JXL point is absent from the CSV."""
            method = str(pt.get("method", "")).strip().lower()
            code = str(pt.get("code", "")).strip()
            source = str(pt.get("source", "")).strip().lower()
            if method == "copiedpoint":
                return ("Present in JXL fieldbook but not found in CSV — "
                        "point was copied from another job, not surveyed in this field session")
            if method == "keyedin":
                return ("Present in JXL fieldbook but not found in CSV — "
                        "point was manually keyed in (not a GPS observation); "
                        "may not have been included in field export")
            if not code:
                return ("Present in JXL fieldbook but not found in CSV — "
                        "no feature code assigned; point may have been excluded from export")
            if source in ("control", "controlpoint", "benchmark", "known"):
                return ("Present in JXL fieldbook but not found in CSV — "
                        "control/benchmark point; typically excluded from feature CSV export")
            if method in ("gnss", "rtkgnss", "rtk"):
                return ("Present in JXL fieldbook but not found in CSV — "
                        "GPS-surveyed point not exported; check field export settings")
            return ("Present in JXL fieldbook but not found in CSV — "
                    "point exists in JXL fieldbook but was not included in the CSV export")

        # Build CSV point name set
        csv_names: set[str] = set()
        if df is not None and self.mapping:
            pn_c = self.mapping.get("pn")
            if pn_c is not None:
                for tup in df.itertuples(index=False, name=None):  # type: ignore[union-attr]
                    v = str(tup[pn_c]).strip()
                    if v:
                        csv_names.add(v)

        for pt_name, pt in points.items():
            if pt.get("deleted"):
                # Always record deleted points — they should appear in the Error Count
                # regardless of whether they also appear in the CSV.
                ws.append([pt_name, "DELETED_IN_FIELD", "Point was deleted in the field (marked deleted in Trimble job file)"])
            else:
                if pt_name not in csv_names:
                    ws.append([pt_name, "NOT_IN_CSV", _not_in_csv_reason(pt)])

        for pt_name in csv_names:
            if pt_name not in points:
                ws.append([pt_name, "NOT_IN_JXL", "Point is in CSV but has no matching record in the JXL fieldbook"])

        ws.sheet_state = "veryHidden"

    @staticmethod
    def _find_sync_folder(start_dir: str) -> str | None:
        """Walk up from start_dir looking for a SYNC subfolder at each level."""
        d = os.path.abspath(start_dir)
        for _ in range(6):
            candidate = os.path.join(d, "SYNC")
            if os.path.isdir(candidate):
                return candidate
            parent = os.path.dirname(d)
            if parent == d:
                break
            d = parent
        return None

    # ------------------------------------------------------------------
    #  Tiered media-index helpers
    # ------------------------------------------------------------------
    _MEDIA_EXTS: set[str] = {
        ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".heic", ".mp4", ".mov",
    }

    def _scan_dir_media(self, directory: str) -> dict[str, str]:
        """Return cached {basename_lower: full_path} for all media under *directory*."""
        abs_dir = os.path.abspath(directory)
        cached = self._media_index_cache.get(abs_dir)
        if cached is not None:
            return cached
        idx: dict[str, str] = {}
        try:
            for dirpath, _, filenames in os.walk(abs_dir):
                for fn in filenames:
                    if os.path.splitext(fn)[1].lower() in self._MEDIA_EXTS:
                        key = fn.lower()
                        if key not in idx:
                            idx[key] = os.path.join(dirpath, fn)
        except OSError:
            pass
        self._media_index_cache[abs_dir] = idx
        return idx

    def _build_jxl_media_index(
        self,
        jxl_paths: list[str],
        jxl_data_map: dict[str, dict[str, Any]] | None = None,
    ) -> dict[str, str]:
        """Build {basename_lower: full_path} using a tiered search strategy.

        For every JXL, scan **only** its ``<stem> Files`` companion folder.
        When a companion folder is missing, fall back to a SYNC-wide scan
        and use file-size comparison to disambiguate duplicate basenames.
        """
        media_index: dict[str, str] = {}
        if not jxl_paths:
            return media_index

        # ── Strict companion-folder search ─────────────────────────────
        missing_jxl_paths: list[str] = []
        for jxl_path in jxl_paths:
            abs_path = os.path.abspath(jxl_path)
            jxl_dir = os.path.dirname(abs_path)
            stem = os.path.splitext(os.path.basename(abs_path))[0]
            companion = os.path.join(jxl_dir, f"{stem} Files")
            if os.path.isdir(companion):
                for k, v in self._scan_dir_media(companion).items():
                    if k not in media_index:
                        media_index[k] = v
            else:
                missing_jxl_paths.append(jxl_path)

        if not missing_jxl_paths:
            return media_index

        # ── Fallback: XML-referenced folders + SYNC with file-size ─────
        # Collect basenames we still need, the subfolder names the JXL XML
        # actually references, and SYNC roots — then scan them all in one
        # pass.  When a basename has multiple candidates, prefer the file
        # from the XML-referenced folder; otherwise use file-size matching.
        needed: set[str] = set()                 # basename_lower values
        xml_folders: set[str] = set()            # folders named in the XML
        xml_ref_paths: dict[str, str] = {}       # basename_lower → full XML-referenced path
        if jxl_data_map:
            for jxl_path in missing_jxl_paths:
                abs_path = os.path.abspath(jxl_path)
                jxl_dir = os.path.dirname(abs_path)
                jdata = jxl_data_map.get(abs_path)
                if not jdata:
                    continue
                for pt in jdata.get("points", {}).values():
                    photo_path_raw: str = str(pt.get("photo_path") or "")
                    photo_name: str = str(pt.get("photo_name") or "")
                    bn = photo_name or os.path.basename(photo_path_raw.replace("\\", "/"))
                    if not bn:
                        continue
                    key = bn.lower()
                    if key in media_index:
                        continue
                    needed.add(key)
                    # Extract the referenced subfolder from the XML path
                    if photo_path_raw:
                        rel = photo_path_raw.replace("\\", "/")
                        parts = rel.split("/")
                        if len(parts) >= 2:
                            ref_folder = os.path.join(jxl_dir, parts[0])
                            if os.path.isdir(ref_folder):
                                xml_folders.add(ref_folder)
                        # Full literal path for file-size reference
                        full_ref = os.path.join(
                            jxl_dir,
                            photo_path_raw.replace("\\", os.sep).replace("/", os.sep),
                        )
                        if key not in xml_ref_paths and os.path.isfile(full_ref):
                            xml_ref_paths[key] = full_ref

        if not needed:
            return media_index

        # Gather SYNC roots
        sync_roots: set[str] = set()
        for jxl_path in missing_jxl_paths:
            s = self._find_sync_folder(os.path.dirname(os.path.abspath(jxl_path)))
            if s:
                sync_roots.add(os.path.abspath(s))

        # Scan XML-referenced folders + SYNC, collecting ALL candidates
        # per basename so we can disambiguate.
        all_candidates: dict[str, list[str]] = {}   # basename_lower → [paths]
        xml_folder_abs: set[str] = {os.path.abspath(f) for f in xml_folders}

        for folder in list(xml_folders) + [s for s in sync_roots]:
            try:
                for dirpath, _, filenames in os.walk(folder):
                    for fn in filenames:
                        key = fn.lower()
                        if key in needed and os.path.splitext(fn)[1].lower() in self._MEDIA_EXTS:
                            full = os.path.join(dirpath, fn)
                            if key not in all_candidates:
                                all_candidates[key] = []
                            if full not in all_candidates[key]:
                                all_candidates[key].append(full)
            except OSError:
                pass

        # Pick the best candidate for each basename
        for key, candidates in all_candidates.items():
            if key in media_index:
                continue
            if len(candidates) == 1:
                media_index[key] = candidates[0]
                continue
            # Prefer a candidate from the XML-referenced folder
            xml_hit: str | None = None
            for c in candidates:
                if any(os.path.abspath(c).startswith(xf + os.sep) for xf in xml_folder_abs):
                    xml_hit = c
                    break
            if xml_hit:
                media_index[key] = xml_hit
                continue
            # Multiple candidates, none in XML folder — use file-size match
            ref_path = xml_ref_paths.get(key)
            if ref_path:
                try:
                    ref_size = os.path.getsize(ref_path)
                    for c in candidates:
                        try:
                            if os.path.getsize(c) == ref_size:
                                media_index[key] = c
                                break
                        except OSError:
                            continue
                except OSError:
                    pass
            # Last resort — take the first candidate
            if key not in media_index:
                media_index[key] = candidates[0]

        return media_index

    def _find_jxl_photos(
        self, jxl_path: str, jxl_data: dict[str, Any]
    ) -> tuple[dict[str, str], dict[str, str]]:
        """Return (found, missing) where found={pt_name: full_path}, missing={pt_name: filename}.

        Uses the tiered media-index helper: single-JXL mode scans the JXL's
        parent directory; multi-JXL mode would use companion folders.
        """
        media_index = self._build_jxl_media_index(
            [jxl_path],
            {os.path.abspath(jxl_path): jxl_data},
        )

        found: dict[str, str] = {}
        missing: dict[str, str] = {}
        points: dict[str, Any] = jxl_data.get("points", {})
        for pt_name, pt in points.items():
            raw_photo: str = str(pt.get("photo_name", ""))
            if not raw_photo:
                continue
            basename = os.path.basename(raw_photo.replace("\\", "/"))
            if not basename:
                continue
            if basename.lower() in media_index:
                found[pt_name] = media_index[basename.lower()]
            else:
                missing[pt_name] = basename

        return found, missing

    def _write_media_audit_sheet(
        self,
        wb: Any,
        found: dict[str, str],
        missing: dict[str, str],
    ) -> None:
        """Write a hidden MEDIA_AUDIT sheet so VBA can report missing media in Error Count."""
        ws = wb.create_sheet("MEDIA_AUDIT")
        ws.append(["PointName", "PhotoFile", "Status"])
        for pt_name in sorted(found):
            ws.append([pt_name, os.path.basename(found[pt_name]), "FOUND"])
        for pt_name in sorted(missing):
            ws.append([pt_name, missing[pt_name], "MISSING"])
        ws.sheet_state = "veryHidden"

    def _update_jxl_photo_names(
        self, jxl_path: str, renames: dict[str, tuple[str, str]]
    ) -> None:
        """Rewrite photo Value elements in the JXL for each renamed file.

        renames = {pt_name: (old_basename, new_basename)}
        The directory prefix inside the JXL value is preserved; only the filename changes.
        """
        try:
            tree = ET.parse(jxl_path)
            root = tree.getroot()
            fieldbook = root.find("FieldBook")
            if fieldbook is None:
                return
            changed = False
            for pr in fieldbook.findall("PointRecord"):
                name = (pr.findtext("Name") or "").strip()
                if name not in renames:
                    continue
                old_ref, new_name = renames[name]
                # old_ref may be the full JXL path ("Folder\IMG.jpg") or just a basename
                old_base = os.path.basename(old_ref.replace("\\", "/"))
                features_el = pr.find("Features")
                if features_el is None:
                    continue
                for feat in features_el.findall("Feature"):
                    for attr in feat.findall("Attribute"):
                        if (attr.findtext("Type") or "").strip().lower() == "photo":
                            val_el = attr.find("Value")
                            if val_el is not None and val_el.text:
                                current = val_el.text.strip()
                                basename = os.path.basename(current.replace("\\", "/"))
                                if basename == old_base:
                                    # Store just the bare new filename — no folder prefix
                                    val_el.text = new_name
                                    changed = True
                            break
            if changed:
                tree.write(jxl_path, encoding="unicode", xml_declaration=True)
        except Exception:
            pass

    def _update_excel_photo_cells(self, renames: dict[str, tuple[str, str]]) -> None:
        """Update photo filename cells in the open Excel Data sheet.

        Uses Range.Find to quickly locate cells containing each old photo
        basename, then sets the entire cell value to the new bare filename
        (dropping any folder prefix like "Folder\\IMG.jpg" → "NewName.jpg").

        renames = {pt_name: (old_ref, new_basename)}
        """
        if not (self._wb_com and renames):
            return
        try:
            ws: Any = self._wb_com.Worksheets("Data")
            used: Any = ws.UsedRange
            # Build lookup: old basename (lowercase) → new bare filename
            basename_map: dict[str, str] = {}
            for old_ref, new_name in renames.values():
                old_base = os.path.basename(old_ref.replace("\\", "/")).lower()
                if old_base:
                    basename_map[old_base] = new_name

            for old_base, new_name in basename_map.items():
                # Find first cell containing the old basename (xlPart=2, xlValues=-4163)
                found_cell: Any = used.Find(old_base, used.Cells(1, 1), -4163, 2, 1, 1, False)
                if found_cell is None:
                    continue
                first_addr: str = str(found_cell.Address)
                while True:
                    val: str = str(found_cell.Value or "")
                    # Verify exact basename match to avoid false positives
                    if os.path.basename(val.replace("\\", "/")).lower() == old_base:
                        found_cell.Value = new_name
                    found_cell = used.FindNext(found_cell)
                    if found_cell is None or str(found_cell.Address) == first_addr:
                        break
        except Exception:
            pass

    def _rename_jxl_photos(self, jxl_paths: list[str]) -> None:
        """Parse multiple JXL files dropped directly and offer to rename auto-generated photo filenames.

        Uses _build_jxl_media_index (companion-folder search with file-size fallback).
        """
        self.status.config(text=f"Scanning {len(jxl_paths)} JXL file(s) for photos…")
        self.update_idletasks()

        # Build jxl_map {abs_path: stem} and collect all points that have photos
        jxl_map: dict[str, str] = {}
        all_matched_pts: dict[str, dict[str, Any]] = {}
        all_jxl_data: dict[str, dict[str, Any]] = {}
        errors: list[str] = []

        for jxl_path in jxl_paths:
            try:
                jxl_data: dict[str, Any] = self._parse_jxl(jxl_path)  # type: ignore[attr-defined]
                abs_path = os.path.abspath(jxl_path)
                stem = os.path.splitext(os.path.basename(jxl_path))[0]
                jxl_map[abs_path] = stem
                all_jxl_data[abs_path] = jxl_data
                for pt_name, pt in jxl_data.get("points", {}).items():
                    if pt.get("photo_name"):
                        all_matched_pts[pt_name.upper()] = pt
            except Exception as e:
                errors.append(f"{os.path.basename(jxl_path)}: {e}")

        if errors:
            messagebox.showwarning("JXL Parse Errors", "\n".join(errors))

        if not all_matched_pts:
            self.status.config(text="Ready")
            messagebox.showinfo(
                "No Photos Found",
                "No photo references were found in the dropped JXL file(s).",
            )
            return

        # Build media index PER JXL — different JXLs can have photos with
        # the same basename (e.g. IMG_14.jpg) that are different files in
        # different companion folders.  A single global index would clobber
        # all but the first.
        jxl_items: list[tuple[str, dict[str, Any], dict[str, str]]] = []
        any_found = False
        for jxl_path in jxl_paths:
            abs_path = os.path.abspath(jxl_path)
            jxl_item_data: dict[str, Any] | None = all_jxl_data.get(abs_path)
            if not jxl_item_data:
                continue
            per_idx = self._build_jxl_media_index(  # type: ignore[attr-defined]
                [jxl_path], {abs_path: jxl_item_data},
            )
            per_jxl_found: dict[str, str] = {}
            for pt_name, pt in jxl_item_data.get("points", {}).items():
                photo_ref: str = str(pt.get("photo_path") or pt.get("photo_name") or "")
                basename = os.path.basename(photo_ref.replace("\\", "/"))
                if basename and basename.lower() in per_idx:
                    per_jxl_found[pt_name] = per_idx[basename.lower()]
            if per_jxl_found:
                any_found = True
                jxl_items.append((jxl_path, jxl_item_data, per_jxl_found))

        self.status.config(text="Ready")

        if not any_found:
            messagebox.showinfo(
                "No Photos Found",
                "Photo references were found in the JXL file(s) but none of the image "
                "files could be located on disk.",
            )
            return

        if not jxl_items:
            messagebox.showinfo(
                "No Photos Found",
                "No photo references with auto-generated names (IMG, DSC, etc.) were found "
                "in the dropped JXL file(s).",
            )
            return

        self._offer_photo_rename_multi(jxl_items, update_excel=False)  # type: ignore[attr-defined]

    def _show_csv_action_dialog(self, n_csvs: int) -> str | None:
        """Ask what to do with dropped CSV file(s). Returns 'validate', 'gnss', or None (cancel)."""
        dlg = tk.Toplevel(self)
        dlg.title("CSV Action")
        dlg.resizable(False, False)
        dlg.grab_set()
        _raise_window(dlg)

        tk.Label(dlg, text=f"{n_csvs} CSV file(s) loaded.",
                 font=("Segoe UI", 10, "bold")).pack(padx=24, pady=(18, 2))
        tk.Label(dlg, text="What would you like to do?",
                 font=("Segoe UI", 9)).pack(padx=24, pady=(0, 14))

        choice: list[str] = [""]

        def _pick(c: str) -> None:
            choice[0] = c
            dlg.destroy()

        bf = tk.Frame(dlg)
        bf.pack(padx=24, pady=(0, 18))
        tk.Button(bf, text="Validate", width=22,
                  command=lambda: _pick("validate")).grid(row=0, column=0, padx=8, pady=4)
        tk.Button(bf, text="Generate GNSS Report", width=22,
                  command=lambda: _pick("gnss")).grid(row=0, column=1, padx=8, pady=4)
        tk.Button(bf, text="Cancel", width=10,
                  command=dlg.destroy).grid(row=1, column=0, columnspan=2, pady=(2, 0))

        dlg.wait_window()
        return choice[0] if choice[0] else None

    def _show_jxl_action_dialog(self, jxl_paths: list[str]) -> None:
        """Ask the user what to do with dropped JXL file(s): rename photos or GNSS report."""
        n = len(jxl_paths)
        dlg = tk.Toplevel(self)
        dlg.title("JXL Action")
        dlg.resizable(False, False)
        dlg.grab_set()
        _raise_window(dlg)

        tk.Label(dlg, text=f"{n} JXL file(s) loaded.",
                 font=("Segoe UI", 10, "bold")).pack(padx=24, pady=(18, 2))
        tk.Label(dlg, text="What would you like to do?",
                 font=("Segoe UI", 9)).pack(padx=24, pady=(0, 14))

        choice: list[str] = [""]

        def _pick(c: str) -> None:
            choice[0] = c
            dlg.destroy()

        btn_frame = tk.Frame(dlg)
        btn_frame.pack(padx=24, pady=(0, 18))
        tk.Button(btn_frame, text="Validate", width=20,
                  command=lambda: _pick("validate")).grid(row=0, column=0, padx=8, pady=4)
        tk.Button(btn_frame, text="Rename Photos", width=20,
                  command=lambda: _pick("rename")).grid(row=0, column=1, padx=8, pady=4)
        tk.Button(btn_frame, text="Generate GNSS Report", width=20,
                  command=lambda: _pick("report")).grid(row=1, column=0, padx=8, pady=4)
        tk.Button(btn_frame, text="Cancel", width=10,
                  command=dlg.destroy).grid(row=1, column=1, padx=8, pady=(4, 0))

        dlg.wait_window()

        if choice[0] == "validate":
            self._validate_jxl(jxl_paths)  # type: ignore[attr-defined]
        elif choice[0] == "rename":
            self._rename_jxl_photos(jxl_paths)  # type: ignore[attr-defined]
        elif choice[0] == "report":
            self._generate_gnss_report(jxl_paths)  # type: ignore[attr-defined]

    def _extract_jxl_attrs_dict(self, jxl_path: str) -> dict[str, list[str]]:
        """Return {point_name: [val1, val2, ...]} — non-photo attribute values in JXL order.

        JXL attributes use generic <Type> tags (Text, Menu, Numeric, Date) rather than
        the descriptive names stored in the FXL.  The Nth JXL attribute value maps
        positionally to the Nth FXL attribute definition for that field code.
        """
        result: dict[str, list[str]] = {}
        try:
            tree = ET.parse(jxl_path)
            root = tree.getroot()
            fieldbook = root.find("FieldBook")
            if fieldbook is None:
                return result
            for pr in fieldbook.findall("PointRecord"):
                name = (pr.findtext("Name") or "").strip()
                if not name:
                    continue
                vals: list[str] = []
                features_el = pr.find("Features")
                if features_el is not None:
                    for feat in features_el.findall("Feature"):
                        for attr in feat.findall("Attribute"):
                            a_type = (attr.findtext("Type") or "").strip().lower()
                            val_text = (attr.findtext("Value") or "").strip()
                            # Photo values carry a folder prefix (e.g. "Job Files\photo.jpg");
                            # strip to bare filename so it matches CSV-loaded photo cells.
                            if a_type == "photo":
                                val_text = os.path.basename(val_text.replace("\\", "/"))
                            vals.append(val_text)
                result[name] = vals
        except Exception:
            pass
        return result

    def _validate_jxl(self, jxl_paths: list[str]) -> None:
        """Produce the same Excel validation report as CSV validation, driven from JXL data.

        Builds a synthetic DataFrame (PointName/N/E/Z/FC/Attrs) from all dropped JXL files,
        sets the session state, and calls _export_and_open_excel() so the full VBA pipeline runs.
        """
        # Resolve the FXL the JXL was actually collected with.
        # Read fxl_filename from the first JXL's <FeatureCodingRecord> and locate the file.
        # Cache the parse result so the main loop below reuses it instead of re-parsing.
        jxl_fxl_name: str = ""
        _parse_cache: dict[str, dict[str, Any]] = {}
        try:
            _meta = self._parse_jxl(jxl_paths[0])  # type: ignore[attr-defined]
            _parse_cache[os.path.abspath(jxl_paths[0])] = _meta
            jxl_fxl_name = str(_meta.get("fxl_filename") or "")
        except Exception:
            pass

        if jxl_fxl_name:
            found_fxl = self._find_fxl_for_jxl(jxl_paths[0], jxl_fxl_name)  # type: ignore[attr-defined]
            if found_fxl:
                # Only prompt if it differs from the currently loaded FXL
                already_loaded = (
                    self.fxl_path
                    and os.path.abspath(found_fxl) == os.path.abspath(self.fxl_path)
                )
                if not already_loaded:
                    confirmed = self._confirm_fxl(  # type: ignore[attr-defined]
                        found_fxl,
                        intro=f"This JXL was collected with:\n  {jxl_fxl_name}\n\nFound it at the path below.",
                    )
                    if confirmed is None:
                        return
                    try:
                        self._load_fxl_path(confirmed, silent=True)  # type: ignore[attr-defined]
                    except Exception as e:
                        messagebox.showerror("FXL Load Error", f"Could not load FXL:\n{e}")
                        return
            else:
                # FXL named in JXL but not found on disk — warn and let user locate it
                if not self.fxl_data:
                    messagebox.showwarning(
                        "FXL Not Found",
                        f"This JXL was collected with:\n  {jxl_fxl_name}\n\n"
                        "That file could not be found. Please drop or load the correct FXL first.",
                    )
                    return
                # FXL already loaded (possibly correct) — confirm before proceeding
                confirmed = self._confirm_fxl(  # type: ignore[attr-defined]
                    self.fxl_path,  # type: ignore[arg-type]
                    intro=(
                        f"This JXL was collected with:\n  {jxl_fxl_name}\n\n"
                        "That file could not be found automatically.\n"
                        "Currently loaded FXL shown below — is this correct?"
                    ),
                )
                if confirmed is None:
                    return
                if confirmed != self.fxl_path:
                    try:
                        self._load_fxl_path(confirmed, silent=True)  # type: ignore[attr-defined]
                    except Exception as e:
                        messagebox.showerror("FXL Load Error", f"Could not load FXL:\n{e}")
                        return
        else:
            # No FXL name in JXL — fall back to existing session FXL or prompt
            if not self.fxl_data:
                if not self._ensure_fxl_after_csv(jxl_paths[0]):  # type: ignore[attr-defined]
                    return
            if not self.fxl_data:
                messagebox.showwarning("No FXL", "An FXL must be loaded before validating a JXL.")
                return

        self.status.config(text=f"Parsing {len(jxl_paths)} JXL file(s)…")
        self.update_idletasks()

        # Max attribute columns needed — determined by the widest field code in the FXL
        max_attrs: int = max((len(v) for v in self.fxl_data.values()), default=0)

        all_rows: list[list[Any]] = []
        errors: list[str] = []
        first_jxl_data: dict[str, Any] | None = None

        # Pre-parse all JXLs so we can build the media index before the main loop.
        _all_jxl_data: dict[str, dict[str, Any]] = {}
        for jxl_path in jxl_paths:
            try:
                _d = _parse_cache.pop(os.path.abspath(jxl_path), None) \
                    or self._parse_jxl(jxl_path)  # type: ignore[attr-defined]
                _all_jxl_data[os.path.abspath(jxl_path)] = _d
            except Exception as e:
                errors.append(f"{os.path.basename(jxl_path)}: {e}")

        # Build the media index once (single JXL → parent dir; multi → companion folders).
        _media_idx: dict[str, str] = self._build_jxl_media_index(  # type: ignore[attr-defined]
            jxl_paths, _all_jxl_data,
        )

        for jxl_path in jxl_paths:
            try:
                jxl_data = _all_jxl_data.get(os.path.abspath(jxl_path))
                if jxl_data is None:
                    continue  # parse failed earlier
                if first_jxl_data is None:
                    first_jxl_data = jxl_data
                attrs_by_pt: dict[str, list[str]] = jxl_data.get("attrs_by_pt") or {}

                pts: dict[str, Any] = jxl_data.get("points") or {}
                for pt_name, pt in pts.items():
                    if pt.get("deleted"):
                        continue
                    fc: str = pt.get("code") or ""
                    # JXL attributes are positional — Nth value maps to Nth FXL attribute.
                    pt_attr_vals: list[str] = attrs_by_pt.get(pt_name, [])
                    # Pad or trim to max_attrs columns
                    attr_vals: list[str] = (pt_attr_vals + [""] * max_attrs)[:max_attrs]

                    # Resolve stale photo filenames inline — same two-pass approach as
                    # _resolve_csv_photo_names so it works regardless of whether _parse_jxl
                    # correctly set photo_name for this point.
                    if _media_idx:
                        _pt_prefix: str = pt_name.lower() + "_"
                        for _i, _v in enumerate(attr_vals):
                            if not _v:
                                continue
                            if os.path.splitext(str(_v).lower())[1] not in self._MEDIA_EXTS:
                                continue
                            if _v.lower() in _media_idx:
                                continue  # Pass 1: exact match — name is still current
                            # Pass 2: file was renamed; look for "{pt_name}_*" on disk
                            for _key in _media_idx:
                                if _key.startswith(_pt_prefix):
                                    attr_vals[_i] = os.path.basename(_media_idx[_key])
                                    break

                    n = pt.get("grid_north")
                    e = pt.get("grid_east")
                    z = pt.get("grid_elev")
                    all_rows.append([
                        pt_name,
                        "" if n is None else n,
                        "" if e is None else e,
                        "" if z is None else z,
                        fc,
                        *attr_vals,
                    ])
            except Exception as ex:
                errors.append(f"{os.path.basename(jxl_path)}: {ex}")

        if errors:
            messagebox.showwarning("JXL Parse Errors", "\n".join(errors))

        if not all_rows:
            self.status.config(text="Ready")
            messagebox.showinfo("No Points", "No point records found in the dropped JXL file(s).")
            return

        # Build DataFrame with fixed known layout
        col_names = ["Point Number", "Northing", "Easting", "Elevation", "Field Code"] + \
                    [f"Attr{i + 1}" for i in range(max_attrs)]
        df = pd.DataFrame(all_rows, columns=col_names, dtype=str)  # type: ignore[reportUnknownMemberType]
        df = df.fillna("")  # type: ignore[reportUnknownMemberType]

        # Wire up session state so the rest of the pipeline works unchanged
        self.df = df
        self.csv_path = jxl_paths[0]          # used for output file naming
        self.jxl_path = jxl_paths[0]
        if first_jxl_data is not None:
            self._jxl_data = first_jxl_data
        self.has_station = False
        self.mapping = {
            "station": None,
            "pn": 0,
            "north": 1,
            "east": 2,
            "elev": 3,
            "fc": 4,
        }
        self.attr_indices = list(range(5, 5 + max_attrs))

        self.status.config(text="Ready")
        self._export_and_open_excel(open_new_excel_instance=(not self.single_excel_instance))  # type: ignore[attr-defined]

    def _extract_jxl_attributes(self, jxl_path: str) -> dict[str, str]:
        """Return {point_name: 'Attr1=Val1; Attr2=Val2'} for all non-photo attributes."""
        result: dict[str, str] = {}
        try:
            tree = ET.parse(jxl_path)
            root = tree.getroot()
            fieldbook = root.find("FieldBook")
            if fieldbook is None:
                return result
            for pr in fieldbook.findall("PointRecord"):
                name = (pr.findtext("Name") or "").strip()
                if not name:
                    continue
                parts: list[str] = []
                features_el = pr.find("Features")
                if features_el is not None:
                    for feat in features_el.findall("Feature"):
                        for attr in feat.findall("Attribute"):
                            a_type = (attr.findtext("Type") or "").strip()
                            a_val = (attr.findtext("Value") or "").strip()
                            if a_type and a_type.lower() != "photo" and a_val:
                                parts.append(f"{a_type}={a_val}")
                if parts:
                    result[name] = "; ".join(parts)
        except Exception:
            pass
        return result

    def _generate_gnss_report(self, jxl_paths: list[str]) -> None:
        """Export a CSV with point data and GNSS quality fields from one or more JXL files."""
        import csv as _csv

        first_dir = os.path.dirname(os.path.abspath(jxl_paths[0]))
        default_name = (
            os.path.splitext(os.path.basename(jxl_paths[0]))[0] + "_GNSS_Report.csv"
            if len(jxl_paths) == 1
            else "GNSS_Report.csv"
        )
        save_path = os.path.join(first_dir, default_name)

        self.status.config(text=f"Generating GNSS report from {len(jxl_paths)} JXL file(s)…")
        self.update_idletasks()

        headers = [
            "Job", "Point Number", "Northing", "Easting", "Elevation",
            "Field Code",
            "H Precision (m)", "V Precision (m)", "PDOP", "Num Satellites", "Survey Method",
            "WGS84 Latitude", "WGS84 Longitude", "WGS84 Height (m)",
            "Media File Name",
        ]

        rows: list[list[Any]] = []
        errors: list[str] = []

        for jxl_path in jxl_paths:
            try:
                jxl_data = self._parse_jxl(jxl_path)  # type: ignore[attr-defined]
                job_name = str(jxl_data.get("job_name", ""))
                pts: dict[str, Any] = jxl_data.get("points") or {}
                for pt_name, pt in pts.items():
                    if pt.get("deleted"):
                        continue
                    rows.append([
                        job_name,
                        pt_name,
                        pt.get("grid_north") if pt.get("grid_north") is not None else "",
                        pt.get("grid_east")  if pt.get("grid_east")  is not None else "",
                        pt.get("grid_elev")  if pt.get("grid_elev")  is not None else "",
                        pt.get("code") or "",
                        pt.get("h_precision")    if pt.get("h_precision")    is not None else "",
                        pt.get("v_precision")    if pt.get("v_precision")    is not None else "",
                        pt.get("pdop")           if pt.get("pdop")           is not None else "",
                        pt.get("num_satellites") if pt.get("num_satellites") is not None else "",
                        pt.get("survey_method") or "",
                        pt.get("wgs84_lat")    if pt.get("wgs84_lat")    is not None else "",
                        pt.get("wgs84_lon")    if pt.get("wgs84_lon")    is not None else "",
                        pt.get("wgs84_height") if pt.get("wgs84_height") is not None else "",
                        pt.get("photo_name") or "",
                    ])
            except Exception as e:
                errors.append(f"{os.path.basename(jxl_path)}: {e}")

        self.status.config(text="Ready")

        if not rows and not errors:
            messagebox.showinfo("GNSS Report", "No points found in the selected JXL file(s).")
            return

        try:
            with open(save_path, "w", newline="", encoding="cp1252", errors="replace") as f:
                writer = _csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            msg = (f"Saved {len(rows)} point(s) from {len(jxl_paths)} JXL file(s).\n"
                   f"{os.path.basename(save_path)}")
            if errors:
                msg += "\n\nErrors:\n" + "\n".join(errors)
            messagebox.showinfo("GNSS Report Complete", msg)
        except Exception as e:
            messagebox.showerror("Save Failed", str(e))

    def _show_rename_result(self, summary: str, errors: list[str]) -> None:
        """Scrollable post-rename result dialog (replaces messagebox for long error lists)."""
        dlg = tk.Toplevel(self)
        dlg.title("Photo Rename Complete")
        dlg.resizable(True, True)
        dlg.geometry("720x420")
        dlg.minsize(500, 260)
        dlg.grab_set()
        _raise_window(dlg)

        tk.Label(dlg, text=summary, font=("Segoe UI", 9, "bold"),
                 justify="left").pack(anchor="w", padx=12, pady=(12, 4))

        if errors:
            tk.Label(dlg, text=f"Skipped / errors ({len(errors)}):",
                     font=("Segoe UI", 8), fg="#555555").pack(anchor="w", padx=12)
            frame = tk.Frame(dlg)
            frame.pack(fill="both", expand=True, padx=12, pady=(2, 4))
            txt = tk.Text(frame, wrap="none", font=("Courier New", 8),
                          relief="flat", bg="#f5f5f5")
            vsb = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)  # type: ignore[arg-type]
            hsb = ttk.Scrollbar(frame, orient="horizontal", command=txt.xview)  # type: ignore[arg-type]
            txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.pack(side="right", fill="y")
            hsb.pack(side="bottom", fill="x")
            txt.pack(side="left", fill="both", expand=True)
            txt.insert("1.0", "\n".join(errors))
            txt.config(state="disabled")

        tk.Button(dlg, text="OK", width=10, command=dlg.destroy).pack(pady=(4, 12))
        dlg.wait_window()

    def _offer_photo_rename(
        self,
        jxl_data: dict[str, Any],
        photo_map: dict[str, str],
        jxl_path: str = "",
        update_excel: bool = True,
    ) -> None:
        """Show a confirmation dialog and rename photos whose names look auto-generated."""
        import re as _re

        job_name: str = str(jxl_data.get("job_name", ""))
        date_str: str = str(jxl_data.get("date_str", ""))

        # Only offer rename for filenames that look auto-generated (e.g. IMG, IMG_1, DSC_1234)
        # Digits are optional so plain "IMG" is caught alongside "IMG_1", "IMG_001", etc.
        _auto_pat = _re.compile(r'^[A-Za-z]{1,5}[_\-]?\d{0,6}$', _re.IGNORECASE)

        rename_items: list[tuple[str, str, str]] = []  # (point_name, old_path, new_path)
        for pt_name, old_path in sorted(photo_map.items()):
            basename = os.path.basename(old_path)
            stem, ext = os.path.splitext(basename)
            if _auto_pat.match(stem):
                new_name = f"{pt_name}_{job_name}_{date_str}{ext}"
                new_path = os.path.join(os.path.dirname(old_path), new_name)
                rename_items.append((pt_name, old_path, new_path))

        if not rename_items:
            return  # nothing needs renaming

        win = tk.Toplevel(self)
        win.title("Photo Rename")
        win.resizable(True, True)
        win.grab_set()
        win.geometry("640x480")
        win.minsize(500, 300)
        _raise_window(win)

        tk.Label(win, text=f"Job: {job_name}", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=(8, 0))
        tk.Label(win, text=f"{len(rename_items)} photo(s) with auto-generated names found.",
                 font=("Segoe UI", 9)).pack(anchor="w", padx=10)
        tk.Label(win, text="New format:  PointName_JobName_Date.ext",
                 font=("Segoe UI", 8), fg="#555555").pack(anchor="w", padx=10, pady=(0, 6))

        # Reserve button row space at bottom BEFORE packing the expanding treeview frame
        btn_row = tk.Frame(win)
        btn_row.pack(side="bottom", pady=8)

        frame = tk.Frame(win)
        frame.pack(fill="both", expand=True, padx=10, pady=4)

        cols = ("Point", "Current Name", "New Name")
        tv = ttk.Treeview(frame, columns=cols, show="headings", height=20)
        tv.heading("Point", text="Point")
        tv.heading("Current Name", text="Current Filename")
        tv.heading("New Name", text="New Filename")
        tv.column("Point", width=90, anchor="w")
        tv.column("Current Name", width=200, anchor="w")
        tv.column("New Name", width=280, anchor="w")

        iids: list[str] = []
        for pt_name, old_path, new_path in rename_items:
            iid = tv.insert("", "end", values=(
                pt_name,
                os.path.basename(old_path),
                os.path.basename(new_path),
            ))
            iids.append(iid)

        sb = ttk.Scrollbar(frame, orient="vertical", command=tv.yview)  # type: ignore[arg-type]
        tv.configure(yscrollcommand=sb.set)
        tv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Select all by default
        for iid in iids:
            tv.selection_add(iid)

        def _select_all() -> None:
            for i in iids:
                tv.selection_add(i)

        def _deselect_all() -> None:
            tv.selection_remove(*iids)

        def _do_rename() -> None:
            selected = set(tv.selection())
            errors: list[str] = []
            # pt_name → (old_basename, new_basename) for JXL update
            jxl_renames: dict[str, tuple[str, str]] = {}
            renamed = 0
            for iid, (pt_name, old_path, new_path) in zip(iids, rename_items):
                if iid not in selected:
                    continue
                try:
                    if os.path.exists(new_path):
                        errors.append(f"{os.path.basename(new_path)} — already exists, skipped")
                        continue
                    os.rename(old_path, new_path)
                    # Use the original JXL photo reference as the cell's current value
                    # (may include a folder prefix like "TWA_022326_TTG Files\IMG.jpg").
                    # Replacing the whole cell value drops the folder prefix so the
                    # Excel cell ends up with just the new bare filename.
                    _pts: dict[str, Any] = jxl_data.get("points") or {}  # type: ignore[union-attr]
                    _pt: dict[str, Any] = _pts.get(pt_name) or {}
                    # photo_path is the full JXL reference ("Folder\IMG.jpg") matching
                    # the exact value written to the Excel cell; fall back to basename.
                    orig_ref: str = str(_pt.get("photo_path") or _pt.get("photo_name") or os.path.basename(old_path))
                    jxl_renames[pt_name] = (orig_ref, os.path.basename(new_path))
                    renamed += 1
                except Exception as ex:
                    errors.append(f"{os.path.basename(old_path)} — {ex}")
            # Update JXL photo Value elements and Excel Data sheet cells
            if jxl_renames and jxl_path:
                self._update_jxl_photo_names(jxl_path, jxl_renames)
            if jxl_renames and update_excel:
                self._update_excel_photo_cells(jxl_renames)
            win.destroy()
            summary = f"Renamed {renamed} file(s)."
            if jxl_renames:
                summary += "\nJXL and Excel sheet updated to match."
            self._show_rename_result(summary, errors)  # type: ignore[attr-defined]

        tk.Button(btn_row, text="Select All", width=10, command=_select_all).grid(row=0, column=0, padx=4)
        tk.Button(btn_row, text="Deselect All", width=10, command=_deselect_all).grid(row=0, column=1, padx=4)
        tk.Button(btn_row, text="Rename", width=10, command=_do_rename,
                  bg="#2E7D32", fg="white").grid(row=0, column=2, padx=4)
        tk.Button(btn_row, text="Skip", width=10, command=win.destroy).grid(row=0, column=3, padx=4)
        win.wait_window()

    def _offer_photo_rename_multi(
        self,
        jxl_items: list[tuple[str, dict[str, Any], dict[str, str]]],
        update_excel: bool = True,
    ) -> None:
        """Single combined photo rename dialog covering multiple JXL files.

        jxl_items: list of (jxl_path, jxl_data, photo_map) — one entry per JXL.
        Collects all auto-named photos across every JXL, shows them in one dialog,
        then groups the confirmed renames back by JXL path for JXL file updates.
        """
        import re as _re
        _auto_pat = _re.compile(r'^[A-Za-z]{1,5}[_\-]?\d{0,6}$', _re.IGNORECASE)

        # (pt_name, old_path, new_path, jxl_path, jxl_data, orig_ref)
        rename_items: list[tuple[str, str, str, str, dict[str, Any], str]] = []
        for jxl_path, jxl_data, photo_map in jxl_items:
            job_name = str(jxl_data.get("job_name", ""))
            date_str = str(jxl_data.get("date_str", ""))
            _pts: dict[str, Any] = jxl_data.get("points") or {}
            for pt_name, old_path in sorted(photo_map.items()):
                basename = os.path.basename(old_path)
                stem, ext = os.path.splitext(basename)
                if _auto_pat.match(stem):
                    new_name = f"{pt_name}_{job_name}_{date_str}{ext}"
                    new_path = os.path.join(os.path.dirname(old_path), new_name)
                    _pt: dict[str, Any] = _pts.get(pt_name) or {}
                    orig_ref = str(_pt.get("photo_path") or _pt.get("photo_name") or basename)
                    rename_items.append((pt_name, old_path, new_path, jxl_path, jxl_data, orig_ref))

        if not rename_items:
            return

        n_jobs = len({item[3] for item in rename_items})

        win = tk.Toplevel(self)
        win.title("Photo Rename")
        win.resizable(True, True)
        win.grab_set()
        win.geometry("780x520")
        win.minsize(600, 350)
        _raise_window(win)

        tk.Label(win,
                 text=f"{len(rename_items)} photo(s) with auto-generated names found across {n_jobs} job(s).",
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=(8, 0))
        tk.Label(win, text="New format:  PointName_JobName_Date.ext",
                 font=("Segoe UI", 8), fg="#555555").pack(anchor="w", padx=10, pady=(0, 6))

        # Reserve button row space at bottom BEFORE packing the expanding treeview frame
        btn_row = tk.Frame(win)
        btn_row.pack(side="bottom", pady=8)

        frame = tk.Frame(win)
        frame.pack(fill="both", expand=True, padx=10, pady=4)

        cols = ("Job", "Point", "Current Filename", "New Filename")
        tv = ttk.Treeview(frame, columns=cols, show="headings", height=20)
        tv.heading("Job", text="Job")
        tv.heading("Point", text="Point")
        tv.heading("Current Filename", text="Current Filename")
        tv.heading("New Filename", text="New Filename")
        tv.column("Job", width=140, anchor="w")
        tv.column("Point", width=80, anchor="w")
        tv.column("Current Filename", width=150, anchor="w")
        tv.column("New Filename", width=260, anchor="w")

        iids: list[str] = []
        for pt_name, old_path, new_path, jxl_path, jxl_data, orig_ref in rename_items:
            iid = tv.insert("", "end", values=(
                str(jxl_data.get("job_name", "")),
                pt_name,
                os.path.basename(old_path),
                os.path.basename(new_path),
            ))
            iids.append(iid)

        sb = ttk.Scrollbar(frame, orient="vertical", command=tv.yview)  # type: ignore[arg-type]
        tv.configure(yscrollcommand=sb.set)
        tv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        for iid in iids:
            tv.selection_add(iid)

        def _select_all() -> None:
            for i in iids:
                tv.selection_add(i)

        def _deselect_all() -> None:
            tv.selection_remove(*iids)

        def _do_rename() -> None:
            selected = set(tv.selection())
            errors: list[str] = []
            by_jxl: dict[str, dict[str, tuple[str, str]]] = {}
            renamed = 0
            for iid, (pt_name, old_path, new_path, jxl_path, _, orig_ref) in zip(iids, rename_items):
                if iid not in selected:
                    continue
                try:
                    if os.path.exists(new_path):
                        errors.append(f"{os.path.basename(new_path)} — already exists, skipped")
                        continue
                    os.rename(old_path, new_path)
                    by_jxl.setdefault(jxl_path, {})[pt_name] = (orig_ref, os.path.basename(new_path))
                    renamed += 1
                except Exception as ex:
                    errors.append(f"{os.path.basename(old_path)} — {ex}")

            for _jpath, _renames in by_jxl.items():
                if _jpath:
                    self._update_jxl_photo_names(_jpath, _renames)  # type: ignore[attr-defined]

            if update_excel and by_jxl:
                all_renames: dict[str, tuple[str, str]] = {}
                for _r in by_jxl.values():
                    all_renames.update(_r)
                self._update_excel_photo_cells(all_renames)  # type: ignore[attr-defined]

            win.destroy()
            summary = f"Renamed {renamed} file(s)."
            if by_jxl:
                summary += f"\n{len(by_jxl)} JXL file(s) updated to match."
            self._show_rename_result(summary, errors)  # type: ignore[attr-defined]

        tk.Button(btn_row, text="Select All", width=10, command=_select_all).grid(row=0, column=0, padx=4)
        tk.Button(btn_row, text="Deselect All", width=10, command=_deselect_all).grid(row=0, column=1, padx=4)
        tk.Button(btn_row, text="Rename", width=10, command=_do_rename,
                  bg="#2E7D32", fg="white").grid(row=0, column=2, padx=4)
        tk.Button(btn_row, text="Skip", width=10, command=win.destroy).grid(row=0, column=3, padx=4)
        win.wait_window()

    # ---------- CRDB → GeoPackage export ----------

    def _on_crdb_button(self) -> None:
        """Handle the 'Export CRDB → GeoPackage…' button click."""
        path = filedialog.askopenfilename(
            title="Select Carlson CRDB file",
            filetypes=[("Carlson CRDB", "*.crdb"), ("All files", "*.*")],
        )
        if path and os.path.isfile(path):
            self._show_crdb_action_dialog([path])  # type: ignore[attr-defined]

    def _load_crdb_rows(self, path: str) -> list[dict[str, Any]]:
        """Read all points from a Carlson CRDB (SQLite) Coordinates table."""
        import sqlite3
        rows: list[dict[str, Any]] = []
        with sqlite3.connect(path) as conn:
            cur = conn.execute("SELECT P, N, E, Z, D FROM Coordinates ORDER BY P")
            for p_val, n_val, e_val, z_val, d_raw in cur.fetchall():
                parts = [x.strip() for x in (d_raw or "").split(",")]
                code = parts[0].upper() if parts else ""
                attrs = parts[1:] if len(parts) > 1 else []
                rows.append({
                    "point_name": str(p_val),
                    "N": float(n_val) if n_val is not None else None,
                    "E": float(e_val) if e_val is not None else None,
                    "Z": float(z_val) if z_val is not None else None,
                    "code": code,
                    "attrs": attrs,
                })
        return rows

    def _extract_jxl_hints(self, rows: list[dict[str, Any]]) -> dict[str, list[str]]:
        """Scan attribute values for Trimble photo paths to find source JXL job names.

        Photo paths like ``JMA_02102026_SAFI Files\\IMG_14.jpg`` encode the JXL
        job name as the folder prefix before ' Files\\'.
        Returns {job_stem: [point_names_that_reference_it]}.
        """
        hints: dict[str, list[str]] = {}
        pat = re.compile(r'^(.+?) [Ff]iles[\\/]', re.IGNORECASE)
        for row in rows:
            for attr_val in row["attrs"]:
                m = pat.match(attr_val.strip())
                if m:
                    stem = m.group(1).strip()
                    if stem:
                        if stem not in hints:
                            hints[stem] = []
                        pname = row["point_name"]
                        if pname not in hints[stem]:
                            hints[stem].append(pname)
        return hints

    def _search_jxl_upward(
        self,
        crdb_path: str,
        target_stems: set[str],
    ) -> dict[str, str]:
        """Search ancestor levels for .jxl files to match against CRDB points.

        Level 0 (parent of CRDB folder) is searched automatically.
        For each level above that, the user is prompted before searching.
        At each level, SYNC-named subfolders are scanned first, then the rest
        of that ancestor's tree (excluding already-searched SYNC folders).
        Returns {stem: absolute_jxl_path}.
        """
        crdb_dir = os.path.dirname(os.path.abspath(crdb_path))
        found: dict[str, str] = {}

        def _collect_jxls_from(root: str) -> dict[str, str]:
            """Walk root for .jxl files, skipping the CRDB dir.

            Returns {abs_path: stem}.
            """
            newly: dict[str, str] = {}
            for dirpath, dirs, filenames in os.walk(root):
                dirs[:] = [dn for dn in dirs
                           if os.path.abspath(os.path.join(dirpath, dn)) != crdb_dir]
                for fn in filenames:
                    if fn.lower().endswith(".jxl"):
                        abs_path = os.path.abspath(os.path.join(dirpath, fn))
                        if abs_path not in found:
                            newly[abs_path] = os.path.splitext(fn)[0]
            return newly

        # Structured search: SYNC → Field Data → prompt to go higher.
        # Locate the Field Data folder (and its SYNC child) by checking known
        # naming variants relative to the CRDB's parent and grandparent.
        _fd_names = ["FIELD_DATA", "Field_Data", "Field Data", "FieldData"]
        _sync_dir: str | None = None
        _field_data_dir: str | None = None
        for _base in [os.path.dirname(crdb_dir), crdb_dir]:
            for _fd_name in _fd_names:
                _fd_candidate = os.path.join(_base, _fd_name)
                if os.path.isdir(_fd_candidate):
                    _field_data_dir = _fd_candidate
                    _sc = os.path.join(_fd_candidate, "SYNC")
                    if os.path.isdir(_sc):
                        _sync_dir = _sc
                    break
            if _field_data_dir:
                break

        # Stage 1: SYNC folder and all subfolders (automatic)
        if _sync_dir:
            found.update(_collect_jxls_from(_sync_dir))

        # Stage 2: Field Data folder and all subfolders (automatic, skips
        #          already-found JXLs from the SYNC walk above)
        if _field_data_dir:
            found.update(_collect_jxls_from(_field_data_dir))

        if found:
            return found

        # Stage 3: Escalate up the directory tree, prompting at each level.
        # Start from the parent of crdb_dir (e.g. ASBUILT) and go up to 6 levels.
        levels: list[str] = []
        d = os.path.dirname(crdb_dir)
        for _ in range(6):
            parent = os.path.dirname(d)
            if not d or d == parent:
                break  # reached filesystem root
            levels.append(d)
            d = parent

        for ancestor in levels:
            ans = messagebox.askyesno(
                "Expand JXL Search",
                f"No JXL files found yet. Search:\n\n"
                f"  {ancestor}\n\n"
                f"JXL files found so far: {len(found)}",
                parent=self,
            )
            if not ans:
                break
            found.update(_collect_jxls_from(ancestor))
            if found:
                break

        return found

    def _match_points_to_jxls(
        self,
        rows: list[dict[str, Any]],
        jxl_map: dict[str, str],
    ) -> tuple[dict[str, dict[str, Any]], list[str], list[str]]:
        """Match each CRDB point to its JXL geodetic record.

        Returns:
            matched   — {UPPER_point_name: jxl_point_dict}
            unresolved — point names with no JXL match
            ambiguous  — point names found in multiple JXLs (first match used)
        """
        # Build {upper_point_name: [(stem, pt_dict)]} across all parsed JXLs
        all_pts: dict[str, list[tuple[str, dict[str, Any]]]] = {}
        for jxl_path, stem in jxl_map.items():
            try:
                jxl_data = self._parse_jxl(jxl_path)  # type: ignore[attr-defined]
                pts_dict: dict[str, Any] = cast(dict[str, Any], jxl_data.get("points", {}))
                for pt_name, pt_dict in pts_dict.items():
                    key = pt_name.strip().upper()
                    if key not in all_pts:
                        all_pts[key] = []
                    all_pts[key].append((stem, cast(dict[str, Any], pt_dict)))
            except Exception:
                pass

        matched: dict[str, dict[str, Any]] = {}
        unresolved: list[str] = []
        ambiguous: list[str] = []

        for row in rows:
            key = row["point_name"].strip().upper()
            matches = all_pts.get(key, [])
            if not matches:
                unresolved.append(row["point_name"])
            else:
                if len(matches) > 1:
                    ambiguous.append(row["point_name"])
                matched[key] = matches[0][1]

        return matched, unresolved, ambiguous

    def _make_gpkg_point_blob(self, lon: float, lat: float, srid: int = 4326) -> bytes:
        """Encode a point as a GPKG geometry blob (ISO WKB)."""
        # GPKG binary header: magic(2) + version(1) + flags(1) + srid(4)
        # flags: 0x01 = little-endian byte order, no envelope, not empty
        header = b"GP" + b"\x00" + b"\x01" + struct.pack("<i", srid)
        # ISO WKB point (little-endian): byte_order(1) + wkb_type(4) + x(8) + y(8)
        wkb = struct.pack("<BIdd", 1, 1, lon, lat)
        return header + wkb

    def _write_gpkg(
        self,
        output_path: str,
        rows: list[dict[str, Any]],
        matched_pts: dict[str, dict[str, Any]],
        fxl_data: dict[str, list[dict[str, Any]]],
        media_found: dict[str, str] | None = None,
    ) -> None:
        """Write a GeoPackage with one feature layer per distinct feature code.

        Geometry is WGS84 lon/lat sourced from JXL.  Attribute column names come
        from the FXL definition for each code; generic Attr1..N are used when the
        FXL has no definition or fewer names than values.  GNSS quality fields and
        media_file (basename only) are appended after the attribute columns.
        """
        _media: dict[str, str] = media_found or {}
        import sqlite3

        # Group rows by code, skipping sentinel elevations
        by_code: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            _z_val = row.get("Z")
            if _z_val is not None and isinstance(_z_val, (int, float)) and _z_val <= -99999999:
                continue
            c = row["code"] or "UNKNOWN"
            by_code.setdefault(c, []).append(row)

        # Determine the maximum number of attribute columns per code
        max_attrs: dict[str, int] = {
            code: max((len(r["attrs"]) for r in code_rows), default=0)
            for code, code_rows in by_code.items()
        }

        def _col_names(code: str) -> list[str]:
            """Return attribute column names for this code from the FXL."""
            n = max_attrs.get(code, 0)
            fxl_attrs = fxl_data.get(code, [])
            names: list[str] = []
            for i in range(n):
                if i < len(fxl_attrs):
                    raw_name = (fxl_attrs[i].get("name") or "").strip()
                    if raw_name:
                        # Sanitize to a valid SQL identifier
                        safe = re.sub(r"[^A-Za-z0-9_]", "_", raw_name)
                        names.append(safe or f"Attr{i + 1}")
                        continue
                names.append(f"Attr{i + 1}")
            return names

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except PermissionError:
                raise PermissionError(
                    f"The output file is open in another application (QGIS, ArcGIS, etc.).\n"
                    f"Close it there first, then click Export again.\n\n"
                    f"{output_path}"
                )

        with sqlite3.connect(output_path) as conn:
            conn.execute("PRAGMA application_id = 0x47503130")  # GP10 — GPKG magic
            conn.execute("PRAGMA user_version = 10200")

            # --- GPKG spec tables ---
            conn.execute("""
                CREATE TABLE gpkg_spatial_ref_sys (
                    srs_name TEXT NOT NULL,
                    srs_id INTEGER NOT NULL PRIMARY KEY,
                    organization TEXT NOT NULL,
                    organization_coordsys_id INTEGER NOT NULL,
                    definition TEXT NOT NULL,
                    description TEXT
                )""")
            # Use parameterized inserts so the WKT string (which contains double-quotes
            # and commas) is passed as a bound parameter rather than inlined into SQL.
            _srs_insert = (
                "INSERT INTO gpkg_spatial_ref_sys"
                "(srs_name,srs_id,organization,organization_coordsys_id,definition,description)"
                " VALUES (?,?,?,?,?,?)"
            )
            _wgs84_wkt = (
                'GEOGCS["WGS 84",DATUM["WGS_1984",'
                'SPHEROID["WGS 84",6378137,298.257223563]],'
                'PRIMEM["Greenwich",0],'
                'UNIT["degree",0.0174532925199433,'
                'AUTHORITY["EPSG","9122"]],'
                'AUTHORITY["EPSG","4326"]]'
            )
            conn.executemany(_srs_insert, [
                ("Undefined Cartesian SRS",  -1, "NONE", -1, "undefined",
                 "undefined cartesian coordinate reference system"),
                ("Undefined Geographic SRS",  0, "NONE",  0, "undefined",
                 "undefined geographic coordinate reference system"),
                ("WGS 84 geographic 2D",  4326, "EPSG", 4326, _wgs84_wkt,
                 "longitude/latitude coordinates in decimal degrees on the WGS 84 spheroid"),
            ])

            conn.execute("""
                CREATE TABLE gpkg_contents (
                    table_name TEXT NOT NULL PRIMARY KEY,
                    data_type TEXT NOT NULL,
                    identifier TEXT,
                    description TEXT DEFAULT '',
                    last_change DATETIME NOT NULL
                        DEFAULT (strftime('%Y-%m-%dT%H:%M:%fZ','now')),
                    min_x REAL, min_y REAL, max_x REAL, max_y REAL,
                    srs_id INTEGER REFERENCES gpkg_spatial_ref_sys(srs_id)
                )""")

            conn.execute("""
                CREATE TABLE gpkg_geometry_columns (
                    table_name TEXT NOT NULL,
                    column_name TEXT NOT NULL,
                    geometry_type_name TEXT NOT NULL,
                    srs_id INTEGER NOT NULL REFERENCES gpkg_spatial_ref_sys(srs_id),
                    z TINYINT NOT NULL,
                    m TINYINT NOT NULL,
                    CONSTRAINT pk_geom_cols PRIMARY KEY (table_name, column_name)
                )""")

            # --- One feature table per code ---
            for code, code_rows in sorted(by_code.items()):
                col_names = _col_names(code)

                # Build CREATE TABLE column list as a Python list to avoid
                # double-comma when col_names is empty.
                # After FXL attribute columns: GNSS quality fields, media_file, source_jxl.
                create_cols = (
                    [
                        "fid INTEGER PRIMARY KEY AUTOINCREMENT",
                        "geom BLOB",
                        "point_name TEXT",
                        "N REAL", "E REAL", "Z REAL",
                    ]
                    + [f'"{c}" TEXT' for c in col_names]
                    + [
                        "h_precision REAL", "v_precision REAL",
                        "pdop REAL", "num_satellites INTEGER",
                        "survey_method TEXT",
                        "media_file TEXT",
                        "source_jxl TEXT",
                    ]
                )
                conn.execute(f'CREATE TABLE "{code}" ({", ".join(create_cols)})')

                conn.execute("""
                    INSERT INTO gpkg_geometry_columns VALUES (?, 'geom', 'POINT', 4326, 0, 0)
                """, (code,))

                # Pre-build INSERT SQL once per code (same for every row in this group)
                ins_cols = (
                    ["geom", "point_name", "N", "E", "Z"]
                    + [f'"{c}"' for c in col_names]
                    + ["h_precision", "v_precision", "pdop", "num_satellites",
                       "survey_method", "media_file", "source_jxl"]
                )
                ins_sql = (f'INSERT INTO "{code}" ({", ".join(ins_cols)}) '
                           f'VALUES ({", ".join(["?"] * len(ins_cols))})')

                # Collect bounding box for gpkg_contents
                lons: list[float] = []
                lats: list[float] = []
                batch_params: list[list[Any]] = []

                for row in code_rows:
                    pkey = row["point_name"].strip().upper()
                    pt_jxl = matched_pts.get(pkey)
                    lon: float | None = None
                    lat: float | None = None
                    h_prec: float | None = None
                    v_prec: float | None = None
                    pdop: float | None = None
                    num_sats: int | None = None
                    survey_method = ""
                    src_jxl = ""
                    if pt_jxl:
                        lon = cast(float | None, pt_jxl.get("wgs84_lon"))
                        lat = cast(float | None, pt_jxl.get("wgs84_lat"))
                        h_prec = cast(float | None, pt_jxl.get("h_precision"))
                        v_prec = cast(float | None, pt_jxl.get("v_precision"))
                        pdop = cast(float | None, pt_jxl.get("pdop"))
                        num_sats = cast(int | None, pt_jxl.get("num_satellites"))
                        survey_method = cast(str, pt_jxl.get("survey_method") or "")
                        src_jxl = cast(str, pt_jxl.get("source", ""))

                    geom_blob: bytes | None = None
                    if lon is not None and lat is not None:
                        geom_blob = self._make_gpkg_point_blob(lon, lat)  # type: ignore[attr-defined]
                        lons.append(lon)
                        lats.append(lat)

                    # media_file — basename only for portability
                    media_path = _media.get(pkey, "")
                    media_file = os.path.basename(media_path) if media_path else ""

                    attrs_padded = row["attrs"] + [""] * max(0, len(col_names) - len(row["attrs"]))
                    attr_vals = attrs_padded[: len(col_names)]

                    batch_params.append(
                        [geom_blob, row["point_name"], row["N"], row["E"], row["Z"]]
                        + attr_vals
                        + [h_prec, v_prec, pdop, num_sats, survey_method, media_file, src_jxl],
                    )

                conn.executemany(ins_sql, batch_params)

                min_x = min(lons) if lons else None
                min_y = min(lats) if lats else None
                max_x = max(lons) if lons else None
                max_y = max(lats) if lats else None
                conn.execute(
                    "INSERT INTO gpkg_contents "
                    "(table_name, data_type, identifier, description, min_x, min_y, max_x, max_y, srs_id) "
                    "VALUES (?, 'features', ?, '', ?, ?, ?, ?, 4326)",
                    (code, code, min_x, min_y, max_x, max_y),
                )

            conn.commit()

    # ---------- CSV export (RAW_POINTS schema) ----------

    # Known field-name mappings from GDB field names to CRDB/JXL source values.
    # Keys are lowercase GDB field names; values are either a CRDB row key,
    # a JXL point dict key, or a special "@" prefix for computed values.
    _CSV_FIELD_MAP: dict[str, str] = {
        "point": "@point_name", "point_name": "@point_name", "pt": "@point_name",
        "point_number": "@point_name", "name": "@point_name",
        "north": "@N", "northing": "@N", "y": "@N", "survey_coord_y": "@wgs84_lat",
        "east": "@E", "easting": "@E", "x": "@E", "survey_coord_x": "@wgs84_lon",
        "elev": "@Z", "elevation": "@Z", "z": "@Z", "survey_coord_z": "@wgs84_height",
        "code": "@code", "feature_code": "@code", "desc": "@code", "description": "@code",
        "photo": "@photo", "media_file": "@photo",
        "pdop": "pdop", "hdop": "hdop", "vdop": "vdop", "tdop": "tdop",
        "sats": "num_satellites", "num_satellites": "num_satellites",
        "precision_horizontal": "h_precision", "hpe": "h_precision", "h_precision": "h_precision",
        "precision_vertical": "v_precision", "vpe": "v_precision", "v_precision": "v_precision",
        "source": "source", "position_source": "survey_method",
        "survey_method": "survey_method", "correction_method": "method",
        "timestamp": "@timestamp", "date_collected": "@timestamp",
        "tag_identifier": "@tag_id",
        "unique_id": "@unique_id",
        "weld_report": "@empty", "file": "@empty", "comments": "@empty",
    }

    def _write_crdb_csv(
        self,
        output_path: str,
        rows: list[dict[str, Any]],
        matched_pts: dict[str, dict[str, Any]],
        fxl_data: dict[str, list[dict[str, Any]]],
        media_found: dict[str, str] | None = None,
        jxl_meta: dict[str, Any] | None = None,
        client_schema: dict[str, Any] | None = None,
    ) -> None:
        """Write CRDB point data as CSV matching a client's GDB schema.

        If client_schema is provided, uses the points_fields from the stored
        schema to determine column names and order.  Otherwise falls back to
        the default RAW_POINTS layout.
        """
        import csv as _csv

        _media: dict[str, str] = media_found or {}
        _meta: dict[str, Any] = jxl_meta or {}
        job_timestamp = _meta.get("timestamp", "")

        # Determine header from client schema or default
        if client_schema and client_schema.get("points_fields"):
            header = [f["name"] for f in client_schema["points_fields"]]
        else:
            max_attrs = 28
            attr_headers = [f"ATT_{i}" for i in range(1, max_attrs + 1)]
            header = (
                ["POINT", "NORTH", "EAST", "ELEV", "CODE"]
                + attr_headers
                + [
                    "PHOTO", "TAG_IDENTIFIER", "Timestamp", "DATE_COLLECTED",
                    "SOURCE", "POSITION_SOURCE",
                    "PDOP", "HDOP", "VDOP", "SATS",
                    "Precision_Horizontal", "Precision_Vertical",
                    "UNIQUE_ID", "SURVEY_COORD_X", "SURVEY_COORD_Y", "SURVEY_COORD_Z",
                    "WELD_REPORT", "FILE", "COMMENTS", "DESCRIPTION",
                ]
            )

        # Detect which header columns are attribute slots (ATT_1, ATTR_2, etc.)
        _attr_pat = re.compile(r"^att(?:r(?:ibute)?)?[_\s]?(\d+)$", re.IGNORECASE)
        attr_col_indices: dict[int, int] = {}  # {1-based attr num: header index}
        for hi, col_name in enumerate(header):
            m = _attr_pat.match(col_name)
            if m:
                attr_col_indices[int(m.group(1))] = hi

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = _csv.writer(f)
            writer.writerow(header)

            for row in rows:
                # Skip points with sentinel "no data" elevations
                _z_val = row.get("Z")
                if _z_val is not None and isinstance(_z_val, (int, float)) and _z_val <= -99999999:
                    continue
                pkey = row["point_name"].strip().upper()
                pt_jxl = matched_pts.get(pkey)
                attrs = row.get("attrs", [])

                # Media / photo
                media_path = _media.get(pkey, "")
                photo = os.path.basename(media_path) if media_path else ""
                if not photo and pt_jxl and pt_jxl.get("photo_name"):
                    photo = pt_jxl["photo_name"]

                # Build a lookup dict for resolving "@" special keys
                _specials: dict[str, Any] = {
                    "@point_name": row["point_name"],
                    "@N": row["N"], "@E": row["E"], "@Z": row["Z"],
                    "@code": row["code"],
                    "@photo": photo,
                    "@timestamp": job_timestamp,
                    "@tag_id": "",
                    "@unique_id": "",
                    "@empty": "",
                    "@wgs84_lat": "", "@wgs84_lon": "", "@wgs84_height": "",
                }
                if pt_jxl:
                    for jk in ("wgs84_lat", "wgs84_lon", "wgs84_height"):
                        v = pt_jxl.get(jk)
                        _specials[f"@{jk}"] = v if v is not None else ""

                # Build output row by mapping each header column
                csv_row: list[Any] = [""] * len(header)
                for hi, col_name in enumerate(header):
                    # Check if this is an attribute slot (ATT_1, ATTR_2, etc.)
                    if hi in attr_col_indices.values():
                        attr_num = next(k for k, v in attr_col_indices.items() if v == hi)
                        csv_row[hi] = attrs[attr_num - 1] if attr_num - 1 < len(attrs) else ""
                        continue
                    # Look up in field map
                    key = col_name.lower()
                    mapped = self._CSV_FIELD_MAP.get(key)
                    if mapped is not None:
                        if mapped.startswith("@"):
                            csv_row[hi] = _specials.get(mapped, "")
                        else:
                            # JXL point field
                            if pt_jxl:
                                v = pt_jxl.get(mapped)
                                csv_row[hi] = v if v is not None else ""
                    # else: leave as empty string (unknown column)

                writer.writerow(csv_row)

    # ---------- Multi-format export helpers ----------

    def _find_gis_output_dir(self, crdb_path: str) -> str:
        """Compute the GIS weekly-update output directory for today.

        Walks up from the CRDB folder looking for an ASBUILT directory,
        then returns ASBUILT/0_GIS/WEEKLY_UPDATE/YYYYMMDD.
        Falls back to <crdb_dir>/0_GIS/WEEKLY_UPDATE/YYYYMMDD if ASBUILT
        is not found within 6 ancestor levels.
        """
        crdb_dir = os.path.dirname(os.path.abspath(crdb_path))
        today = datetime.date.today().strftime("%Y%m%d")

        # Walk up looking for ASBUILT
        d = crdb_dir
        for _ in range(7):
            if os.path.basename(d).upper() == "ASBUILT":
                return os.path.join(d, "0_GIS", "WEEKLY_UPDATE", today)
            parent = os.path.dirname(d)
            if parent == d:
                break
            d = parent

        # Also check if ASBUILT is a child of crdb_dir or any ancestor
        d = crdb_dir
        for _ in range(7):
            candidate = os.path.join(d, "ASBUILT")
            if os.path.isdir(candidate):
                return os.path.join(candidate, "0_GIS", "WEEKLY_UPDATE", today)
            parent = os.path.dirname(d)
            if parent == d:
                break
            d = parent

        # Fallback
        return os.path.join(crdb_dir, "0_GIS", "WEEKLY_UPDATE", today)

    def _write_crdb_shp(
        self,
        output_path: str,
        rows: list[dict[str, Any]],
        matched_pts: dict[str, dict[str, Any]],
        media_found: dict[str, str] | None = None,
    ) -> None:
        """Write a single combined Shapefile (.shp/.shx/.dbf/.prj) with all points.

        Pure-Python writer using struct — no external dependencies.
        Geometry is WGS84 Point from JXL geodetic data.
        """
        _media: dict[str, str] = media_found or {}
        max_attrs = 28

        # --- Define DBF field descriptors ---
        # (name, type, size, decimal)
        dbf_fields: list[tuple[str, str, int, int]] = [
            ("point_name", "C", 50, 0),
            ("N", "N", 20, 10),
            ("E", "N", 20, 10),
            ("Z", "N", 20, 10),
            ("code", "C", 30, 0),
        ]
        for i in range(1, max_attrs + 1):
            dbf_fields.append((f"ATT_{i}", "C", 80, 0))
        dbf_fields += [
            ("h_prec", "N", 12, 4),
            ("v_prec", "N", 12, 4),
            ("pdop", "N", 8, 2),
            ("sats", "N", 4, 0),
            ("method", "C", 20, 0),
            ("media", "C", 100, 0),
            ("src_jxl", "C", 100, 0),
        ]

        record_size = 1  # deletion flag byte
        for _, _, sz, _ in dbf_fields:
            record_size += sz

        # --- Collect records and geometry ---
        shp_records: list[tuple[float | None, float | None, list[bytes]]] = []
        dbf_records: list[bytes] = []

        for row in rows:
            _z_val = row.get("Z")
            if _z_val is not None and isinstance(_z_val, (int, float)) and _z_val <= -99999999:
                continue
            pkey = row["point_name"].strip().upper()
            pt_jxl = matched_pts.get(pkey)
            lon: float | None = None
            lat: float | None = None
            h_prec = ""
            v_prec = ""
            pdop_val = ""
            sats_val = ""
            method = ""
            src_jxl = ""
            if pt_jxl:
                lon = cast(float | None, pt_jxl.get("wgs84_lon"))
                lat = cast(float | None, pt_jxl.get("wgs84_lat"))
                _hp = pt_jxl.get("h_precision")
                h_prec = f"{_hp:.4f}" if _hp is not None else ""
                _vp = pt_jxl.get("v_precision")
                v_prec = f"{_vp:.4f}" if _vp is not None else ""
                _pd = pt_jxl.get("pdop")
                pdop_val = f"{_pd:.2f}" if _pd is not None else ""
                _ns = pt_jxl.get("num_satellites")
                sats_val = str(_ns) if _ns is not None else ""
                method = str(pt_jxl.get("survey_method") or "")
                src_jxl = str(pt_jxl.get("source") or "")

            media_path = _media.get(pkey, "")
            media_file = os.path.basename(media_path) if media_path else ""

            attrs = row.get("attrs", [])
            attr_vals = [attrs[i] if i < len(attrs) else "" for i in range(max_attrs)]

            # Build DBF record
            rec = b"\x20"  # not deleted
            for (fname, ftype, fsize, fdec), val in zip(
                dbf_fields,
                [row["point_name"], row["N"], row["E"], row["Z"], row.get("code", "")]
                + attr_vals
                + [h_prec, v_prec, pdop_val, sats_val, method, media_file, src_jxl],
            ):
                if ftype == "N":
                    s = str(val).strip() if val not in (None, "") else ""
                    rec += s.rjust(fsize)[:fsize].encode("latin-1")
                else:
                    s = str(val) if val is not None else ""
                    rec += s.ljust(fsize)[:fsize].encode("latin-1", errors="replace")
            dbf_records.append(rec)
            shp_records.append((lon, lat, []))

        num_records = len(dbf_records)
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        base = os.path.splitext(output_path)[0]

        # --- Write .prj (WGS84) ---
        with open(base + ".prj", "w", encoding="utf-8") as f:
            f.write(
                'GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",'
                'SPHEROID["WGS_1984",6378137.0,298.257223563]],'
                'PRIMEM["Greenwich",0.0],'
                'UNIT["Degree",0.0174532925199433]]'
            )

        # --- Compute bounding box ---
        valid_pts = [(lon, lat) for lon, lat, _ in shp_records
                     if lon is not None and lat is not None]
        if valid_pts:
            xmin = min(p[0] for p in valid_pts)  # type: ignore[arg-type]
            ymin = min(p[1] for p in valid_pts)  # type: ignore[arg-type]
            xmax = max(p[0] for p in valid_pts)  # type: ignore[arg-type]
            ymax = max(p[1] for p in valid_pts)  # type: ignore[arg-type]
        else:
            xmin = ymin = xmax = ymax = 0.0

        # --- Write .shp ---
        # Each point record: record header (8) + shape type (4) + X,Y (16) = 28 bytes
        # Null shape record: record header (8) + shape type (4) = 12 bytes
        shp_content_length = 50  # file header = 100 bytes = 50 16-bit words
        offsets: list[int] = []
        content_lengths: list[int] = []
        for lon, lat, _ in shp_records:
            offsets.append(shp_content_length)
            if lon is not None and lat is not None:
                rec_len = 10  # (4+16) / 2 = 10 words
            else:
                rec_len = 2   # 4 / 2 = 2 words
            content_lengths.append(rec_len)
            shp_content_length += 4 + rec_len  # record header (4 words) + content

        with open(base + ".shp", "wb") as f:
            # File header
            f.write(struct.pack(">I", 9994))           # file code
            f.write(b"\x00" * 20)                      # unused
            f.write(struct.pack(">I", shp_content_length))  # file length in 16-bit words
            f.write(struct.pack("<I", 1000))            # version
            f.write(struct.pack("<I", 1))               # shape type = Point
            f.write(struct.pack("<d", xmin))
            f.write(struct.pack("<d", ymin))
            f.write(struct.pack("<d", xmax))
            f.write(struct.pack("<d", ymax))
            f.write(struct.pack("<d", 0.0))             # zmin
            f.write(struct.pack("<d", 0.0))             # zmax
            f.write(struct.pack("<d", 0.0))             # mmin
            f.write(struct.pack("<d", 0.0))             # mmax

            for i, (lon, lat, _) in enumerate(shp_records):
                f.write(struct.pack(">II", i + 1, content_lengths[i]))
                if lon is not None and lat is not None:
                    f.write(struct.pack("<I", 1))       # Point
                    f.write(struct.pack("<dd", lon, lat))
                else:
                    f.write(struct.pack("<I", 0))       # Null shape

        # --- Write .shx ---
        shx_length = 50 + num_records * 4  # header + 8 bytes per record (in 16-bit words: 4)
        with open(base + ".shx", "wb") as f:
            f.write(struct.pack(">I", 9994))
            f.write(b"\x00" * 20)
            f.write(struct.pack(">I", shx_length))
            f.write(struct.pack("<I", 1000))
            f.write(struct.pack("<I", 1))
            f.write(struct.pack("<d", xmin))
            f.write(struct.pack("<d", ymin))
            f.write(struct.pack("<d", xmax))
            f.write(struct.pack("<d", ymax))
            f.write(struct.pack("<d", 0.0))
            f.write(struct.pack("<d", 0.0))
            f.write(struct.pack("<d", 0.0))
            f.write(struct.pack("<d", 0.0))
            for offset, clen in zip(offsets, content_lengths):
                f.write(struct.pack(">II", offset, clen))

        # --- Write .dbf ---
        today = datetime.date.today()
        header_size = 32 + len(dbf_fields) * 32 + 1  # +1 for header terminator
        with open(base + ".dbf", "wb") as f:
            # DBF header
            f.write(struct.pack("<B", 3))               # version
            f.write(struct.pack("<3B", today.year - 1900, today.month, today.day))
            f.write(struct.pack("<I", num_records))
            f.write(struct.pack("<H", header_size))
            f.write(struct.pack("<H", record_size))
            f.write(b"\x00" * 20)                       # reserved

            # Field descriptors
            for fname, ftype, fsize, fdec in dbf_fields:
                f.write(fname.encode("latin-1").ljust(11, b"\x00")[:11])
                f.write(ftype.encode("latin-1"))
                f.write(b"\x00" * 4)                    # reserved
                f.write(struct.pack("<B", fsize))
                f.write(struct.pack("<B", fdec))
                f.write(b"\x00" * 14)                   # reserved

            f.write(b"\r")  # header terminator

            for rec in dbf_records:
                f.write(rec)

            f.write(b"\x1a")  # EOF marker

    def _write_crdb_landxml(
        self,
        output_path: str,
        rows: list[dict[str, Any]],
        matched_pts: dict[str, dict[str, Any]],
    ) -> None:
        """Write a LandXML 1.2 file with CgPoints for all survey points.

        Coordinates are WGS84 lat/lon/height from JXL. Points without
        geodetic data use local N/E/Z.
        """
        ns = "http://www.landxml.org/schema/LandXML-1.2"
        now = datetime.datetime.now()
        stem = os.path.splitext(os.path.basename(output_path))[0]

        root = ET.Element("LandXML", attrib={
            "xmlns": ns,
            "version": "1.2",
            "date": now.strftime("%Y-%m-%d"),
            "time": now.strftime("%H:%M:%S"),
        })
        ET.SubElement(root, "Project", attrib={"name": stem})

        cg_points = ET.SubElement(root, "CgPoints")

        for row in rows:
            _z_val = row.get("Z")
            if _z_val is not None and isinstance(_z_val, (int, float)) and _z_val <= -99999999:
                continue
            pkey = row["point_name"].strip().upper()
            pt_jxl = matched_pts.get(pkey)

            attribs: dict[str, str] = {"name": row["point_name"]}
            code = row.get("code", "")
            if code:
                attribs["code"] = code

            if pt_jxl:
                lat = pt_jxl.get("wgs84_lat")
                lon = pt_jxl.get("wgs84_lon")
                ht = pt_jxl.get("wgs84_height")
                if lat is not None and lon is not None:
                    coord_text = f"{lat} {lon}"
                    if ht is not None:
                        coord_text += f" {ht}"
                else:
                    coord_text = f"{row['N']} {row['E']} {row['Z']}"
            else:
                coord_text = f"{row['N']} {row['E']} {row['Z']}"

            cg_pt = ET.SubElement(cg_points, "CgPoint", attrib=attribs)
            cg_pt.text = coord_text

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ")
        tree.write(output_path, xml_declaration=True, encoding="UTF-8")

    def _write_crdb_kmz(
        self,
        output_path: str,
        rows: list[dict[str, Any]],
        matched_pts: dict[str, dict[str, Any]],
        fxl_data: dict[str, list[dict[str, Any]]],
        media_found: dict[str, str] | None = None,
    ) -> None:
        """Write a KMZ file (zipped KML) with one Folder per feature code.

        Each Placemark includes point name, code, attributes, GNSS quality,
        and media filename in the description.
        """
        import zipfile

        _media: dict[str, str] = media_found or {}
        stem = os.path.splitext(os.path.basename(output_path))[0]

        # Group rows by code, skipping sentinel elevations
        by_code: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            _z_val = row.get("Z")
            if _z_val is not None and isinstance(_z_val, (int, float)) and _z_val <= -99999999:
                continue
            c = row.get("code") or "UNKNOWN"
            by_code.setdefault(c, []).append(row)

        kml_ns = "http://www.opengis.net/kml/2.2"
        kml = ET.Element("kml", attrib={"xmlns": kml_ns})
        document = ET.SubElement(kml, "Document")
        ET.SubElement(document, "name").text = stem

        for code in sorted(by_code.keys()):
            code_rows = by_code[code]
            folder = ET.SubElement(document, "Folder")
            ET.SubElement(folder, "name").text = code

            # Get FXL attribute names for this code
            fxl_attrs = fxl_data.get(code, [])

            for row in code_rows:
                pkey = row["point_name"].strip().upper()
                pt_jxl = matched_pts.get(pkey)

                lon: float | None = None
                lat: float | None = None
                alt: float | None = None
                if pt_jxl:
                    lon = cast(float | None, pt_jxl.get("wgs84_lon"))
                    lat = cast(float | None, pt_jxl.get("wgs84_lat"))
                    alt = cast(float | None, pt_jxl.get("wgs84_height"))

                if lon is None or lat is None:
                    continue  # Skip points without geodetic position

                pm = ET.SubElement(folder, "Placemark")
                ET.SubElement(pm, "name").text = row["point_name"]

                # Build description with attributes and GNSS data
                desc_parts: list[str] = [f"Code: {code}"]
                attrs = row.get("attrs", [])
                for i, val in enumerate(attrs):
                    if val:
                        attr_name = (fxl_attrs[i].get("name", f"Attr{i+1}")
                                     if i < len(fxl_attrs) else f"Attr{i+1}")
                        desc_parts.append(f"{attr_name}: {val}")
                if pt_jxl:
                    hp = pt_jxl.get("h_precision")
                    vp = pt_jxl.get("v_precision")
                    if hp is not None:
                        desc_parts.append(f"H Precision: {hp:.4f}")
                    if vp is not None:
                        desc_parts.append(f"V Precision: {vp:.4f}")
                    pd_val = pt_jxl.get("pdop")
                    if pd_val is not None:
                        desc_parts.append(f"PDOP: {pd_val:.2f}")
                    ns_val = pt_jxl.get("num_satellites")
                    if ns_val is not None:
                        desc_parts.append(f"Satellites: {ns_val}")
                    sm = pt_jxl.get("survey_method")
                    if sm:
                        desc_parts.append(f"Method: {sm}")
                media_path = _media.get(pkey, "")
                if media_path:
                    desc_parts.append(f"Photo: {os.path.basename(media_path)}")

                ET.SubElement(pm, "description").text = "\n".join(desc_parts)

                point = ET.SubElement(pm, "Point")
                coord_str = f"{lon},{lat}"
                if alt is not None:
                    coord_str += f",{alt}"
                ET.SubElement(point, "coordinates").text = coord_str

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        kml_bytes = ET.tostring(kml, xml_declaration=True, encoding="UTF-8")

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("doc.kml", kml_bytes)

    # ---------- DWG geometry export ----------

    def _find_dwg_for_crdb(self, crdb_path: str) -> str | None:
        """Auto-detect a DWG file in the same directory as the CRDB."""
        crdb_dir = os.path.dirname(crdb_path)
        crdb_stem = os.path.splitext(os.path.basename(crdb_path))[0].lower()
        try:
            dwg_files = [f for f in os.listdir(crdb_dir) if f.lower().endswith(".dwg")]
        except OSError:
            return None
        if not dwg_files:
            return None
        if len(dwg_files) == 1:
            return os.path.join(crdb_dir, dwg_files[0])
        # Multiple DWGs — prefer exact stem match, then partial match
        for dwg in dwg_files:
            if os.path.splitext(dwg)[0].lower() == crdb_stem:
                return os.path.join(crdb_dir, dwg)
        for dwg in dwg_files:
            if crdb_stem in os.path.splitext(dwg)[0].lower():
                return os.path.join(crdb_dir, dwg)
        return os.path.join(crdb_dir, sorted(dwg_files)[0])

    def _make_gpkg_linestring_blob(self, coords: list[tuple[float, float]],
                                    srid: int = 4326) -> bytes:
        """Encode a LINESTRING as a GPKG geometry blob (ISO WKB)."""
        header = b"GP\x00\x01" + struct.pack("<i", srid)
        wkb = struct.pack("<BII", 1, 2, len(coords))
        for x, y in coords:
            wkb += struct.pack("<dd", x, y)
        return header + wkb

    def _make_gpkg_polygon_blob(self, rings: list[list[tuple[float, float]]],
                                 srid: int = 4326) -> bytes:
        """Encode a POLYGON as a GPKG geometry blob (ISO WKB)."""
        header = b"GP\x00\x01" + struct.pack("<i", srid)
        wkb = struct.pack("<BI", 1, 3)
        wkb += struct.pack("<I", len(rings))
        for ring in rings:
            wkb += struct.pack("<I", len(ring))
            for x, y in ring:
                wkb += struct.pack("<dd", x, y)
        return header + wkb

    def _open_dwg_readonly(self, dwg_path: str) -> tuple[Any, Any, bool]:
        """Open a DWG via AutoCAD COM (invisible).

        Returns ``(acad_app, document, we_started_acad)``.
        *we_started_acad* is True when a new AutoCAD process was launched so
        the caller knows to call ``acad.Quit()`` during cleanup.
        """
        if win32 is None:
            raise RuntimeError("win32com is not available")
        we_started = False
        try:
            acad: Any = win32.GetActiveObject("AutoCAD.Application")  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
        except Exception:
            acad = win32.Dispatch("AutoCAD.Application")
            we_started = True
            # Some AutoCAD versions need time to initialize before
            # accepting property changes — retry Visible with a brief wait.
            import time
            for _attempt in range(10):
                try:
                    acad.Visible = False
                    break
                except Exception:
                    time.sleep(1)
            else:
                try:
                    acad.Visible = False
                except Exception:
                    pass
        doc: Any = acad.Documents.Open(dwg_path, True)  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]  # ReadOnly
        return acad, doc, we_started  # type: ignore[reportUnknownVariableType]

    def _read_dwg_cs_code(self, doc: Any) -> str:
        """Best-effort read of the coordinate system code from a Map 3D drawing."""
        for var_name in ("CGEOCS", "MAPCSCODE", "GEOGRAPHICCS"):
            try:
                val = doc.GetVariable(var_name)
                if val and str(val).strip():
                    return str(val).strip()
            except Exception:
                continue
        # Try Map 3D Object Model (version-specific ProgIDs)
        try:
            app = doc.Application
            for progid in ("AutoCADMap.Application",
                           "AutoCADMap.Application.26",
                           "AutoCADMap.Application.25",
                           "AutoCADMap.Application.24"):
                try:
                    mapi = app.GetInterfaceObject(progid)
                    if mapi:
                        proj = mapi.GetProject()
                        cs = (getattr(proj, "Projection", "")
                              or getattr(proj, "CoordinateSystem", ""))
                        if cs:
                            return str(cs).strip()
                except Exception:
                    continue
        except Exception:
            pass
        return ""

    def _convert_dwg_to_dxf(self, dwg_path: str) -> tuple[str, str]:
        """Convert a DWG file to DXF via AutoCAD COM SaveAs.

        Copies the DWG to a temp directory first (so it is never locked or
        read-only), opens the copy in AutoCAD, queries the Map 3D coordinate
        system, does a single SaveAs to DXF, then closes the document.

        Returns ``(dxf_path, cs_code)`` where *cs_code* is the Map 3D
        coordinate system string (e.g. ``"TX83-SF"``) or empty if not set.
        Raises RuntimeError if conversion fails.
        """
        import shutil
        import tempfile
        import win32com.client as win32  # type: ignore[import-untyped]

        output_dir = tempfile.mkdtemp(prefix="dvt_dxf_")
        stem = os.path.splitext(os.path.basename(dwg_path))[0]
        dxf_path = os.path.join(output_dir, stem + ".dxf")
        cs_code = ""

        # Copy the DWG so we always have a writable, unlocked copy
        tmp_dwg = os.path.join(output_dir, os.path.basename(dwg_path))
        shutil.copy2(dwg_path, tmp_dwg)

        acad: Any = None  # type: ignore[reportUnknownVariableType]
        doc: Any = None  # type: ignore[reportUnknownVariableType]
        docs_before = 0
        try:
            try:
                acad = win32.GetActiveObject("AutoCAD.Application")  # type: ignore[reportUnknownMemberType]
            except Exception:
                acad = win32.Dispatch("AutoCAD.Application")  # type: ignore[reportUnknownMemberType]

            # Wait for AutoCAD to be ready (visibility not required)
            import time
            for _ in range(15):
                try:
                    _ = acad.Documents  # type: ignore[reportUnknownMemberType]
                    break
                except Exception:
                    time.sleep(1)

            # Remember how many docs were open before we add ours
            try:
                docs_before = int(acad.Documents.Count)  # type: ignore[reportUnknownMemberType,reportUnknownArgumentType]
            except Exception:
                pass

            # Try to read the CS code from any already-open document first.
            # The original DWG (open by the user) has the Map 3D CS assigned;
            # the temp copy we're about to open may not.
            for _di in range(docs_before):
                if cs_code:
                    break
                try:
                    _existing_doc: Any = acad.Documents.Item(_di)  # type: ignore[reportUnknownMemberType,reportUnknownVariableType]
                    for var_name in ("CGEOCS", "MAPCSCODE"):
                        try:
                            val = _existing_doc.GetVariable(var_name)  # type: ignore[reportUnknownMemberType]
                            if val and str(val).strip():  # type: ignore[reportUnknownArgumentType]
                                cs_code = str(val).strip()  # type: ignore[reportUnknownArgumentType]
                                break
                        except Exception:
                            continue
                except Exception:
                    continue

            # Open the COPY (not the original) — no lock conflicts
            doc = acad.Documents.Open(tmp_dwg, False)  # type: ignore[reportUnknownMemberType]  # False = not ReadOnly

            # If CS not found from existing docs, try the temp copy
            if not cs_code:
                for var_name in ("CGEOCS", "MAPCSCODE"):
                    try:
                        val = doc.GetVariable(var_name)  # type: ignore[reportUnknownMemberType]
                        if val and str(val).strip():  # type: ignore[reportUnknownArgumentType]
                            cs_code = str(val).strip()  # type: ignore[reportUnknownArgumentType]
                            break
                    except Exception:
                        continue

            # acR2018_dxf = 61
            AC_SAVE_AS_R2018_DXF = 61
            doc.SaveAs(dxf_path, AC_SAVE_AS_R2018_DXF)  # type: ignore[reportUnknownMemberType]

        finally:
            if doc is not None:
                try:
                    doc.Close(False)  # type: ignore[reportUnknownMemberType]  # False = don't save changes
                except Exception:
                    pass
            # Quit AutoCAD if it had no documents open before we started
            if acad is not None and docs_before == 0:
                import time as _t
                _t.sleep(1)  # give AutoCAD a moment to finish closing the doc
                try:
                    # Double-check no documents remain
                    _remaining = int(acad.Documents.Count)  # type: ignore[reportUnknownMemberType,reportUnknownArgumentType]
                    if _remaining == 0:
                        acad.Quit()  # type: ignore[reportUnknownMemberType]
                except Exception:
                    # If Documents.Count fails, AutoCAD may already be shutting down
                    try:
                        acad.Quit()  # type: ignore[reportUnknownMemberType]
                    except Exception:
                        pass
            # Remove the temp DWG copy
            try:
                os.remove(tmp_dwg)
            except Exception:
                pass

        if not os.path.isfile(dxf_path):
            raise RuntimeError(
                "AutoCAD SaveAs did not produce a DXF file.\n"
                "Ensure AutoCAD is installed and the DWG is valid.")

        # If AutoCAD didn't provide a CS code, try to find a .prj file
        # near the DWG and resolve it to an Autodesk CS code.
        # Walk up to 3 levels from the DWG to find project-level .prj files
        # (e.g. DWG in ASBUILT/WORKING/, .prj in PROJECT_DATA/SHP/)
        if not cs_code:
            import glob as _glob
            dwg_dir = os.path.dirname(os.path.abspath(dwg_path))
            _search_dirs = [dwg_dir]
            _d = dwg_dir
            for _ in range(3):
                _d = os.path.dirname(_d)
                if _d and _d != os.path.dirname(_d):  # stop at root
                    _search_dirs.append(_d)
            for _search_dir in _search_dirs:
                for _prj in _glob.glob(os.path.join(_search_dir, "**", "*.prj"), recursive=True):
                    try:
                        with open(_prj, encoding="utf-8") as _pf:
                            _wkt = _pf.read().strip()
                        if _wkt and "PROJCS" in _wkt:
                            from pyproj import CRS as _CRS  # type: ignore[import-untyped]
                            _prj_crs: Any = _CRS.from_wkt(_wkt)  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                            _prj_epsg: Any = _prj_crs.to_epsg()  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                            if _prj_epsg:
                                # Reverse-lookup the Autodesk code from the EPSG
                                _ad_map = self._load_autodesk_cs_map()  # type: ignore[attr-defined]
                                for _ak, _av in _ad_map.items():
                                    if _av == int(_prj_epsg):  # type: ignore[reportUnknownArgumentType]
                                        cs_code = _ak
                                        break
                                if not cs_code:
                                    cs_code = str(_prj_epsg)  # type: ignore[reportUnknownArgumentType]
                                break
                    except Exception:
                        continue
                if cs_code:
                    break

        return dxf_path, cs_code

    def _read_dwg_geometry_ezdxf(
        self, dwg_path: str,
    ) -> tuple[dict[str, int], dict[str, list[dict[str, Any]]], str]:
        """Read DWG/DXF geometry using ezdxf.

        If the file is a .dwg, it is first converted to .dxf via AutoCAD COM
        SaveAs, and the Map 3D coordinate system is queried at that time.

        Returns ``(layer_counts, geometry_dict, cs_code)`` where *cs_code*
        is the coordinate-system identifier from AutoCAD Map 3D (empty
        string if not detected).
        """
        import ezdxf  # type: ignore[import-untyped]

        read_path = dwg_path
        tmp_dxf: str | None = None
        cs_code = ""

        # ezdxf can only read DXF — convert DWG first via AutoCAD COM
        if dwg_path.lower().endswith(".dwg"):
            tmp_dxf, cs_code = self._convert_dwg_to_dxf(dwg_path)
            read_path = tmp_dxf

        try:
            doc = ezdxf.readfile(read_path)  # type: ignore[reportUnknownMemberType,reportPrivateImportUsage]
            msp = doc.modelspace()

            layer_counts: dict[str, int] = {}
            geom: dict[str, list[dict[str, Any]]] = {}

            # Sentinel elevation value — skip points with this Z
            _NO_ELEV = -99999999.0

            def _clean_z(z: float) -> float:
                """Replace sentinel elevations with 0."""
                return 0.0 if z <= _NO_ELEV else z

            # Build layer color lookup for ByLayer (ACI 256) resolution
            _layer_colors: dict[str, int] = {}
            try:
                for _lyr in doc.layers:  # type: ignore[reportUnknownMemberType]
                    _ln: str = _lyr.dxf.name  # type: ignore[reportUnknownMemberType]
                    _lc: int = _lyr.color  # type: ignore[reportUnknownMemberType]
                    _layer_colors[_ln] = _lc
            except Exception:
                pass

            for entity in msp:
                dxf = entity.dxf
                layer = dxf.layer if dxf.hasattr("layer") else "0"
                color = dxf.color if dxf.hasattr("color") else 256
                etype = entity.dxftype()

                # Resolve ByLayer (256) and ByBlock (0) to actual layer color
                if color in (0, 256):
                    color = _layer_colors.get(layer, 7)

                if etype == "LWPOLYLINE":
                    pts: list[tuple[float, float]] = [(p[0], p[1]) for p in entity.get_points(format="xy")]  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                    pts3: list[tuple[float, float, float]] = [(p[0], p[1], _clean_z(p[2] if len(p) > 2 else 0.0)) for p in entity.get_points(format="xyz")]  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                    if len(pts) < 2:
                        continue
                    closed: bool = entity.closed  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                    layer_counts[layer] = layer_counts.get(layer, 0) + 1
                    geom.setdefault(layer, [])
                    if closed:
                        if pts[-1] != pts[0]:
                            pts.append(pts[0])
                            pts3.append(pts3[0])
                        geom[layer].append({"geom_type": "POLYGON", "coords": [pts],
                                            "coords_3d": [pts3],
                                            "entity_type": "AcDbLWPolyline", "color": color})
                    else:
                        geom[layer].append({"geom_type": "LINESTRING", "coords": pts,
                                            "coords_3d": pts3,
                                            "entity_type": "AcDbLWPolyline", "color": color})

                elif etype == "POLYLINE":
                    pts: list[tuple[float, float]] = [(v.dxf.location.x, v.dxf.location.y)  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                           for v in entity.vertices if v.dxf.hasattr("location")]  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                    pts3: list[tuple[float, float, float]] = [  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                        (v.dxf.location.x, v.dxf.location.y,  # type: ignore[reportUnknownMemberType]
                         _clean_z(v.dxf.location.z if hasattr(v.dxf.location, "z") else 0.0))  # type: ignore[reportUnknownMemberType,reportUnknownArgumentType]
                        for v in entity.vertices if v.dxf.hasattr("location")]  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                    if len(pts) < 2:
                        continue
                    closed: bool = entity.is_closed  # type: ignore[reportUnknownMemberType,reportAttributeAccessIssue]
                    layer_counts[layer] = layer_counts.get(layer, 0) + 1
                    geom.setdefault(layer, [])
                    if closed:
                        if pts[-1] != pts[0]:
                            pts.append(pts[0])
                            pts3.append(pts3[0])
                        geom[layer].append({"geom_type": "POLYGON", "coords": [pts],
                                            "coords_3d": [pts3],
                                            "entity_type": "AcDbPolyline", "color": color})
                    else:
                        geom[layer].append({"geom_type": "LINESTRING", "coords": pts,
                                            "coords_3d": pts3,
                                            "entity_type": "AcDbPolyline", "color": color})

                elif etype == "LINE":
                    sp = dxf.start
                    ep = dxf.end
                    _sz = _clean_z(sp.z if hasattr(sp, "z") else 0.0)
                    _ez = _clean_z(ep.z if hasattr(ep, "z") else 0.0)
                    layer_counts[layer] = layer_counts.get(layer, 0) + 1
                    geom.setdefault(layer, [])
                    geom[layer].append({"geom_type": "LINESTRING",
                                        "coords": [(sp.x, sp.y), (ep.x, ep.y)],
                                        "coords_3d": [(sp.x, sp.y, _sz), (ep.x, ep.y, _ez)],
                                        "entity_type": "AcDbLine", "color": color})

                elif etype == "POINT":
                    loc = dxf.location
                    _pz = loc.z if hasattr(loc, "z") else 0.0
                    # Skip points with sentinel "no data" elevations
                    if _pz <= _NO_ELEV or loc.x <= _NO_ELEV or loc.y <= _NO_ELEV:
                        continue
                    _pz = _clean_z(_pz)
                    layer_counts[layer] = layer_counts.get(layer, 0) + 1
                    geom.setdefault(layer, [])
                    geom[layer].append({"geom_type": "POINT",
                                        "coords": (loc.x, loc.y),
                                        "coords_3d": (loc.x, loc.y, _pz),
                                        "entity_type": "AcDbPoint", "color": color})

            # cs_code was already obtained from AutoCAD COM during conversion

        finally:
            # Clean up temp DXF directory
            if tmp_dxf:
                try:
                    import shutil
                    shutil.rmtree(os.path.dirname(tmp_dxf), ignore_errors=True)
                except Exception:
                    pass

        return layer_counts, geom, cs_code

    def _parse_dxf_geometry(
        self, dxf_path: str,
    ) -> tuple[dict[str, int], dict[str, list[dict[str, Any]]]]:
        """Parse a DXF file and extract geometry by layer.

        Returns ``(layer_counts, geometry_dict)`` where geometry_dict is
        ``{layer_name: [entity_dict, ...]}``, same format as _extract_dwg_geometry.
        This is much faster than COM iteration since DXF is a text file.
        """
        layer_counts: dict[str, int] = {}
        geom: dict[str, list[dict[str, Any]]] = {}

        _supported = {"LWPOLYLINE", "POLYLINE", "LINE", "POINT"}

        with open(dxf_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()

        # DXF is pairs of (group_code, value) lines
        i = 0
        n = len(lines)

        def _next_pair() -> tuple[int, str] | None:
            nonlocal i
            while i + 1 < n:
                try:
                    code = int(lines[i].strip())
                except (ValueError, IndexError):
                    i += 2
                    continue
                val = lines[i + 1].strip()
                i += 2
                return code, val
            return None

        # Skip to ENTITIES section
        in_entities = False
        while i < n:
            pair = _next_pair()
            if pair is None:
                break
            code, val = pair
            if code == 2 and val == "ENTITIES":
                in_entities = True
                break

        if not in_entities:
            return layer_counts, geom

        # Parse entities
        while i < n:
            pair = _next_pair()
            if pair is None:
                break
            code, val = pair

            # End of ENTITIES section
            if code == 0 and val == "ENDSEC":
                break

            if code != 0 or val not in _supported:
                continue

            ent_type = val
            layer = "0"
            color = 256
            x_vals: list[float] = []
            y_vals: list[float] = []
            closed = False
            # For LINE entities
            x1 = y1 = x2 = y2 = 0.0

            # Read entity properties until next entity (code 0)
            while i < n:
                pair = _next_pair()
                if pair is None:
                    break
                gc, gv = pair

                if gc == 0:
                    # We've hit the next entity — back up
                    i -= 2
                    break

                if gc == 8:
                    layer = gv
                elif gc == 62:
                    try:
                        color = int(gv)
                    except ValueError:
                        pass
                elif gc == 70 and ent_type in ("LWPOLYLINE", "POLYLINE"):
                    try:
                        flags = int(gv)
                        closed = bool(flags & 1)
                    except ValueError:
                        pass
                elif ent_type == "LWPOLYLINE":
                    if gc == 10:
                        x_vals.append(float(gv))
                    elif gc == 20:
                        y_vals.append(float(gv))
                elif ent_type == "LINE":
                    if gc == 10:
                        x1 = float(gv)
                    elif gc == 20:
                        y1 = float(gv)
                    elif gc == 11:
                        x2 = float(gv)
                    elif gc == 21:
                        y2 = float(gv)
                elif ent_type == "POINT":
                    if gc == 10:
                        x1 = float(gv)
                    elif gc == 20:
                        y1 = float(gv)

            # Build geometry entry
            layer_counts[layer] = layer_counts.get(layer, 0) + 1
            geom.setdefault(layer, [])

            if ent_type == "LWPOLYLINE":
                coords = list(zip(x_vals, y_vals))
                if len(coords) < 2:
                    continue
                if closed:
                    if coords[-1] != coords[0]:
                        coords.append(coords[0])
                    geom[layer].append({
                        "geom_type": "POLYGON", "coords": [coords],
                        "entity_type": "AcDbLWPolyline", "color": color,
                    })
                else:
                    geom[layer].append({
                        "geom_type": "LINESTRING", "coords": coords,
                        "entity_type": "AcDbLWPolyline", "color": color,
                    })
            elif ent_type == "LINE":
                geom[layer].append({
                    "geom_type": "LINESTRING",
                    "coords": [(x1, y1), (x2, y2)],
                    "entity_type": "AcDbLine", "color": color,
                })
            elif ent_type == "POINT":
                geom[layer].append({
                    "geom_type": "POINT",
                    "coords": (x1, y1),
                    "entity_type": "AcDbPoint", "color": color,
                })
            # POLYLINE (heavy/3D) vertices are sub-entities — skip for now,
            # fallback COM path handles them.

        return layer_counts, geom

    def _extract_dwg_geometry(
        self, doc: Any, selected_layers: set[str],
    ) -> dict[str, list[dict[str, Any]]]:
        """Extract geometry from selected layers in an open AutoCAD document.

        Returns ``{layer_name: [entity_dict, ...]}``.
        Each *entity_dict* has keys ``geom_type``, ``coords``,
        ``entity_type``, ``color``.
        """
        result: dict[str, list[dict[str, Any]]] = {ly: [] for ly in selected_layers}
        ms = doc.ModelSpace

        for i in range(ms.Count):
            try:
                ent = ms.Item(i)
            except Exception:
                continue
            layer = ent.Layer
            if layer not in selected_layers:
                continue

            ent_name = ent.EntityName
            color = 256
            try:
                color = ent.color
            except Exception:
                pass

            try:
                if ent_name == "AcDbLWPolyline":
                    raw = list(ent.Coordinates)
                    coords = [(raw[j], raw[j + 1])
                              for j in range(0, len(raw), 2)]
                    if len(coords) < 2:
                        continue
                    try:
                        closed = bool(ent.Closed)
                    except Exception:
                        closed = False
                    if closed:
                        if coords[-1] != coords[0]:
                            coords.append(coords[0])
                        result[layer].append({
                            "geom_type": "POLYGON", "coords": [coords],
                            "entity_type": ent_name, "color": color,
                        })
                    else:
                        result[layer].append({
                            "geom_type": "LINESTRING", "coords": coords,
                            "entity_type": ent_name, "color": color,
                        })

                elif ent_name in ("AcDb3dPolyline", "AcDbPolyline",
                                  "AcDb2dPolyline"):
                    raw = list(ent.Coordinates)
                    # Heavy / 3D polylines store x, y, z triples
                    stride = 3
                    coords = [(raw[j], raw[j + 1])
                              for j in range(0, len(raw), stride)]
                    if len(coords) < 2:
                        continue
                    try:
                        closed = bool(ent.Closed)
                    except Exception:
                        closed = False
                    if closed:
                        if coords[-1] != coords[0]:
                            coords.append(coords[0])
                        result[layer].append({
                            "geom_type": "POLYGON", "coords": [coords],
                            "entity_type": ent_name, "color": color,
                        })
                    else:
                        result[layer].append({
                            "geom_type": "LINESTRING", "coords": coords,
                            "entity_type": ent_name, "color": color,
                        })

                elif ent_name == "AcDbLine":
                    sp = ent.StartPoint
                    ep = ent.EndPoint
                    result[layer].append({
                        "geom_type": "LINESTRING",
                        "coords": [(sp[0], sp[1]), (ep[0], ep[1])],
                        "entity_type": "AcDbLine", "color": color,
                    })

                elif ent_name == "AcDbPoint":
                    pt = ent.Coordinates
                    result[layer].append({
                        "geom_type": "POINT",
                        "coords": (pt[0], pt[1]),
                        "entity_type": "AcDbPoint", "color": color,
                    })
            except Exception:
                continue

        return {ly: ents for ly, ents in result.items() if ents}

    def _write_dwg_geometry_to_gpkg(
        self,
        gpkg_path: str,
        layer_data: dict[str, list[dict[str, Any]]],
        srid: int = 0,
        cs_description: str = "",
        transformer: Any = None,
    ) -> int:
        """Append DWG geometry as new tables in an existing GeoPackage.

        Returns the total number of features written.
        """
        import sqlite3
        total = 0

        with sqlite3.connect(gpkg_path) as conn:
            # Ensure the SRS is registered
            cur = conn.execute(
                "SELECT 1 FROM gpkg_spatial_ref_sys WHERE srs_id = ?", (srid,))
            if not cur.fetchone():
                conn.execute(
                    "INSERT INTO gpkg_spatial_ref_sys "
                    "(srs_name, srs_id, organization, "
                    "organization_coordsys_id, definition, description) "
                    "VALUES (?,?,?,?,?,?)",
                    (cs_description or f"EPSG:{srid}", srid,
                     "EPSG" if srid > 0 else "NONE", srid,
                     "undefined",
                     cs_description or "DWG coordinate system"))

            for layer_name, entities in sorted(layer_data.items()):
                if not entities:
                    continue

                safe_name = "dwg_" + re.sub(r"[^A-Za-z0-9_]", "_", layer_name)

                # Dominant geometry type for GPKG metadata
                tcnt: dict[str, int] = {}
                for e in entities:
                    tcnt[e["geom_type"]] = tcnt.get(e["geom_type"], 0) + 1
                gpkg_geom = (max(tcnt, key=lambda k: tcnt[k])
                             if len(tcnt) == 1 else "GEOMETRY")

                conn.execute(f'''CREATE TABLE IF NOT EXISTS "{safe_name}" (
                    fid INTEGER PRIMARY KEY AUTOINCREMENT,
                    geom BLOB,
                    entity_type TEXT,
                    geom_type TEXT,
                    color INTEGER,
                    vertices INTEGER,
                    horiz_dist REAL,
                    slope_dist REAL,
                    source_layer TEXT
                )''')

                conn.execute(
                    "INSERT INTO gpkg_geometry_columns "
                    "VALUES (?,?,?,?,0,0)",
                    (safe_name, "geom", gpkg_geom, srid))

                min_x = min_y = float("inf")
                max_x = max_y = float("-inf")

                for ent in entities:
                    geom_blob: bytes | None = None

                    if ent["geom_type"] == "LINESTRING":
                        pts: list[tuple[float, float]] = ent["coords"]
                        if transformer:
                            pts = [transformer(x, y) for x, y in pts]
                        geom_blob = self._make_gpkg_linestring_blob(  # type: ignore[attr-defined]
                            pts, srid)
                        for x, y in pts:
                            min_x = min(min_x, x)
                            min_y = min(min_y, y)
                            max_x = max(max_x, x)
                            max_y = max(max_y, y)

                    elif ent["geom_type"] == "POLYGON":
                        rings = ent["coords"]
                        if transformer:
                            rings = [[transformer(x, y) for x, y in r]
                                     for r in rings]
                        geom_blob = self._make_gpkg_polygon_blob(  # type: ignore[attr-defined]
                            rings, srid)
                        for ring in rings:
                            for x, y in ring:
                                min_x = min(min_x, x)
                                min_y = min(min_y, y)
                                max_x = max(max_x, x)
                                max_y = max(max_y, y)

                    elif ent["geom_type"] == "POINT":
                        px, py = ent["coords"]
                        if transformer:
                            px, py = transformer(px, py)
                        geom_blob = self._make_gpkg_point_blob(  # type: ignore[attr-defined]
                            px, py, srid)
                        min_x = min(min_x, px)
                        min_y = min(min_y, py)
                        max_x = max(max_x, px)
                        max_y = max(max_y, py)

                    if geom_blob:
                        import math as _math
                        _gt = ent["geom_type"]
                        _c3d: list[Any] = ent.get("coords_3d", [])
                        if _gt == "LINESTRING":
                            _nv = len(ent["coords"])
                        elif _gt == "POLYGON":
                            _nv = sum(len(r) for r in ent["coords"])
                        else:
                            _nv = 1
                        # Compute distances from 3D coords
                        _hd: float | None = None
                        _sd: float | None = None
                        if _gt == "LINESTRING" and _c3d and len(_c3d) >= 2:
                            _hd = 0.0
                            _sd = 0.0
                            for _di in range(len(_c3d) - 1):
                                _dx = _c3d[_di+1][0] - _c3d[_di][0]
                                _dy = _c3d[_di+1][1] - _c3d[_di][1]
                                _dz = _c3d[_di+1][2] - _c3d[_di][2]
                                _hd += _math.sqrt(_dx*_dx + _dy*_dy)
                                _sd += _math.sqrt(_dx*_dx + _dy*_dy + _dz*_dz)
                        conn.execute(
                            f'INSERT INTO "{safe_name}" '
                            f"(geom, entity_type, geom_type, color, vertices, "
                            f"horiz_dist, slope_dist, source_layer) "
                            f"VALUES (?,?,?,?,?,?,?,?)",
                            (geom_blob, ent["entity_type"], _gt,
                             ent["color"], _nv, _hd, _sd, layer_name))
                        total += 1

                conn.execute(
                    "INSERT INTO gpkg_contents "
                    "(table_name, data_type, identifier, description, "
                    "min_x, min_y, max_x, max_y, srs_id) "
                    "VALUES (?,'features',?,?,?,?,?,?,?)",
                    (safe_name, safe_name, f"DWG layer: {layer_name}",
                     min_x if min_x != float("inf") else None,
                     min_y if min_y != float("inf") else None,
                     max_x if max_x != float("-inf") else None,
                     max_y if max_y != float("-inf") else None,
                     srid))

            conn.commit()
        return total

    def _write_dwg_geometry_to_shp(
        self,
        output_path: str,
        geometry: dict[str, list[dict[str, Any]]],
        srid: int = 0,
    ) -> None:
        """Append DWG line/polygon geometry to a Shapefile.

        Creates a single combined Shapefile with all DWG entities.
        Since shapefiles support only one geometry type, writes
        POLYLINE (type 3) which can hold both lines and polygon rings.
        """
        # Collect all entities as polylines (polygons become closed polylines)
        all_entities: list[dict[str, Any]] = []
        for layer_name, ents in geometry.items():
            for ent in ents:
                all_entities.append({**ent, "layer": layer_name})

        if not all_entities:
            return

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        base = os.path.splitext(output_path)[0]

        # .prj — write the CRS definition so GIS software knows the coordinate system
        prj_wkt = ""
        if srid:
            try:
                from pyproj import CRS as _CRS  # type: ignore[import-untyped]
                _prj_crs: Any = _CRS.from_epsg(srid)  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                prj_wkt = str(_prj_crs.to_wkt("WKT1_ESRI"))  # type: ignore[reportUnknownMemberType]
            except Exception:
                if srid == 4326:
                    prj_wkt = (
                        'GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",'
                        'SPHEROID["WGS_1984",6378137.0,298.257223563]],'
                        'PRIMEM["Greenwich",0.0],'
                        'UNIT["Degree",0.0174532925199433]]'
                    )
        if prj_wkt:
            with open(base + ".prj", "w", encoding="utf-8") as f:
                f.write(prj_wkt)

        # DBF fields
        dbf_fields: list[tuple[str, str, int, int]] = [
            ("layer", "C", 50, 0),
            ("geom_type", "C", 15, 0),
            ("ent_type", "C", 30, 0),
            ("color", "N", 6, 0),
            ("color_rgb", "C", 12, 0),
            ("vertices", "N", 8, 0),
            ("horiz_dist", "N", 14, 2),
            ("slope_dist", "N", 14, 2),
        ]
        # ACI → RGB name for the shapefile attribute
        _ACI_NAMES: dict[int, str] = {
            1: "Red", 2: "Yellow", 3: "Green", 4: "Cyan",
            5: "Blue", 6: "Magenta", 7: "White", 8: "DkGray",
            9: "LtGray", 256: "ByLayer", 0: "ByBlock",
        }
        record_size = 1
        for _, _, sz, _ in dbf_fields:
            record_size += sz

        # Compute bounding box
        all_x: list[float] = []
        all_y: list[float] = []
        for ent in all_entities:
            gt = ent["geom_type"]
            if gt == "POINT":
                all_x.append(ent["coords"][0])
                all_y.append(ent["coords"][1])
            elif gt == "LINESTRING":
                for x, y in ent["coords"]:
                    all_x.append(x)
                    all_y.append(y)
            elif gt == "POLYGON":
                for ring in ent["coords"]:
                    for x, y in ring:
                        all_x.append(x)
                        all_y.append(y)

        xmin = min(all_x) if all_x else 0.0
        ymin = min(all_y) if all_y else 0.0
        xmax = max(all_x) if all_x else 0.0
        ymax = max(all_y) if all_y else 0.0

        # Build SHP records
        num_records = len(all_entities)
        shp_data = bytearray()
        shx_offsets: list[tuple[int, int]] = []
        shp_file_len = 50  # header in 16-bit words

        for ent in all_entities:
            gt = ent["geom_type"]
            offset = shp_file_len

            if gt == "POINT":
                px, py = ent["coords"]
                rec_content = struct.pack("<I", 1)  # Point
                rec_content += struct.pack("<dd", px, py)
            elif gt == "LINESTRING":
                coords = ent["coords"]
                ex = [c[0] for c in coords]
                ey = [c[1] for c in coords]
                rec_content = struct.pack("<I", 3)  # PolyLine
                rec_content += struct.pack("<dddd", min(ex), min(ey), max(ex), max(ey))
                rec_content += struct.pack("<II", 1, len(coords))  # 1 part
                rec_content += struct.pack("<I", 0)  # part index
                for x, y in coords:
                    rec_content += struct.pack("<dd", x, y)
            elif gt == "POLYGON":
                rings = ent["coords"]
                total_pts = sum(len(r) for r in rings)
                ex = [pt[0] for r in rings for pt in r]
                ey = [pt[1] for r in rings for pt in r]
                rec_content = struct.pack("<I", 5)  # Polygon
                rec_content += struct.pack("<dddd", min(ex), min(ey), max(ex), max(ey))
                rec_content += struct.pack("<II", len(rings), total_pts)
                idx = 0
                for ring in rings:
                    rec_content += struct.pack("<I", idx)
                    idx += len(ring)
                for ring in rings:
                    for x, y in ring:
                        rec_content += struct.pack("<dd", x, y)
            else:
                continue

            content_words = len(rec_content) // 2
            rec_header = struct.pack(">II", len(shx_offsets) + 1, content_words)
            shp_data += rec_header + rec_content
            shx_offsets.append((offset, content_words))
            shp_file_len += 4 + content_words  # header (4 words) + content

        # Determine dominant shape type for the file header
        has_poly = any(e["geom_type"] in ("LINESTRING", "POLYGON") for e in all_entities)
        has_polygon = any(e["geom_type"] == "POLYGON" for e in all_entities)
        if has_polygon:
            shape_type = 5  # Polygon
        elif has_poly:
            shape_type = 3  # PolyLine
        else:
            shape_type = 1  # Point

        # Write .shp
        with open(base + ".shp", "wb") as f:
            f.write(struct.pack(">I", 9994))
            f.write(b"\x00" * 20)
            f.write(struct.pack(">I", shp_file_len))
            f.write(struct.pack("<I", 1000))
            f.write(struct.pack("<I", shape_type))
            f.write(struct.pack("<dddd", xmin, ymin, xmax, ymax))
            f.write(struct.pack("<dddd", 0.0, 0.0, 0.0, 0.0))  # z,m ranges
            f.write(bytes(shp_data))

        # Write .shx
        shx_len = 50 + num_records * 4
        with open(base + ".shx", "wb") as f:
            f.write(struct.pack(">I", 9994))
            f.write(b"\x00" * 20)
            f.write(struct.pack(">I", shx_len))
            f.write(struct.pack("<I", 1000))
            f.write(struct.pack("<I", shape_type))
            f.write(struct.pack("<dddd", xmin, ymin, xmax, ymax))
            f.write(struct.pack("<dddd", 0.0, 0.0, 0.0, 0.0))
            for off, clen in shx_offsets:
                f.write(struct.pack(">II", off, clen))

        # Write .dbf
        today = datetime.date.today()
        header_size = 32 + len(dbf_fields) * 32 + 1
        with open(base + ".dbf", "wb") as f:
            f.write(struct.pack("<B", 3))
            f.write(struct.pack("<3B", today.year - 1900, today.month, today.day))
            f.write(struct.pack("<I", num_records))
            f.write(struct.pack("<H", header_size))
            f.write(struct.pack("<H", record_size))
            f.write(b"\x00" * 20)
            for fname, ftype, fsize, fdec in dbf_fields:
                f.write(fname.encode("latin-1").ljust(11, b"\x00")[:11])
                f.write(ftype.encode("latin-1"))
                f.write(b"\x00" * 4)
                f.write(struct.pack("<B", fsize))
                f.write(struct.pack("<B", fdec))
                f.write(b"\x00" * 14)
            f.write(b"\r")
            for ent in all_entities:
                rec = b"\x20"
                import math as _math
                _aci = ent.get("color", 256)
                _cname = _ACI_NAMES.get(_aci, f"ACI {_aci}")
                _gt = ent["geom_type"]
                _c3d: list[Any] = ent.get("coords_3d", [])
                if _gt == "LINESTRING":
                    _nv = len(ent["coords"])
                elif _gt == "POLYGON":
                    _nv = sum(len(r) for r in ent["coords"])
                elif _gt == "POINT":
                    _nv = 1
                else:
                    _nv = 0
                # Compute distances from 3D coords
                _hd = 0.0
                _sd = 0.0
                if _gt == "LINESTRING" and _c3d and len(_c3d) >= 2:
                    for _di in range(len(_c3d) - 1):
                        _dx = _c3d[_di+1][0] - _c3d[_di][0]
                        _dy = _c3d[_di+1][1] - _c3d[_di][1]
                        _dz = _c3d[_di+1][2] - _c3d[_di][2]
                        _hd += _math.sqrt(_dx*_dx + _dy*_dy)
                        _sd += _math.sqrt(_dx*_dx + _dy*_dy + _dz*_dz)
                vals: list[str] = [ent.get("layer", ""), _gt,
                        ent.get("entity_type", ""), str(_aci), _cname, str(_nv),
                        f"{_hd:.2f}", f"{_sd:.2f}"]
                for (_, ftype, fsize, _), val in zip(dbf_fields, vals):  # type: ignore[reportUnknownVariableType]
                    s = str(val)
                    if ftype == "N":
                        rec += s.rjust(fsize)[:fsize].encode("latin-1")
                    else:
                        rec += s.ljust(fsize)[:fsize].encode("latin-1", errors="replace")
                f.write(rec)
            f.write(b"\x1a")

    # Lazily-loaded Autodesk CS code → EPSG mapping from NameMapper.csv
    _AUTODESK_CS_TO_EPSG: dict[str, int] | None = None

    @staticmethod
    def _load_autodesk_cs_map() -> dict[str, int]:
        """Load the Autodesk→EPSG mapping from NameMapper.csv (4000+ entries).

        The file ships with AutoCAD Map 3D at:
        C:\\ProgramData\\Autodesk\\Geospatial Coordinate Systems <ver>\\NameMapper.csv
        """
        if DataValidationTool._AUTODESK_CS_TO_EPSG is not None:
            return DataValidationTool._AUTODESK_CS_TO_EPSG

        import csv
        import glob as _glob
        from collections import defaultdict

        mapping: dict[str, int] = {}

        # Find the NameMapper.csv
        nm_path = ""
        for pattern in [
            r"C:\ProgramData\Autodesk\Geospatial Coordinate Systems*\NameMapper.csv",
        ]:
            hits = _glob.glob(pattern)
            if hits:
                nm_path = hits[-1]  # latest version
                break

        if nm_path and os.path.isfile(nm_path):
            try:
                with open(nm_path, encoding="utf-8") as f:
                    rows = list(csv.DictReader(f))

                # Group by GenericId
                groups: dict[str, list[dict[str, str]]] = defaultdict(list)
                for r in rows:
                    groups[r["GenericId"]].append(r)

                for _gid, members in groups.items():
                    ad_name: str | None = None
                    epsg_code: int | None = None
                    for m in members:
                        if m["Flavor"] == "Autodesk":
                            ad_name = m["NameId"]
                        if m["Flavor"] == "EPSG" and m["NumericId"] != "0":
                            epsg_code = int(m["NumericId"])
                    if ad_name and epsg_code:
                        mapping[ad_name] = epsg_code
            except Exception:
                pass

        DataValidationTool._AUTODESK_CS_TO_EPSG = mapping
        return mapping

    @staticmethod
    def _resolve_autodesk_cs(cs_code: str) -> int:
        """Resolve an Autodesk Map 3D CS code to an EPSG integer.

        Uses the NameMapper.csv from the Autodesk install (4000+ mappings),
        then falls back to pyproj and direct integer parsing.
        Returns 0 if unresolvable.
        """
        if not cs_code or not cs_code.strip():
            return 0
        code = cs_code.strip()

        # 1. Direct EPSG integer (user typed a number)
        try:
            return int(code)
        except ValueError:
            pass

        # 2. Autodesk NameMapper.csv (CO83-NF → 2231, etc.)
        mapping = DataValidationTool._load_autodesk_cs_map()
        epsg = mapping.get(code, 0)
        if epsg:
            return epsg
        # Try case-insensitive
        code_upper = code.upper()
        for k, v in mapping.items():
            if k.upper() == code_upper:
                return v

        # 3. Try pyproj from_user_input (handles full CRS names, WKT, etc.)
        try:
            from pyproj import CRS as _CRS  # type: ignore[import-untyped]
            _crs: Any = _CRS.from_user_input(code)  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
            _e: Any = _crs.to_epsg()  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
            if _e:
                return int(_e)  # type: ignore[reportUnknownArgumentType]
        except Exception:
            pass

        return 0

    def _write_dwg_geometry_to_kml(
        self,
        output_path: str,
        geometry: dict[str, list[dict[str, Any]]],
        srid: int = 0,
        source_crs: str = "",
    ) -> None:
        """Write DWG geometry (lines/polygons) as KML appended into an existing KMZ,
        or as a standalone KML file.

        Coordinates are reprojected to WGS 84 lon/lat via pyproj.  The source
        CRS is resolved from *source_crs* (raw Autodesk/Map 3D CS code) or
        *srid* (EPSG integer), whichever is available.

        If output_path is a .kmz that already exists, the DWG geometry is added
        to the existing KML inside it. Otherwise writes a new .kmz.
        """
        import zipfile

        # Build a reprojection function — resolve CRS via mapping table + pyproj
        _reproject: Any = None  # type: ignore[reportUnknownVariableType]
        _resolved_epsg = 0
        if source_crs:
            _resolved_epsg = self._resolve_autodesk_cs(source_crs)  # type: ignore[attr-defined]
        if not _resolved_epsg and srid not in (0, 4326):
            _resolved_epsg = srid
        if _resolved_epsg and _resolved_epsg != 4326:
            try:
                from pyproj import Transformer as _PT  # type: ignore[import-untyped]
                _t: Any = _PT.from_crs(  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                    f"EPSG:{_resolved_epsg}", "EPSG:4326", always_xy=True)
                _reproject = lambda x, y: _t.transform(x, y)  # type: ignore[reportUnknownVariableType,reportUnknownLambdaType]
            except Exception:
                pass

        def _tx(x: float, y: float) -> tuple[float, float]:
            """Transform coords to lon/lat if reprojection is available."""
            if _reproject:
                return _reproject(x, y)  # type: ignore[reportUnknownVariableType]
            return x, y

        stem = os.path.splitext(os.path.basename(output_path))[0]
        kml_ns = "http://www.opengis.net/kml/2.2"

        # Try to load existing KML from KMZ
        existing_kml: bytes | None = None
        if os.path.isfile(output_path) and output_path.lower().endswith(".kmz"):
            try:
                with zipfile.ZipFile(output_path, "r") as zf:
                    existing_kml = zf.read("doc.kml")
            except Exception:
                pass

        if existing_kml:
            root = ET.fromstring(existing_kml)
            # Find the Document element
            doc_el = root.find(f"{{{kml_ns}}}Document")
            if doc_el is None:
                doc_el = root
        else:
            root = ET.Element("kml", attrib={"xmlns": kml_ns})
            doc_el = ET.SubElement(root, "Document")
            ET.SubElement(doc_el, "name").text = stem

        # AutoCAD Color Index (ACI) → KML color (aaBBGGRR hex)
        # Standard ACI: 1=red,2=yellow,3=green,4=cyan,5=blue,6=magenta,7=white
        _ACI_RGB: dict[int, tuple[int, int, int]] = {
            1: (255, 0, 0), 2: (255, 255, 0), 3: (0, 255, 0),
            4: (0, 255, 255), 5: (0, 0, 255), 6: (255, 0, 255),
            7: (255, 255, 255), 8: (128, 128, 128), 9: (192, 192, 192),
            10: (255, 0, 0), 11: (255, 127, 127), 12: (204, 0, 0),
            13: (204, 102, 102), 14: (153, 0, 0), 15: (153, 76, 76),
            20: (255, 63, 0), 30: (255, 127, 0), 40: (255, 191, 0),
            50: (255, 255, 0), 60: (191, 255, 0), 70: (127, 255, 0),
            80: (63, 255, 0), 90: (0, 255, 0), 100: (0, 255, 63),
            110: (0, 255, 127), 120: (0, 255, 191), 130: (0, 255, 255),
            140: (0, 191, 255), 150: (0, 127, 255), 160: (0, 63, 255),
            170: (0, 0, 255), 180: (63, 0, 255), 190: (127, 0, 255),
            200: (191, 0, 255), 210: (255, 0, 255), 220: (255, 0, 191),
            230: (255, 0, 127), 240: (255, 0, 63), 250: (51, 51, 51),
            251: (80, 80, 80), 252: (105, 105, 105), 253: (130, 130, 130),
            254: (190, 190, 190), 255: (255, 255, 255),
        }

        def _aci_to_kml_color(aci: int, alpha: int = 255) -> str:
            """Convert ACI color to KML aaBBGGRR hex string."""
            r, g, b = _ACI_RGB.get(aci, (255, 255, 255))
            return f"{alpha:02x}{b:02x}{g:02x}{r:02x}"

        # Create shared styles for each unique color (avoids bloat)
        _style_ids: dict[int, str] = {}

        # Add a folder for DWG geometry
        dwg_folder = ET.SubElement(doc_el, "Folder")
        ET.SubElement(dwg_folder, "name").text = "DWG Geometry"

        _ent_counter = 0
        for layer_name, ents in sorted(geometry.items()):
            layer_folder = ET.SubElement(dwg_folder, "Folder")
            ET.SubElement(layer_folder, "name").text = layer_name

            for ent in ents:
                _ent_counter += 1
                gt = ent["geom_type"]
                color_idx = ent.get("color", 256)
                if color_idx == 256 or color_idx == 0:
                    color_idx = 7  # ByLayer/ByBlock → white

                # Create/reuse style for this color
                if color_idx not in _style_ids:
                    sid = f"dwg_color_{color_idx}"
                    _style_ids[color_idx] = sid
                    style = ET.SubElement(doc_el, "Style", attrib={"id": sid})
                    ls_style = ET.SubElement(style, "LineStyle")
                    ET.SubElement(ls_style, "color").text = _aci_to_kml_color(color_idx)
                    ET.SubElement(ls_style, "width").text = "2"
                    ps_style = ET.SubElement(style, "PolyStyle")
                    ET.SubElement(ps_style, "color").text = _aci_to_kml_color(color_idx, alpha=80)
                    ET.SubElement(ps_style, "outline").text = "1"
                    icon_style = ET.SubElement(style, "IconStyle")
                    ET.SubElement(icon_style, "color").text = _aci_to_kml_color(color_idx)

                pm = ET.SubElement(layer_folder, "Placemark")
                ET.SubElement(pm, "name").text = f"{layer_name} ({gt.lower()}) #{_ent_counter}"
                ET.SubElement(pm, "styleUrl").text = f"#{_style_ids[color_idx]}"

                # Extended data — attributes visible when clicking in Google Earth
                ext = ET.SubElement(pm, "ExtendedData")
                _ext_pairs: list[tuple[str, str]] = [
                    ("Layer", layer_name),
                    ("Geometry Type", gt),
                    ("Entity Type", ent.get("entity_type", "")),
                    ("ACI Color", str(color_idx)),
                ]
                for _dk, _dv in _ext_pairs:
                    data = ET.SubElement(ext, "Data", attrib={"name": _dk})
                    ET.SubElement(data, "value").text = _dv

                # Add vertex count and line length info
                import math as _math
                if gt == "LINESTRING":
                    n_verts = len(ent["coords"])
                    data = ET.SubElement(ext, "Data", attrib={"name": "Vertices"})
                    ET.SubElement(data, "value").text = str(n_verts)
                    # Compute horizontal and slope distances from 3D coords (in source CRS units)
                    _c3d: list[tuple[float, float, float]] = ent.get("coords_3d", [])
                    if _c3d and len(_c3d) >= 2:
                        _h_dist = 0.0
                        _s_dist = 0.0
                        for _i in range(len(_c3d) - 1):
                            _dx = _c3d[_i+1][0] - _c3d[_i][0]
                            _dy = _c3d[_i+1][1] - _c3d[_i][1]
                            _dz = _c3d[_i+1][2] - _c3d[_i][2]
                            _h_dist += _math.sqrt(_dx*_dx + _dy*_dy)
                            _s_dist += _math.sqrt(_dx*_dx + _dy*_dy + _dz*_dz)
                        data = ET.SubElement(ext, "Data", attrib={"name": "Horizontal Distance (ft)"})
                        ET.SubElement(data, "value").text = f"{_h_dist:.2f}"
                        data = ET.SubElement(ext, "Data", attrib={"name": "Slope Distance (ft)"})
                        ET.SubElement(data, "value").text = f"{_s_dist:.2f}"
                elif gt == "POLYGON":
                    n_verts = sum(len(r) for r in ent["coords"])
                    data = ET.SubElement(ext, "Data", attrib={"name": "Vertices"})
                    ET.SubElement(data, "value").text = str(n_verts)

                # Geometry
                if gt == "LINESTRING":
                    ls = ET.SubElement(pm, "LineString")
                    ET.SubElement(ls, "tessellate").text = "1"
                    coord_str = " ".join(
                        f"{_tx(x, y)[0]},{_tx(x, y)[1]},0" for x, y in ent["coords"])
                    ET.SubElement(ls, "coordinates").text = coord_str

                elif gt == "POLYGON":
                    poly = ET.SubElement(pm, "Polygon")
                    ET.SubElement(poly, "tessellate").text = "1"
                    for ri, ring in enumerate(ent["coords"]):
                        boundary = ET.SubElement(
                            poly,
                            "outerBoundaryIs" if ri == 0 else "innerBoundaryIs")
                        lr = ET.SubElement(boundary, "LinearRing")
                        coord_str = " ".join(
                            f"{_tx(x, y)[0]},{_tx(x, y)[1]},0" for x, y in ring)
                        ET.SubElement(lr, "coordinates").text = coord_str

                elif gt == "POINT":
                    pt = ET.SubElement(pm, "Point")
                    lon, lat = _tx(*ent["coords"])
                    ET.SubElement(pt, "coordinates").text = f"{lon},{lat},0"

        kml_bytes = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("doc.kml", kml_bytes)

    def _write_dwg_geometry_to_landxml(
        self,
        output_path: str,
        geometry: dict[str, list[dict[str, Any]]],
        srid: int = 0,
        source_crs: str = "",
    ) -> None:
        """Append DWG geometry to an existing LandXML file as PlanFeatures.

        Coordinates are reprojected to WGS 84 lat/lon if *srid* or
        *source_crs* identifies a projected CRS.

        If the file exists, parses it and adds geometry elements.
        Otherwise creates a minimal LandXML with just the geometry.
        """
        # Build reprojection function (LandXML uses lat/lon like KML)
        _reproject: Any = None  # type: ignore[reportUnknownVariableType]
        _resolved_epsg = 0
        if source_crs:
            _resolved_epsg = self._resolve_autodesk_cs(source_crs)  # type: ignore[attr-defined]
        if not _resolved_epsg and srid not in (0, 4326):
            _resolved_epsg = srid
        if _resolved_epsg and _resolved_epsg != 4326:
            try:
                from pyproj import Transformer as _PT  # type: ignore[import-untyped]
                _t: Any = _PT.from_crs(  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                    f"EPSG:{_resolved_epsg}", "EPSG:4326", always_xy=True)
                _reproject = lambda x, y: _t.transform(x, y)  # type: ignore[reportUnknownVariableType,reportUnknownLambdaType]
            except Exception:
                pass

        def _tx(x: float, y: float) -> tuple[float, float]:
            if _reproject:
                return _reproject(x, y)  # type: ignore[reportUnknownVariableType]
            return x, y

        ns = "http://www.landxml.org/schema/LandXML-1.2"

        if os.path.isfile(output_path):
            tree = ET.parse(output_path)
            root_el = tree.getroot()
            # Strip namespace prefix for easier element creation
            # Re-register namespace to avoid ns0: prefixes
            ET.register_namespace("", ns)
        else:
            ET.register_namespace("", ns)
            now = datetime.datetime.now()
            root_el = ET.Element("LandXML", attrib={
                "xmlns": ns,
                "version": "1.2",
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%H:%M:%S"),
            })
            stem = os.path.splitext(os.path.basename(output_path))[0]
            ET.SubElement(root_el, "Project", attrib={"name": stem})
            tree = ET.ElementTree(root_el)

        # Add PlanFeatures for each layer
        _ent_n = 0
        for layer_name, ents in sorted(geometry.items()):
            plan_feat = ET.SubElement(root_el, "PlanFeatures",
                                      attrib={"name": layer_name})
            for ent in ents:
                _ent_n += 1
                import math as _math
                gt = ent["geom_type"]
                _color = ent.get("color", 256)
                _etype = ent.get("entity_type", "")
                _c3d: list[Any] = ent.get("coords_3d", [])
                if gt == "LINESTRING":
                    _nv = len(ent["coords"])
                    _hd = 0.0
                    _sd = 0.0
                    if _c3d and len(_c3d) >= 2:
                        for _di in range(len(_c3d) - 1):
                            _dx = _c3d[_di+1][0] - _c3d[_di][0]
                            _dy = _c3d[_di+1][1] - _c3d[_di][1]
                            _dz = _c3d[_di+1][2] - _c3d[_di][2]
                            _hd += _math.sqrt(_dx*_dx + _dy*_dy)
                            _sd += _math.sqrt(_dx*_dx + _dy*_dy + _dz*_dz)
                    pf = ET.SubElement(plan_feat, "PlanFeature",
                                       attrib={"name": f"{layer_name}_line_{_ent_n}",
                                               "desc": f"layer={layer_name} type={_etype} color={_color} vertices={_nv} horiz_dist={_hd:.2f} slope_dist={_sd:.2f}"})
                    coord_geom = ET.SubElement(pf, "CoordGeom")
                    coords = ent["coords"]
                    for j in range(len(coords) - 1):
                        lon1, lat1 = _tx(coords[j][0], coords[j][1])
                        lon2, lat2 = _tx(coords[j+1][0], coords[j+1][1])
                        line = ET.SubElement(coord_geom, "Line")
                        start = ET.SubElement(line, "Start")
                        start.text = f"{lat1} {lon1}"
                        end = ET.SubElement(line, "End")
                        end.text = f"{lat2} {lon2}"

                elif gt == "POLYGON":
                    _nv = sum(len(r) for r in ent["coords"])
                    pf = ET.SubElement(plan_feat, "PlanFeature",
                                       attrib={"name": f"{layer_name}_polygon_{_ent_n}",
                                               "desc": f"layer={layer_name} type={_etype} color={_color} vertices={_nv}"})
                    coord_geom = ET.SubElement(pf, "CoordGeom")
                    for ring in ent["coords"]:
                        for j in range(len(ring) - 1):
                            lon1, lat1 = _tx(ring[j][0], ring[j][1])
                            lon2, lat2 = _tx(ring[j+1][0], ring[j+1][1])
                            line = ET.SubElement(coord_geom, "Line")
                            start = ET.SubElement(line, "Start")
                            start.text = f"{lat1} {lon1}"
                            end = ET.SubElement(line, "End")
                            end.text = f"{lat2} {lon2}"

                elif gt == "POINT":
                    pf = ET.SubElement(plan_feat, "PlanFeature",
                                       attrib={"name": f"{layer_name}_point_{_ent_n}",
                                               "desc": f"layer={layer_name} type={_etype} color={_color}"})
                    coord_geom = ET.SubElement(pf, "CoordGeom")
                    lon, lat = _tx(*ent["coords"])
                    il = ET.SubElement(coord_geom, "IrregularLine")
                    start = ET.SubElement(il, "Start")
                    start.text = f"{lat} {lon}"
                    end = ET.SubElement(il, "End")
                    end.text = f"{lat} {lon}"

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        ET.indent(tree, space="  ")
        tree.write(output_path, xml_declaration=True, encoding="UTF-8")

    def _show_dwg_layer_dialog(
        self, dwg_path: str, gpkg_path: str,
        previous_layers: list[str] | None = None,
    ) -> tuple[list[str], dict[str, list[dict[str, Any]]], int, str]:
        """Show a layer-selection popup, extract DWG geometry, append to GPKG.

        *previous_layers*, if provided, pre-checks those layers in the dialog
        (from a prior export stored in the watchlist).  All other layers
        default to unchecked.

        Returns ``(exported_layer_names, geometry_dict, write_srid, cs_code)``.
        """
        self.status.config(text="Reading DWG file…")
        self.update_idletasks()

        exported_layers: list[str] = []
        extracted_geom: dict[str, list[dict[str, Any]]] = {}
        final_srid: list[int] = [0]
        cs_code = ""

        try:
            layer_counts, _prescan_geom, cs_code = self._read_dwg_geometry_ezdxf(dwg_path)  # type: ignore[attr-defined]
        except Exception as exc:
            self.status.config(text="Ready")
            messagebox.showerror(
                "DWG Read Error",
                f"Could not read DWG geometry:\n{exc}\n\n"
                "Ensure the DWG file is accessible.")
            return exported_layers, extracted_geom, final_srid[0], cs_code

        self.status.config(text="Ready")

        if not layer_counts:
            messagebox.showinfo(
                "No Geometry",
                "No exportable geometry found in the DWG model space.\n"
                "(Supported: polylines, lines, points)")
            return exported_layers, extracted_geom, final_srid[0], cs_code

        # --- Build layer-selection dialog ---
        win = tk.Toplevel(self)
        win.title("Select DWG Layers for Export")
        win.resizable(False, False)
        win.grab_set()
        win.lift()  # type: ignore[arg-type]
        win.focus_force()

        tk.Label(
            win, text="Select DWG Layers to Export",
            font=("Helvetica", 11, "bold"),
        ).pack(padx=10, pady=(10, 4), anchor="w")
        tk.Label(
            win, text=f"DWG: {os.path.basename(dwg_path)}",
            fg="#1F4E79",
        ).pack(padx=10, pady=2, anchor="w")
        if cs_code:
            tk.Label(
                win, text=f"Map 3D CS code: {cs_code}",
            ).pack(padx=10, pady=2, anchor="w")
        else:
            tk.Label(
                win, text="Map 3D CS code: not detected — enter EPSG below",
                fg="#CC0000",
            ).pack(padx=10, pady=2, anchor="w")

        tk.Label(
            win, text="Check the layers to include as geometry:",
        ).pack(padx=10, pady=(8, 4), anchor="w")

        # Scrollable checkbox list
        list_frame = tk.Frame(win)
        list_frame.pack(padx=10, pady=4, fill="both", expand=True)

        canvas = tk.Canvas(
            list_frame,
            height=min(300, len(layer_counts) * 24 + 10),
            width=380,
        )
        scrollbar = ttk.Scrollbar(
            list_frame, orient="vertical", command=canvas.yview)  # type: ignore[arg-type]
        inner = tk.Frame(canvas)
        inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        if len(layer_counts) > 12:
            scrollbar.pack(side="right", fill="y")

        _prev_set: set[str] = set(previous_layers) if previous_layers else set()
        layer_vars: dict[str, tk.BooleanVar] = {}
        for lname in sorted(layer_counts):
            var = tk.BooleanVar(value=(lname in _prev_set))
            layer_vars[lname] = var
            tk.Checkbutton(
                inner,
                text=f"{lname}  ({layer_counts[lname]} entities)",
                variable=var, anchor="w",
            ).pack(padx=4, pady=1, anchor="w", fill="x")

        # Select all / none
        sel_frame = tk.Frame(win)
        sel_frame.pack(padx=10, pady=2)
        tk.Button(
            sel_frame, text="Select All",
            command=lambda: [v.set(True) for v in layer_vars.values()],
        ).pack(side="left", padx=4)
        tk.Button(
            sel_frame, text="Select None",
            command=lambda: [v.set(False) for v in layer_vars.values()],
        ).pack(side="left", padx=4)

        # CRS / EPSG dropdown — auto-fill from header if detected, else user picks
        crs_frame = tk.Frame(win)
        crs_frame.pack(padx=10, pady=(6, 2), fill="x")
        tk.Label(
            crs_frame, text="Coordinate System:",
            font=("Helvetica", 9, "bold"),
        ).pack(side="left")

        # Build the CRS choices list from NameMapper.csv (Autodesk install)
        _crs_choices: list[str] = []
        _ad_map = self._load_autodesk_cs_map()  # type: ignore[attr-defined]
        if _ad_map:
            # Build "EPSG — Autodesk_Code — EPSG_Name" for each entry
            # We need EPSG names too — build a reverse lookup
            import csv as _csv
            import glob as _glob
            _epsg_names: dict[int, str] = {}
            for _pat in [r"C:\ProgramData\Autodesk\Geospatial Coordinate Systems*\NameMapper.csv"]:
                for _nmf in _glob.glob(_pat):
                    try:
                        with open(_nmf, encoding="utf-8") as _f:
                            for _row in _csv.DictReader(_f):
                                if _row["Flavor"] == "EPSG" and _row["NumericId"] != "0":
                                    _epsg_names[int(_row["NumericId"])] = _row["NameId"]
                    except Exception:
                        pass
                    break
            # Build display strings — filter to relevant US CRS only
            # (full 4000+ list overloads the tkinter Combobox)
            _seen: set[int] = set()
            _entries: list[tuple[str, int, str]] = []
            for _ad_code, _epsg in _ad_map.items():
                if _epsg not in _seen:
                    _ename = _epsg_names.get(_epsg, "")
                    # Keep: NAD83/NAD27 State Plane, UTM, WGS84
                    if not _ename:
                        continue
                    _keep = ("NAD83 /" in _ename or "NAD27 /" in _ename
                             or "WGS 84 / UTM" in _ename)
                    if not _keep:
                        continue
                    _seen.add(_epsg)
                    _entries.append((_ename, _epsg, _ad_code))
            _entries.sort(key=lambda x: x[0])
            for _ename, _epsg, _ad_code in _entries:
                _crs_choices.append(f"{_epsg} — {_ename} [{_ad_code}]")
        if not _crs_choices:
            # Fallback if NameMapper.csv not found
            _crs_choices = [
                "2231 — NAD83 / Colorado North (ftUS) [CO83-NF]",
                "2232 — NAD83 / Colorado Central (ftUS) [CO83-CF]",
                "2233 — NAD83 / Colorado South (ftUS) [CO83-SF]",
                "2275 — NAD83 / Texas North (ftUS) [TX83-NF]",
                "2276 — NAD83 / Texas North Central (ftUS) [TX83-NCF]",
                "2277 — NAD83 / Texas Central (ftUS) [TX83-CF]",
                "2278 — NAD83 / Texas South Central (ftUS) [TX83-SCF]",
                "2279 — NAD83 / Texas South (ftUS) [TX83-SF]",
            ]

        # Try to resolve auto-detected cs_code to an EPSG number
        _auto_epsg = ""
        if cs_code:
            _resolved = self._resolve_autodesk_cs(cs_code)  # type: ignore[attr-defined]
            if _resolved:
                _auto_epsg = str(_resolved)
                # Find the matching display string
                for _ch in _crs_choices:
                    if _ch.startswith(f"{_resolved} — "):
                        _auto_epsg = _ch
                        break
        epsg_var = tk.StringVar(value=_auto_epsg)
        epsg_combo = ttk.Combobox(crs_frame, textvariable=epsg_var,
                                  values=_crs_choices, width=48)
        epsg_combo.pack(side="left", padx=6, fill="x", expand=True)

        dwg_status_lbl = tk.Label(win, text="", anchor="w")
        dwg_status_lbl.pack(padx=10, pady=2, fill="x")

        def _do_dwg_export() -> None:
            selected = {n for n, v in layer_vars.items() if v.get()}
            if not selected:
                messagebox.showwarning(
                    "No Layers", "Select at least one layer.", parent=win)
                return

            dwg_status_lbl.config(
                text="Preparing geometry…", fg="#1F4E79")
            win.update_idletasks()

            geometry = {ly: ents for ly, ents in _prescan_geom.items()
                        if ly in selected and ents}

            if not geometry:
                messagebox.showinfo(
                    "No Geometry",
                    "No exportable geometry on selected layers.",
                    parent=win)
                dwg_status_lbl.config(text="")
                return

            dwg_status_lbl.config(
                text="Writing geometry to GeoPackage…")
            win.update_idletasks()

            # Determine SRS from the CRS dropdown/entry
            write_srid = 0
            _epsg_input = epsg_var.get().strip()
            if _epsg_input:
                # Handle combo format "2231 — NAD83 / ..." or plain "2231"
                _epsg_num_str = _epsg_input.split(" — ")[0].strip() if " — " in _epsg_input else _epsg_input
                try:
                    write_srid = int(_epsg_num_str)
                except ValueError:
                    write_srid = self._resolve_autodesk_cs(_epsg_num_str)  # type: ignore[attr-defined]

            # Try pyproj transformation to WGS84 (optional)
            transformer = None
            native_srid = write_srid  # preserve original SRID for KML/LandXML writers
            if write_srid not in (0, 4326):
                try:
                    from pyproj import Transformer as _PT  # type: ignore[import-untyped]
                    _t: Any = _PT.from_crs(  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                        f"EPSG:{write_srid}", "EPSG:4326",
                        always_xy=True)
                    transformer: Any = lambda x, y: _t.transform(x, y)  # type: ignore[reportUnknownVariableType,reportUnknownLambdaType]
                    write_srid = 4326
                except Exception:
                    pass  # keep native CRS

            try:
                count = self._write_dwg_geometry_to_gpkg(  # type: ignore[attr-defined]
                    gpkg_path, geometry,
                    srid=write_srid, cs_description=cs_code,
                    transformer=transformer)
            except Exception as exc:
                messagebox.showerror(
                    "Write Error",
                    f"Failed to write DWG geometry:\n{exc}",
                    parent=win)
                dwg_status_lbl.config(text="")
                return

            n_layers = len(geometry)
            if write_srid == 4326:
                crs_msg = "CRS: WGS 84 (EPSG:4326)"
            elif cs_code:
                crs_msg = (f"CRS: {cs_code} (native — "
                           f"assign EPSG in GIS if needed)")
            else:
                crs_msg = "CRS: undefined — assign in GIS"

            dwg_status_lbl.config(text="Done!", fg="#2E7D32")
            exported_layers.extend(selected)
            extracted_geom.update(geometry)
            final_srid[0] = native_srid  # pass original CRS so KML/LandXML can reproject
            messagebox.showinfo(
                "DWG Geometry Added",
                f"Added {count} entities from {n_layers} layer(s) to:\n"
                f"  {os.path.basename(gpkg_path)}\n\n{crs_msg}",
                parent=win)
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=(6, 10))
        tk.Button(
            btn_frame, text="Export Geometry", width=16,
            bg="#1F4E79", fg="white", command=_do_dwg_export,
        ).pack(side="left", padx=8)
        tk.Button(
            btn_frame, text="Skip", width=10, command=win.destroy,
        ).pack(side="left", padx=8)

        win.wait_window()

        return exported_layers, extracted_geom, final_srid[0], cs_code

    def _write_unresolved_excel(
        self,
        unresolved: list[str],
        ambiguous: list[str],
        media_missing: dict[str, str],
        output_path: str,
    ) -> None:
        """Write an Excel issues report with sheets for JXL mismatches and missing media."""
        if openpyxl is None:
            return
        from openpyxl.styles import Font as _URFont, PatternFill as _URPF  # type: ignore[import]
        hdr_fill = _URPF(fgColor="1F4E79", fill_type="solid")
        hdr_font = _URFont(bold=True, color="FFFFFF")

        def _hdr(ws: Any, cols: list[str]) -> None:
            ws.append(cols)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font

        wb = openpyxl.Workbook()

        # Sheet 1: JXL match issues
        ws_jxl = wb.active
        ws_jxl.title = "JXL Issues"  # type: ignore[union-attr]
        _hdr(ws_jxl, ["Point Name", "Issue"])
        for pname in sorted(unresolved):
            ws_jxl.append([pname, "No matching JXL record found"])  # type: ignore[union-attr]
        for pname in sorted(ambiguous):
            ws_jxl.append([pname, "Multiple JXL matches — first used"])  # type: ignore[union-attr]
        ws_jxl.column_dimensions["A"].width = 22  # type: ignore[index]
        ws_jxl.column_dimensions["B"].width = 38  # type: ignore[index]

        # Sheet 2: Missing media
        if media_missing:
            ws_media = wb.create_sheet("Missing Media")
            _hdr(ws_media, ["Point Name", "Expected File"])
            for pname in sorted(media_missing):
                ws_media.append([pname, media_missing[pname]])  # type: ignore[union-attr]
            ws_media.column_dimensions["A"].width = 22  # type: ignore[index]
            ws_media.column_dimensions["B"].width = 48  # type: ignore[index]

        wb.save(output_path)

    def _find_crdb_media(
        self,
        crdb_path: str,
        jxl_map: dict[str, str],
        matched_pts: dict[str, dict[str, Any]],
    ) -> tuple[dict[str, str], dict[str, str]]:
        """Locate media files referenced in JXL data for matched CRDB points.

        Uses the tiered media-index helper:
          • Single JXL  — scans the JXL's parent directory tree.
          • Multiple JXLs — scans each ``<stem> Files`` companion folder,
            with a file-size SYNC fallback when a companion is missing.

        Returns:
            found   — {UPPER_point_name: absolute_file_path}
            missing — {UPPER_point_name: expected_filename}
        """
        jxl_paths = list(jxl_map.keys())

        # Build a synthetic jxl_data_map from matched_pts so the fallback
        # can resolve XML-referenced subfolder paths for file-size matching.
        # Group points by the JXL directory they came from (best guess: use
        # the first JXL whose directory is an ancestor of the photo_path).
        jxl_data_map: dict[str, dict[str, Any]] = {}
        for jp in jxl_paths:
            jxl_data_map[jp] = {"points": {}}
        # Distribute matched_pts across JXLs (simple: all go to every JXL
        # for fallback purposes — the basename-level dedup inside the helper
        # keeps this correct).
        for jp in jxl_paths:
            jxl_data_map[jp]["points"] = dict(matched_pts)

        media_index = self._build_jxl_media_index(jxl_paths, jxl_data_map)

        found: dict[str, str] = {}
        missing: dict[str, str] = {}
        for upper_name, pt_data in matched_pts.items():
            photo_ref: str = str(pt_data.get("photo_path") or pt_data.get("photo_name") or "")
            if not photo_ref:
                continue
            basename = os.path.basename(photo_ref.replace("\\", "/"))
            if not basename:
                continue
            if basename.lower() in media_index:
                found[upper_name] = media_index[basename.lower()]
            else:
                missing[upper_name] = basename
        return found, missing

    # ---------- CRDB action chooser ----------

    def _show_crdb_action_dialog(self, crdb_paths: list[str]) -> None:
        """Handle one or more dropped CRDB files — go straight to export."""
        for cp in crdb_paths:
            self._show_crdb_export_dialog(cp)  # type: ignore[attr-defined]

    def _validate_crdb(self, crdb_paths: list[str]) -> None:
        """Validate CRDB point data through the same Excel pipeline as CSV validation.

        Builds a synthetic DataFrame from CRDB rows (N/E/Z/code/attrs), finds companion
        JXLs so the Geodetic Info sheet is populated, then calls _export_and_open_excel.
        """
        if not self.fxl_data:
            if not self._ensure_fxl_after_csv(crdb_paths[0]):  # type: ignore[attr-defined]
                return
        if not self.fxl_data:
            messagebox.showwarning("No FXL", "An FXL must be loaded before validating a CRDB.")
            return

        self.status.config(text=f"Parsing {len(crdb_paths)} CRDB file(s)…")
        self.update_idletasks()

        max_attrs: int = max((len(v) for v in self.fxl_data.values()), default=0)
        all_rows: list[list[Any]] = []
        errors: list[str] = []
        first_crdb = crdb_paths[0]

        for crdb_path in crdb_paths:
            try:
                for row in self._load_crdb_rows(crdb_path):  # type: ignore[attr-defined]
                    attr_vals = (row["attrs"] + [""] * max_attrs)[:max_attrs]
                    n, e, z = row.get("N"), row.get("E"), row.get("Z")
                    all_rows.append([
                        row["point_name"],
                        "" if n is None else n,
                        "" if e is None else e,
                        "" if z is None else z,
                        row["code"],
                        *attr_vals,
                    ])
            except Exception as ex:
                errors.append(f"{os.path.basename(crdb_path)}: {ex}")

        if errors:
            messagebox.showwarning("CRDB Load Errors", "\n".join(errors))
        if not all_rows:
            self.status.config(text="Ready")
            messagebox.showinfo("No Points", "No point records found in the dropped CRDB file(s).")
            return

        col_names = ["Point Number", "Northing", "Easting", "Elevation", "Field Code"] + \
                    [f"Attr{i + 1}" for i in range(max_attrs)]
        df = pd.DataFrame(all_rows, columns=col_names, dtype=str)  # type: ignore[reportUnknownMemberType]
        df = df.fillna("")  # type: ignore[reportUnknownMemberType]

        # Find companion JXLs so the Geodetic Info sheet can be written
        try:
            jxl_map = self._search_jxl_upward(first_crdb, set())  # type: ignore[attr-defined]
            if jxl_map:
                # Merge all JXL points into one combined data dict for the sheet
                combined_pts: dict[str, Any] = {}
                first_jxl_meta: dict[str, Any] = {}
                for jxl_path in jxl_map.keys():
                    jd: dict[str, Any] = self._parse_jxl(jxl_path)  # type: ignore[attr-defined]
                    if not first_jxl_meta:
                        first_jxl_meta = {k: v for k, v in jd.items() if k != "points"}
                    combined_pts.update(jd.get("points") or {})
                combined_jxl: dict[str, Any] = {**first_jxl_meta, "points": combined_pts}
                self.jxl_path = next(iter(jxl_map.keys()))
                self._jxl_data = combined_jxl
            else:
                self.jxl_path = None
                self._jxl_data = None
        except Exception:
            self.jxl_path = None
            self._jxl_data = None

        self.df = df
        self.csv_path = first_crdb
        self.has_station = False
        self.mapping = {"station": None, "pn": 0, "north": 1, "east": 2, "elev": 3, "fc": 4}
        self.attr_indices = list(range(5, 5 + max_attrs))

        self.status.config(text="Ready")
        self._export_and_open_excel(open_new_excel_instance=(not self.single_excel_instance))  # type: ignore[attr-defined]

    def _crdb_rename_media(self, crdb_path: str) -> None:
        """Find JXLs for a CRDB, match points, then offer photo rename for found media."""
        self.status.config(text=f"Scanning {os.path.basename(crdb_path)} for media…")
        self.update_idletasks()
        try:
            rows = self._load_crdb_rows(crdb_path)  # type: ignore[attr-defined]
            target_stems: set[str] = set()
            jxl_map = self._search_jxl_upward(crdb_path, target_stems)  # type: ignore[attr-defined]
            if not jxl_map:
                messagebox.showwarning("No JXL Files",
                    f"No JXL files found near:\n{crdb_path}", parent=self)
                return
            matched_pts, _, _ = self._match_points_to_jxls(rows, jxl_map)  # type: ignore[attr-defined]
            media_found, media_missing = self._find_crdb_media(  # type: ignore[attr-defined]
                crdb_path, jxl_map, matched_pts)
            if not media_found:
                messagebox.showinfo("No Media Found",
                    "No media files were found for points in this CRDB.", parent=self)
                return
            # Build per-JXL photo items for the rename dialog
            jxl_items: list[tuple[str, dict[str, Any], dict[str, str]]] = []
            for jxl_path, _ in jxl_map.items():
                try:
                    jxl_d = self._parse_jxl(jxl_path)  # type: ignore[attr-defined]
                    jxl_pts: dict[str, Any] = jxl_d.get("points") or {}
                    photo_map: dict[str, str] = {
                        pt_name: media_found[pt_name.upper()]
                        for pt_name in jxl_pts
                        if pt_name.upper() in media_found
                    }
                    if photo_map:
                        jxl_items.append((jxl_path, jxl_d, photo_map))
                except Exception:
                    pass
            if jxl_items:
                self._offer_photo_rename_multi(jxl_items, update_excel=False)  # type: ignore[attr-defined]
            elif media_missing:
                messagebox.showinfo("No Renameable Photos",
                    f"Media found but none have auto-generated names.\n"
                    f"{len(media_missing)} file(s) could not be located.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)
        finally:
            self.status.config(text="Ready")

    def _write_crdb_gnss_csv(
        self,
        save_path: str,
        rows: list[dict[str, Any]],
        matched_pts: dict[str, dict[str, Any]],
        media_found: dict[str, str],
        crdb_path: str,
    ) -> None:
        """Write a GNSS report CSV using pre-matched CRDB/JXL data."""
        import csv as _csv

        headers = [
            "Job", "Point Number", "Northing", "Easting", "Elevation",
            "Field Code",
            "H Precision (m)", "V Precision (m)", "PDOP", "Num Satellites", "Survey Method",
            "WGS84 Latitude", "WGS84 Longitude", "WGS84 Height (m)",
            "Media File Name",
        ]
        job_label = os.path.splitext(os.path.basename(crdb_path))[0]
        report_rows: list[list[Any]] = []
        for row in rows:
            pt_name: str = row["point_name"]
            pt_jxl: dict[str, Any] = matched_pts.get(pt_name.upper()) or {}
            media_path = media_found.get(pt_name.upper(), "")
            media_name = os.path.basename(media_path) if media_path else ""
            report_rows.append([
                job_label,
                pt_name,
                row.get("N") if row.get("N") is not None else "",
                row.get("E") if row.get("E") is not None else "",
                row.get("Z") if row.get("Z") is not None else "",
                row.get("code") or "",
                pt_jxl.get("h_precision")    if pt_jxl.get("h_precision")    is not None else "",
                pt_jxl.get("v_precision")    if pt_jxl.get("v_precision")    is not None else "",
                pt_jxl.get("pdop")           if pt_jxl.get("pdop")           is not None else "",
                pt_jxl.get("num_satellites") if pt_jxl.get("num_satellites") is not None else "",
                pt_jxl.get("survey_method") or "",
                pt_jxl.get("wgs84_lat")    if pt_jxl.get("wgs84_lat")    is not None else "",
                pt_jxl.get("wgs84_lon")    if pt_jxl.get("wgs84_lon")    is not None else "",
                pt_jxl.get("wgs84_height") if pt_jxl.get("wgs84_height") is not None else "",
                media_name,
            ])
        with open(save_path, "w", newline="", encoding="cp1252", errors="replace") as f:
            w = _csv.writer(f)
            w.writerow(headers)
            w.writerows(report_rows)

    def _crdb_gnss_report(self, crdb_paths: list[str]) -> None:
        """Generate a CSV GNSS report from one or more CRDBs, combining CRDB coords with JXL GNSS data."""
        import csv as _csv

        first_dir = os.path.dirname(os.path.abspath(crdb_paths[0]))
        default_name = (
            os.path.splitext(os.path.basename(crdb_paths[0]))[0] + "_GNSS_Report.csv"
            if len(crdb_paths) == 1 else "GNSS_Report.csv"
        )
        save_path = os.path.join(first_dir, default_name)

        self.status.config(text="Generating CRDB GNSS report…")
        self.update_idletasks()

        headers = [
            "Job", "Point Number", "Northing", "Easting", "Elevation",
            "Field Code",
            "H Precision (m)", "V Precision (m)", "PDOP", "Num Satellites", "Survey Method",
            "WGS84 Latitude", "WGS84 Longitude", "WGS84 Height (m)",
            "Media File Name",
        ]
        report_rows: list[list[Any]] = []
        errors: list[str] = []

        for crdb_path in crdb_paths:
            try:
                rows = self._load_crdb_rows(crdb_path)  # type: ignore[attr-defined]
                jxl_map = self._search_jxl_upward(crdb_path, set())  # type: ignore[attr-defined]
                matched_pts, _, _ = self._match_points_to_jxls(rows, jxl_map)  # type: ignore[attr-defined]
                media_found, _ = self._find_crdb_media(crdb_path, jxl_map, matched_pts)  # type: ignore[attr-defined]
                job_label = os.path.splitext(os.path.basename(crdb_path))[0]
                for row in rows:
                    pt_name: str = row["point_name"]
                    pt_jxl: dict[str, Any] = matched_pts.get(pt_name.upper()) or {}
                    media_path = media_found.get(pt_name.upper(), "")
                    media_name = os.path.basename(media_path) if media_path else ""
                    report_rows.append([
                        job_label,
                        pt_name,
                        row.get("N") if row.get("N") is not None else "",
                        row.get("E") if row.get("E") is not None else "",
                        row.get("Z") if row.get("Z") is not None else "",
                        row.get("code") or "",
                        pt_jxl.get("h_precision")    if pt_jxl.get("h_precision")    is not None else "",
                        pt_jxl.get("v_precision")    if pt_jxl.get("v_precision")    is not None else "",
                        pt_jxl.get("pdop")           if pt_jxl.get("pdop")           is not None else "",
                        pt_jxl.get("num_satellites") if pt_jxl.get("num_satellites") is not None else "",
                        pt_jxl.get("survey_method") or "",
                        pt_jxl.get("wgs84_lat")    if pt_jxl.get("wgs84_lat")    is not None else "",
                        pt_jxl.get("wgs84_lon")    if pt_jxl.get("wgs84_lon")    is not None else "",
                        pt_jxl.get("wgs84_height") if pt_jxl.get("wgs84_height") is not None else "",
                        media_name,
                    ])
            except Exception as e:
                errors.append(f"{os.path.basename(crdb_path)}: {e}")

        self.status.config(text="Ready")

        try:
            with open(save_path, "w", newline="", encoding="cp1252", errors="replace") as f:
                _csv.writer(f).writerow(headers)
                _csv.writer(f).writerows(report_rows)
            msg = f"Saved {len(report_rows)} point(s).\n{os.path.basename(save_path)}"
            if errors:
                msg += "\n\nErrors:\n" + "\n".join(errors)
            messagebox.showinfo("GNSS Report Complete", msg)
        except Exception as e:
            messagebox.showerror("Save Failed", str(e))

    def _show_crdb_export_dialog(self, crdb_path: str) -> None:
        """Main wizard for CRDB → GeoPackage export."""
        self.crdb_path = crdb_path

        # 1. Load CRDB rows
        try:
            rows = self._load_crdb_rows(crdb_path)  # type: ignore[attr-defined]
        except Exception as exc:
            messagebox.showerror("CRDB Load Error", f"Could not read {os.path.basename(crdb_path)}:\n{exc}")
            return
        if not rows:
            messagebox.showwarning("Empty CRDB", "No point records found in this CRDB file.")
            return

        # 2. Extract JXL hints from photo paths in the D column (used as a search hint only)
        hints = self._extract_jxl_hints(rows)  # type: ignore[attr-defined]
        target_stems: set[str] = set(hints.keys())

        # 3. Search for JXL files — always runs regardless of photo hints.
        #    Priority locations (Field Data/SYNC, Field Data) are scanned for ALL JXLs;
        #    target_stems is only used as a filter when expanding to broader ancestor trees.
        jxl_map: dict[str, str] = self._search_jxl_upward(crdb_path, target_stems)  # type: ignore[attr-defined]

        # 4. Match points to JXL geodetic data
        matched_pts, unresolved, ambiguous = self._match_points_to_jxls(rows, jxl_map)  # type: ignore[attr-defined]

        # 4b. Locate media files for matched points that have JXL photo references
        media_found, media_missing = self._find_crdb_media(crdb_path, jxl_map, matched_pts)  # type: ignore[attr-defined]

        # 5. Try to auto-load FXL from the first JXL's <SourceFilename> tag if not already loaded
        fxl_data: dict[str, list[dict[str, Any]]] = dict(self.fxl_data) if self.fxl_data else {}
        fxl_source_label = os.path.basename(self.fxl_path) if self.fxl_path else "(none loaded)"
        if not fxl_data and jxl_map:
            first_jxl = next(iter(jxl_map.keys()))
            try:
                jxl_meta = self._parse_jxl(first_jxl)  # type: ignore[attr-defined]
                fxl_fname: str = cast(str, jxl_meta.get("fxl_filename") or "")
                if fxl_fname:
                    crdb_dir = os.path.dirname(crdb_path)
                    cand = os.path.join(crdb_dir, fxl_fname)
                    if not os.path.isfile(cand):
                        for lib_p in self._get_fxl_library_candidates():
                            if os.path.basename(lib_p).lower() == fxl_fname.lower():
                                cand = lib_p
                                break
                    if os.path.isfile(cand):
                        fxl_data = self.parse_fxl(cand)
                        fxl_source_label = os.path.basename(cand) + " (auto)"
            except Exception:
                pass

        # 5b. Auto-detect DWG file in same folder as CRDB
        dwg_path_auto = self._find_dwg_for_crdb(crdb_path)  # type: ignore[attr-defined]

        # 6. Build the export dialog
        win = tk.Toplevel(self)
        win.title("Export CRDB Data")
        win.resizable(False, False)
        win.grab_set()
        _raise_window(win)

        pad: dict[str, Any] = dict(padx=10, pady=3, sticky="w")

        tk.Label(win, text="CRDB Data Export", font=("Helvetica", 11, "bold")).grid(
            row=0, column=0, columnspan=2, padx=10, pady=(10, 4), sticky="w"
        )

        tk.Label(win, text="CRDB file:").grid(row=1, column=0, **pad)
        tk.Label(win, text=os.path.basename(crdb_path), fg="#1F4E79").grid(row=1, column=1, **pad)

        tk.Label(win, text="Points loaded:").grid(row=2, column=0, **pad)
        codes_found = sorted({r["code"] for r in rows if r["code"]})
        tk.Label(win, text=f"{len(rows)} ({len(codes_found)} feature code(s): {', '.join(codes_found[:6])}"
                           + (" …" if len(codes_found) > 6 else "") + ")").grid(row=2, column=1, **pad)

        tk.Label(win, text="JXL geodetic:").grid(row=3, column=0, **pad)
        geo_txt = (f"{len(matched_pts)} matched"
                   + (f", {len(unresolved)} unresolved" if unresolved else "")
                   + (f", {len(ambiguous)} ambiguous" if ambiguous else ""))
        # Diagnostic hint: show JXL file count and a sample CRDB point name so the
        # user can immediately see whether files weren't found or names don't align.
        if not matched_pts:
            if not jxl_map:
                geo_txt += "  —  no JXL files found in search path"
            else:
                sample_crdb = rows[0]["point_name"] if rows else "?"
                sample_jxl_pts: list[str] = []
                for _fp in list(jxl_map.keys())[:1]:
                    try:
                        _d = self._parse_jxl(_fp)  # type: ignore[attr-defined]
                        sample_jxl_pts = [str(k) for k in cast(dict[str, Any], _d.get("points") or {}).keys()][:3]
                    except Exception:
                        pass
                geo_txt += (f"  —  {len(jxl_map)} JXL file(s) loaded"
                            f";  CRDB pt: '{sample_crdb}'"
                            + (f";  JXL pts: {sample_jxl_pts}" if sample_jxl_pts else ""))
        geo_color = "#C62828" if unresolved else "#2E7D32"
        geo_lbl = tk.Label(win, text=geo_txt, fg=geo_color)
        geo_lbl.grid(row=3, column=1, **pad)

        # Mutable containers so the Browse callback can update them
        jxl_map_ref: list[dict[str, str]] = [jxl_map]
        matched_ref: list[dict[str, dict[str, Any]]] = [matched_pts]
        unresolved_ref: list[list[str]] = [unresolved]
        ambiguous_ref: list[list[str]] = [ambiguous]
        media_found_ref: list[dict[str, str]] = [media_found]
        media_missing_ref: list[dict[str, str]] = [media_missing]

        def _browse_jxl_folder() -> None:
            folder = filedialog.askdirectory(
                title="Select folder containing JXL files",
                parent=win,
            )
            if not folder:
                return
            # Scan the chosen folder (SYNC subfolders first, then the rest)
            new_map: dict[str, str] = {}
            sync_dirs: list[str] = []
            for dp, dns, _ in os.walk(folder):
                for dn in dns:
                    if dn.upper() == "SYNC":
                        sync_dirs.append(os.path.join(dp, dn))
            for sdir in sync_dirs:
                for dp2, _, fns in os.walk(sdir):
                    for fn in fns:
                        if fn.lower().endswith(".jxl"):
                            abs_path = os.path.abspath(os.path.join(dp2, fn))
                            if abs_path not in new_map:
                                new_map[abs_path] = os.path.splitext(fn)[0]
            excl = {os.path.abspath(s) for s in sync_dirs}
            for dp2, dns2, fns in os.walk(folder):
                dns2[:] = [dn for dn in dns2
                           if os.path.abspath(os.path.join(dp2, dn)) not in excl]
                for fn in fns:
                    if fn.lower().endswith(".jxl"):
                        abs_path = os.path.abspath(os.path.join(dp2, fn))
                        if abs_path not in new_map:
                            new_map[abs_path] = os.path.splitext(fn)[0]
            if not new_map:
                messagebox.showwarning("No JXL Files", "No .jxl files found in the selected folder.", parent=win)
                return
            # Re-run matching with the new JXL set
            new_matched, new_unres, new_amb = self._match_points_to_jxls(rows, new_map)  # type: ignore[attr-defined]
            new_mf, new_mm = self._find_crdb_media(crdb_path, new_map, new_matched)  # type: ignore[attr-defined]
            jxl_map_ref[0] = new_map
            matched_ref[0] = new_matched
            unresolved_ref[0] = new_unres
            ambiguous_ref[0] = new_amb
            media_found_ref[0] = new_mf
            media_missing_ref[0] = new_mm
            # Refresh the geo label
            new_txt = (f"{len(new_matched)} matched"
                       + (f", {len(new_unres)} unresolved" if new_unres else "")
                       + (f", {len(new_amb)} ambiguous" if new_amb else "")
                       + f"  —  {len(new_map)} JXL file(s) from {os.path.basename(folder)}")
            new_color = "#C62828" if new_unres else "#2E7D32"
            geo_lbl.config(text=new_txt, fg=new_color)

        tk.Button(win, text="Browse for JXL folder…", command=_browse_jxl_folder).grid(
            row=3, column=1, padx=(310, 8), pady=3, sticky="w"
        )

        tk.Label(win, text="FXL attributes:").grid(row=4, column=0, **pad)
        fxl_color = "#2E7D32" if fxl_data else "#E65100"
        tk.Label(win, text=fxl_source_label, fg=fxl_color).grid(row=4, column=1, **pad)

        # FXL load button (shown when no FXL auto-detected)
        fxl_data_ref: list[dict[str, list[dict[str, Any]]]] = [fxl_data]
        fxl_lbl_ref: list[tk.Label] = []

        def _load_fxl_manually() -> None:
            path = filedialog.askopenfilename(
                title="Select FXL for attribute names",
                filetypes=[("FXL / XML", "*.fxl *.xml"), ("All files", "*.*")],
                parent=win,
            )
            if path and os.path.isfile(path):
                try:
                    fxl_data_ref[0] = self.parse_fxl(path)
                    if fxl_lbl_ref:
                        fxl_lbl_ref[0].config(
                            text=os.path.basename(path) + " (manual)", fg="#2E7D32"
                        )
                except Exception as exc:
                    messagebox.showerror("FXL Error", str(exc), parent=win)

        if not fxl_data:
            btn_load_fxl = tk.Button(win, text="Load FXL…", command=_load_fxl_manually)
            btn_load_fxl.grid(row=4, column=1, padx=(180, 10), pady=3, sticky="w")

        # Media row
        tk.Label(win, text="Media files:").grid(row=5, column=0, **pad)
        _pts_with_media = len(media_found) + len(media_missing)
        if _pts_with_media == 0:
            media_txt, media_color = "none referenced in JXL", "#888888"
        elif media_missing:
            media_txt = f"{len(media_found)} found, {len(media_missing)} missing"
            media_color = "#C62828"
        else:
            media_txt, media_color = f"{len(media_found)} found", "#2E7D32"
        tk.Label(win, text=media_txt, fg=media_color).grid(row=5, column=1, **pad)

        # DWG file row
        dwg_var = tk.StringVar(value=dwg_path_auto or "")
        if dwg_path_auto:
            _dwg_label = os.path.basename(dwg_path_auto) + "  (auto-detected)"
            _dwg_color = "#2E7D32"
        else:
            _dwg_label = "(none found in CRDB folder)"
            _dwg_color = "#888888"
        tk.Label(win, text="DWG file:").grid(row=6, column=0, **pad)
        _dwg_row = tk.Frame(win)
        _dwg_row.grid(row=6, column=1, padx=10, pady=3, sticky="w")
        dwg_lbl = tk.Label(_dwg_row, text=_dwg_label, fg=_dwg_color)
        dwg_lbl.pack(side="left")

        def _browse_dwg() -> None:
            p = filedialog.askopenfilename(
                title="Select DWG file for geometry export",
                initialdir=os.path.dirname(crdb_path),
                filetypes=[("DWG files", "*.dwg"), ("All files", "*.*")],
                parent=win,
            )
            if p:
                dwg_var.set(p)
                dwg_lbl.config(text=os.path.basename(p), fg="#2E7D32")
                dwg_chk_var.set(True)

        tk.Button(_dwg_row, text="Browse…", command=_browse_dwg).pack(
            side="left", padx=(8, 0))

        # Client schema selector
        client_schemas = self._load_client_schemas()  # type: ignore[attr-defined]
        tk.Label(win, text="Client schema:").grid(row=7, column=0, **pad)
        client_names = sorted(client_schemas.keys())
        client_var = tk.StringVar(value=client_names[0] if client_names else "(none)")
        client_values = client_names + ["(none — generic export)"]
        client_combo = ttk.Combobox(win, textvariable=client_var,
                                    values=client_values, width=30, state="readonly")
        client_combo.grid(row=7, column=1, padx=10, pady=3, sticky="w")

        # DWG geometry checkbox
        dwg_chk_var = tk.BooleanVar(value=bool(dwg_path_auto))
        dwg_chk = tk.Checkbutton(
            win,
            text="Include DWG geometry (select layers after export)",
            variable=dwg_chk_var,
        )
        dwg_chk.grid(row=8, column=0, columnspan=2, padx=10, pady=2, sticky="w")

        # Output folder (deterministic: ASBUILT/0_GIS/WEEKLY_UPDATE/YYYYMMDD)
        crdb_stem = os.path.splitext(os.path.basename(crdb_path))[0]
        output_dir = self._find_gis_output_dir(crdb_path)  # type: ignore[attr-defined]
        tk.Label(win, text="Output folder:").grid(row=9, column=0, **pad)
        tk.Label(win, text=output_dir, fg="#1F4E79", wraplength=420, justify="left").grid(
            row=9, column=1, **pad)
        tk.Label(win, text="Export formats:", fg="#555555").grid(row=10, column=0, **pad)
        tk.Label(win, text="GPKG, CSV, Shapefile, LandXML, KMZ, GNSS Report", fg="#555555").grid(
            row=10, column=1, **pad)

        # Unresolved-report checkbox
        report_var = tk.BooleanVar(value=bool(unresolved or ambiguous or media_missing))
        if unresolved or ambiguous or media_missing:
            issues = len(unresolved) + len(ambiguous) + len(media_missing)
            tk.Checkbutton(
                win,
                text=f"Write issues report ({issues} entries)",
                variable=report_var,
            ).grid(row=11, column=0, columnspan=2, padx=10, pady=2, sticky="w")

        # Status label for after export
        status_lbl = tk.Label(win, text="", anchor="w")
        status_lbl.grid(row=12, column=0, columnspan=2, padx=10, pady=2, sticky="ew")

        def _do_export() -> None:
            # Always use the ref containers — Browse may have updated them
            _matched   = matched_ref[0]
            _unres     = unresolved_ref[0]
            _amb       = ambiguous_ref[0]
            _mf        = media_found_ref[0]
            _mm        = media_missing_ref[0]
            _jxl_map   = jxl_map_ref[0]

            # Create the output directory
            os.makedirs(output_dir, exist_ok=True)

            # Collect JXL metadata for CSV timestamp fields
            _jxl_meta: dict[str, Any] = {}
            if _jxl_map:
                first_jxl = next(iter(_jxl_map.keys()))
                try:
                    _jxl_meta = self._parse_jxl(first_jxl)  # type: ignore[attr-defined]
                except Exception:
                    pass

            # Resolve client schema for CSV export
            _client_name = client_var.get().strip()
            _client_schema: dict[str, Any] | None = None
            if _client_name and _client_name in client_schemas:
                _client_schema = client_schemas[_client_name]

            # --- Field code selection dialog ---
            # Show a dialog for the user to pick which codes to export
            # (CSV always gets all points; other formats get filtered)
            _all_codes = sorted(set(r.get("code", "UNKNOWN") or "UNKNOWN" for r in rows
                                    if not (isinstance(r.get("Z"), (int, float)) and r["Z"] <= -99999999)))
            _code_counts: dict[str, int] = {}
            for _r in rows:
                _zv = _r.get("Z")
                if _zv is not None and isinstance(_zv, (int, float)) and _zv <= -99999999:
                    continue
                _cc = _r.get("code", "UNKNOWN") or "UNKNOWN"
                _code_counts[_cc] = _code_counts.get(_cc, 0) + 1

            # Look up previously selected codes from watchlist
            _prev_codes: list[str] = []
            try:
                _wl_tmp = self._load_watchlist()  # type: ignore[attr-defined]
                for _we in _wl_tmp.get("entries", []):
                    if os.path.normcase(_we.get("crdb_path", "")) == os.path.normcase(crdb_path):
                        _prev_codes = _we.get("export_codes", [])
                        break
            except Exception:
                pass

            _prev_code_set: set[str] = set(_prev_codes) if _prev_codes else set()
            _selected_codes: list[str] = []

            # Build the code selection dialog
            _code_win = tk.Toplevel(self)
            _code_win.title("Select Field Codes for Export")
            _code_win.resizable(False, False)
            _code_win.grab_set()
            _code_win.lift()  # type: ignore[arg-type]
            _code_win.focus_force()

            tk.Label(
                _code_win, text="Select Field Codes to Export",
                font=("Helvetica", 11, "bold"),
            ).pack(padx=10, pady=(10, 4), anchor="w")
            tk.Label(
                _code_win, text=f"CRDB: {os.path.basename(crdb_path)}",
                fg="#1F4E79",
            ).pack(padx=10, pady=2, anchor="w")
            tk.Label(
                _code_win, text="CSV always exports all points. Other formats export selected codes only.",
                fg="#666666", font=("Helvetica", 8),
            ).pack(padx=10, pady=2, anchor="w")

            tk.Label(
                _code_win, text="Check the field codes to include:",
            ).pack(padx=10, pady=(8, 4), anchor="w")

            # Scrollable checkbox list
            _code_list_frame = tk.Frame(_code_win)
            _code_list_frame.pack(padx=10, pady=4, fill="both", expand=True)
            _code_canvas = tk.Canvas(
                _code_list_frame,
                height=min(300, len(_all_codes) * 24 + 10),
                width=380,
            )
            _code_scrollbar = ttk.Scrollbar(
                _code_list_frame, orient="vertical", command=_code_canvas.yview)  # type: ignore[arg-type]
            _code_inner = tk.Frame(_code_canvas)
            _code_inner.bind(
                "<Configure>",
                lambda e: _code_canvas.configure(scrollregion=_code_canvas.bbox("all")))
            _code_canvas.create_window((0, 0), window=_code_inner, anchor="nw")
            _code_canvas.configure(yscrollcommand=_code_scrollbar.set)
            _code_canvas.pack(side="left", fill="both", expand=True)
            if len(_all_codes) > 12:
                _code_scrollbar.pack(side="right", fill="y")

            _code_vars: dict[str, tk.BooleanVar] = {}
            for _cname in _all_codes:
                _cv = tk.BooleanVar(value=(_cname in _prev_code_set))
                _code_vars[_cname] = _cv
                tk.Checkbutton(
                    _code_inner,
                    text=f"{_cname}  ({_code_counts.get(_cname, 0)} points)",
                    variable=_cv, anchor="w",
                ).pack(padx=4, pady=1, anchor="w", fill="x")

            # Select all / none
            _code_sel_frame = tk.Frame(_code_win)
            _code_sel_frame.pack(padx=10, pady=2)
            tk.Button(
                _code_sel_frame, text="Select All",
                command=lambda: [v.set(True) for v in _code_vars.values()],
            ).pack(side="left", padx=4)
            tk.Button(
                _code_sel_frame, text="Select None",
                command=lambda: [v.set(False) for v in _code_vars.values()],
            ).pack(side="left", padx=4)

            _code_confirmed = [False]

            def _on_code_export() -> None:
                sel = [c for c, v in _code_vars.items() if v.get()]
                if not sel:
                    messagebox.showwarning(
                        "No Codes", "Select at least one field code.", parent=_code_win)
                    return
                _selected_codes.extend(sel)
                _code_confirmed[0] = True
                _code_win.destroy()

            _code_btn_frame = tk.Frame(_code_win)
            _code_btn_frame.pack(pady=(6, 10))
            tk.Button(
                _code_btn_frame, text="Export Selected", width=16,
                bg="#1F4E79", fg="white", command=_on_code_export,
            ).pack(side="left", padx=8)
            tk.Button(
                _code_btn_frame, text="Cancel", width=10,
                command=_code_win.destroy,
            ).pack(side="left", padx=8)

            _code_win.wait_window()

            if not _code_confirmed[0]:
                return  # user cancelled

            # Filter rows for non-CSV formats
            _selected_code_set = set(_selected_codes)
            _filtered_rows = [r for r in rows
                              if (r.get("code", "UNKNOWN") or "UNKNOWN") in _selected_code_set]

            # Write all formats — each is independent, failures don't block the rest
            gpkg_path = os.path.join(output_dir, crdb_stem + ".gpkg")
            csv_path = os.path.join(output_dir, crdb_stem + ".csv")
            shp_path = os.path.join(output_dir, crdb_stem + ".shp")
            xml_path = os.path.join(output_dir, crdb_stem + ".xml")
            kmz_path = os.path.join(output_dir, crdb_stem + ".kmz")
            gnss_path = os.path.join(output_dir, crdb_stem + "_GNSS_Report.csv")

            written: list[str] = []
            errors: list[str] = []

            # CSV gets ALL rows; other formats get filtered rows
            for label, func in [
                ("GeoPackage", lambda: self._write_gpkg(gpkg_path, _filtered_rows, _matched, fxl_data_ref[0], _mf)),  # type: ignore[attr-defined]
                ("CSV", lambda: self._write_crdb_csv(csv_path, rows, _matched, fxl_data_ref[0], _mf, _jxl_meta, _client_schema)),  # type: ignore[attr-defined]
                ("Shapefile", lambda: self._write_crdb_shp(shp_path, _filtered_rows, _matched, _mf)),  # type: ignore[attr-defined]
                ("LandXML", lambda: self._write_crdb_landxml(xml_path, _filtered_rows, _matched)),  # type: ignore[attr-defined]
                ("KMZ", lambda: self._write_crdb_kmz(kmz_path, _filtered_rows, _matched, fxl_data_ref[0], _mf)),  # type: ignore[attr-defined]
                ("GNSS Report", lambda: self._write_crdb_gnss_csv(gnss_path, rows, _matched, _mf, crdb_path)),  # type: ignore[attr-defined]
            ]:
                try:
                    func()
                    written.append(label)
                except Exception as exc:
                    errors.append(f"{label}: {exc}")

            if not written:
                messagebox.showerror("Export Error",
                                     "All exports failed:\n" + "\n".join(errors), parent=win)
                return

            msgs = [f"Exported {len(_filtered_rows)} of {len(rows)} points "
                    f"({len(_selected_codes)} of {len(_all_codes)} codes) to:",
                    f"  {output_dir}",
                    f"  Formats: {', '.join(written)}"]
            if errors:
                msgs.append(f"\nFailed: {'; '.join(errors)}")

            if report_var.get() and (_unres or _amb or _mm):
                report_path = os.path.join(os.path.dirname(crdb_path), crdb_stem + "_issues.xlsx")
                try:
                    self._write_unresolved_excel(_unres, _amb, _mm, report_path)  # type: ignore[attr-defined]
                    msgs.append(f"Issues report: {os.path.basename(report_path)}")
                except Exception:
                    msgs.append("(could not write issues report)")

            status_lbl.config(text="  Export complete.", fg="#2E7D32")
            messagebox.showinfo("Export Complete", "\n".join(msgs), parent=win)
            win.destroy()

            # Launch DWG layer-selection popup if user opted in
            _dwg_exported_layers: list[str] = []
            _dwg_geom: dict[str, list[dict[str, Any]]] = {}
            _dwg_srid: int = 0
            _dwg_cs: str = ""
            if dwg_chk_var.get() and dwg_var.get().strip():
                # Look up previously selected layers from watchlist
                _prev_layers: list[str] = []
                try:
                    _wl = self._load_watchlist()  # type: ignore[attr-defined]
                    for _wentry in _wl.get("entries", []):
                        if os.path.normcase(_wentry.get("crdb_path", "")) == os.path.normcase(crdb_path):
                            _prev_layers = _wentry.get("dwg_layers", [])
                            break
                except Exception:
                    pass
                _dwg_exported_layers, _dwg_geom, _dwg_srid, _dwg_cs = self._show_dwg_layer_dialog(  # type: ignore[attr-defined]
                    dwg_var.get().strip(), gpkg_path, previous_layers=_prev_layers)

                # Write DWG geometry to the other export formats
                # KMZ and LandXML writers handle their own reprojection via srid/source_crs
                if _dwg_geom:
                    _dwg_shp = os.path.join(output_dir, crdb_stem + "_geometry.shp")
                    _dwg_kmz = os.path.join(output_dir, crdb_stem + ".kmz")
                    _dwg_xml = os.path.join(output_dir, crdb_stem + ".xml")
                    for _label, _func in [
                        ("SHP geometry", lambda: self._write_dwg_geometry_to_shp(_dwg_shp, _dwg_geom, _dwg_srid)),  # type: ignore[attr-defined]
                        ("KMZ geometry", lambda: self._write_dwg_geometry_to_kml(_dwg_kmz, _dwg_geom, _dwg_srid, _dwg_cs)),  # type: ignore[attr-defined]
                        ("LandXML geometry", lambda: self._write_dwg_geometry_to_landxml(_dwg_xml, _dwg_geom, _dwg_srid, _dwg_cs)),  # type: ignore[attr-defined]
                    ]:
                        try:
                            _func()
                        except Exception:
                            pass  # geometry in non-GPKG formats is best-effort

            # Auto-register this CRDB in the daily watch list (after DWG
            # dialog so we can persist the selected layer names)
            try:
                self._register_crdb_watch(  # type: ignore[attr-defined]
                    crdb_path=crdb_path,
                    gpkg_output_path=gpkg_path,
                    fxl_path=self.fxl_path if fxl_data_ref[0] else None,
                    jxl_map=_jxl_map,
                    dwg_path=dwg_var.get().strip() if _dwg_exported_layers else None,
                    dwg_layers=_dwg_exported_layers,
                    dwg_srid=_dwg_srid if _dwg_exported_layers else 0,
                    dwg_cs=_dwg_cs if _dwg_exported_layers else "",
                    export_codes=_selected_codes,
                    client_name=_client_name if _client_name and _client_name in client_schemas else None,
                )
            except Exception:
                pass  # watch-list failure must never block the export result

            # Alert about missing media (shown after dialog closes so it isn't obscured)
            if _mm:
                lines = [f"  {pt}: {fn}" for pt, fn in sorted(_mm.items())[:15]]
                if len(_mm) > 15:
                    lines.append(f"  … and {len(_mm) - 15} more (see Issues Report)")
                messagebox.showwarning(
                    "Missing Media Files",
                    f"{len(_mm)} point(s) have media referenced in JXL "
                    f"but the files could not be located:\n\n" + "\n".join(lines),
                )


        btn_row2 = tk.Frame(win)
        btn_row2.grid(row=13, column=0, columnspan=2, pady=(6, 10))
        tk.Button(btn_row2, text="Export", width=14, bg="#1F4E79", fg="white",
                  command=_do_export).grid(row=0, column=0, padx=8)
        tk.Button(btn_row2, text="Cancel", width=10, command=win.destroy).grid(row=0, column=1, padx=8)

        win.wait_window()

    # ---------- Client GDB schema management ----------

    def _get_client_schemas_path(self) -> str:
        """Return the path to client_schemas.json, co-located with config.json."""
        base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
        return os.path.join(base, "client_schemas.json")

    def _load_client_schemas(self) -> dict[str, Any]:
        path = self._get_client_schemas_path()  # type: ignore[attr-defined]
        if os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    return cast(dict[str, Any], data)
            except Exception:
                pass
        return {}

    def _save_client_schema(self, client_name: str, schema: dict[str, Any]) -> None:
        path = self._get_client_schemas_path()  # type: ignore[attr-defined]
        all_schemas = self._load_client_schemas()  # type: ignore[attr-defined]
        all_schemas[client_name] = schema
        with open(path, "w", encoding="utf-8") as f:
            json.dump(all_schemas, f, indent=2, default=str)

    def _read_gdb_schema(self, gdb_path: str) -> dict[str, Any]:
        """Read an Esri File Geodatabase and extract its schema.

        Returns a dict with 'crs', 'layers' (list of layer defs), and
        'points_layer' (auto-detected survey points layer name or None).
        """
        try:
            import pyogrio  # type: ignore[import-untyped]
        except ImportError:
            raise ImportError("pyogrio is required to read GDB files.\nInstall: pip install geopandas")

        raw_layers: Any = pyogrio.list_layers(gdb_path)  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
        crs_str = ""
        layers: list[dict[str, Any]] = []

        for layer_name, geom_type in raw_layers:  # type: ignore[reportUnknownVariableType]
            try:
                info: Any = pyogrio.read_info(gdb_path, layer=str(layer_name))  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                fields: list[str] = list(info.get("fields", []))  # type: ignore[reportUnknownMemberType,reportUnknownArgumentType]
                dtypes: list[str] = [str(d) for d in info.get("dtypes", [])]  # type: ignore[reportUnknownVariableType,reportUnknownMemberType,reportUnknownArgumentType]
                count: int = int(info.get("features", 0))  # type: ignore[reportUnknownMemberType,reportUnknownArgumentType]
                layer_crs: str = str(info.get("crs", ""))  # type: ignore[reportUnknownMemberType,reportUnknownArgumentType]
                if layer_crs and not crs_str:
                    crs_str = layer_crs
                layers.append({
                    "name": str(layer_name),  # type: ignore[reportUnknownArgumentType]
                    "geometry_type": str(geom_type) if geom_type else "Table",  # type: ignore[reportUnknownArgumentType]
                    "feature_count": count,
                    "fields": [{"name": fn, "type": t} for fn, t in zip(fields, dtypes)],
                })
            except Exception:
                layers.append({
                    "name": str(layer_name),  # type: ignore[reportUnknownArgumentType]
                    "geometry_type": str(geom_type) if geom_type else "Table",  # type: ignore[reportUnknownArgumentType]
                    "feature_count": 0,
                    "fields": [],
                })

        # Auto-detect the survey points layer by looking for columns that
        # match the pattern: point name + northing/easting + code + attributes.
        points_layer: str | None = None
        best_score = 0
        _point_keywords = {"point", "pt", "point_number", "point_name", "name"}
        _north_keywords = {"north", "northing", "y", "survey_coord_y"}
        _east_keywords = {"east", "easting", "x", "survey_coord_x"}
        _code_keywords = {"code", "feature_code", "desc", "description"}
        _attr_pattern = re.compile(r"^att(?:r(?:ibute)?)?[_\s]?\d+$", re.IGNORECASE)

        for lyr in layers:
            if lyr["geometry_type"] == "Table":
                continue
            field_names_lower = {f["name"].lower() for f in lyr["fields"]}
            score = 0
            if field_names_lower & _point_keywords:
                score += 2
            if field_names_lower & _north_keywords:
                score += 2
            if field_names_lower & _east_keywords:
                score += 2
            if field_names_lower & _code_keywords:
                score += 2
            attr_count = sum(1 for f in lyr["fields"] if _attr_pattern.match(f["name"]))
            if attr_count >= 3:
                score += 3
            if lyr["geometry_type"] and "point" in lyr["geometry_type"].lower():
                score += 1
            if score > best_score:
                best_score = score
                points_layer = lyr["name"]

        return {
            "crs": crs_str[:200],
            "layers": layers,
            "points_layer": points_layer,
        }

    def _show_gdb_import_dialog(self, gdb_path: str) -> None:
        """Show dialog to import a GDB schema and associate it with a client name."""
        self.status.config(text="Reading GDB schema…")
        self.update_idletasks()

        try:
            schema = self._read_gdb_schema(gdb_path)  # type: ignore[attr-defined]
        except Exception as exc:
            self.status.config(text="Ready")
            messagebox.showerror("GDB Read Error", f"Could not read GDB:\n{exc}")
            return

        self.status.config(text="Ready")
        layers = schema.get("layers", [])
        points_layer = schema.get("points_layer")

        # Count layer types
        point_layers = [l for l in layers if l["geometry_type"] and "point" in l["geometry_type"].lower()]
        line_layers = [l for l in layers if l["geometry_type"] and "line" in l["geometry_type"].lower()]
        poly_layers = [l for l in layers if l["geometry_type"] and "polygon" in l["geometry_type"].lower()]
        table_layers = [l for l in layers if l["geometry_type"] == "Table"]

        win = tk.Toplevel(self)
        win.title("Import Client GDB Schema")
        win.resizable(False, False)
        win.grab_set()
        _raise_window(win)

        pad: dict[str, Any] = dict(padx=10, pady=3, sticky="w")

        tk.Label(win, text="Import GDB Schema", font=("Helvetica", 11, "bold")).grid(
            row=0, column=0, columnspan=2, padx=10, pady=(10, 4), sticky="w")

        tk.Label(win, text="GDB file:").grid(row=1, column=0, **pad)
        tk.Label(win, text=os.path.basename(gdb_path), fg="#1F4E79").grid(row=1, column=1, **pad)

        tk.Label(win, text="Layers found:").grid(row=2, column=0, **pad)
        summary = (f"{len(layers)} total — "
                   f"{len(point_layers)} point, {len(line_layers)} line, "
                   f"{len(poly_layers)} polygon, {len(table_layers)} table")
        tk.Label(win, text=summary).grid(row=2, column=1, **pad)

        tk.Label(win, text="CRS:").grid(row=3, column=0, **pad)
        crs_display = schema.get("crs", "")[:80] or "(unknown)"
        tk.Label(win, text=crs_display, fg="#555555").grid(row=3, column=1, **pad)

        # Survey points layer selector
        tk.Label(win, text="Survey points layer:").grid(row=4, column=0, **pad)
        layer_names = [l["name"] for l in layers if l["geometry_type"] != "Table"]
        pts_var = tk.StringVar(value=points_layer or (layer_names[0] if layer_names else ""))
        pts_combo = ttk.Combobox(win, textvariable=pts_var, values=layer_names,
                                 width=30, state="readonly")
        pts_combo.grid(row=4, column=1, padx=10, pady=3, sticky="w")
        if points_layer:
            auto_txt = f"  (auto-detected: {points_layer})"
            tk.Label(win, text=auto_txt, fg="#2E7D32", font=("Segoe UI", 8)).grid(
                row=5, column=1, padx=10, sticky="w")

        # Client name
        tk.Label(win, text="Client name:").grid(row=6, column=0, **pad)
        client_var = tk.StringVar()
        existing_schemas = self._load_client_schemas()  # type: ignore[attr-defined]
        client_combo = ttk.Combobox(win, textvariable=client_var,
                                    values=sorted(existing_schemas.keys()),
                                    width=30)
        client_combo.grid(row=6, column=1, padx=10, pady=3, sticky="w")
        client_combo.focus_set()

        status_lbl = tk.Label(win, text="", anchor="w")
        status_lbl.grid(row=7, column=0, columnspan=2, padx=10, pady=2, sticky="ew")

        def _do_save() -> None:
            client_name = client_var.get().strip()
            if not client_name:
                messagebox.showwarning("No Client Name", "Enter a client name.", parent=win)
                return

            selected_pts_layer = pts_var.get().strip()

            # Check for existing
            if client_name in existing_schemas:
                ans = messagebox.askyesno(
                    "Client Exists",
                    f"A schema for '{client_name}' already exists.\n\n"
                    f"Overwrite with this GDB's schema?",
                    parent=win,
                )
                if not ans:
                    return

            # Build the stored schema
            stored: dict[str, Any] = {
                "source_gdb": os.path.basename(gdb_path),
                "crs": schema.get("crs", ""),
                "points_layer": selected_pts_layer,
                "points_fields": [],
                "layers": [],
            }
            for lyr in layers:
                lyr_entry: dict[str, Any] = {
                    "name": lyr["name"],
                    "geometry_type": lyr["geometry_type"],
                    "fields": [f["name"] for f in lyr["fields"]],
                }
                stored["layers"].append(lyr_entry)
                if lyr["name"] == selected_pts_layer:
                    stored["points_fields"] = [
                        {"name": f["name"], "type": f["type"]} for f in lyr["fields"]
                    ]

            self._save_client_schema(client_name, stored)  # type: ignore[attr-defined]
            pts_count = len(stored["points_fields"])
            messagebox.showinfo(
                "Schema Saved",
                f"Client '{client_name}' schema saved.\n\n"
                f"  Survey points layer: {selected_pts_layer} ({pts_count} fields)\n"
                f"  Total layers: {len(layers)}",
                parent=win,
            )
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.grid(row=8, column=0, columnspan=2, pady=(6, 10))
        tk.Button(btn_frame, text="Save Schema", width=14, bg="#1F4E79", fg="white",
                  command=_do_save).grid(row=0, column=0, padx=8)
        tk.Button(btn_frame, text="Cancel", width=10, command=win.destroy).grid(row=0, column=1, padx=8)

        win.wait_window()

    def _show_schema_manager(self) -> None:
        """Show a dialog to manage client GDB schemas — list, import, delete."""
        schemas = self._load_client_schemas()  # type: ignore[attr-defined]

        win = tk.Toplevel(self)
        win.title("Client GDB Schemas")
        win.resizable(False, False)
        win.grab_set()
        _raise_window(win)

        tk.Label(win, text="Client GDB Schemas", font=("Helvetica", 11, "bold")).pack(
            padx=14, pady=(10, 4), anchor="w")
        tk.Label(win, text="Import a client's GDB to store their field schema for CRDB exports.",
                 fg="#555555").pack(padx=14, anchor="w")

        # Schema list
        list_frame = tk.Frame(win)
        list_frame.pack(padx=14, pady=8, fill="both")
        cols = ("client", "source", "points_layer", "fields")
        tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=8)
        tree.heading("client", text="Client Name")
        tree.heading("source", text="Source GDB")
        tree.heading("points_layer", text="Points Layer")
        tree.heading("fields", text="Fields")
        tree.column("client", width=150)
        tree.column("source", width=180)
        tree.column("points_layer", width=130)
        tree.column("fields", width=60)
        tree.pack(side="left", fill="both")
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=tree.yview)  # type: ignore[arg-type]
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)

        def _refresh() -> None:
            nonlocal schemas
            schemas = self._load_client_schemas()  # type: ignore[attr-defined]
            tree.delete(*tree.get_children())
            for name, s in sorted(schemas.items()):
                n_fields = len(s.get("points_fields", []))
                tree.insert("", "end", values=(
                    name,
                    s.get("source_gdb", ""),
                    s.get("points_layer", ""),
                    n_fields,
                ))

        _refresh()

        btn_frame = tk.Frame(win)
        btn_frame.pack(padx=14, pady=(0, 10))

        def _import_gdb() -> None:
            folder = filedialog.askdirectory(
                title="Select a .gdb folder to import",
                parent=win,
            )
            if not folder:
                return
            if not folder.lower().endswith(".gdb"):
                messagebox.showwarning("Not a GDB",
                    "Please select an Esri File Geodatabase folder (*.gdb).", parent=win)
                return
            win.withdraw()
            self._show_gdb_import_dialog(folder)  # type: ignore[attr-defined]
            win.deiconify()
            _refresh()

        def _delete_schema() -> None:
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            client_name = vals[0]
            ans = messagebox.askyesno(
                "Delete Schema",
                f"Delete the schema for '{client_name}'?",
                parent=win,
            )
            if ans:
                schemas.pop(client_name, None)
                path = self._get_client_schemas_path()  # type: ignore[attr-defined]
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(schemas, f, indent=2, default=str)
                _refresh()

        tk.Button(btn_frame, text="Import GDB…", width=14, command=_import_gdb).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Delete", width=10, command=_delete_schema).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Close", width=10, command=win.destroy).pack(side="left", padx=6)

        win.wait_window()

    # ---------- CRDB watch list ----------

    def _get_watchlist_path(self) -> str:
        """Return the path to crdb_watchlist.json, co-located with config.json."""
        base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
        return os.path.join(base, "crdb_watchlist.json")

    def _load_watchlist(self) -> dict[str, Any]:
        path = self._get_watchlist_path()  # type: ignore[attr-defined]
        if os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    return cast(dict[str, Any], data)
            except Exception:
                pass
        return {"entries": [], "pending_notifications": []}

    def _save_watchlist(self, wl: dict[str, Any]) -> None:
        path = self._get_watchlist_path()  # type: ignore[attr-defined]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(wl, f, indent=2)

    @staticmethod
    def _crdb_file_hash(path: str) -> str:
        """Return a hex MD5 of the CRDB file for change detection."""
        h = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _register_crdb_watch(
        self,
        crdb_path: str,
        gpkg_output_path: str,
        fxl_path: str | None,
        jxl_map: dict[str, str],
        dwg_path: str | None = None,
        dwg_layers: list[str] | None = None,
        dwg_srid: int = 0,
        dwg_cs: str = "",
        export_codes: list[str] | None = None,
        client_name: str | None = None,
    ) -> None:
        """Add or refresh a CRDB in the watch list, then ensure the Task Scheduler job exists."""
        wl = self._load_watchlist()  # type: ignore[attr-defined]
        entries: list[Any] = cast(list[Any], wl.get("entries", []))
        now_str = datetime.datetime.now().isoformat(timespec="seconds")

        try:
            file_hash = self._crdb_file_hash(crdb_path)
            file_mtime = os.path.getmtime(crdb_path)
        except Exception:
            file_hash = ""
            file_mtime = 0.0

        # Find existing entry for this crdb_path (update) or create new one
        norm_path = os.path.normcase(os.path.abspath(crdb_path))
        updated = False
        for entry in entries:
            if os.path.normcase(os.path.abspath(cast(str, entry.get("crdb_path", "")))) == norm_path:
                entry["gpkg_output_path"] = gpkg_output_path
                entry["fxl_path"] = fxl_path
                entry["jxl_map"] = jxl_map
                entry["dwg_path"] = dwg_path
                entry["dwg_layers"] = dwg_layers or []
                entry["dwg_srid"] = dwg_srid
                entry["dwg_cs"] = dwg_cs
                entry["export_codes"] = export_codes or []
                if client_name is not None:
                    entry["client_name"] = client_name
                entry["last_crdb_upload"] = now_str   # reset the 3-month clock
                entry["last_file_hash"] = file_hash
                entry["last_file_mtime"] = file_mtime
                updated = True
                break

        if not updated:
            entries.append({
                "crdb_path": crdb_path,
                "gpkg_output_path": gpkg_output_path,
                "fxl_path": fxl_path,
                "jxl_map": jxl_map,
                "dwg_path": dwg_path,
                "dwg_layers": dwg_layers or [],
                "dwg_srid": dwg_srid,
                "dwg_cs": dwg_cs,
                "export_codes": export_codes or [],
                "client_name": client_name or "",
                "first_registered": now_str,
                "last_crdb_upload": now_str,
                "last_checked": now_str,
                "last_file_hash": file_hash,
                "last_file_mtime": file_mtime,
            })

        wl["entries"] = entries
        self._save_watchlist(wl)  # type: ignore[attr-defined]

        # Register the Task Scheduler job (idempotent — /f overwrites if exists)
        self._ensure_task_scheduler()  # type: ignore[attr-defined]

    def _ensure_task_scheduler(self) -> None:
        """Create (or silently overwrite) the daily 2 AM Windows Task Scheduler job.

        Uses PowerShell to set ``StartWhenAvailable = $true`` so that if the
        computer was off at 2 AM, the task runs as soon as the user logs in.
        Falls back to plain ``schtasks`` if PowerShell fails.
        """
        try:
            if getattr(sys, "frozen", False):
                task_exe = sys.executable
                task_args = "--crdb-check"
            else:
                py_dir = os.path.dirname(sys.executable)
                pythonw = os.path.join(py_dir, "pythonw.exe")
                if not os.path.isfile(pythonw):
                    pythonw = sys.executable
                script = os.path.abspath(__file__)
                task_exe = pythonw
                task_args = f'"{script}" --crdb-check'

            task_name = "DataValidationTool_CRDBCheck"

            # Try PowerShell first — supports StartWhenAvailable
            ps_script = (
                f"$action = New-ScheduledTaskAction -Execute '{task_exe}' -Argument '{task_args}';"
                f"$trigger = New-ScheduledTaskTrigger -Daily -At '2:00AM';"
                f"$settings = New-ScheduledTaskSettingsSet -StartWhenAvailable -AllowStartIfOnBatteries -DontStopIfGoingOnBatteries;"
                f"Register-ScheduledTask -TaskName '{task_name}' -Action $action -Trigger $trigger -Settings $settings -Force"
            )
            result = subprocess.run(
                ["powershell", "-WindowStyle", "Hidden", "-NonInteractive", "-Command", ps_script],
                capture_output=True, check=False,
                creationflags=0x08000000,  # CREATE_NO_WINDOW
            )
            if result.returncode == 0:
                return  # success

            # Fallback: plain schtasks (no StartWhenAvailable support)
            subprocess.run(
                [
                    "schtasks", "/create",
                    "/tn", task_name,
                    "/tr", f'"{task_exe}" {task_args}',
                    "/sc", "DAILY",
                    "/st", "02:00",
                    "/f",
                    "/rl", "HIGHEST",
                ],
                capture_output=True, check=False,
            )
        except Exception:
            pass  # Task Scheduler failure is non-fatal

    def _send_windows_toast(self, title: str, body: str) -> None:
        """Fire a Windows balloon-tip / toast notification via PowerShell (no extra deps)."""
        ps = (
            "Add-Type -AssemblyName System.Windows.Forms; "
            "$n = New-Object System.Windows.Forms.NotifyIcon; "
            "$n.Icon = [System.Drawing.SystemIcons]::Information; "
            "$n.BalloonTipIcon = 'Info'; "
            f"$n.BalloonTipTitle = '{title.replace(chr(39), '')}'; "
            f"$n.BalloonTipText  = '{body.replace(chr(39), '')}'; "
            "$n.Visible = $true; "
            "$n.ShowBalloonTip(9000); "
            "Start-Sleep -Milliseconds 10000; "
            "$n.Dispose()"
        )
        try:
            subprocess.Popen(
                ["powershell", "-WindowStyle", "Hidden", "-NonInteractive", "-Command", ps],
                creationflags=0x08000000,  # CREATE_NO_WINDOW
            )
        except Exception:
            pass

    def _check_pending_notifications(self) -> None:
        """Called ~0.8 s after startup — show any notifications queued by background checks."""
        try:
            wl = self._load_watchlist()  # type: ignore[attr-defined]
            pending: list[Any] = cast(list[Any], wl.get("pending_notifications", []))
            if not pending:
                return

            lines: list[str] = []
            gpkg_paths: list[str] = []
            for note in pending:
                msg = cast(str, note.get("message", ""))
                gp = cast(str, note.get("gpkg_path", ""))
                if msg:
                    lines.append(f"• {msg}")
                if gp and gp not in gpkg_paths:
                    gpkg_paths.append(gp)

            wl["pending_notifications"] = []
            self._save_watchlist(wl)  # type: ignore[attr-defined]

            messagebox.showinfo(
                "CRDB Watch List — Data Exported",
                "The following CRDB files had changes and were automatically re-exported:\n\n"
                + "\n".join(lines)
                + "\n\nAll export files (GPKG, CSV, SHP, LandXML, KMZ) are ready.",
            )
        except Exception:
            pass

    def _startup_watchlist_check(self) -> None:
        """Called ~2 s after startup — run a background watchlist check.

        This catches any missed scheduled runs (e.g. computer was off at 2 AM).
        Runs in a background thread so it doesn't block the UI.
        """
        import threading

        watchlist_path = self._get_watchlist_path()  # type: ignore[attr-defined]
        if not os.path.isfile(watchlist_path):
            return

        def _bg() -> None:
            try:
                _run_background_check(watchlist_path)
            except Exception:
                pass
            # After the check, schedule notification display on the main thread
            try:
                self.after(500, self._check_pending_notifications)  # type: ignore[attr-defined]
            except Exception:
                pass

        t = threading.Thread(target=_bg, daemon=True)
        t.start()

    def _show_watchlist_dialog(self) -> None:
        """Display and manage the list of watched CRDB files."""
        wl = self._load_watchlist()  # type: ignore[attr-defined]

        win = tk.Toplevel(self)
        win.title("CRDB Watch List")
        _place_window(win, 760, 340)
        win.grab_set()
        _raise_window(win)

        tk.Label(win, text="CRDB files monitored for daily changes (checked at 2:00 AM)",
                 font=("Helvetica", 10, "bold")).pack(anchor="w", padx=10, pady=(8, 2))

        cols = ("crdb", "gpkg", "last_upload", "last_checked", "expires")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=8)
        tree.heading("crdb", text="CRDB File")
        tree.heading("gpkg", text="Output GPKG")
        tree.heading("last_upload", text="Last Export")
        tree.heading("last_checked", text="Last Checked")
        tree.heading("expires", text="Expires")
        tree.column("crdb", width=180)
        tree.column("gpkg", width=160)
        tree.column("last_upload", width=110)
        tree.column("last_checked", width=110)
        tree.column("expires", width=90)
        tree.pack(fill="both", expand=True, padx=8, pady=4)

        def _populate() -> None:
            for iid in tree.get_children():
                tree.delete(iid)
            for e in cast(list[Any], wl.get("entries", [])):
                upload_str = cast(str, e.get("last_crdb_upload", ""))[:19].replace("T", " ")
                checked_str = cast(str, e.get("last_checked", ""))[:19].replace("T", " ")
                try:
                    exp_dt = datetime.datetime.fromisoformat(
                        cast(str, e.get("last_crdb_upload", ""))
                    ) + datetime.timedelta(days=90)
                    exp_str = exp_dt.strftime("%Y-%m-%d")
                except Exception:
                    exp_str = "?"
                tree.insert("", "end", values=(
                    os.path.basename(cast(str, e.get("crdb_path", ""))),
                    os.path.basename(cast(str, e.get("gpkg_output_path", ""))),
                    upload_str, checked_str, exp_str,
                ))

        _populate()

        def _remove_selected() -> None:
            sel = tree.selection()
            if not sel:
                return
            idx = tree.index(sel[0])
            entries_now: list[Any] = cast(list[Any], wl.get("entries", []))
            if 0 <= idx < len(entries_now):
                removed = os.path.basename(cast(str, entries_now[idx].get("crdb_path", "")))
                entries_now.pop(idx)
                wl["entries"] = entries_now
                self._save_watchlist(wl)  # type: ignore[attr-defined]
                _populate()
                self.status.config(text=f"Removed '{removed}' from watch list.")

        def _check_now() -> None:
            """Trigger an immediate background check (runs in a thread, shows result)."""
            import threading
            watchlist_path = self._get_watchlist_path()  # type: ignore[attr-defined]
            win.config(cursor="wait")

            def _worker() -> None:
                try:
                    _run_background_check(watchlist_path)
                except Exception:
                    pass
                self.after(0, _done)

            def _done() -> None:
                win.config(cursor="")
                updated_wl = self._load_watchlist()  # type: ignore[attr-defined]
                wl.update(updated_wl)
                _populate()
                pending_count = len(cast(list[Any], wl.get("pending_notifications", [])))
                if pending_count:
                    self._check_pending_notifications()  # type: ignore[attr-defined]
                else:
                    messagebox.showinfo("Check Complete",
                                        "All watched CRDBs checked — no changes detected.",
                                        parent=win)

            threading.Thread(target=_worker, daemon=True).start()

        btn_row = tk.Frame(win)
        btn_row.pack(pady=6)
        tk.Button(btn_row, text="Remove Selected", width=16, command=_remove_selected).grid(row=0, column=0, padx=6)
        tk.Button(btn_row, text="Check Now", width=14, command=_check_now).grid(row=0, column=1, padx=6)
        tk.Button(btn_row, text="Close", width=10, command=win.destroy).grid(row=0, column=2, padx=6)
        win.wait_window()

    # ---------- validation log ----------
    def _write_validation_log(self) -> None:
        """Append a JSON entry to validation_log.json in the CSV's folder."""
        try:
            csv_dir = os.path.dirname(self.csv_path) if self.csv_path else None
            if not csv_dir:
                return
            log_path = os.path.join(csv_dir, "validation_log.json")
            entry = {
                "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
                "user": os.environ.get("USERNAME", os.environ.get("USER", "unknown")),
                "csv_file": os.path.basename(self.csv_path or ""),
                "fxl_file": os.path.basename(self.fxl_path or ""),
                "mtr_file": os.path.basename(self.mtr_path or "") if self.mtr_path else "",
            }
            entries: list[Any] = []
            if os.path.isfile(log_path):
                try:
                    with open(log_path, "r", encoding="utf-8") as f:
                        existing = json.load(f)
                    if isinstance(existing, list):
                        entries = cast(list[Any], existing)
                except Exception:
                    pass
            entries.append(entry)
            with open(log_path, "w", encoding="utf-8") as f:
                json.dump(entries, f, indent=2)
        except Exception:
            pass  # log failures must never break the main flow

    # ---------- close / cleanup ----------
    def _on_close(self):
        try:
            if self._excel and self._wb_com:
                try:
                    wbname = self._wb_com.Name
                    self._excel.Run(f"'{wbname}'!ValidationModule.PromptSaveAndClose")
                except Exception:
                    pass
        except Exception:
            pass
        self.after(400, self._finalize_and_exit)

    def _finalize_and_exit(self):
        try:
            if self._wb_com:
                try:
                    self._wb_com.Close(False)
                except Exception:
                    pass
            if self._excel:
                try:
                    if self._excel.Workbooks.Count == 0:
                        self._excel.Quit()
                except Exception:
                    pass
        except Exception:
            pass
        self._excel = None
        self._wb_com = None
        self._excel_opened = False
        try:
            shutil.rmtree(self._tmpdir, ignore_errors=True)
        except Exception:
            pass
        self.destroy()


def _raise_window(win: Any) -> None:
    """Bring *win* to the front and give it keyboard focus.

    Uses a brief -topmost flash so Windows actually respects the z-order
    request even when another application currently owns focus.
    """
    win.lift()
    win.focus_force()
    try:
        win.attributes("-topmost", True)
        win.after(10, lambda: win.attributes("-topmost", False))
    except Exception:
        pass


def _place_window(win: Any, width: int, height: int) -> None:
    """Center *win* on the monitor that contains the mouse cursor.

    Uses the Windows API (ctypes) so the window lands on whichever physical
    screen the user is currently working on, including secondary monitors.
    Falls back to basic tkinter centering when the API is unavailable.
    """
    import ctypes
    import ctypes.wintypes

    placed = False
    try:
        class _MONITORINFO(ctypes.Structure):
            _fields_ = [
                ("cbSize",    ctypes.c_ulong),
                ("rcMonitor", ctypes.wintypes.RECT),
                ("rcWork",    ctypes.wintypes.RECT),   # excludes taskbar
                ("dwFlags",   ctypes.c_ulong),
            ]

        pt = ctypes.wintypes.POINT()
        ctypes.windll.user32.GetCursorPos(ctypes.byref(pt))
        # MONITOR_DEFAULTTONEAREST = 2 → nearest monitor when cursor is off-screen
        h_mon = ctypes.windll.user32.MonitorFromPoint(pt, 2)
        mi = _MONITORINFO()
        mi.cbSize = ctypes.sizeof(_MONITORINFO)
        ctypes.windll.user32.GetMonitorInfoW(h_mon, ctypes.byref(mi))

        mon_w = mi.rcWork.right  - mi.rcWork.left
        mon_h = mi.rcWork.bottom - mi.rcWork.top
        x = mi.rcWork.left + max(0, (mon_w - width)  // 2)
        y = mi.rcWork.top  + max(0, (mon_h - height) // 2)
        win.geometry(f"{width}x{height}+{x}+{y}")
        placed = True
    except Exception:
        pass

    if not placed:
        # Generic fallback: use tkinter's reported screen dimensions
        win.update_idletasks()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = max(0, (sw - width)  // 2)
        y = max(0, (sh - height) // 2)
        win.geometry(f"{width}x{height}+{x}+{y}")

    _raise_window(win)


class _HeadlessCRDBRunner:
    """Minimal stub that borrows App's pure-logic CRDB methods for headless background use.

    None of the borrowed methods access tkinter state — they only call each other
    via self.method(), so binding them onto this plain object works correctly.
    """
    pass


# Bind the pure-logic methods from DataValidationTool onto the headless stub.
# This is done after the class is defined so the names are resolved.
for _m in (
    "_load_crdb_rows", "_extract_jxl_hints", "_match_points_to_jxls",
    "_parse_jxl", "_make_gpkg_point_blob", "_write_gpkg", "parse_fxl",
    "_crdb_file_hash",
    # Multi-format export methods (headless re-export)
    "_write_crdb_csv", "_write_crdb_shp", "_write_crdb_landxml", "_write_crdb_kmz",
    "_find_gis_output_dir",
    "_parse_dxf_geometry",
    "_read_dwg_geometry_ezdxf", "_convert_dwg_to_dxf",
    "_resolve_autodesk_cs", "_load_autodesk_cs_map",
    # DWG geometry to non-GPKG formats
    "_write_dwg_geometry_to_shp", "_write_dwg_geometry_to_kml",
    "_write_dwg_geometry_to_landxml",
    # DWG geometry methods (headless re-export)
    "_open_dwg_readonly", "_read_dwg_cs_code", "_extract_dwg_geometry",
    "_write_dwg_geometry_to_gpkg", "_make_gpkg_linestring_blob",
    "_make_gpkg_polygon_blob",
):
    setattr(_HeadlessCRDBRunner, _m, getattr(DataValidationTool, _m))


def _run_background_check(watchlist_path: str) -> None:
    """Headless CRDB change check — called by Task Scheduler or the Watch List dialog.

    For each watched CRDB:
    - Computes MD5 and compares against the stored hash.
    - If changed: re-exports the GeoPackage using the stored JXL map and FXL path.
    - Queues a pending notification and sends a Windows toast.
    - Removes entries whose last_crdb_upload is older than 90 days.
    Updates crdb_watchlist.json in place.
    """
    if not os.path.isfile(watchlist_path):
        return

    try:
        with open(watchlist_path, "r", encoding="utf-8") as f:
            wl: dict[str, Any] = cast(dict[str, Any], json.load(f))
    except Exception:
        return

    entries: list[Any] = cast(list[Any], wl.get("entries", []))
    pending: list[Any] = cast(list[Any], wl.get("pending_notifications", []))
    runner = _HeadlessCRDBRunner()
    now = datetime.datetime.now()
    updated_entries: list[Any] = []
    newly_exported: list[str] = []

    for entry in entries:
        crdb_path: str = cast(str, entry.get("crdb_path", ""))
        gpkg_path: str = cast(str, entry.get("gpkg_output_path", ""))
        fxl_path: str | None = cast(str | None, entry.get("fxl_path"))
        jxl_map: dict[str, str] = cast(dict[str, str], entry.get("jxl_map", {}))
        last_upload_str: str = cast(str, entry.get("last_crdb_upload", now.isoformat()))

        # Drop entries older than 90 days since the last manual upload
        try:
            last_upload_dt = datetime.datetime.fromisoformat(last_upload_str)
            if (now - last_upload_dt).days > 90:
                continue  # expired — do not carry forward
        except Exception:
            pass

        if not os.path.isfile(crdb_path):
            updated_entries.append(entry)  # keep entry; file may be temporarily unavailable
            continue

        # Check for changes via MD5
        try:
            current_hash: str = runner._crdb_file_hash(crdb_path)  # type: ignore[attr-defined]
        except Exception:
            updated_entries.append(entry)
            continue

        stored_hash: str = cast(str, entry.get("last_file_hash", ""))
        entry["last_checked"] = now.isoformat(timespec="seconds")

        if current_hash == stored_hash:
            updated_entries.append(entry)
            continue  # no change

        # CRDB changed — re-export all formats into a date-stamped GIS folder
        try:
            rows: list[dict[str, Any]] = runner._load_crdb_rows(crdb_path)  # type: ignore[attr-defined]
            matched, _, _ = runner._match_points_to_jxls(rows, jxl_map)  # type: ignore[attr-defined]

            fxl_data: dict[str, list[dict[str, Any]]] = {}
            if fxl_path and os.path.isfile(fxl_path):
                try:
                    fxl_data = runner.parse_fxl(fxl_path)  # type: ignore[attr-defined]
                except Exception:
                    pass

            # Compute output directory (ASBUILT/0_GIS/WEEKLY_UPDATE/YYYYMMDD)
            _out_dir: str = runner._find_gis_output_dir(crdb_path)  # type: ignore[attr-defined]
            os.makedirs(_out_dir, exist_ok=True)  # type: ignore[reportUnknownArgumentType]
            _stem = os.path.splitext(os.path.basename(crdb_path))[0]

            # Filter rows by saved export_codes (CSV gets all, others get filtered)
            _export_codes_w: list[str] = cast(list[str], entry.get("export_codes") or [])
            _filtered: list[dict[str, Any]]
            if _export_codes_w:
                _ec_set = set(_export_codes_w)
                _filtered = [r for r in cast(list[dict[str, Any]], rows)
                             if (r.get("code", "UNKNOWN") or "UNKNOWN") in _ec_set]
            else:
                _filtered = cast(list[dict[str, Any]], rows)

            # Write all five formats (best-effort each)
            gpkg_path = os.path.join(_out_dir, _stem + ".gpkg")  # type: ignore[reportUnknownArgumentType]
            runner._write_gpkg(gpkg_path, _filtered, matched, fxl_data, {})  # type: ignore[attr-defined]
            for _wfunc in [
                lambda: runner._write_crdb_csv(os.path.join(_out_dir, _stem + ".csv"), rows, matched, fxl_data, {}),  # type: ignore[attr-defined]
                lambda: runner._write_crdb_shp(os.path.join(_out_dir, _stem + ".shp"), _filtered, matched),  # type: ignore[attr-defined]
                lambda: runner._write_crdb_landxml(os.path.join(_out_dir, _stem + ".xml"), _filtered, matched),  # type: ignore[attr-defined]
                lambda: runner._write_crdb_kmz(os.path.join(_out_dir, _stem + ".kmz"), _filtered, matched, fxl_data),  # type: ignore[attr-defined]
            ]:
                try:
                    _wfunc()
                except Exception:
                    pass  # non-GPKG formats are best-effort in headless mode

            # Re-export DWG geometry if configured — uses saved layers and CRS
            dwg_path_w: str = cast(str, entry.get("dwg_path") or "")
            dwg_layers_w: list[str] = cast(list[str], entry.get("dwg_layers") or [])
            dwg_srid_w: int = int(entry.get("dwg_srid", 0) or 0)
            dwg_cs_w: str = cast(str, entry.get("dwg_cs", "") or "")
            if dwg_path_w and dwg_layers_w and os.path.isfile(dwg_path_w):
                try:
                    # Use ezdxf approach (DWG→DXF via AutoCAD, read with ezdxf)
                    _layer_counts: dict[str, int]
                    _all_geom: dict[str, list[dict[str, Any]]]
                    _cs: str
                    _layer_counts, _all_geom, _cs = runner._read_dwg_geometry_ezdxf(dwg_path_w)  # type: ignore[attr-defined]
                    _all_geom = cast(dict[str, list[dict[str, Any]]], _all_geom)
                    _cs = cast(str, _cs)
                    # Use saved CS if auto-detection failed
                    if not _cs:
                        _cs = dwg_cs_w
                    # Filter to saved layers only
                    _lset = set(dwg_layers_w)
                    _geom: dict[str, list[dict[str, Any]]] = {
                        ly: ents for ly, ents in _all_geom.items()
                        if ly in _lset and ents}
                    # Use saved SRID if auto-detection failed
                    _srid: int = dwg_srid_w
                    if _cs and not _srid:
                        _srid = int(runner._resolve_autodesk_cs(_cs))  # type: ignore[attr-defined]

                    if _geom:
                        # GPKG: write with transformer
                        _gpkg_srid = _srid
                        _transformer: Any = None
                        if _srid and _srid != 4326:
                            try:
                                from pyproj import Transformer as _HPT  # type: ignore[import-untyped]
                                _ht: Any = _HPT.from_crs(  # type: ignore[reportUnknownVariableType,reportUnknownMemberType]
                                    f"EPSG:{_srid}", "EPSG:4326", always_xy=True)
                                _transformer = lambda x, y: _ht.transform(x, y)  # type: ignore[reportUnknownVariableType,reportUnknownLambdaType]
                                _gpkg_srid = 4326
                            except Exception:
                                pass
                        runner._write_dwg_geometry_to_gpkg(  # type: ignore[attr-defined]
                            gpkg_path, _geom, srid=_gpkg_srid,
                            cs_description=_cs, transformer=_transformer)
                        # Other formats
                        for _dwg_wf in [
                            lambda: runner._write_dwg_geometry_to_shp(  # type: ignore[attr-defined]
                                os.path.join(_out_dir, _stem + "_geometry.shp"), _geom, _srid),
                            lambda: runner._write_dwg_geometry_to_kml(  # type: ignore[attr-defined]
                                os.path.join(_out_dir, _stem + ".kmz"), _geom, _srid, _cs),
                            lambda: runner._write_dwg_geometry_to_landxml(  # type: ignore[attr-defined]
                                os.path.join(_out_dir, _stem + ".xml"), _geom, _srid, _cs),
                        ]:
                            try:
                                _dwg_wf()
                            except Exception:
                                pass
                except Exception:
                    pass  # DWG geometry is best-effort in headless mode

            entry["last_file_hash"] = current_hash
            entry["last_file_mtime"] = os.path.getmtime(crdb_path)
            newly_exported.append(os.path.basename(crdb_path))

            pending.append({
                "message": (f"{os.path.basename(crdb_path)} changed — "
                            f"all exports regenerated to:\n"
                            f"  {_out_dir}\n"
                            f"  (GPKG, CSV, SHP, LandXML, KMZ"
                            + (", DWG geometry" if dwg_path_w and dwg_layers_w else "")
                            + f")\n"
                            f"  {now.strftime('%Y-%m-%d %H:%M')}"),
                "timestamp": now.isoformat(timespec="seconds"),
                "gpkg_path": gpkg_path,
                "output_dir": _out_dir,
            })
        except Exception as exc:
            # Log the error but keep the entry so it's retried tomorrow
            pending.append({
                "message": (f"Auto-export failed for {os.path.basename(crdb_path)}: {exc}"),
                "timestamp": now.isoformat(timespec="seconds"),
                "gpkg_path": gpkg_path,
            })

        updated_entries.append(entry)

    wl["entries"] = updated_entries
    wl["pending_notifications"] = pending

    try:
        with open(watchlist_path, "w", encoding="utf-8") as f:
            json.dump(wl, f, indent=2)
    except Exception:
        pass

    # Fire a Windows toast if any exports were regenerated
    if newly_exported:
        names = ", ".join(newly_exported[:3])
        if len(newly_exported) > 3:
            names += f" (+{len(newly_exported) - 3} more)"
        ps = (
            "Add-Type -AssemblyName System.Windows.Forms; "
            "$n = New-Object System.Windows.Forms.NotifyIcon; "
            "$n.Icon = [System.Drawing.SystemIcons]::Information; "
            "$n.BalloonTipIcon = 'Info'; "
            "$n.BalloonTipTitle = 'Data Validation Tool'; "
            f"$n.BalloonTipText = 'CRDB data changed and re-exported: {names.replace(chr(39), '')}'; "
            "$n.Visible = $true; "
            "$n.ShowBalloonTip(9000); "
            "Start-Sleep -Milliseconds 10000; "
            "$n.Dispose()"
        )
        try:
            subprocess.Popen(
                ["powershell", "-WindowStyle", "Hidden", "-NonInteractive", "-Command", ps],
                creationflags=0x08000000,
            )
        except Exception:
            pass


def main() -> None:
    if "--crdb-check" in sys.argv:
        # Headless background mode — no GUI, called by Task Scheduler at 2 AM
        base = os.path.dirname(sys.executable if getattr(sys, "frozen", False)
                               else os.path.abspath(__file__))
        watchlist_path = os.path.join(base, "crdb_watchlist.json")
        _run_background_check(watchlist_path)
        return

    app = DataValidationTool()
    try:
        app.mainloop()
    except KeyboardInterrupt:
        pass


if __name__ == "__main__":
    main()
