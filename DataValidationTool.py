import sys
import site
import os
import traceback
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Menu, Toplevel, Listbox, Scrollbar, BooleanVar
import pandas as pd

start_time = time.time()
print(f"Start imports: {start_time}")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    print("openpyxl imported successfully.")
except ImportError as e:
    print(f"Failed to import openpyxl: {e}")
    openpyxl = None

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, Menu
    print("tkinter imported successfully.")
except ImportError as e:
    print(f"Warning: Failed to import tkinter: {e}")
    tk = ttk = filedialog = messagebox = Menu = None

try:
    print(f"User site-packages: {site.getusersitepackages()}")
    print(f"System site-packages: {site.getsitepackages()}")
    import tkinterdnd2
    print("tkinterdnd2 module imported successfully.")
    from tkinterdnd2 import Tk as TkinterDnD2, DND_FILES
    print("TkinterDnD2 and DND_FILES imported successfully.")
except ImportError as e:
    print(f"Warning: Failed to import TkinterDnD2: {e}")
    tkinterdnd2 = TkinterDnD2 = DND_FILES = None

try:
    import xml.etree.ElementTree as ET
    print("xml.etree.ElementTree imported successfully.")
except ImportError as e:
    print(f"Failed to import xml.etree.ElementTree: {e}")
    sys.exit(1)

try:
    import re
    print("re imported successfully.")
except ImportError as e:
    print(f"Failed to import re: {e}")
    sys.exit(1)

try:
    import enchant
    print("enchant imported successfully.")
except ImportError as e:
    print(f"Failed to import enchant: {e}")
    enchant = None

try:
    import datetime
    print("datetime imported successfully.")
except ImportError as e:
    print(f"Failed to import datetime: {e}")
    datetime = None

print(f"Imports completed: {time.time() - start_time} seconds")

if tk:
    class Tooltip:
        def __init__(self, widget, text_func):
            self.widget = widget
            self.text_func = text_func
            self.tooltip = None
            self.id = None
            widget.bind("<Enter>", self.show)
            widget.bind("<Leave>", self.hide)

        def show(self, event=None):
            if self.tooltip or not self.text_func:
                return
            text = self.text_func()
            if not text:
                text = "No information available"
            x, y, _, _ = self.widget.bbox("insert")
            x += self.widget.winfo_rootx() + 25
            y += self.widget.winfo_rooty() + 20
            self.tooltip = tk.Toplevel(self.widget)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip.wm_geometry(f"+{x}+{y}")
            label = tk.Label(self.tooltip, text=text, justify="left",
                            background="#FFFFE0", relief="solid", borderwidth=1,
                            font=("Helvetica", "8"), wraplength=300)
            label.pack(ipadx=5, ipady=5)

        def hide(self, event=None):
            if self.tooltip:
                self.tooltip.destroy()
                self.tooltip = None

    class DataValidationTool(TkinterDnD2 if TkinterDnD2 else tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Data Validation Tool")
            self.geometry("1200x600")
            self.fxl_data = {}
            self.df = None
            self.csv_path = None
            self.dict_en = enchant.Dict("en_US") if enchant else None
            self.custom_dict = self.load_custom_dict()
            self.max_attributes = 0
            self.error_counts = {
                "Point Number Duplicates": 0,
                "XRAY Duplicates": 0,
                "Misspelled": 0,
                "Invalid: N/A": 0,
                "Invalid: NA": 0,
                "Invalid: UNK": 0,
                "Invalid: Space": 0,
                "Invalid: Zero": 0,
                "Invalid: Dash": 0,
                "Invalid: Underscore": 0,
                "Invalid: NaN": 0,
                "Invalid: Blank": 0,
                "Not in FXL: Field Code": 0,
                "Not in FXL: Attribute": 0,
                "Invalid Bend Angle": 0,
                "Not All Caps": 0
            }
            self.error_state = {}  # (row, col) -> (error_type, field_code, attr_name)
            self.error_count_label = None
            self.ignored_errors = {}
            self.ignore_all_prompted = set()
            self.undo_stack = []
            self.redo_stack = []

            style = ttk.Style(self)
            style.theme_use("default")
            style.layout("Custom.TEntry", [("Entry.border", {"sticky": "nswe", "children": [
                ("Entry.background", {"sticky": "nswe", "children": [
                    ("Entry.padding", {"sticky": "nswe", "children": [
                        ("Entry.textarea", {"sticky": "nswe"})
                    ]})
                ]})
            ]})])

            for state in ("Normal", "Duplicate", "Misspelled", "Invalid", "Extra", "NotInFXL"):
                style.configure(f"{state}.TEntry", font=("Helvetica", "8"), background="white" if state == "Normal" else "yellow" if state == "Duplicate" else "orange" if state == "Misspelled" else "#FFFFE0" if state == "Invalid" else "#D3D3D3" if state == "Extra" else "red",
                                fieldbackground="white" if state == "Normal" else "yellow" if state == "Duplicate" else "orange" if state == "Misspelled" else "#FFFFE0" if state == "Invalid" else "#D3D3D3" if state == "Extra" else "red",
                                bordercolor="black", relief="ridge", borderwidth=1)
                style.configure(f"{state}.TCombobox", font=("Helvetica", "8"), background="white" if state == "Normal" else "yellow" if state == "Duplicate" else "orange" if state == "Misspelled" else "#FFFFE0" if state == "Invalid" else "#D3D3D3" if state == "Extra" else "red",
                                fieldbackground="white" if state == "Normal" else "yellow" if state == "Duplicate" else "orange" if state == "Misspelled" else "#FFFFE0" if state == "Invalid" else "#D3D3D3" if state == "Extra" else "red",
                                bordercolor="black", relief="ridge", borderwidth=1, arrowcolor="black", arrowsize=20)
                style.layout(f"{state}.TEntry", style.layout("Custom.TEntry"))

            style.configure("Header.TLabel", font=("Helvetica", "8"), bordercolor="black", relief="ridge", borderwidth=1)
            self.setup_ui()
            self.setup_drag_and_drop()

        def load_custom_dict(self):
            custom_dict_file = "custom_dict.txt"
            custom_dict = set()
            if os.path.exists(custom_dict_file):
                with open(custom_dict_file, "r") as f:
                    for line in f:
                        word = line.strip()
                        if word:
                            custom_dict.add(word)
            return custom_dict

        def save_custom_dict(self):
            custom_dict_file = "custom_dict.txt"
            with open(custom_dict_file, "w") as f:
                for word in self.custom_dict:
                    f.write(f"{word}\n")

        def get_base_field_code(self, field_code):
            match = re.match(r"^(.*?)\d+$", field_code)
            if match:
                return match.group(1)
            return field_code

        def setup_ui(self):
            start_time = time.time()
            print(f"Start setup_ui: {start_time}")

            btn_frame = ttk.Frame(self)
            btn_frame.pack(side="top", fill="x", pady=5)
            ttk.Button(btn_frame, text="Load FXL", command=self.load_fxl).pack(side="left", padx=5)
            self.csv_btn = ttk.Button(btn_frame, text="Load CSV", command=self.load_csv, state="disabled")
            self.csv_btn.pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Export CSV", command=self.export_csv).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Export Custom CSV", command=self.export_custom_csv).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Generate Error Report", command=lambda: self.generate_error_report()).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Undo", command=self.undo).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="Redo", command=self.redo).pack(side="left", padx=5)

            self.error_count_label = ttk.Label(btn_frame, text="Error Counts: None", font=("Helvetica", "8"))
            self.error_count_label.pack(side="left", padx=10)
            print(f"Button frame setup done: {time.time() - start_time} seconds")

            self.canvas = tk.Canvas(self)
            self.v_scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
            self.h_scrollbar = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
            self.main_frame = ttk.Frame(self.canvas)
            self.header_frame = ttk.Frame(self.main_frame)
            self.header_frame.grid(row=0, column=0, sticky="nsew")
            self.table = ttk.Frame(self.main_frame)
            self.table.grid(row=1, column=0, sticky="nsew")

            self.v_scrollbar.pack(side="right", fill="y")
            self.h_scrollbar.pack(side="bottom", fill="x")
            self.canvas.pack(side="top", fill="both", expand=True)
            self.canvas.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)

            self.canvas_frame = self.canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
            self.main_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
            self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
            self.canvas.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel)
            self.canvas.bind_all("<Button-4>", self._on_mousewheel_up)
            self.canvas.bind_all("<Button-5>", self._on_mousewheel_down)
            print(f"Canvas setup done: {time.time() - start_time} seconds")

        def _on_mousewheel(self, event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        def _on_shift_mousewheel(self, event):
            self.canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        def _on_mousewheel_up(self, event):
            self.canvas.yview_scroll(-1, "units")
            return "break"

        def _on_mousewheel_down(self, event):
            self.canvas.yview_scroll(1, "units")
            return "break"

        def setup_drag_and_drop(self):
            if TkinterDnD2:
                try:
                    print("Registering drop target...")
                    self.drop_target_register(DND_FILES)
                    self.dnd_bind('<<Drop>>', self.handle_drop)
                    print("Drag-and-drop enabled.")
                except Exception as e:
                    print(f"Warning: Failed to setup drag-and-drop: {e}")
            else:
                print("Drag-and-drop not supported.")

        def handle_drop(self, event):
            files = self.splitlist(event.data)
            for file_path in files:
                if os.path.isfile(file_path):
                    ext = os.path.splitext(file_path)[1].lower()
                    state = str(self.csv_btn["state"])
                    print(f"File dropped: {file_path}, Extension: {ext}, CSV Button State: {state}, Raw State: '{state}'")
                    print(f"Condition check: ext == '.csv': {ext == '.csv'}, self.csv_btn['state'] == 'normal': {state == 'normal'}")
                    if ext == ".fxl":
                        print("Processing .fxl file...")
                        self.load_fxl_from_path(file_path)
                    elif ext == ".csv" and state == "normal":
                        print("Processing .csv file...")
                        self.load_csv_from_path(file_path)
                    else:
                        print(f"File not processed: {file_path}, Extension: {ext}, CSV Button State: {state}")

        def load_fxl_from_path(self, file_path):
            print(f"Loading FXL from: {file_path}")
            try:
                self.fxl_data = self.parse_fxl(file_path)
                self.max_attributes = max(len(attrs) for attrs in self.fxl_data.values()) if self.fxl_data else 0
                print(f"Maximum number of attributes in FXL: {self.max_attributes}")
                self.csv_btn.config(state="normal")
                state = str(self.csv_btn["state"])
                print(f"CSV Button State set to: {state}, Raw State: '{state}'")
                messagebox.showinfo("Success", "FXL loaded.")
            except Exception as e:
                print(f"Error loading FXL: {e}")
                messagebox.showerror("Error", str(e))

        def load_csv_from_path(self, file_path):
            import csv
            print(f"Loading CSV from: {file_path}")
            try:
                self.csv_path = file_path
                self.df = None
                self.raw_df = None  # Store the raw CSV data
                # Read the CSV as raw text to preprocess
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()

                # Skip empty lines
                lines = [line.strip() for line in lines if line.strip()]
                if not lines:
                    messagebox.showerror("Error", f"Failed to load CSV: '{file_path}'\n\nFile is empty.")
                    return

                # Attempt to detect delimiter using csv.Sniffer
                try:
                    sample = '\n'.join(lines[:5])
                    sniffer = csv.Sniffer()
                    dialect = sniffer.sniff(sample)
                    delimiter = dialect.delimiter
                    print(f"Detected delimiter: '{delimiter}'")
                except Exception as e:
                    print(f"Delimiter detection failed: {e}, defaulting to comma")
                    delimiter = ','

                # Parse the CSV lines
                reader = csv.reader(lines, delimiter=delimiter)
                data = list(reader)

                # Check if the first row is a metadata row (contains Job:, Version:, Units:)
                if any('Job:' in cell or 'Version:' in cell or 'Units:' in cell for cell in data[0]):
                    print("Metadata row detected in first row, removing it...")
                    data = data[1:]

                # Create raw DataFrame with all columns
                max_cols_raw = max(len(row) for row in data)
                processed_data_raw = []
                for row in data:
                    processed_row = row + [''] * (max_cols_raw - len(row))  # Pad with empty strings
                    processed_data_raw.append(processed_row)
                self.raw_df = pd.DataFrame(processed_data_raw)

                # Create processed DataFrame with only the first 5 columns for validation
                max_cols = 5  # Point Number, Easting, Northing, Elevation, Field Code
                processed_data = []
                for row in data:
                    processed_row = row[:max_cols] + [''] * (max_cols - min(len(row), max_cols))
                    processed_data.append(processed_row)
                self.df = pd.DataFrame(processed_data)

                if self.df is None or self.df.empty:
                    messagebox.showerror("Error", f"Failed to load CSV: '{file_path}'\n\nNo data after processing. Please check the file format.")
                    return

                print("Setting headers...")
                self.set_headers()
                print("Headers set, showing table...")
                self.show_table()
                print("Table updated with CSV data.")
            except Exception as e:
                print(f"Unexpected error: {e}")
                messagebox.showerror("Error", f"Failed to process CSV: '{file_path}'\n\nUnexpected error: {e}")

        def load_fxl(self):
            file = filedialog.askopenfilename(filetypes=[("FXL files", "*.fxl")])
            if file:
                self.load_fxl_from_path(file)

        def parse_fxl(self, path):
            tree = ET.parse(path)
            ns = {"fxl": "http://trimble.com/schema/fxl"}
            data = {}
            feature_types = [
                "PointFeatureDefinition",
                "LineFeatureDefinition",
                "BlockFeatureDefinition",
                "PolygonFeatureDefinition"
            ]
            for feature_type in feature_types:
                for feature in tree.findall(f".//fxl:{feature_type}", ns):
                    code = feature.get("Code")
                    if not code:
                        continue
                    attrs = []
                    for attr in feature.findall(".//fxl:Attributes/*", ns):
                        attr_info = {"name": attr.get("Name"), "type": "list" if attr.tag.endswith("ListAttribute") else "other"}
                        if attr_info["type"] == "list":
                            attr_info["items"] = [item.text for item in attr.findall(".//fxl:ListItems/fxl:Item", ns) if item.text]
                        attrs.append(attr_info)
                    data[code] = attrs
                    print(f"Loaded {feature_type} with Code: {code}, Attributes: {attrs}")
            return data

        def load_csv(self):
            file = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
            print(f"Selected CSV file: '{file}'")
            if not file:
                messagebox.showerror("Error", "No CSV file selected. Please select a valid CSV file.")
                return
            if not file.lower().endswith('.csv'):
                messagebox.showerror("Error", f"Invalid file selected: '{file}'. Please select a file with a .csv extension.")
                return
            self.load_csv_from_path(file)

        def set_headers(self):
            print("Setting headers...")
            base = ["Point Number", "Northing", "Easting", "Elevation", "Field Code"]
            num_attr_columns = max(self.df.shape[1] - len(base), self.max_attributes) if self.df.shape[1] > len(base) else self.max_attributes
            while self.df.shape[1] < len(base) + num_attr_columns:
                self.df[self.df.shape[1]] = ""
            self.df.columns = base + [f"Attr{i}" for i in range(1, num_attr_columns + 1)]
            print(f"Headers set: {self.df.columns.tolist()}")

        def show_table(self):
            # Clear existing widgets
            for widget in self.canvas_frame.winfo_children():
                widget.destroy()
            self.cells = {}

            if self.df is None or self.df.empty:
                return

            # Create headers
            headers = ["Point Number", "Easting", "Northing", "Elevation", "Field Code"]
            # Add attribute headers from raw_df
            if self.raw_df is not None:
                for col in range(5, self.raw_df.shape[1]):
                    headers.append(f"Attribute {col - 4}")

            for col, header in enumerate(headers):
                label = ttk.Label(self.canvas_frame, text=header, borderwidth=1, relief="solid", anchor="center")
                label.grid(row=0, column=col, sticky="nsew")

            # Populate the table
            for row in range(self.df.shape[0]):
                for col in range(len(headers)):
                    if col < 5:  # Use self.df for the first 5 columns
                        value = self.df.iloc[row, col] if col < self.df.shape[1] else ""
                    else:  # Use self.raw_df for attribute columns
                        raw_col = col
                        value = self.raw_df.iloc[row, raw_col] if raw_col < self.raw_df.shape[1] else ""

                    entry = ttk.Entry(self.canvas_frame, justify="center")
                    entry.insert(0, value)
                    entry.grid(row=row + 1, column=col, sticky="nsew")
                    self.cells[(row, col)] = entry

            # Configure grid weights
            for col in range(len(headers)):
                self.canvas_frame.grid_columnconfigure(col, weight=1)
            for row in range(self.df.shape[0] + 1):
                self.canvas_frame.grid_rowconfigure(row, weight=1)

            # Update canvas scrolling
            self.canvas_frame.update_idletasks()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        def on_cell_edit(self, row, col):
            if hasattr(self, 'is_validating') and self.is_validating:
                print(f"Skipping on_cell_edit validation at Row {row}, Col {col} to prevent loop")
                return

            widget = self.cells[(row, col)]
            new_value = widget.get()
            old_value = self.df.iat[row, col] if self.df.iat[row, col] is not None else ""
            self.df.iat[row, col] = new_value
            action = {
                "type": "edit",
                "row": row,
                "col": col,
                "old_value": old_value,
                "new_value": new_value
            }
            self.undo_stack.append(action)
            self.redo_stack.clear()
            self.validate_all()

        def undo(self):
            if not self.undo_stack:
                return
            action = self.undo_stack.pop()
            self.redo_stack.append(action)
            if action["type"] == "edit":
                row, col = action["row"], action["col"]
                widget = self.cells[(row, col)]
                widget.delete(0, tk.END)
                widget.insert(0, action["old_value"])
                self.df.iat[row, col] = action["old_value"]
            elif action["type"] == "ignore":
                row, col = action["row"], action["col"]
                del self.ignored_errors[(row, col)]
            elif action["type"] == "ignore_all":
                error_type = action["error_type"]
                cells_to_restore = action["cells"]
                for r, c in cells_to_restore:
                    del self.ignored_errors[(r, c)]
                self.ignore_all_prompted.remove(error_type)
            self.validate_all()

        def redo(self):
            if not self.redo_stack:
                return
            action = self.redo_stack.pop()
            self.undo_stack.append(action)
            if action["type"] == "edit":
                row, col = action["row"], action["col"]
                widget = self.cells[(row, col)]
                widget.delete(0, tk.END)
                widget.insert(0, action["new_value"])
                self.df.iat[row, col] = action["new_value"]
            elif action["type"] == "ignore":
                row, col = action["row"], action["col"]
                error_type = action["error_type"]
                self.ignored_errors[(row, col)] = error_type
            elif action["type"] == "ignore_all":
                error_type = action["error_type"]
                cells_to_ignore = action["cells"]
                for r, c in cells_to_ignore:
                    self.ignored_errors[(r, c)] = error_type
                self.ignore_all_prompted.add(error_type)
            self.validate_all()

        def validate_bend_attributes(self, values):
            errors = []
            def is_valid_angle(value):
                if not value or str(value).strip() == "":
                    return False
                try:
                    val = float(value)
                    return val > 0  # Valid angles must be positive
                except (ValueError, TypeError):
                    return False

            def is_zero_or_blank(value):
                if not value or str(value).strip() == "":
                    return True
                try:
                    val = float(value)
                    return val == 0  # "0" is explicitly invalid for required angles
                except (ValueError, TypeError):
                    return False

            bend_field_codes = [
                "PIPE_BEND", "PT_OF_INTERSECTION", "PIPE", "BURIED_FITTINGS", "BEND",
                "FITTING", "FIELD BEND", "ELBOW", "FACTORY BEND", "BENDS", "ELBOWS",
                "FITTINGS", "PI", "BEND_TYPE", "OVERBEND", "SAG_BEND", "OVB", "SAG", "ELL"
            ]
            direction_attrs = [
                "PIPE_BEND_TYPE_D", "BEND_TYPE", "TURN_DIRECTION", "BEND_DIRECTION",
                "BEND TYPE", "DIRECTION", "TYPE"
            ]
            horz_angle_attrs = [
                "TURN_(DEG_OR_RADII)", "BEND ANGLE", "ANGLE", "HORIZONTAL_ANGL",
                "HORZ", "HORIZONTAL", "H", "H ANGLE", "H ANG", "HORZ ANG", "HORZ ANGLE"
            ]
            vert_angle_attrs = [
                "BEND ANGLE", "ANGLE", "VERTICAL_ANGL", "VERT ANGLE",
                "VERT", "VERTICAL", "V", "V ANGLE", "V ANG", "VERT ANG"
            ]
            no_type_angle_attrs = ["HORIZONTAL_ANGLE", "VERTICAL_ANGLE", "ANGLE", "COMMENT", "DEGREE"]

            horz_direction_values = [
                "SIDE BEND LEFT", "SIDE BEND RIGHT", "OTHER", "REV UNKNOWN", "UNKNOWN",
                "ROLLED", "SBL", "SBR", "LT", "LFT", "LEFT", "RGT", "RGHT", "RT",
                "RIGHT", "UND", "OTHER (SEE REMARK)"
            ]
            vert_direction_values = [
                "OVER BEND", "SAG BEND", "OTHER", "REV UNKNOWN", "UNKNOWN", "ROLLED",
                "OB", "OVB", "SAG", "OVER", "SG", "OVR", "OVERBEND", "UND", "OTHER (SEE REMARK)"
            ]
            combo_direction_values = [
                "LEFT & OVER", "RIGHT & OVER", "LEFT & SAG", "RIGHT & SAG", "OTHER",
                "REV UNKNOWN", "UNKNOWN", "SBLT+OB", "SBLT+SAG", "SBRT+OB", "SBRT+SAG",
                "ROLLED", "RT_SAG", "RT_OVERBEND", "LT_SAG", "LT_OVERBEND", "OVR_LFT",
                "OVR_RGHT", "SAG_LFT", "SAG_RGHT", "UND", "LEFT AND OVER", "RIGHT AND OVER",
                "SAG AND RIGHT", "SAG AND LEFT", "LEFT AND SAG", "RIGHT AND SAG",
                "OVER AND RIGHT", "OVER AND LEFT", "COMBO", "COMBINATION",
                "SAG RIGHT COMBO", "SAG LEFT COMBO", "OVERBEND RIGHT COMBO",
                "OVERBEND LEFT COMBO", "COMBINATION SIDE BEND LEFT SAG",
                "COMBINATION SIDE BEND LEFT AND OVER BEND",
                "COMBINATION SIDE BEND RIGHT AND OVER BEND",
                "COMBINATION SIDE BEND RIGHT AND SAG", "OTHER (SEE REMARK)"
            ]

            seen_cells = set()
            for row in range(self.df.shape[0]):
                field_code = values.get((row, 4), "").strip()
                base_field_code = self.get_base_field_code(field_code)
                if base_field_code.upper() not in [fc.upper() for fc in bend_field_codes]:
                    continue
                attrs = self.fxl_data.get(base_field_code, [])
                if not attrs:
                    continue

                direction_col = None
                horz_angle_col = None
                vert_angle_col = None
                no_type_angle_col = None
                for col in range(5, self.df.shape[1]):
                    attr_idx = col - 5
                    if attr_idx >= len(attrs):
                        continue
                    attr_name = attrs[attr_idx]["name"].upper()
                    if attr_name in [da.upper() for da in direction_attrs]:
                        direction_col = col
                    elif attr_name in [ha.upper() for ha in horz_angle_attrs]:
                        horz_angle_col = col
                    elif attr_name in [va.upper() for va in vert_angle_attrs]:
                        vert_angle_col = col
                    elif attr_name in [nta.upper() for nta in no_type_angle_attrs]:
                        no_type_angle_col = col

                if no_type_angle_col and base_field_code.upper() in [fc.upper() for fc in ["BEND", "ELBOW", "OVERBEND", "SAG_BEND", "OVB", "SAG", "ELL"]]:
                    angle_val = values.get((row, no_type_angle_col), "").strip()
                    if not is_valid_angle(angle_val):
                        error_key = (row, no_type_angle_col)
                        if error_key not in seen_cells:
                            errors.append((row, no_type_angle_col, f"Invalid Bend Angle: {attrs[no_type_angle_col - 5]['name']} must be > 0 for Field Code {field_code}"))
                            seen_cells.add(error_key)
                    continue

                if not direction_col:
                    continue

                direction_val = values.get((row, direction_col), "").strip().upper()
                if direction_val in [v.upper() for v in horz_direction_values]:
                    if horz_angle_col:
                        horz_val = values.get((row, horz_angle_col), "").strip()
                        if not is_valid_angle(horz_val):
                            error_key = (row, horz_angle_col)
                            if error_key not in seen_cells:
                                errors.append((row, horz_angle_col, f"Invalid Bend Angle: {attrs[horz_angle_col - 5]['name']} must be > 0 for {attrs[direction_col - 5]['name']}={direction_val}"))
                                seen_cells.add(error_key)
                    if vert_angle_col:
                        vert_val = values.get((row, vert_angle_col), "").strip()
                        if not is_zero_or_blank(vert_val):
                            error_key = (row, vert_angle_col)
                            if error_key not in seen_cells:
                                errors.append((row, vert_angle_col, f"Invalid Bend Angle: {attrs[vert_angle_col - 5]['name']} must be 0 or blank for {attrs[direction_col - 5]['name']}={direction_val}"))
                                seen_cells.add(error_key)

                if direction_val in [v.upper() for v in vert_direction_values]:
                    if vert_angle_col:
                        vert_val = values.get((row, vert_angle_col), "").strip()
                        if not is_valid_angle(vert_val):
                            error_key = (row, vert_angle_col)
                            if error_key not in seen_cells:
                                errors.append((row, vert_angle_col, f"Invalid Bend Angle: {attrs[vert_angle_col - 5]['name']} must be > 0 for {attrs[direction_col - 5]['name']}={direction_val}"))
                                seen_cells.add(error_key)
                    if horz_angle_col:
                        horz_val = values.get((row, horz_angle_col), "").strip()
                        if not is_zero_or_blank(horz_val):
                            error_key = (row, horz_angle_col)
                            if error_key not in seen_cells:
                                errors.append((row, horz_angle_col, f"Invalid Bend Angle: {attrs[horz_angle_col - 5]['name']} must be 0 or blank for {attrs[direction_col - 5]['name']}={direction_val}"))
                                seen_cells.add(error_key)

                if direction_val in [v.upper() for v in combo_direction_values]:
                    if horz_angle_col:
                        horz_val = values.get((row, horz_angle_col), "").strip()
                        if not is_valid_angle(horz_val):
                            error_key = (row, horz_angle_col)
                            if error_key not in seen_cells:
                                errors.append((row, horz_angle_col, f"Invalid Bend Angle: {attrs[horz_angle_col - 5]['name']} must be > 0 for {attrs[direction_col - 5]['name']}={direction_val}"))
                                seen_cells.add(error_key)
                    if vert_angle_col:
                        vert_val = values.get((row, vert_angle_col), "").strip()
                        if not is_valid_angle(vert_val):
                            error_key = (row, vert_angle_col)
                            if error_key not in seen_cells:
                                errors.append((row, vert_angle_col, f"Invalid Bend Angle: {attrs[vert_angle_col - 5]['name']} must be > 0 for {attrs[direction_col - 5]['name']}={direction_val}"))
                                seen_cells.add(error_key)

            print(f"Found {len(errors)} Invalid Bend Angle errors: {errors}")
            return errors

        def get_cell_tooltip(self, widget, row, col):
            val = widget.get()
            tooltip_text = f"Value: {val}"
            is_combobox = isinstance(widget, ttk.Combobox)

            # Initialize default values
            code = ""
            attr_name = ""
            attrs = []

            # Field Code column
            if col == 4:
                tooltip_text += f"\nField Code: {val}"
            # Attribute columns
            if col > 4:
                code = self.get_base_field_code(self.df.iat[row, 4]) if self.df.shape[1] > 4 else ""
                attrs = self.fxl_data.get(code, [])
                if (col - 5) < len(attrs):
                    attr_name = attrs[col - 5]["name"]
                    tooltip_text += f"\nAttribute: {attr_name}"
                else:
                    tooltip_text += "\nExtra Attribute"

            # Get validation context
            values = {(r, c): w.get() for (r, c), w in self.cells.items()}
            points = [v for (r, c), v in values.items() if c == 0]
            dup_points = {v for v in points if points.count(v) > 1}
            xrays = ["XRAY", "X-RAY", "XRAY NO.", "XRAY NO", "X-RAY NO", "X-RAY NO."]
            xray_cells = [(r, 5 + i) for r in range(self.df.shape[0]) for i, a in enumerate(self.fxl_data.get(self.get_base_field_code(values.get((r, 4), "")), []))
                        if a["name"].upper().replace("-", "").replace(" ", "").replace(".", "") in [n.upper().replace("-", "").replace(" ", "").replace(".", "") for n in xrays]]
            xray_vals = [values[p] for p in xray_cells if p in values]
            non_duplicate_xray_vals = ["0", "NA", "N/A", "UNK", "UNKNOWN", "-", " ", "_", "NaN"]
            dup_xrays = {v for v in xray_vals if xray_vals.count(v) > 1 and v not in non_duplicate_xray_vals}
            bend_errors = self.validate_bend_attributes(values)

            # Check for errors
            bend_error = None
            for bend_r, bend_c, error_type in bend_errors:
                if (row, col) == (bend_r, bend_c):
                    bend_error = error_type
                    break

            if (row, col) in self.ignored_errors:
                tooltip_text += f"\nError Ignored: {self.ignored_errors[(row, col)]}"
            else:
                if bend_error:
                    tooltip_text += f"\nError: {bend_error}"
                elif col == 0 and val in dup_points:
                    tooltip_text += "\nError: Duplicate Point Number"
                elif (row, col) in xray_cells and val in dup_xrays:
                    tooltip_text += "\nError: Duplicate XRAY Value"
                elif self.is_misspelled(val):
                    tooltip_text += "\nError: Misspelled"
                elif col == 4 and val:
                    base_code = self.get_base_field_code(val)
                    if base_code not in self.fxl_data.keys():
                        tooltip_text += f"\nError: Field Code Not in FXL (Incorrect Value: {val})"
                elif col > 4:
                    if (col - 5) < len(attrs):
                        if attrs[col - 5]["type"] == "list" and val and val not in attrs[col - 5]["items"]:
                            tooltip_text += f"\nError: Value Not in FXL List (Incorrect Value: {val})"
                        else:
                            val_upper = val.strip().upper()
                            invalid_values = ["0", "N/A", "NA", "UNK", " ", "-", "_", "NaN"]
                            if val_upper == "0" and attr_name not in {'Turn_(Deg_or_Radii)', 'BEND ANGLE', 'ANGLE', 'HORIZONTAL_ANGL', 'HORZ', 'HORIZONTAL',
                                                                    'H', 'H ANGLE', 'H ANG', 'HORZ ANG', 'HORZ ANGLE', 'VERTICAL_ANGL', 'VERT ANGLE',
                                                                    'VERT', 'VERTICAL', 'V', 'V ANGLE', 'V ANG', 'VERT ANG'}:
                                tooltip_text += f"\nError: Invalid Value ({val})"
                            elif val_upper in ["N/A", "NA", "UNK", " ", "-", "_", "NaN"]:
                                tooltip_text += f"\nError: Invalid Value ({val})"
                            elif val == "" and re.sub(r'[^a-zA-Z]', '', attr_name.upper()) not in ["REMARK", "REMARKS", "COMMENT", "COMMENTS", "NOTE", "NOTES"]:
                                tooltip_text += "\nError: Invalid Value (Blank)"
                    elif (col - 5) >= len(attrs):
                        tooltip_text += "\nExtra Attribute"
                if not is_combobox and val and val != val.upper():
                    tooltip_text += f"\nError: Not All Caps (Value: {val})"

            # Ensure tooltip is not empty
            if tooltip_text == f"Value: {val}":
                tooltip_text += "\nNo errors"
            return tooltip_text

        def is_misspelled(self, text):
            if not enchant or not self.dict_en:
                return False
            if not isinstance(text, str) or text in self.custom_dict:
                return False
            if '-' in text:
                return False
            words = re.split(r'\s+', text.strip())
            words = [word for word in words if re.match(r'^[a-zA-Z]{2,}$', word)]
            return any(not self.dict_en.check(word) and word not in self.custom_dict for word in words)

        def update_error_count_label(self):
            if self.error_count_label:
                error_text = "Error Counts:\n"
                for error_type, count in self.error_counts.items():
                    if count > 0:
                        error_text += f"{error_type}: {count}\n"
                if error_text == "Error Counts:\n":
                    error_text = "Error Counts: None"
                self.error_count_label.config(text=error_text)

        def show_error_menu(self, event, widget, row, col):
            # Get the current style of the widget
            current_style = widget.cget("style")
            style_name = current_style.split('.')[0]  # e.g., "Invalid" from "Invalid.TEntry"
            error_type = None

            # If the cell is already ignored, don't show the menu
            if (row, col) in self.ignored_errors:
                return

            # Map styles to error types
            if style_name == "Invalid":
                # Determine the specific error type by re-checking the value
                val = widget.get().strip().upper()
                angle_attrs = {'Turn_(Deg_or_Radii)', 'BEND ANGLE', 'ANGLE', 'HORIZONTAL_ANGL', 'HORZ', 'HORIZONTAL',
                            'H', 'H ANGLE', 'H ANG', 'HORZ ANG', 'HORZ ANGLE', 'VERTICAL_ANGL', 'VERT ANGLE',
                            'VERT', 'VERTICAL', 'V', 'V ANGLE', 'V ANG', 'VERT ANG'}
                invalid_values = ["0", "N/A", "NA", "UNK", " ", "-", "_", "NaN"]
                code = self.get_base_field_code(self.df.iat[row, 4]) if self.df.shape[1] > 4 else ""
                attrs = self.fxl_data.get(code, [])
                if (col - 5) < len(attrs):
                    attr_name = attrs[col - 5]["name"]
                    normalized_attr_name = re.sub(r'[^a-zA-Z]', '', attr_name.upper())
                    if val == "0" and attr_name not in angle_attrs:
                        error_type = "Invalid Value (0, {})".format(attr_name)
                    elif val in invalid_values:
                        error_type = "Invalid Value ({}, {})".format(val, attr_name)
                    elif val == "" and normalized_attr_name not in ["REMARK", "REMARKS", "COMMENT", "COMMENTS", "NOTE", "NOTES"]:
                        error_type = "Invalid Value (Blank, {})".format(attr_name)
                elif not isinstance(widget, ttk.Combobox) and val and val != val.upper():
                    error_type = "Not All Caps"
            elif style_name == "Duplicate":
                if col == 0:
                    error_type = "Duplicate Point Number"
                elif (row, col) in self.xray_cells and val in self.dup_xrays:
                    error_type = "Duplicate XRAY Value"
            elif style_name == "Misspelled":
                error_type = "Misspelled"
            elif style_name == "NotInFXL":
                if col == 4:
                    error_type = "Field Code Not in FXL"
                else:
                    code = self.get_base_field_code(self.df.iat[row, 4]) if self.df.shape[1] > 4 else ""
                    attrs = self.fxl_data.get(code, [])
                    if (col - 5) < len(attrs):
                        attr_name = attrs[col - 5]["name"]
                        error_type = f"Value Not in FXL List ({attr_name})"

            # If an error is detected, show the context menu
            if error_type:
                menu = Menu(self, tearoff=0)
                menu.add_command(label="Ignore Error", command=lambda: self.ignore_error(row, col, error_type))
                menu.tk_popup(event.x_root, event.y_root)
            else:
                print(f"No error detected for right-click at Row {row}, Col {col}, Style: {style_name}")

        def ignore_error(self, row, col, error_type):
            self.ignored_errors[(row, col)] = error_type
            action = {"type": "ignore", "row": row, "col": col, "error_type": error_type}
            self.undo_stack.append(action)
            self.redo_stack.clear()
            error_count = 0
            cells_with_error = []
            code = self.get_base_field_code(self.df.iat[row, 4]) if self.df.shape[1] > 4 else ""
            attrs = self.fxl_data.get(code, [])
            attr_name = attrs[col - 5]["name"] if col > 4 and (col - 5) < len(attrs) else ""

            # Use self.error_state to find matching errors without re-validating
            for (r, c), (err_type, field_code, cell_attr_name) in self.error_state.items():
                if (r, c) == (row, col) or (r, c) in self.ignored_errors:
                    continue
                if err_type == error_type and field_code == code and (c <= 4 or cell_attr_name == attr_name):
                    error_count += 1
                    cells_with_error.append((r, c))

            # Temporarily unbind events to prevent recursive calls
            for (r, c), w in self.cells.items():
                w.unbind("<FocusOut>")
                if isinstance(w, ttk.Combobox):
                    w.unbind("<<ComboboxSelected>>")

            try:
                if error_count > 0 and error_type not in self.ignore_all_prompted:
                    response = messagebox.askyesno("Ignore All", f"There are {error_count} more errors of type '{error_type}' for Field Code '{code}' and attribute '{attr_name}'. Ignore all errors of this type?")
                    if response:
                        action = {"type": "ignore_all", "error_type": error_type, "cells": cells_with_error}
                        self.undo_stack.append(action)
                        self.redo_stack.clear()
                        for r, c in cells_with_error:
                            self.ignored_errors[(r, c)] = error_type
                        self.ignore_all_prompted.add(error_type)
                    else:
                        self.ignore_all_prompted.add(error_type)
                self.validate_all()
            finally:
                # Re-bind events after validation
                for (r, c), w in self.cells.items():
                    w.bind("<FocusOut>", lambda e, r=r, c=c: self.on_cell_edit(r, c))
                    if isinstance(w, ttk.Combobox):
                        w.bind("<<ComboboxSelected>>", lambda e, r=r, c=c: self.on_cell_edit(r, c))

        def validate_all(self):
            if hasattr(self, 'is_validating') and self.is_validating:
                print("Skipping re-entrant validation to prevent loop")
                return self.cells

            self.is_validating = True
            start_time = time.time()
            print(f"Start validate_all: {start_time}")

            self.error_counts = {
                "Point Number Duplicates": 0,
                "XRAY Duplicates": 0,
                "Misspelled": 0,
                "Invalid: N/A": 0,
                "Invalid: NA": 0,
                "Invalid: UNK": 0,
                "Invalid: Space": 0,
                "Invalid: Zero": 0,
                "Invalid: Dash": 0,
                "Invalid: Underscore": 0,
                "Invalid: NaN": 0,
                "Invalid: Blank": 0,
                "Not in FXL: Field Code": 0,
                "Not in FXL: Attribute": 0,
                "Invalid Bend Angle": 0,
                "Not All Caps": 0
            }
            self.error_state = {}
            values = {(r, c): w.get() for (r, c), w in self.cells.items()}
            for w in self.cells.values():
                w.configure(style="Normal.TEntry" if isinstance(w, ttk.Entry) else "Normal.TCombobox")
            print(f"Reset styles: {time.time() - start_time} seconds")

            # Point Number duplicates
            point_counts = {}
            first_points = {}
            for r in range(self.df.shape[0]):
                val = values.get((r, 0), "")
                if val:
                    point_counts[val] = point_counts.get(val, 0) + 1
                    if val not in first_points:
                        first_points[val] = r
            dup_points = {v for v, count in point_counts.items() if count > 1}

            # XRAY duplicates
            xrays = ["XRAY", "X-RAY", "XRAY NO.", "XRAY NO", "X-RAY NO", "X-RAY NO."]
            xray_cells = [(r, 5 + i) for r in range(self.df.shape[0]) for i, a in enumerate(self.fxl_data.get(self.get_base_field_code(values.get((r, 4), "")), []))
                        if a["name"].upper().replace("-", "").replace(" ", "").replace(".", "") in [n.upper().replace("-", "").replace(" ", "").replace(".", "") for n in xrays]]
            xray_counts = {}
            first_xrays = {}
            for (r, c) in xray_cells:
                val = values.get((r, c), "")
                if val:
                    xray_counts[val] = xray_counts.get(val, 0) + 1
                    if val not in first_xrays:
                        first_xrays[val] = (r, c)
            non_duplicate_xray_vals = ["0", "NA", "N/A", "UNK", "UNKNOWN", "-", " ", "_", "NaN"]
            dup_xrays = {v for v, count in xray_counts.items() if count > 1 and v not in non_duplicate_xray_vals}
            print(f"Duplicate checks done: {time.time() - start_time} seconds")

            bend_errors = self.validate_bend_attributes(values)
            print(f"Bend validation done: {time.time() - start_time} seconds")

            angle_attrs = {'Turn_(Deg_or_Radii)', 'BEND ANGLE', 'ANGLE', 'HORIZONTAL_ANGL', 'HORZ', 'HORIZONTAL',
                        'H', 'H ANGLE', 'H ANG', 'HORZ ANG', 'HORZ ANGLE', 'VERTICAL_ANGL', 'VERT ANGLE',
                        'VERT', 'VERTICAL', 'V', 'V ANGLE', 'V ANG', 'VERT ANG'}

            try:
                for (r, c), w in self.cells.items():
                    val = values[(r, c)]
                    style = "Normal"
                    error_type = None
                    is_combobox = isinstance(w, ttk.Combobox)
                    attrs = []  # Initialize attrs to avoid UnboundLocalError
                    if (r, c) in self.ignored_errors:
                        continue
                    # Check bend errors first
                    for bend_r, bend_c, err_type in bend_errors:
                        if (r, c) == (bend_r, bend_c):
                            style = "Invalid"
                            error_type = err_type
                            self.error_counts["Invalid Bend Angle"] += 1
                            break
                    if style == "Normal":
                        if c == 0 and val in dup_points:
                            style = "Duplicate"
                            error_type = "Duplicate Point Number"
                            if r != first_points.get(val):
                                self.error_counts["Point Number Duplicates"] += 1
                        elif (r, c) in xray_cells and val in dup_xrays:
                            style = "Duplicate"
                            error_type = "Duplicate XRAY Value"
                            if (r, c) != first_xrays.get(val):
                                self.error_counts["XRAY Duplicates"] += 1
                        elif self.is_misspelled(val):
                            style = "Misspelled"
                            error_type = "Misspelled"
                            self.error_counts["Misspelled"] += 1
                        elif c == 4 and val:
                            base_code = self.get_base_field_code(val)
                            if base_code not in self.fxl_data.keys():
                                style = "NotInFXL"
                                error_type = "Field Code Not in FXL"
                                self.error_counts["Not in FXL: Field Code"] += 1
                        elif c > 4:
                            code = self.get_base_field_code(self.df.iat[r, 4]) if self.df.shape[1] > 4 else ""
                            attrs = self.fxl_data.get(code, [])  # Set attrs here
                            if (c - 5) < len(attrs):
                                attr_name = attrs[c - 5]["name"]
                                normalized_attr_name = re.sub(r'[^a-zA-Z]', '', attr_name.upper())
                                print(f"Row {r}, Col {c}, Attribute Name: '{attr_name}', Normalized: '{normalized_attr_name}', Value: '{val}'")
                                if attrs[c - 5]["type"] == "list" and val and val not in attrs[c - 5]["items"]:
                                    style = "NotInFXL"
                                    error_type = f"Value Not in FXL List ({attr_name})"
                                    self.error_counts["Not in FXL: Attribute"] += 1
                                elif val == "" and normalized_attr_name in ["REMARK", "REMARKS", "COMMENT", "COMMENTS", "NOTE", "NOTES"]:
                                    style = "Normal"
                                else:
                                    val_upper = val.strip().upper()
                                    invalid_values = ["0", "N/A", "NA", "UNK", " ", "-", "_", "NaN"]
                                    if val_upper == "0" and attr_name not in angle_attrs:
                                        style = "Invalid"
                                        error_type = f"Invalid Value (0, {attr_name})"
                                        self.error_counts["Invalid: Zero"] += 1
                                    elif val_upper in ["N/A", "NA", "UNK", " ", "-", "_", "NaN"]:
                                        style = "Invalid"
                                        error_type = f"Invalid Value ({val_upper}, {attr_name})"
                                        self.error_counts[invalid_values[invalid_values.index(val_upper)]] += 1
                                    elif val == "" and normalized_attr_name not in ["REMARK", "REMARKS", "COMMENT", "COMMENTS", "NOTE", "NOTES"]:
                                        style = "Invalid"
                                        error_type = f"Invalid Value (Blank, {attr_name})"
                                        self.error_counts["Invalid: Blank"] += 1
                            else:
                                style = "Extra"
                                print(f"Applying Extra style at Row {r}, Col {c}")
                        if not is_combobox and val and val != val.upper():
                            style = "Invalid"
                            error_type = "Not All Caps"
                            self.error_counts["Not All Caps"] += 1

                    if style != "Normal" and error_type:
                        field_code = self.get_base_field_code(self.df.iat[r, 4]) if self.df.shape[1] > 4 and c >= 4 else ""
                        attr_name = attrs[c - 5]["name"] if c > 4 and (c - 5) < len(attrs) else ""
                        self.error_state[(r, c)] = (error_type, field_code, attr_name)

                    w.configure(style=f"{style}.{'TEntry' if isinstance(w, ttk.Entry) else 'TCombobox'}")
                    if style == "Misspelled":
                        w.bind("<Button-3>", lambda e, v=val: self.show_spell_menu(e, v))
            except Exception as e:
                print(f"Error during validation: {e}")
                messagebox.showerror("Error", f"Validation error: {e}\n\nAn error occurred while validating the CSV data. Please check the console output for details.")
            finally:
                self.is_validating = False

            print(f"Validation loop done: {time.time() - start_time} seconds")
            self.update_error_count_label()
            print(f"Error count label updated: {time.time() - start_time} seconds")
            return values

        def show_spell_menu(self, event, text):
            menu = Menu(self, tearoff=0)
            menu.add_command(label=f"Add '{text}' to dictionary", command=lambda: self.add_word(text))
            menu.tk_popup(event.x_root, event.y_root)

        def add_word(self, word):
            self.custom_dict.add(word)
            self.save_custom_dict()
            self.validate_all()

        def export_custom_csv(self):
            if not self.csv_path:
                messagebox.showerror("Error", "No CSV file loaded. Please load a CSV file first.")
                return
            export_window = Toplevel(self)
            export_window.title("Export Custom CSV")
            export_window.geometry("600x500")
            main_frame = ttk.Frame(export_window)
            main_frame.pack(fill="both", expand=True, padx=5, pady=5)
            field_code_frame = ttk.LabelFrame(main_frame, text="Select Field Codes (Rows)")
            field_code_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
            field_code_listbox = Listbox(field_code_frame, selectmode="multiple", height=10, exportselection=False)
            field_code_scrollbar = Scrollbar(field_code_frame, orient="vertical")
            field_code_listbox.config(yscrollcommand=field_code_scrollbar.set)
            field_code_scrollbar.config(command=field_code_listbox.yview)
            field_code_scrollbar.pack(side="right", fill="y")
            field_code_listbox.pack(fill="both", expand=True)
            field_codes = sorted(set(self.df["Field Code"].dropna()))
            for code in field_codes:
                field_code_listbox.insert(tk.END, code)
            field_code_btn_frame = ttk.Frame(field_code_frame)
            field_code_btn_frame.pack(fill="x", pady=5)
            ttk.Button(field_code_btn_frame, text="Select All", command=lambda: field_code_listbox.select_set(0, tk.END)).pack(side="left", padx=5)
            ttk.Button(field_code_btn_frame, text="Deselect All", command=lambda: field_code_listbox.selection_clear(0, tk.END)).pack(side="left", padx=5)
            column_frame = ttk.LabelFrame(main_frame, text="Select Columns")
            column_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
            column_listbox = Listbox(column_frame, selectmode="multiple", height=10, exportselection=False)
            column_scrollbar = Scrollbar(column_frame, orient="vertical")
            column_listbox.config(yscrollcommand=column_scrollbar.set)
            column_scrollbar.config(command=column_listbox.yview)
            column_scrollbar.pack(side="right", fill="y")
            column_listbox.pack(fill="both", expand=True)
            columns = list(self.df.columns)
            for col in columns:
                column_listbox.insert(tk.END, col)
            column_btn_frame = ttk.Frame(column_frame)
            column_btn_frame.pack(fill="x", pady=5)
            ttk.Button(column_btn_frame, text="Select All", command=lambda: column_listbox.select_set(0, tk.END)).pack(side="left", padx=5)
            ttk.Button(column_btn_frame, text="Deselect All", command=lambda: column_listbox.selection_clear(0, tk.END)).pack(side="left", padx=5)
            include_header_var = BooleanVar(value=True)
            ttk.Checkbutton(main_frame, text="Include Header", variable=include_header_var).grid(row=1, column=0, columnspan=2, pady=5)
            main_frame.grid_columnconfigure(0, weight=1)
            main_frame.grid_columnconfigure(1, weight=1)
            main_frame.grid_rowconfigure(0, weight=1)

            def do_export():
                selected_field_codes = [field_code_listbox.get(i) for i in field_code_listbox.curselection()]
                if not selected_field_codes:
                    messagebox.showerror("Error", "Please select at least one Field Code.")
                    return
                selected_columns = [column_listbox.get(i) for i in column_listbox.curselection()]
                if not selected_columns:
                    messagebox.showerror("Error", "Please select at least one column.")
                    return
                data = [[self.cells[(r, c)].get() for c in range(self.df.shape[1]) if self.df.columns[c] in selected_columns]
                        for r in range(self.df.shape[0]) if self.cells[(r, 4)].get() in selected_field_codes]
                custom_df = pd.DataFrame(data, columns=selected_columns)
                csv_dir = os.path.dirname(self.csv_path)
                csv_name = os.path.splitext(os.path.basename(self.csv_path))[0]
                custom_csv_file = os.path.join(csv_dir, f"{csv_name}_custom.csv")
                custom_df.to_csv(custom_csv_file, index=False, header=include_header_var.get())
                messagebox.showinfo("Success", f"Custom CSV exported to {custom_csv_file}")
                export_window.destroy()

            ttk.Button(main_frame, text="Export", command=do_export).grid(row=2, column=0, columnspan=2, pady=10)

        def generate_error_report(self):
            if not self.csv_path:
                messagebox.showerror("Error", "No CSV file loaded. Please load a CSV file first.")
                return
            if not openpyxl:
                messagebox.showerror("Error", "Excel report generation requires the 'openpyxl' module, which is not available.")
                return
            csv_dir = os.path.dirname(self.csv_path)
            csv_name = os.path.splitext(os.path.basename(self.csv_path))[0]
            excel_file = os.path.join(csv_dir, f"{csv_name}_error_report.xlsx")

            # Ensure error_state is up-to-date
            if not self.error_state:
                values = self.validate_all()
            else:
                values = {(r, c): w.get() for (r, c), w in self.cells.items()}

            # Temporarily unbind events to prevent recursive calls
            for (r, c), w in self.cells.items():
                w.unbind("<FocusOut>")
                if isinstance(w, ttk.Combobox):
                    w.unbind("<<ComboboxSelected>>")

            try:
                data = []
                error_flags = []
                tooltips = []
                for row in range(self.df.shape[0]):
                    row_data = []
                    row_errors = []
                    row_tooltips = []
                    for col in range(self.df.shape[1]):
                        val = values[(row, col)]
                        row_data.append(val)
                        error_flag = ""
                        if (row, col) in self.ignored_errors:
                            error_flag = "Ignored"
                        elif (row, col) in self.error_state:
                            error_type, _, _ = self.error_state[(row, col)]
                            if "Invalid" in error_type:
                                error_flag = "Invalid"
                            elif "Duplicate" in error_type:
                                error_flag = "Duplicate"
                            elif "Misspelled" in error_type:
                                error_flag = "Misspelled"
                            elif "Not in FXL" in error_type:
                                error_flag = "NotInFXL"
                        else:
                            # Check for extra attributes
                            if col > 4:
                                code = self.get_base_field_code(self.df.iat[row, 4]) if self.df.shape[1] > 4 else ""
                                attrs = self.fxl_data.get(code, [])
                                if (col - 5) >= len(attrs):
                                    error_flag = "Extra"
                        row_errors.append(error_flag)
                        tooltip = None
                        if error_flag and error_flag != "Ignored" and error_flag != "Extra":
                            widget = self.cells[(row, col)]
                            tooltip = self.get_cell_tooltip(widget, row, col)
                        row_tooltips.append(tooltip)
                    data.append(row_data)
                    error_flags.append(row_errors)
                    tooltips.append(row_tooltips)
                df_data = pd.DataFrame(data, columns=self.df.columns)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Error Report"
                summary_rows = 0
                ws.append(["Error Report"])
                summary_rows += 1
                ws.append(["------------"])
                summary_rows += 1
                for error_type, count in self.error_counts.items():
                    if count > 0:
                        ws.append([f"{error_type}: {count}"])
                        summary_rows += 1
                ws.append([])
                summary_rows += 1
                headers = list(self.df.columns)
                ws.append(headers)
                header_row = summary_rows + 1
                numeric_columns = ["Point Number", "Northing", "Easting", "Elevation"]
                date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
                for row in range(df_data.shape[0]):
                    row_data = df_data.iloc[row].tolist()
                    ws.append(row_data)
                for col_idx, col_name in enumerate(headers, start=1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    if col_name in numeric_columns:
                        for row_idx in range(header_row + 1, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            try:
                                cell.value = float(cell.value) if cell.value else None
                                cell.number_format = '0.0000000000'
                            except (ValueError, TypeError):
                                pass
                    else:
                        for row_idx in range(header_row + 1, ws.max_row + 1):
                            cell = ws[f"{col_letter}{row_idx}"]
                            if cell.value:
                                if date_pattern.match(str(cell.value)):
                                    try:
                                        cell.value = datetime.datetime.strptime(str(cell.value), '%Y-%m-%d').date()
                                        cell.number_format = 'YYYY-MM-DD'
                                    except ValueError:
                                        cell.number_format = '@'
                                else:
                                    try:
                                        cell.value = float(cell.value)
                                        cell.number_format = '0.00'
                                    except (ValueError, TypeError):
                                        cell.number_format = '@'
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                color_start_row = header_row + 1
                for row in range(len(data)):
                    adjusted_row = color_start_row + row
                    for col in range(len(headers)):
                        cell = ws.cell(row=adjusted_row, column=col + 1)
                        error_value = error_flags[row][col]
                        if error_value == "Duplicate":
                            cell.fill = yellow_fill
                        elif error_value == "Misspelled":
                            cell.fill = orange_fill
                        elif error_value == "Invalid":
                            cell.fill = light_yellow_fill
                        elif error_value == "Extra":
                            cell.fill = gray_fill
                        elif error_value == "NotInFXL":
                            cell.fill = red_fill
                        tooltip = tooltips[row][col]
                        if tooltip:
                            cell.comment = Comment(tooltip, "Data Validation Tool")
                wb.save(excel_file)
                messagebox.showinfo("Success", f"Excel error report saved to {excel_file}")
            except Exception as e:
                print(f"Error generating Excel report: {e}")
                messagebox.showerror("Error", f"Failed to generate Excel report: {e}")
            finally:
                # Re-bind events after report generation
                for (r, c), w in self.cells.items():
                    w.bind("<FocusOut>", lambda e, r=r, c=c: self.on_cell_edit(r, c))
                    if isinstance(w, ttk.Combobox):
                        w.bind("<<ComboboxSelected>>", lambda e, r=r, c=c: self.on_cell_edit(r, c))

        def export_csv(self):
            if not self.csv_path:
                messagebox.showerror("Error", "No CSV file loaded. Please load a CSV file first.")
                return
            csv_dir = os.path.dirname(self.csv_path)
            csv_name = os.path.splitext(os.path.basename(self.csv_path))[0]
            corrected_csv_file = os.path.join(csv_dir, f"{csv_name}_corrected.csv")
            data = [[self.cells[(r, c)].get() for c in range(self.df.shape[1])] for r in range(self.df.shape[0])]
            pd.DataFrame(data, columns=self.df.columns).to_csv(corrected_csv_file, index=False)
            messagebox.showinfo("Success", f"CSV exported to {corrected_csv_file}")

    if __name__ == "__main__":
        try:
            app = DataValidationTool()
            app.mainloop()
        except Exception as e:
            print(f"Failed to launch application: {e}")
            traceback.print_exc()
else:
    print("GUI unavailable. Run in a Tkinter-enabled environment.")